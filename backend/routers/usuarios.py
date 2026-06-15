from fastapi import APIRouter, Depends, HTTPException, Request, status, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel, ConfigDict, EmailStr, Field
from typing import Optional, List
from database import get_db
from models.usuario import Usuario, RolUsuario
from models.laboratorio import Laboratorio
from models.departamento import Departamento
from dependencies import hashear_password, verificar_password, get_current_user, require_roles
from rls import usuario_lab_filter, resolve_lab_id
import secrets
from services.auditoria import registrar, Accion, Recurso
import string
import io

router = APIRouter(prefix="/usuarios", tags=["Usuarios"])


# ─── Schemas ───────────────────────────────────────────────────────────────────

class UsuarioCreate(BaseModel):
    nombre: str = Field(..., min_length=2, max_length=100)
    email: EmailStr
    numero_empleado: Optional[str] = None
    password: str = Field(..., min_length=6)
    rol: RolUsuario = RolUsuario.DOCENTE
    laboratorio_id: Optional[int] = None
    departamento_id: Optional[int] = None

class UsuarioUpdate(BaseModel):
    nombre: Optional[str] = Field(None, min_length=2, max_length=100)
    email: Optional[EmailStr] = None
    numero_empleado: Optional[str] = None
    rol: Optional[RolUsuario] = None
    laboratorio_id: Optional[int] = None
    departamento_id: Optional[int] = None
    activo: Optional[bool] = None
    acceso_consultorio: Optional[bool] = None

class UsuarioResponse(BaseModel):
    id: int
    nombre: str
    email: str
    numero_empleado: Optional[str]
    rol: str
    laboratorio_id: Optional[int]
    laboratorio_nombre: Optional[str] = None
    departamento_id: Optional[int] = None
    departamento_nombre: Optional[str] = None
    departamento_clave: Optional[str] = None
    activo: bool
    acceso_consultorio: bool = False

    model_config = ConfigDict(from_attributes=True)

class ResetPasswordResponse(BaseModel):
    mensaje: str
    password_temporal: str

class CambiarPasswordPropio(BaseModel):
    password_actual: str = Field(..., min_length=1, description="Contraseña actual")
    password_nuevo:  str = Field(..., min_length=6, description="Nueva contraseña (mín. 6 caracteres)")


# ─── Helper ────────────────────────────────────────────────────────────────────

def _serializar(u: Usuario, db: Session) -> dict:
    lab_nombre = None
    dep = None
    if u.laboratorio_id:
        lab = db.query(Laboratorio).filter(Laboratorio.id == u.laboratorio_id).first()
        lab_nombre = lab.nombre if lab else None
    if u.departamento_id:
        dep = db.query(Departamento).filter(Departamento.id == u.departamento_id).first()
    return {
        "id": u.id,
        "nombre": u.nombre,
        "email": u.email,
        "numero_empleado": u.numero_empleado,
        "rol": u.rol.value,
        "laboratorio_id": u.laboratorio_id,
        "laboratorio_nombre": lab_nombre,
        "departamento_id": u.departamento_id,
        "departamento_nombre": dep.nombre if dep else None,
        "departamento_clave": dep.clave if dep else None,
        "activo": u.activo,
        "acceso_consultorio": bool(u.acceso_consultorio),
    }

def _generar_password(longitud: int = 10) -> str:
    alfabeto = string.ascii_letters + string.digits + "!@#$"
    return ''.join(secrets.choice(alfabeto) for _ in range(longitud))

def _validar_laboratorio(laboratorio_id: Optional[int], rol: RolUsuario, db: Session):
    """LAB_ADMIN debe tener laboratorio. Verifica que el lab exista."""
    if rol == RolUsuario.LAB_ADMIN and not laboratorio_id:
        raise HTTPException(status_code=422, detail="Un LAB_ADMIN debe tener laboratorio asignado")
    if laboratorio_id:
        lab = db.query(Laboratorio).filter(Laboratorio.id == laboratorio_id, Laboratorio.activo == True).first()
        if not lab:
            raise HTTPException(status_code=404, detail="Laboratorio no encontrado o inactivo")

def _validar_departamento(departamento_id: Optional[int], rol: RolUsuario, db: Session):
    if rol == RolUsuario.ADMINISTRATIVO and not departamento_id:
        raise HTTPException(status_code=422, detail="Un usuario ADMINISTRATIVO debe tener departamento asignado")
    if departamento_id:
        dep = db.query(Departamento).filter(Departamento.id == departamento_id, Departamento.activo == True).first()
        if not dep:
            raise HTTPException(status_code=404, detail="Departamento no encontrado o inactivo")


# ─── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("", summary="Listar usuarios")
def listar_usuarios(
    rol: Optional[str] = None,
    activo: Optional[bool] = None,
    laboratorio_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(
        RolUsuario.SUPER_ADMIN,
        RolUsuario.LAB_ADMIN,
        RolUsuario.TUTORIA_ADMIN,
    ))
):
    query = db.query(Usuario)

    if current_user.rol == RolUsuario.TUTORIA_ADMIN and rol != RolUsuario.DOCENTE.value:
        raise HTTPException(
            status_code=403,
            detail="El Responsable de Tutoría solo puede consultar docentes para asignación de grupos"
        )

    # RLS: LAB_ADMIN solo ve usuarios de su lab (+ todos los docentes)
    query = usuario_lab_filter(query, Usuario, current_user)

    if rol:
        try:
            query = query.filter(Usuario.rol == RolUsuario(rol))
        except ValueError:
            raise HTTPException(status_code=422, detail=f"Rol inválido: {rol}")
    if activo is not None:
        query = query.filter(Usuario.activo == activo)

    # El filtro manual por laboratorio_id solo aplica si SUPER_ADMIN lo pide explícitamente
    # Para LAB_ADMIN ya está forzado por usuario_lab_filter; ignoramos el parámetro
    if laboratorio_id and current_user.rol == RolUsuario.SUPER_ADMIN:
        query = query.filter(Usuario.laboratorio_id == laboratorio_id)
    if departamento_id:
        query = query.filter(Usuario.departamento_id == departamento_id)

    usuarios = query.order_by(Usuario.nombre).all()
    return [_serializar(u, db) for u in usuarios]


@router.post("", status_code=status.HTTP_201_CREATED, summary="Crear usuario")
def crear_usuario(
    request: Request,
    data: UsuarioCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN))
):
    if db.query(Usuario).filter(Usuario.email == data.email).first():
        raise HTTPException(status_code=409, detail="Ya existe un usuario con ese email")
    if data.numero_empleado and db.query(Usuario).filter(Usuario.numero_empleado == data.numero_empleado).first():
        raise HTTPException(status_code=409, detail="Ya existe un usuario con ese número de empleado")

    _validar_laboratorio(data.laboratorio_id, data.rol, db)
    _validar_departamento(data.departamento_id, data.rol, db)

    usuario = Usuario(
        nombre=data.nombre,
        email=data.email,
        numero_empleado=data.numero_empleado,
        password_hash=hashear_password(data.password),
        rol=data.rol,
        laboratorio_id=data.laboratorio_id,
        departamento_id=data.departamento_id,
        activo=True,
    )
    db.add(usuario)
    db.commit()
    db.refresh(usuario)
    registrar(db, accion=Accion.CREAR_USUARIO, recurso=Recurso.USUARIO,
              usuario=current_user, recurso_id=usuario.id,
              detalle={"nombre": usuario.nombre, "email": usuario.email, "rol": usuario.rol.value},
              request=request)
    return _serializar(usuario, db)


@router.get("/me", summary="Perfil del usuario actual")
def mi_perfil(
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    return _serializar(current_user, db)


@router.get("/{usuario_id}", summary="Detalle de usuario")
def obtener_usuario(
    usuario_id: int,
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN))
):
    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return _serializar(u, db)


@router.put("/{usuario_id}", summary="Editar usuario")
def editar_usuario(
    request: Request,
    usuario_id: int,
    data: UsuarioUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN))
):
    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # Proteger la cuenta SUPER_ADMIN: solo él mismo puede modificarla
    if u.rol == RolUsuario.SUPER_ADMIN and u.id != current_user.id:
        raise HTTPException(status_code=403, detail="No se pueden modificar los datos del Super Administrador")

    # No permitir editar el propio rol/estado (evitar accidente)
    if u.id == current_user.id and data.activo is False:
        raise HTTPException(status_code=400, detail="No puedes desactivarte a ti mismo")

    campos = data.model_dump(exclude_none=True)

    # Verificar email único si cambia
    if "email" in campos and campos["email"] != u.email:
        if db.query(Usuario).filter(Usuario.email == campos["email"]).first():
            raise HTTPException(status_code=409, detail="Ese email ya está en uso")

    # Verificar lab si rol cambia a LAB_ADMIN
    nuevo_rol = RolUsuario(campos["rol"]) if "rol" in campos else u.rol
    nuevo_lab = campos.get("laboratorio_id", u.laboratorio_id)
    nuevo_dep = campos.get("departamento_id", u.departamento_id)
    _validar_laboratorio(nuevo_lab, nuevo_rol, db)
    _validar_departamento(nuevo_dep, nuevo_rol, db)

    for campo, valor in campos.items():
        setattr(u, campo, valor)

    db.commit()
    db.refresh(u)
    registrar(db, accion=Accion.EDITAR_USUARIO, recurso=Recurso.USUARIO,
              usuario=current_user, recurso_id=u.id,
              detalle={
                  "usuario_afectado": u.nombre,
                  "email_afectado":   u.email,
                  "campos_modificados": list(campos.keys()),
              },
              request=request)
    return _serializar(u, db)


@router.delete("/{usuario_id}", summary="Eliminar usuario permanentemente")
def eliminar_usuario(
    request: Request,
    usuario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN))
):
    """
    Elimina un usuario de forma permanente.
    Bloquea si tiene horarios, sesiones, reservaciones o activos asociados.
    En ese caso usa desactivar (activo=False) en su lugar.
    """


    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if u.id == current_user.id:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")

    if u.rol == RolUsuario.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="No se puede eliminar al Super Administrador")

    # Desactivación lógica (soft delete): conserva historial e integridad referencial
    u.activo = False
    db.commit()

    registrar(db, accion=Accion.ELIMINAR_USUARIO, recurso=Recurso.USUARIO,
              usuario=current_user, recurso_id=usuario_id,
              detalle={"nombre": u.nombre, "email": u.email},
              request=request)

    return {"ok": True, "mensaje": f"Usuario '{u.nombre}' desactivado"}


@router.post("/{usuario_id}/reset-password", response_model=ResetPasswordResponse, summary="Resetear contraseña (admin)")
def reset_password(
    request: Request,
    usuario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN))
):
    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # Solo el SUPER_ADMIN puede resetear su propia contraseña, nadie más
    if u.rol == RolUsuario.SUPER_ADMIN and u.id != current_user.id:
        raise HTTPException(status_code=403, detail="No se puede resetear la contraseña del Super Administrador")

    nueva = _generar_password()
    u.password_hash = hashear_password(nueva)
    u.debe_cambiar_password = True  # contraseña temporal: forzar cambio al entrar
    db.commit()

    registrar(db, accion=Accion.CAMBIAR_PASSWORD, recurso=Recurso.USUARIO,
              usuario=current_user, recurso_id=u.id,
              detalle={"tipo": "reset_admin", "afectado": u.email},
              request=request)
    return {
        "mensaje": f"Contrasena reseteada para {u.nombre}",
        "password_temporal": nueva
    }


@router.put("/me/password", summary="Cambiar mi propia contraseña")
def cambiar_mi_password(
    data: CambiarPasswordPropio,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Cualquier usuario autenticado puede cambiar su propia contraseña.
    Requiere la contraseña actual para confirmar identidad.
    """
    # Verificar contraseña actual
    if not verificar_password(data.password_actual, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")

    # Evitar que sea la misma
    if data.password_actual == data.password_nuevo:
        raise HTTPException(status_code=400, detail="La nueva contraseña debe ser diferente a la actual")

    u = db.query(Usuario).filter(Usuario.id == current_user.id).first()
    u.password_hash = hashear_password(data.password_nuevo)
    u.debe_cambiar_password = False  # ya cumplió el cambio obligatorio
    db.commit()

    return {"mensaje": "Contraseña actualizada correctamente"}


@router.post("/bulk-excel", status_code=status.HTTP_201_CREATED, summary="Carga masiva desde Excel")
async def bulk_excel(
    request: Request,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN))
):
    """
    Carga masiva de usuarios desde Excel.

    El archivo debe tener las columnas:
    nombre | email | numero_empleado (opcional) | rol | laboratorio_id (opcional) | departamento_id/departamento_clave (opcional)

    La contraseña inicial se genera automáticamente y se devuelve en la respuesta.
    Rol aceptado: SUPER_ADMIN, LAB_ADMIN, DOCENTE, ADMINISTRATIVO, TUTORIA_ADMIN
    """
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="pandas no disponible")

    if not archivo.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    contenido = await archivo.read()
    try:
        df = pd.read_excel(io.BytesIO(contenido))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el Excel: {str(e)}")

    # Normalizar columnas
    df.columns = [c.strip().lower() for c in df.columns]
    requeridas = {'nombre', 'email', 'rol'}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise HTTPException(
            status_code=422,
            detail=f"Columnas requeridas faltantes: {', '.join(faltantes)}. "
                   f"Se necesita: nombre, email, rol. Opcional: numero_empleado, laboratorio_id, departamento_id, departamento_clave"
        )

    creados   = []
    errores   = []
    omitidos  = []

    for i, row in df.iterrows():
        fila = i + 2  # Número de fila en Excel (header = 1)
        nombre = str(row.get('nombre', '')).strip()
        email  = str(row.get('email', '')).strip().lower()
        rol_str = str(row.get('rol', '')).strip().upper()
        num_emp = str(row.get('numero_empleado', '')).strip() if 'numero_empleado' in df.columns else None
        lab_id  = row.get('laboratorio_id', None)
        dep_id  = row.get('departamento_id', None)
        dep_clave = str(row.get('departamento_clave', '')).strip() if 'departamento_clave' in df.columns else ''

        # Validar campos
        if not nombre or not email or not rol_str:
            errores.append({"fila": fila, "error": "nombre, email y rol son requeridos"})
            continue

        try:
            rol = RolUsuario(rol_str)
        except ValueError:
            errores.append({"fila": fila, "email": email, "error": f"Rol invalido: '{rol_str}'. Use: SUPER_ADMIN, LAB_ADMIN, DOCENTE, ADMINISTRATIVO, TUTORIA_ADMIN"})
            continue

        # Verificar duplicados
        if db.query(Usuario).filter(Usuario.email == email).first():
            omitidos.append({"fila": fila, "email": email, "razon": "Email ya registrado"})
            continue

        lab_id_int = None
        if lab_id and str(lab_id).strip() not in ('', 'nan', 'None'):
            try:
                lab_id_int = int(float(str(lab_id)))
                if not db.query(Laboratorio).filter(Laboratorio.id == lab_id_int, Laboratorio.activo == True).first():
                    errores.append({"fila": fila, "email": email, "error": f"Laboratorio {lab_id_int} no existe o está inactivo"})
                    continue
            except (ValueError, TypeError):
                errores.append({"fila": fila, "email": email, "error": f"laboratorio_id inválido: '{lab_id}'"})
                continue

        if rol == RolUsuario.LAB_ADMIN and not lab_id_int:
            errores.append({"fila": fila, "email": email, "error": "LAB_ADMIN requiere laboratorio_id"})
            continue

        dep_id_int = None
        if dep_id and str(dep_id).strip() not in ('', 'nan', 'None'):
            try:
                dep_id_int = int(float(str(dep_id)))
            except (ValueError, TypeError):
                errores.append({"fila": fila, "email": email, "error": f"departamento_id inválido: '{dep_id}'"})
                continue
        elif dep_clave and dep_clave.lower() != 'nan':
            dep = db.query(Departamento).filter(Departamento.clave.ilike(dep_clave), Departamento.activo == True).first()
            if not dep:
                errores.append({"fila": fila, "email": email, "error": f"departamento_clave no encontrada: '{dep_clave}'"})
                continue
            dep_id_int = dep.id

        if dep_id_int and not db.query(Departamento).filter(Departamento.id == dep_id_int, Departamento.activo == True).first():
            errores.append({"fila": fila, "email": email, "error": f"Departamento {dep_id_int} no existe o está inactivo"})
            continue
        if rol == RolUsuario.ADMINISTRATIVO and not dep_id_int:
            errores.append({"fila": fila, "email": email, "error": "ADMINISTRATIVO requiere departamento_id o departamento_clave"})
            continue

        password_tmp = _generar_password()
        usuario = Usuario(
            nombre=nombre,
            email=email,
            numero_empleado=num_emp if num_emp and num_emp != 'nan' else None,
            password_hash=hashear_password(password_tmp),
            rol=rol,
            laboratorio_id=lab_id_int,
            departamento_id=dep_id_int,
            activo=True,
            debe_cambiar_password=True,  # contraseña temporal: forzar cambio
        )
        db.add(usuario)
        creados.append({"fila": fila, "nombre": nombre, "email": email, "rol": rol_str, "password_temporal": password_tmp})

    if creados:
        db.commit()

    return {
        "resumen": {
            "procesados": len(df),
            "creados": len(creados),
            "omitidos": len(omitidos),
            "errores": len(errores),
        },
        "creados": creados,
        "omitidos": omitidos,
        "errores": errores,
    }


@router.post("/importar-docentes", status_code=status.HTTP_201_CREATED,
             summary="Importar docentes desde Plantilla_Docentes_UTECAN.xlsx")
async def importar_docentes(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN))
):
    """
    Importa docentes desde la plantilla oficial UTECAN.
    Columnas (hoja «Docentes»):
      Grado | Apellido Paterno | Apellido Materno | Nombre(s) | Email |
      Periodo activo | Laboratorio (opcional) | Teléfono (opcional) | Contraseña temporal (opcional)

    - Si el email ya existe → actualiza nombre y reactiva si estaba inactivo.
    - Si la contraseña está vacía → usa «Utecan2026» como temporal.
    """
    import openpyxl

    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx o .xls")

    contenido = await archivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        raise HTTPException(400, "Archivo Excel inválido o dañado")

    ws = wb["Docentes"] if "Docentes" in wb.sheetnames else wb.active

    creados     = []
    actualizados = []
    errores     = []

    # Plantilla: fila 1=título, 2=leyenda, 3=cabeceras, 4=ejemplo → datos desde fila 5
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        if all(v is None or str(v).strip() == "" for v in row):
            break

        grado    = str(row[0]).strip() if row[0] else ""
        ap_pat   = str(row[1]).strip() if row[1] else ""
        ap_mat   = str(row[2]).strip() if row[2] else ""
        nombres  = str(row[3]).strip() if row[3] else ""
        email    = str(row[4]).strip().lower() if row[4] else ""
        pwd_raw  = str(row[8]).strip() if len(row) > 8 and row[8] else ""

        # Validar
        fila_errores = []
        if not email:    fila_errores.append("email vacío")
        if not ap_pat:   fila_errores.append("apellido paterno vacío")
        if not nombres:  fila_errores.append("nombre(s) vacío")

        if fila_errores:
            errores.append({"fila": row_idx, "datos": email or "?", "errores": fila_errores})
            continue

        nombre_completo = f"{grado} {ap_pat} {ap_mat} {nombres}".strip()
        password_final  = pwd_raw if len(pwd_raw) >= 6 else "Utecan2026"

        existente = db.query(Usuario).filter(Usuario.email == email).first()
        if existente:
            existente.nombre = nombre_completo
            existente.activo = True
            actualizados.append({"fila": row_idx, "nombre": nombre_completo, "email": email})
        else:
            db.add(Usuario(
                nombre        = nombre_completo,
                email         = email,
                password_hash = hashear_password(password_final),
                rol           = RolUsuario.DOCENTE,
                activo        = True,
                debe_cambiar_password = True,  # contraseña temporal: forzar cambio
            ))
            creados.append({
                "fila":              row_idx,
                "nombre":            nombre_completo,
                "email":             email,
                "password_temporal": password_final,
            })

    db.commit()
    return {
        "resumen": {
            "creados":     len(creados),
            "actualizados": len(actualizados),
            "errores":     len(errores),
        },
        "creados":     creados,
        "actualizados": actualizados,
        "errores":     errores,
    }
