from fastapi import APIRouter, Depends, HTTPException, Request, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from services.auditoria import registrar, Accion, Recurso
import openpyxl, io, unicodedata
import uuid
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional, List
from xml.sax.saxutils import escape
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode.qr import QrCodeWidget
from database import get_db
from models.departamento import Departamento
from models.inventario import (
    Activo, Prestamo, Incidente, MantenimientoPreventivo, UbicacionInventario,
    MovimientoInventario, SolicitudBajaInventario, LevantamientoInventario,
    RevisionLevantamientoInventario, SeguimientoIncidente,
)
from models.laboratorio import Laboratorio, Computadora
from models.usuario import Usuario, RolUsuario
from models.usuario_permiso import UsuarioPermiso
from models.adeudo import Adeudo
from models.auditoria import AuditLog
from models.catalogo import CatalogoInventarioItem
from dependencies import get_current_user, require_roles
from rls import assert_lab_write, assert_resource_access
from services.user_permissions import (
    PERM_INVENTARIO_VALIDATE,
    departamentos_inventario,
    puede_gestionar_inventario,
    puede_validar_inventario,
)
import datetime


def _utcnow() -> datetime.datetime:
    return datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None)

router = APIRouter(prefix="/inventario", tags=["Inventario y Préstamos"])

CATEGORIAS = [
    "COMPUTADORA", "IMPRESORA_3D", "BRAZO_ROBOTICO", "SCANNER", "IOT",
    "HERRAMIENTA", "MOBILIARIO", "AUDIOVISUAL", "REDES", "MEDICO",
    "CRISTALERIA", "REACTIVO", "INSTRUMENTO_MEDICION", "EQUIPO_LABORATORIO",
    "MATERIAL_CONSUMIBLE", "SEGURIDAD_EPP", "ALMACENAMIENTO",
    "OFICINA", "VEHICULO", "OTRO",
]
ESTADOS_ACTIVO = ["OPERATIVO", "MANTENIMIENTO", "DAÑADO", "BAJA"]
ALCANCES_ACTIVO = ["LABORATORIO", "INSTITUCIONAL"]
TIPOS_INVENTARIO = ["ACTIVO"]
ESTADOS_ADMIN_ACTIVO = ["BORRADOR", "EN_REVISION", "OBSERVADO", "VALIDADO", "RECHAZADO", "BAJA_SOLICITADA", "BAJA_EJECUTADA"]
ESTADOS_BAJA = ["SOLICITADA", "EN_REVISION", "VALIDADA_FISICAMENTE", "AUTORIZADA", "RECHAZADA", "EJECUTADA", "CANCELADA"]
ESTADOS_LEVANTAMIENTO = ["ABIERTO", "CERRADO", "CANCELADO"]
ESTADOS_REVISION_LEVANTAMIENTO = ["LOCALIZADO", "NO_LOCALIZADO", "OTRA_UBICACION", "DANADO", "PROPUESTO_BAJA", "DATOS_INCOMPLETOS"]
UNIDADES_MEDIDA = ["PIEZA", "CAJA", "PAQUETE", "JUEGO", "METRO", "LITRO", "KILO", "SERVICIO", "OTRO"]
TIPOS_UBICACION = ["EDIFICIO", "OFICINA", "AULA", "LABORATORIO", "ALMACEN", "BIBLIOTECA", "CONSULTORIO", "TALLER", "EXTERIOR", "OTRO"]
TIPOS_MOVIMIENTO = ["TRANSFERENCIA_DEPARTAMENTO", "CAMBIO_UBICACION", "CAMBIO_RESGUARDANTE", "PRESTAMO_TEMPORAL", "RETORNO", "BAJA", "MANTENIMIENTO", "AJUSTE_INVENTARIO"]
ESTADOS_MOVIMIENTO = ["SOLICITADO", "AUTORIZADO", "RECHAZADO", "ENTREGADO", "RECIBIDO", "CANCELADO"]
ESTADOS_PRESTAMO = ["ACTIVO", "DEVUELTO", "VENCIDO"]
CONDICIONES = ["EXCELENTE", "BUENO", "REGULAR", "DAÑADO"]
ESTADOS_INCIDENTE_CERRADOS = ("REPARADO", "DADO_DE_BAJA", "CERRADO_SIN_ADEUDO")
CATALOGO_TIPO_CATEGORIA = "CATEGORIA_ACTIVO"
CATALOGO_TIPO_UBICACION = "TIPO_UBICACION"
CATALOGO_TIPOS = [CATALOGO_TIPO_CATEGORIA, CATALOGO_TIPO_UBICACION]
CATALOGO_ALCANCES = ["LABORATORIO", "INSTITUCIONAL", "AMBOS"]


def _es_rol_laboratorio(usuario: Usuario) -> bool:
    return usuario.rol in (RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB)


def _es_admin_inventario_global(usuario: Usuario) -> bool:
    return usuario.rol in (RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB)


def _filtrar_lab_asignado(query, model, usuario: Usuario, laboratorio_id: int | None):
    if _es_rol_laboratorio(usuario):
        if not usuario.laboratorio_id:
            raise HTTPException(status_code=403, detail="No tienes laboratorio asignado")
        if laboratorio_id and laboratorio_id != usuario.laboratorio_id:
            raise HTTPException(status_code=403, detail="Solo puedes consultar inventario de tu laboratorio")
        return query.filter(model.laboratorio_id == usuario.laboratorio_id)
    if laboratorio_id:
        return query.filter(model.laboratorio_id == laboratorio_id)
    return query


def _asegurar_acceso_incidente(incidente: Incidente, usuario: Usuario) -> None:
    if _es_rol_laboratorio(usuario):
        if not usuario.laboratorio_id or incidente.laboratorio_id != usuario.laboratorio_id:
            raise HTTPException(
                status_code=403,
                detail="Solo puedes gestionar incidentes de tu laboratorio",
            )


def _departamentos_visibles_inventario(db: Session, usuario: Usuario) -> list[int] | None:
    if _es_admin_inventario_global(usuario) or puede_validar_inventario(db, usuario):
        return None
    ids = set(departamentos_inventario(db, usuario))
    if usuario.rol == RolUsuario.ADMINISTRATIVO and usuario.departamento_id:
        ids.add(usuario.departamento_id)
    return sorted(ids)


def _resolver_departamento_inventario(
    db: Session,
    usuario: Usuario,
    departamento_id: int | None,
) -> int:
    if _es_admin_inventario_global(usuario) or puede_validar_inventario(db, usuario):
        if departamento_id is None:
            raise HTTPException(status_code=422, detail="departamento_id es requerido")
        return departamento_id
    permitidos = departamentos_inventario(db, usuario)
    if departamento_id:
        if departamento_id not in permitidos:
            raise HTTPException(status_code=403, detail="Solo puedes gestionar inventario de tu departamento")
        return departamento_id
    if usuario.departamento_id and usuario.departamento_id in permitidos:
        return usuario.departamento_id
    if len(permitidos) == 1:
        return permitidos[0]
    raise HTTPException(status_code=422, detail="Selecciona el departamento del activo")


def _asegurar_write_inventario_departamento(
    db: Session,
    usuario: Usuario,
    departamento_id: int | None,
) -> None:
    if _es_admin_inventario_global(usuario) or puede_validar_inventario(db, usuario):
        return
    if not departamento_id or not puede_gestionar_inventario(db, usuario, departamento_id):
        raise HTTPException(status_code=403, detail="No tienes permiso para gestionar inventario de este departamento")


def _asegurar_acceso_activo_departamental(
    db: Session,
    activo: Activo | None,
    usuario: Usuario,
) -> None:
    assert_resource_access(activo, usuario)
    if _es_admin_inventario_global(usuario) or puede_validar_inventario(db, usuario):
        return
    permitidos = departamentos_inventario(db, usuario)
    if not activo or activo.departamento_id not in permitidos:
        raise HTTPException(status_code=404, detail="Activo no encontrado")


def _asegurar_activo_validado(activo: Activo | None) -> None:
    if not activo:
        raise HTTPException(status_code=404, detail="Activo no encontrado")
    estado = (activo.estado_admin or "VALIDADO").upper()
    if estado != "VALIDADO":
        raise HTTPException(
            status_code=409,
            detail=(
                f"El activo esta en estado {estado.replace('_', ' ')}. "
                "Solo puede consultarse o corregirse hasta que Super Admin lo valide."
            ),
        )


# Códigos cortos por categoría para armar el número de inventario
TIPO_CODIGO = {
    "COMPUTADORA":    "PC",
    "IMPRESORA_3D":   "IMP",
    "BRAZO_ROBOTICO": "ROB",
    "SCANNER":        "SCN",
    "IOT":            "IOT",
    "HERRAMIENTA":    "HER",
    "MOBILIARIO":     "MOB",
    "AUDIOVISUAL":    "AUD",
    "REDES":          "RED",
    "MEDICO":         "MED",
    "CRISTALERIA":    "CRI",
    "REACTIVO":       "REA",
    "INSTRUMENTO_MEDICION": "IME",
    "EQUIPO_LABORATORIO": "LAB",
    "MATERIAL_CONSUMIBLE": "CON",
    "SEGURIDAD_EPP":  "EPP",
    "ALMACENAMIENTO": "ALM",
    "OFICINA":        "OFI",
    "VEHICULO":       "VEH",
    "OTRO":           "OTR",
}


# ─── Schemas ───────────────────────────────────────────────────────────────────

class ActivoCreate(BaseModel):
    laboratorio_id: Optional[int] = None
    departamento_id: Optional[int] = None
    ubicacion_id: Optional[int] = None
    responsable_id: Optional[int] = None
    alcance: str = Field(default="LABORATORIO", description=f"Uno de: {ALCANCES_ACTIVO}")
    tipo_inventario: str = Field(default="ACTIVO", description=f"Uno de: {TIPOS_INVENTARIO}")
    estado_admin: str = Field(default="VALIDADO", description=f"Uno de: {ESTADOS_ADMIN_ACTIVO}")
    # Opcional: si no se envía, el sistema genera el código automáticamente
    codigo_inventario: Optional[str] = Field(None, min_length=2, max_length=50)
    numero_oficial: Optional[str] = Field(None, max_length=80)
    nombre: str                      = Field(..., min_length=2, max_length=100)
    categoria: str                   = Field(..., description=f"Una de: {CATEGORIAS}")
    area: Optional[str]              = None   # Prefijo de área, p.ej. "LTI", "LINF"
    marca: Optional[str]             = None
    modelo: Optional[str]            = None
    numero_serie: Optional[str]      = None
    valor: Optional[float]           = Field(None, ge=0)
    # Patrimonial: siempre cantidad=1, PIEZA, sin stock mínimo
    cantidad: float                          = Field(default=1.0, ge=1.0, le=1.0)
    unidad_medida: str                       = Field(default="PIEZA", pattern="^PIEZA$")
    stock_minimo: Optional[float]            = Field(None, ge=0)
    estado: str                              = "OPERATIVO"
    especificaciones: Optional[str]          = None
    observaciones: Optional[str]             = None
    resguardante_externo_nombre: Optional[str] = None   # cuando no es usuario del sistema
    ubicacion_tipo: Optional[str]            = None
    ubicacion_nombre: Optional[str]          = None

class ActivoUpdate(BaseModel):
    laboratorio_id: Optional[int]            = None
    departamento_id: Optional[int]           = None
    ubicacion_id: Optional[int]              = None
    responsable_id: Optional[int]            = None
    alcance: Optional[str]                   = None
    tipo_inventario: Optional[str]           = None
    estado_admin: Optional[str]              = None
    numero_oficial: Optional[str]            = Field(None, max_length=80)
    nombre: Optional[str]                    = Field(None, min_length=2, max_length=100)
    categoria: Optional[str]                 = None
    area: Optional[str]                      = None
    marca: Optional[str]                     = None
    modelo: Optional[str]                    = None
    numero_serie: Optional[str]              = None
    valor: Optional[float]                   = Field(None, ge=0)
    # cantidad y unidad_medida no se exponen en edición — siempre son 1/PIEZA
    estado: Optional[str]                    = None
    especificaciones: Optional[str]          = None
    observaciones: Optional[str]             = None
    resguardante_externo_nombre: Optional[str] = None
    ubicacion_tipo: Optional[str]            = None
    ubicacion_nombre: Optional[str]          = None
    activo: Optional[bool]                   = None


class ActivoValidacionUpdate(BaseModel):
    estado_admin: str
    observaciones: Optional[str] = None


class UbicacionInventarioIn(BaseModel):
    nombre: str = Field(..., min_length=2, max_length=150)
    tipo: str = Field(default="OFICINA")
    edificio: Optional[str] = Field(None, max_length=120)
    piso: Optional[str] = Field(None, max_length=40)
    referencia: Optional[str] = Field(None, max_length=250)
    departamento_id: Optional[int] = None
    activo: bool = True


class CatalogoInventarioIn(BaseModel):
    tipo: str = Field(..., description=f"Uno de: {CATALOGO_TIPOS}")
    nombre: str = Field(..., min_length=2, max_length=150)
    clave: Optional[str] = Field(None, max_length=50)
    prefijo_codigo: Optional[str] = Field(None, max_length=12)
    alcance: str = Field(default="AMBOS", description=f"Uno de: {CATALOGO_ALCANCES}")
    activo: bool = True


class CatalogoInventarioUpdate(BaseModel):
    nombre: Optional[str] = Field(None, min_length=2, max_length=150)
    prefijo_codigo: Optional[str] = Field(None, max_length=12)
    alcance: Optional[str] = Field(None, description=f"Uno de: {CATALOGO_ALCANCES}")
    activo: Optional[bool] = None


class MovimientoInventarioCreate(BaseModel):
    tipo: str = Field(..., description=f"Uno de: {TIPOS_MOVIMIENTO}")
    departamento_destino_id: Optional[int] = None
    ubicacion_destino_id: Optional[int] = None
    resguardante_destino_id: Optional[int] = None
    ubicacion_destino_nombre: Optional[str] = None
    resguardante_destino_nombre: Optional[str] = None
    cantidad: Optional[float] = Field(None, gt=0)
    observaciones: Optional[str] = None


class MovimientoInventarioEstado(BaseModel):
    observaciones: Optional[str] = None


class SolicitudBajaCreate(BaseModel):
    motivo: str = Field(..., min_length=4)
    diagnostico: Optional[str] = None
    evidencia_url: Optional[str] = None
    destino_final: Optional[str] = None
    observaciones: Optional[str] = None


class SolicitudBajaAccion(BaseModel):
    observaciones: Optional[str] = None
    destino_final: Optional[str] = None


class LevantamientoCreate(BaseModel):
    nombre: str = Field(..., min_length=3, max_length=150)
    departamento_id: Optional[int] = None
    laboratorio_id: Optional[int] = None
    observaciones: Optional[str] = None


class RevisionLevantamientoIn(BaseModel):
    activo_id: int
    estado: str = Field(..., description=f"Uno de: {ESTADOS_REVISION_LEVANTAMIENTO}")
    ubicacion_reportada: Optional[str] = None
    resguardante_reportado: Optional[str] = None
    observaciones: Optional[str] = None
    evidencia_url: Optional[str] = None

class PrestamoCreate(BaseModel):
    activo_id: Optional[int] = None
    activo_ids: List[int] = Field(default_factory=list, max_length=50)
    receptor_nombre: str              = Field(..., min_length=2, max_length=100)
    receptor_matricula: Optional[str] = None
    receptor_tipo: str                = Field(default="ALUMNO")
    proposito: Optional[str]          = None
    fecha_devolucion_esperada: Optional[str] = None   # ISO date string YYYY-MM-DD
    notas: Optional[str]              = None

class PrestamoDevolver(BaseModel):
    condicion_devolucion: str         = Field(..., description="BUENO, REGULAR, MALO, DAÑADO")
    notas_devolucion: Optional[str]   = None


class PrestamoGrupoDevolver(PrestamoDevolver):
    prestamo_ids: List[int] = Field(default_factory=list, max_length=50)

class IncidenteCreate(BaseModel):
    activo_id:       Optional[int]   = None
    computadora_id:  Optional[int]   = None
    laboratorio_id:  Optional[int]   = None
    origen:          str             = Field(default="MANUAL")   # PRESTAMO | SESION | MANUAL
    origen_id:       Optional[int]   = None
    tipo:            str             = Field(default="DAÑO")     # DAÑO | MANTENIMIENTO | PERDIDA | OTRO
    descripcion:     Optional[str]   = None
    prioridad:       str             = Field(default="MEDIA")    # ALTA | MEDIA | BAJA

class IncidenteUpdate(BaseModel):
    estado:            Optional[str]   = None   # PENDIENTE | EN_REVISION | REPARADO | DADO_DE_BAJA
    prioridad:         Optional[str]   = None
    notas_seguimiento: Optional[str]   = None
    costo_reparacion:  Optional[float] = None
    descripcion:       Optional[str]   = None
    activo_id:         Optional[int]   = None
    computadora_id:    Optional[int]   = None
    motivo_vinculacion: Optional[str]  = Field(default=None, max_length=250)


class IncidenteSeguimientoCreate(BaseModel):
    texto: str = Field(..., min_length=2, max_length=2000)


class IncidenteReabrir(BaseModel):
    motivo: str = Field(..., min_length=5, max_length=1000)


# ─── Helpers ───────────────────────────────────────────────────────────────────

def _normalizar(texto: str) -> str:
    """Quita acentos y convierte a mayúsculas para comparaciones flexibles."""
    txt = texto.upper().strip()
    return "".join(
        c for c in unicodedata.normalize("NFD", txt)
        if unicodedata.category(c) != "Mn"
    )

def _normalizar_clave_catalogo(texto: str) -> str:
    base = _normalizar(texto or "")
    chars = []
    anterior_sep = False
    for char in base:
        if char.isalnum():
            chars.append(char)
            anterior_sep = False
        elif not anterior_sep:
            chars.append("_")
            anterior_sep = True
    clave = "".join(chars).strip("_")
    return (clave or "OTRO")[:50]


def _catalogo_base_items(tipo: str) -> list[dict]:
    if tipo == CATALOGO_TIPO_CATEGORIA:
        return [
            {
                "id": None,
                "tipo": tipo,
                "clave": clave,
                "nombre": clave.replace("_", " ").title(),
                "prefijo_codigo": TIPO_CODIGO.get(clave, "OTR"),
                "alcance": "AMBOS",
                "activo": True,
                "protegido": True,
                "base": True,
            }
            for clave in CATEGORIAS
        ]
    if tipo == CATALOGO_TIPO_UBICACION:
        return [
            {
                "id": None,
                "tipo": tipo,
                "clave": clave,
                "nombre": clave.replace("_", " ").title(),
                "prefijo_codigo": None,
                "alcance": "AMBOS",
                "activo": True,
                "protegido": True,
                "base": True,
            }
            for clave in TIPOS_UBICACION
        ]
    return []


def _serializar_catalogo_inventario(item: CatalogoInventarioItem) -> dict:
    return {
        "id": item.id,
        "tipo": item.tipo,
        "clave": item.clave,
        "nombre": item.nombre,
        "prefijo_codigo": item.prefijo_codigo,
        "alcance": item.alcance,
        "activo": item.activo,
        "protegido": item.protegido,
        "base": False,
        "creado_por_id": item.creado_por_id,
        "creado_en": item.creado_en.isoformat() if item.creado_en else None,
        "actualizado_en": item.actualizado_en.isoformat() if item.actualizado_en else None,
    }


def _catalogo_items(db: Session, tipo: str, solo_activos: bool = True) -> list[dict]:
    items = _catalogo_base_items(tipo)
    q = db.query(CatalogoInventarioItem).filter(CatalogoInventarioItem.tipo == tipo)
    if solo_activos:
        q = q.filter(CatalogoInventarioItem.activo == True)
    custom = [_serializar_catalogo_inventario(i) for i in q.order_by(CatalogoInventarioItem.nombre).all()]
    por_clave = {i["clave"]: i for i in items}
    for item in custom:
        por_clave[item["clave"]] = item
    return sorted(por_clave.values(), key=lambda i: (not i.get("base"), i["nombre"]))


def _catalogo_claves(db: Session, tipo: str) -> list[str]:
    return [i["clave"] for i in _catalogo_items(db, tipo, solo_activos=True)]


def _catalogo_prefijo_categoria(db: Session, categoria: str) -> str:
    clave = (categoria or "").upper()
    item = db.query(CatalogoInventarioItem).filter(
        CatalogoInventarioItem.tipo == CATALOGO_TIPO_CATEGORIA,
        CatalogoInventarioItem.clave == clave,
        CatalogoInventarioItem.activo == True,
    ).first()
    if item and item.prefijo_codigo:
        return item.prefijo_codigo.upper()
    return TIPO_CODIGO.get(clave, "OTR")


def _asegurar_permiso_catalogo_inventario(db: Session, usuario: Usuario) -> None:
    if not puede_validar_inventario(db, usuario):
        raise HTTPException(
            status_code=403,
            detail="Solo Inventario Institucional puede administrar este catalogo",
        )


def _buscar_lab(nombre_excel: str, labs_dict: dict, db, current_user):
    """
    Busca un laboratorio con tolerancia a acentos y diferencias parciales.
    labs_dict: {nombre_normalizado: objeto Laboratorio}
    Retorna el lab o None.
    """
    clave = _normalizar(nombre_excel)
    if _es_rol_laboratorio(current_user):
        if not current_user.laboratorio_id:
            return None
        lab_asignado = db.query(Laboratorio).filter(Laboratorio.id == current_user.laboratorio_id).first()
        if not lab_asignado:
            return None
        if not clave:
            return lab_asignado
        clave_asignada = _normalizar(lab_asignado.nombre)
        if clave == clave_asignada or clave in clave_asignada or clave_asignada in clave:
            return lab_asignado
        return None
    if not clave:
        return None
    # 1. Coincidencia exacta (sin acentos)
    if clave in labs_dict:
        return labs_dict[clave]
    # 2. El nombre del Excel está contenido en el nombre del lab o viceversa
    for key, lab in labs_dict.items():
        if clave in key or key in clave:
            return lab
    # 3. Coincidencia por palabras clave (al menos la mitad de palabras coinciden)
    palabras = set(clave.split())
    for key, lab in labs_dict.items():
        palabras_lab = set(key.split())
        comunes = palabras & palabras_lab
        if len(comunes) >= max(1, len(palabras) // 2):
            return lab
    return None


def _generar_codigo(db: Session, categoria: str, area: str = None) -> str:
    """Genera el siguiente número de inventario libre en formato UTC-[ÁREA]-[TIPO]-[SEQ]."""
    tipo      = _catalogo_prefijo_categoria(db, categoria)
    area_code = (area or "UTC").upper().strip().replace(" ", "")[:10]
    prefix    = f"UTC-{area_code}-{tipo}-"
    existentes = db.query(Activo.codigo_inventario).filter(
        Activo.codigo_inventario.like(f"{prefix}%")
    ).all()
    max_seq = 0
    for (code,) in existentes:
        try:
            seq = int(code[len(prefix):])
            max_seq = max(max_seq, seq)
        except (ValueError, IndexError):
            pass
    return f"{prefix}{(max_seq + 1):03d}"


def _actualizar_estado_prestamos(db: Session):
    """Marca como VENCIDO cualquier préstamo activo con fecha de retorno pasada."""
    ahora = _utcnow()
    vencidos = db.query(Prestamo).filter(
        Prestamo.estado == "ACTIVO",
        Prestamo.fecha_retorno_esperada < ahora,
    ).all()
    for p in vencidos:
        p.estado = "VENCIDO"
    if vencidos:
        db.commit()


def _meta_mantenimiento_activo(db: Session, activo_id: int) -> dict:
    ahora = _utcnow()
    en_7 = ahora + datetime.timedelta(days=7)
    pendientes = db.query(MantenimientoPreventivo).filter(
        MantenimientoPreventivo.activo_id == activo_id,
        MantenimientoPreventivo.estado.in_(["PENDIENTE", "EN_PROCESO"]),
    ).order_by(
        MantenimientoPreventivo.fecha_limite.is_(None),
        MantenimientoPreventivo.fecha_limite,
        MantenimientoPreventivo.fecha_programada,
    ).all()
    vencidos = [m for m in pendientes if m.fecha_limite and m.fecha_limite < ahora]
    proximos = [m for m in pendientes if m.fecha_limite and ahora <= m.fecha_limite <= en_7]
    siguiente = pendientes[0] if pendientes else None
    estado_alerta = "OK"
    if vencidos:
        estado_alerta = "VENCIDO"
    elif proximos:
        estado_alerta = "PROXIMO"
    elif pendientes:
        estado_alerta = "PROGRAMADO"
    return {
        "pendientes": len(pendientes),
        "vencidos": len(vencidos),
        "proximos_7": len(proximos),
        "estado_alerta": estado_alerta,
        "siguiente_id": siguiente.id if siguiente else None,
        "siguiente_tipo": siguiente.tipo if siguiente else None,
        "siguiente_fecha": siguiente.fecha_programada.isoformat() if siguiente and siguiente.fecha_programada else None,
        "siguiente_limite": siguiente.fecha_limite.isoformat() if siguiente and siguiente.fecha_limite else None,
    }


def _serializar_activo(a: Activo, db: Session) -> dict:
    lab = db.query(Laboratorio).filter(Laboratorio.id == a.laboratorio_id).first() if a.laboratorio_id else None
    dep = db.query(Departamento).filter(Departamento.id == a.departamento_id).first() if a.departamento_id else None
    ubicacion = db.query(UbicacionInventario).filter(UbicacionInventario.id == a.ubicacion_id).first() if a.ubicacion_id else None
    responsable = db.query(Usuario).filter(Usuario.id == a.responsable_id).first() if a.responsable_id else None
    mantenimiento = _meta_mantenimiento_activo(db, a.id)
    validacion = _meta_validacion_activo(a, db)
    computadora = db.query(Computadora).filter(Computadora.activo_id == a.id).first()
    prestamo_activo = db.query(Prestamo).filter(
        Prestamo.activo_id == a.id,
        Prestamo.estado.in_(["ACTIVO", "VENCIDO"])
    ).first()
    ubicacion_label = None
    if ubicacion:
        ubicacion_label = " / ".join([p for p in [ubicacion.edificio, ubicacion.nombre] if p])
    elif a.ubicacion_nombre:
        ubicacion_label = a.ubicacion_nombre
    return {
        "id": a.id,
        "alcance": a.alcance or ("LABORATORIO" if a.laboratorio_id else "INSTITUCIONAL"),
        "tipo_inventario": a.tipo_inventario or "ACTIVO",
        "estado_admin": a.estado_admin or "VALIDADO",
        "laboratorio_id": a.laboratorio_id,
        "laboratorio_nombre": lab.nombre if lab else None,
        "departamento_id": a.departamento_id,
        "departamento_nombre": dep.nombre if dep else None,
        "departamento_clave": dep.clave if dep else None,
        "ubicacion_id": a.ubicacion_id,
        "ubicacion_tipo": a.ubicacion_tipo or (ubicacion.tipo if ubicacion else None),
        "ubicacion_nombre": a.ubicacion_nombre,
        "ubicacion_label": ubicacion_label,
        "responsable_id": a.responsable_id,
        "responsable_nombre": responsable.nombre if responsable else None,
        "codigo_inventario": a.codigo_inventario,
        "numero_oficial": a.numero_oficial,
        "nombre": a.nombre,
        "categoria": a.categoria,
        "marca": a.marca,
        "modelo": a.modelo,
        "numero_serie": a.numero_serie,
        "fecha_adquisicion": a.fecha_adquisicion.isoformat() if a.fecha_adquisicion else None,
        "valor": a.valor,
        "cantidad": a.cantidad,
        "unidad_medida": a.unidad_medida,
        "stock_minimo": a.stock_minimo,
        "estado": a.estado,
        "especificaciones": a.especificaciones,
        "observaciones": a.observaciones,
        "area": a.area,
        "resguardante_externo_nombre": a.resguardante_externo_nombre,
        "activo": a.activo,
        "prestado": prestamo_activo is not None,
        "prestamo_estado": prestamo_activo.estado if prestamo_activo else None,
        "computadora_id": computadora.id if computadora else None,
        "computadora_codigo": computadora.codigo if computadora else None,
        "computadora_numero": computadora.numero if computadora else None,
        "computadora_fila": computadora.fila if computadora else None,
        "mantenimiento": mantenimiento,
        **validacion,
    }


def _serializar_ubicacion(u: UbicacionInventario, db: Session) -> dict:
    dep = db.query(Departamento).filter(Departamento.id == u.departamento_id).first() if u.departamento_id else None
    return {
        "id": u.id,
        "nombre": u.nombre,
        "tipo": u.tipo,
        "edificio": u.edificio,
        "piso": u.piso,
        "referencia": u.referencia,
        "departamento_id": u.departamento_id,
        "departamento_nombre": dep.nombre if dep else None,
        "activo": u.activo,
        "creado_en": u.creado_en.isoformat() if u.creado_en else None,
        "label": " / ".join([p for p in [u.edificio, u.nombre] if p]),
    }


def _query_activos_filtrados(
    db: Session,
    current_user: Usuario,
    laboratorio_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    ubicacion_id: Optional[int] = None,
    alcance: Optional[str] = None,
    tipo_inventario: Optional[str] = None,
    estado_admin: Optional[str] = None,
    categoria: Optional[str] = None,
    estado: Optional[str] = None,
    solo_activos: bool = True,
):
    q = db.query(Activo)
    if solo_activos:
        q = q.filter(Activo.activo == True)
    q = _filtrar_lab_asignado(q, Activo, current_user, laboratorio_id)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None:
        if not departamentos_visibles:
            return None
        if departamento_id:
            if departamento_id not in departamentos_visibles:
                raise HTTPException(status_code=403, detail="Solo puedes consultar inventario de tu departamento")
        else:
            q = q.filter(Activo.departamento_id.in_(departamentos_visibles))
    if departamento_id:
        q = q.filter(Activo.departamento_id == departamento_id)
    if ubicacion_id:
        q = q.filter(Activo.ubicacion_id == ubicacion_id)
    if alcance:
        q = q.filter(Activo.alcance == alcance.upper())
    if tipo_inventario:
        q = q.filter(Activo.tipo_inventario == tipo_inventario.upper())
    if estado_admin:
        q = q.filter(Activo.estado_admin == estado_admin.upper())
    if categoria:
        q = q.filter(Activo.categoria == categoria.upper())
    if estado:
        q = q.filter(Activo.estado == estado.upper())
    return q


def _serializar_departamento_opcion(dep: Departamento) -> dict:
    return {
        "id": dep.id,
        "nombre": dep.nombre,
        "clave": dep.clave,
        "activo": dep.activo,
    }


def _serializar_usuario_resguardo(u: Usuario, db: Session) -> dict:
    dep = db.query(Departamento).filter(Departamento.id == u.departamento_id).first() if u.departamento_id else None
    lab = db.query(Laboratorio).filter(Laboratorio.id == u.laboratorio_id).first() if u.laboratorio_id else None
    return {
        "id": u.id,
        "nombre": u.nombre,
        "email": u.email,
        "numero_empleado": u.numero_empleado,
        "rol": u.rol.value if hasattr(u.rol, "value") else str(u.rol),
        "departamento_id": u.departamento_id,
        "departamento_nombre": dep.nombre if dep else None,
        "laboratorio_id": u.laboratorio_id,
        "laboratorio_nombre": lab.nombre if lab else None,
    }


def _texto_pdf(value) -> str:
    return escape(str(value or ""))


def _qr_drawing(payload: str, size: float):
    qr = QrCodeWidget(payload)
    qr.barWidth = size
    qr.barHeight = size
    drawing = Drawing(size, size)
    drawing.add(qr)
    return drawing


def _etiqueta_activo_pdf(activo_data: dict, styles: dict):
    codigo = activo_data.get("codigo_inventario") or f"ACT-{activo_data.get('id')}"
    numero = activo_data.get("numero_oficial")
    ubicacion = activo_data.get("ubicacion_label") or activo_data.get("ubicacion_nombre") or "Sin ubicacion"
    responsable = activo_data.get("responsable_nombre") or activo_data.get("resguardante_externo_nombre") or "Sin resguardante"
    qr_payload = "|".join([
        "SIGA-UTECAN",
        "ACTIVO",
        str(activo_data.get("id") or ""),
        str(codigo),
        str(numero or ""),
    ])
    qr = _qr_drawing(qr_payload, 2.15 * cm)
    texto = Table(
        [
            [Paragraph(_texto_pdf(codigo), styles["codigo"])],
            [Paragraph(_texto_pdf(numero or "Sin numero oficial"), styles["muted"])],
            [Paragraph(_texto_pdf(activo_data.get("nombre")), styles["nombre"])],
            [Paragraph(_texto_pdf((activo_data.get("categoria") or "").replace("_", " ")), styles["detalle"])],
            [Paragraph(_texto_pdf(ubicacion), styles["detalle"])],
            [Paragraph(_texto_pdf(responsable), styles["muted"])],
        ],
        colWidths=[6.1 * cm],
        style=TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]),
    )
    return Table(
        [[qr, texto]],
        colWidths=[2.45 * cm, 6.25 * cm],
        style=TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#0f766e")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]),
    )


def _generar_pdf_etiquetas(activos_data: list[dict]) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=1.0 * cm,
        rightMargin=1.0 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )
    base = getSampleStyleSheet()
    styles = {
        "codigo": ParagraphStyle("codigo", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=10.5, leading=12, textColor=colors.HexColor("#0f172a"), alignment=TA_LEFT),
        "nombre": ParagraphStyle("nombre", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=8.2, leading=9.5, textColor=colors.HexColor("#1e293b"), alignment=TA_LEFT),
        "detalle": ParagraphStyle("detalle", parent=base["Normal"], fontSize=7.2, leading=8.4, textColor=colors.HexColor("#334155"), alignment=TA_LEFT),
        "muted": ParagraphStyle("muted", parent=base["Normal"], fontSize=6.8, leading=8, textColor=colors.HexColor("#64748b"), alignment=TA_LEFT),
    }
    etiquetas = [_etiqueta_activo_pdf(a, styles) for a in activos_data]
    rows = []
    for i in range(0, len(etiquetas), 2):
        rows.append([etiquetas[i], etiquetas[i + 1] if i + 1 < len(etiquetas) else ""])
    grid = Table(
        rows,
        colWidths=[9.1 * cm, 9.1 * cm],
        rowHeights=[3.35 * cm for _ in rows],
        hAlign="LEFT",
        style=TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]),
    )
    story = [grid] if rows else [Spacer(1, 1 * cm), Paragraph("No hay activos para etiquetar.", base["Normal"])]
    doc.build(story)
    buffer.seek(0)
    return buffer


def _iso(dt) -> str | None:
    return dt.isoformat() if dt else None


def _serializar_audit_log(log: AuditLog) -> dict:
    return {
        "id": log.id,
        "fecha": _iso(log.timestamp),
        "accion": log.accion,
        "recurso": log.recurso,
        "usuario_nombre": log.usuario_nombre,
        "usuario_email": log.usuario_email,
        "detalle": log.detalle or {},
        "exito": log.exito,
    }


def _meta_validacion_activo(a: Activo, db: Session) -> dict:
    logs = db.query(AuditLog).filter(
        AuditLog.recurso == Recurso.ACTIVO,
        AuditLog.recurso_id == a.id,
    ).order_by(AuditLog.timestamp.desc()).limit(40).all()
    validacion = next(
        (log for log in logs if (log.detalle or {}).get("flujo") == "VALIDACION_INVENTARIO"),
        None,
    )
    creacion = next((log for log in reversed(logs) if log.accion == Accion.CREAR_ACTIVO), None)
    detalle = validacion.detalle if validacion else {}
    return {
        "registrado_por_id": creacion.usuario_id if creacion else None,
        "registrado_por_nombre": creacion.usuario_nombre if creacion else None,
        "registrado_fecha": _iso(creacion.timestamp) if creacion else None,
        "validacion_motivo": detalle.get("observaciones"),
        "validacion_revisor": validacion.usuario_nombre if validacion else None,
        "validacion_fecha": _iso(validacion.timestamp) if validacion else None,
    }


def _destinatarios_validacion_activo(a: Activo, db: Session) -> set[int]:
    destinatarios: set[int] = set()
    creacion = db.query(AuditLog).filter(
        AuditLog.recurso == Recurso.ACTIVO,
        AuditLog.recurso_id == a.id,
        AuditLog.accion == Accion.CREAR_ACTIVO,
    ).order_by(AuditLog.timestamp.asc()).first()
    if creacion and creacion.usuario_id:
        destinatarios.add(creacion.usuario_id)

    if a.laboratorio_id:
        usuarios_lab = db.query(Usuario.id).filter(
            Usuario.laboratorio_id == a.laboratorio_id,
            Usuario.activo == True,
            Usuario.rol.in_([RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB]),
        ).all()
        destinatarios.update(uid for (uid,) in usuarios_lab)
    elif a.departamento_id:
        dep = db.query(Departamento).filter(Departamento.id == a.departamento_id).first()
        if dep and dep.responsable_id:
            destinatarios.add(dep.responsable_id)
    return destinatarios


def _destinatarios_revision_institucional(db: Session, excluir_usuario_id: int | None = None) -> set[int]:
    destinatarios = {
        uid
        for (uid,) in db.query(Usuario.id).filter(
            Usuario.activo == True,
            Usuario.rol == RolUsuario.SUPER_ADMIN,
        ).all()
    }
    destinatarios.update(
        uid
        for (uid,) in db.query(UsuarioPermiso.usuario_id).join(
            Usuario,
            Usuario.id == UsuarioPermiso.usuario_id,
        ).filter(
            UsuarioPermiso.permiso == PERM_INVENTARIO_VALIDATE,
            UsuarioPermiso.activo == True,
            Usuario.activo == True,
        ).all()
    )
    if excluir_usuario_id:
        destinatarios.discard(excluir_usuario_id)
    return destinatarios


def _evento_expediente(tipo: str, fecha, titulo: str, descripcion: str | None = None, estado: str | None = None, actor: str | None = None, detalle: dict | None = None) -> dict:
    return {
        "tipo": tipo,
        "fecha": _iso(fecha),
        "titulo": titulo,
        "descripcion": descripcion,
        "estado": estado,
        "actor": actor,
        "detalle": detalle or {},
    }


def _serializar_movimiento(m: MovimientoInventario, db: Session) -> dict:
    activo = db.query(Activo).filter(Activo.id == m.activo_id).first()
    dep_o = db.query(Departamento).filter(Departamento.id == m.departamento_origen_id).first() if m.departamento_origen_id else None
    dep_d = db.query(Departamento).filter(Departamento.id == m.departamento_destino_id).first() if m.departamento_destino_id else None
    ubi_o = db.query(UbicacionInventario).filter(UbicacionInventario.id == m.ubicacion_origen_id).first() if m.ubicacion_origen_id else None
    ubi_d = db.query(UbicacionInventario).filter(UbicacionInventario.id == m.ubicacion_destino_id).first() if m.ubicacion_destino_id else None
    solicitado = db.query(Usuario).filter(Usuario.id == m.solicitado_por_id).first() if m.solicitado_por_id else None
    autorizado = db.query(Usuario).filter(Usuario.id == m.autorizado_por_id).first() if m.autorizado_por_id else None
    entregado = db.query(Usuario).filter(Usuario.id == m.entregado_por_id).first() if m.entregado_por_id else None
    recibido = db.query(Usuario).filter(Usuario.id == m.recibido_por_id).first() if m.recibido_por_id else None
    return {
        "id": m.id,
        "activo_id": m.activo_id,
        "activo_nombre": activo.nombre if activo else None,
        "activo_codigo": activo.codigo_inventario if activo else None,
        "tipo": m.tipo,
        "estado": m.estado,
        "departamento_origen_id": m.departamento_origen_id,
        "departamento_origen_nombre": dep_o.nombre if dep_o else None,
        "departamento_destino_id": m.departamento_destino_id,
        "departamento_destino_nombre": dep_d.nombre if dep_d else None,
        "ubicacion_origen_id": m.ubicacion_origen_id,
        "ubicacion_origen_nombre": m.ubicacion_origen_nombre or (ubi_o.nombre if ubi_o else None),
        "ubicacion_destino_id": m.ubicacion_destino_id,
        "ubicacion_destino_nombre": m.ubicacion_destino_nombre or (ubi_d.nombre if ubi_d else None),
        "resguardante_origen_id": m.resguardante_origen_id,
        "resguardante_origen_nombre": m.resguardante_origen_nombre,
        "resguardante_destino_id": m.resguardante_destino_id,
        "resguardante_destino_nombre": m.resguardante_destino_nombre,
        "cantidad": m.cantidad,
        "solicitado_por": solicitado.nombre if solicitado else None,
        "autorizado_por": autorizado.nombre if autorizado else None,
        "entregado_por": entregado.nombre if entregado else None,
        "recibido_por": recibido.nombre if recibido else None,
        "fecha_solicitud": m.fecha_solicitud.isoformat() if m.fecha_solicitud else None,
        "fecha_autorizacion": m.fecha_autorizacion.isoformat() if m.fecha_autorizacion else None,
        "fecha_entrega": m.fecha_entrega.isoformat() if m.fecha_entrega else None,
        "fecha_recepcion": m.fecha_recepcion.isoformat() if m.fecha_recepcion else None,
        "observaciones": m.observaciones,
        "evidencia_url": m.evidencia_url,
    }


def _serializar_solicitud_baja(s: SolicitudBajaInventario, db: Session) -> dict:
    activo     = db.query(Activo).filter(Activo.id == s.activo_id).first()
    solicitado = db.query(Usuario).filter(Usuario.id == s.solicitado_por_id).first()  if s.solicitado_por_id  else None
    revisado   = db.query(Usuario).filter(Usuario.id == s.revisado_por_id).first()    if s.revisado_por_id    else None
    validado   = db.query(Usuario).filter(Usuario.id == s.validado_por_id).first()    if s.validado_por_id    else None
    autorizado = db.query(Usuario).filter(Usuario.id == s.autorizado_por_id).first()  if s.autorizado_por_id  else None
    ejecutado  = db.query(Usuario).filter(Usuario.id == s.ejecutado_por_id).first()   if s.ejecutado_por_id   else None

    # Cuando no hay autorizador y el registro es anterior a v1.3, mostrar etiqueta informativa
    autorizado_label = (
        autorizado.nombre if autorizado
        else ("Previo a trazabilidad v1.3" if s.migrado_version else None)
    )
    return {
        "id": s.id,
        "activo_id": s.activo_id,
        "activo_nombre": activo.nombre if activo else None,
        "activo_codigo": activo.codigo_inventario if activo else None,
        "estado": s.estado,
        "motivo": s.motivo,
        "diagnostico": s.diagnostico,
        "evidencia_url": s.evidencia_url,
        "destino_final": s.destino_final,
        "observaciones": s.observaciones,
        "solicitado_por": solicitado.nombre if solicitado else None,
        "revisado_por": revisado.nombre if revisado else None,
        "validado_por": validado.nombre if validado else None,
        "autorizado_por": autorizado_label,
        "ejecutado_por": ejecutado.nombre if ejecutado else None,
        "fecha_solicitud":    s.fecha_solicitud.isoformat()    + "Z" if s.fecha_solicitud    else None,
        "fecha_revision":     s.fecha_revision.isoformat()     + "Z" if s.fecha_revision     else None,
        "fecha_validacion":   s.fecha_validacion.isoformat()   + "Z" if s.fecha_validacion   else None,
        "fecha_autorizacion": s.fecha_autorizacion.isoformat() + "Z" if s.fecha_autorizacion else None,
        "fecha_ejecucion":    s.fecha_ejecucion.isoformat()    + "Z" if s.fecha_ejecucion    else None,
        "migrado_version": s.migrado_version,
    }


def _serializar_levantamiento(l: LevantamientoInventario, db: Session) -> dict:
    dep = db.query(Departamento).filter(Departamento.id == l.departamento_id).first() if l.departamento_id else None
    lab = db.query(Laboratorio).filter(Laboratorio.id == l.laboratorio_id).first() if l.laboratorio_id else None
    creado = db.query(Usuario).filter(Usuario.id == l.creado_por_id).first() if l.creado_por_id else None
    revisiones = db.query(RevisionLevantamientoInventario).filter(RevisionLevantamientoInventario.levantamiento_id == l.id).all()
    return {
        "id": l.id,
        "nombre": l.nombre,
        "estado": l.estado,
        "departamento_id": l.departamento_id,
        "departamento_nombre": dep.nombre if dep else None,
        "laboratorio_id": l.laboratorio_id,
        "laboratorio_nombre": lab.nombre if lab else None,
        "creado_por": creado.nombre if creado else None,
        "fecha_inicio": l.fecha_inicio.isoformat() if l.fecha_inicio else None,
        "fecha_cierre": l.fecha_cierre.isoformat() if l.fecha_cierre else None,
        "observaciones": l.observaciones,
        "revisados": len(revisiones),
        "no_localizados": sum(1 for r in revisiones if r.estado == "NO_LOCALIZADO"),
        "propuestos_baja": sum(1 for r in revisiones if r.estado == "PROPUESTO_BAJA"),
        "revisiones": [_serializar_revision(r, db) for r in revisiones],
    }


def _serializar_revision(r: RevisionLevantamientoInventario, db: Session) -> dict:
    activo = db.query(Activo).filter(Activo.id == r.activo_id).first()
    usuario = db.query(Usuario).filter(Usuario.id == r.revisado_por_id).first() if r.revisado_por_id else None
    return {
        "id": r.id,
        "levantamiento_id": r.levantamiento_id,
        "activo_id": r.activo_id,
        "activo_nombre": activo.nombre if activo else None,
        "activo_codigo": activo.codigo_inventario if activo else None,
        "estado": r.estado,
        "ubicacion_reportada": r.ubicacion_reportada,
        "resguardante_reportado": r.resguardante_reportado,
        "observaciones": r.observaciones,
        "evidencia_url": r.evidencia_url,
        "revisado_por": usuario.nombre if usuario else None,
        "fecha_revision": r.fecha_revision.isoformat() if r.fecha_revision else None,
    }


def _query_activos_levantamiento(db: Session, l: LevantamientoInventario, current_user: Usuario):
    q = _query_activos_filtrados(
        db,
        current_user,
        laboratorio_id=l.laboratorio_id,
        departamento_id=l.departamento_id,
        solo_activos=True,
    )
    if q is None:
        return []
    return q.filter(
        or_(Activo.estado_admin == "VALIDADO", Activo.estado_admin.is_(None))
    ).order_by(Activo.categoria, Activo.nombre).all()


def _asegurar_acceso_levantamiento(db: Session, l: LevantamientoInventario | None, current_user: Usuario) -> None:
    if not l:
        raise HTTPException(status_code=404, detail="Levantamiento no encontrado")
    if _es_rol_laboratorio(current_user) and l.laboratorio_id != current_user.laboratorio_id:
        raise HTTPException(status_code=403, detail="No puedes gestionar levantamientos de otro laboratorio")
    _asegurar_write_inventario_departamento(db, current_user, l.departamento_id)


def _validar_destinos_movimiento(data: MovimientoInventarioCreate, db: Session):
    if data.tipo.upper() not in TIPOS_MOVIMIENTO:
        raise HTTPException(status_code=422, detail=f"Tipo de movimiento invalido. Use: {TIPOS_MOVIMIENTO}")
    if data.departamento_destino_id and not db.query(Departamento).filter(Departamento.id == data.departamento_destino_id).first():
        raise HTTPException(status_code=404, detail="Departamento destino no encontrado")
    if data.ubicacion_destino_id and not db.query(UbicacionInventario).filter(UbicacionInventario.id == data.ubicacion_destino_id, UbicacionInventario.activo == True).first():
        raise HTTPException(status_code=404, detail="Ubicacion destino no encontrada")
    if data.resguardante_destino_id and not db.query(Usuario).filter(Usuario.id == data.resguardante_destino_id, Usuario.activo == True).first():
        raise HTTPException(status_code=404, detail="Resguardante destino no encontrado")


def _aplicar_movimiento_recibido(activo: Activo, mov: MovimientoInventario):
    tipo = (mov.tipo or "").upper()
    if tipo == "BAJA":
        activo.estado_admin = "BAJA_SOLICITADA"
        activo.estado = "BAJA"
        activo.activo = False
        return
    if tipo == "MANTENIMIENTO":
        activo.estado = "MANTENIMIENTO"
        return
    if tipo in ("TRANSFERENCIA_DEPARTAMENTO", "CAMBIO_UBICACION", "CAMBIO_RESGUARDANTE", "PRESTAMO_TEMPORAL", "RETORNO", "AJUSTE_INVENTARIO"):
        if mov.departamento_destino_id is not None:
            activo.departamento_id = mov.departamento_destino_id
        if mov.ubicacion_destino_id is not None:
            activo.ubicacion_id = mov.ubicacion_destino_id
            activo.ubicacion_nombre = None
        elif mov.ubicacion_destino_nombre:
            activo.ubicacion_id = None
            activo.ubicacion_nombre = mov.ubicacion_destino_nombre
        if mov.resguardante_destino_id is not None:
            # Resguardante con cuenta SIGA: usar FK, limpiar texto libre
            activo.responsable_id = mov.resguardante_destino_id
            activo.resguardante_externo_nombre = None
        elif mov.resguardante_destino_nombre:
            # Resguardante externo sin cuenta: usar texto libre, limpiar FK
            activo.resguardante_externo_nombre = mov.resguardante_destino_nombre
            activo.responsable_id = None

def _serializar_prestamo(p: Prestamo, db: Session) -> dict:
    activo   = db.query(Activo).filter(Activo.id == p.activo_id).first()
    lab      = db.query(Laboratorio).filter(Laboratorio.id == activo.laboratorio_id).first() if activo else None
    ahora    = _utcnow()
    dias_vencido = None
    if p.estado == "VENCIDO" and p.fecha_retorno_esperada:
        dias_vencido = (ahora - p.fecha_retorno_esperada).days

    # Decodificar datos extras guardados en observaciones_salida como JSON simple
    receptor_tipo = "ALUMNO"
    proposito     = None
    notas         = None
    if p.observaciones_salida and p.observaciones_salida.startswith("__meta__"):
        try:
            import json
            meta = json.loads(p.observaciones_salida[8:])
            receptor_tipo = meta.get("receptor_tipo", "ALUMNO")
            proposito     = meta.get("proposito")
            notas         = meta.get("notas")
        except Exception:
            proposito = p.observaciones_salida

    return {
        "id": p.id,
        "folio": p.folio or f"PRE-{p.id:06d}",
        "activo_id": p.activo_id,
        "activo_nombre": activo.nombre if activo else None,
        "activo_codigo": activo.codigo_inventario if activo else None,
        "activo_categoria": activo.categoria if activo else None,
        "activo_lab": lab.nombre if lab else None,
        "laboratorio_id": activo.laboratorio_id if activo else None,
        # Nombres compatibles con el frontend
        "receptor_nombre": p.solicitante_nombre,
        "receptor_matricula": p.solicitante_id_escolar,
        "receptor_tipo": receptor_tipo,
        "proposito": proposito,
        "notas": notas,
        "fecha_prestamo": p.fecha_salida.isoformat() if p.fecha_salida else None,
        "fecha_devolucion_esperada": p.fecha_retorno_esperada.isoformat() if p.fecha_retorno_esperada else None,
        "fecha_devolucion_real": p.fecha_retorno_real.isoformat() if p.fecha_retorno_real else None,
        "estado": p.estado,
        "condicion_devolucion": p.condicion_retorno,
        "dias_vencido": dias_vencido,
    }


def _meta_prestamo(p: Prestamo) -> dict:
    """Extrae metadatos legibles guardados en observaciones_salida."""
    if not p.observaciones_salida:
        return {"receptor_tipo": "ALUMNO", "proposito": None, "descripcion": None}
    if not p.observaciones_salida.startswith("__meta__"):
        return {
            "receptor_tipo": "ALUMNO",
            "proposito": None,
            "descripcion": p.observaciones_salida,
        }
    try:
        import json
        meta = json.loads(p.observaciones_salida[8:])
        return {
            "receptor_tipo": meta.get("receptor_tipo", "ALUMNO"),
            "proposito": meta.get("proposito"),
            "descripcion": None,
        }
    except Exception:
        return {"receptor_tipo": "ALUMNO", "proposito": None, "descripcion": None}


def _serializar_incidente(i: Incidente, db: Session) -> dict:
    from models.laboratorio import Computadora
    from models.sesion import AsignacionPC
    activo    = db.query(Activo).filter(Activo.id == i.activo_id).first() if i.activo_id else None
    lab       = db.query(Laboratorio).filter(Laboratorio.id == i.laboratorio_id).first() if i.laboratorio_id else None
    # Para computadoras, intentar obtener datos básicos
    pc_info   = None
    if i.computadora_id:
        try:
            pc = db.query(Computadora).filter(Computadora.id == i.computadora_id).first()
            if pc:
                pc_info = {"codigo": pc.codigo, "fila": getattr(pc, "fila", None)}
                if not lab:
                    lab = db.query(Laboratorio).filter(Laboratorio.id == pc.laboratorio_id).first()
        except Exception:
            pass

    reporter  = db.query(Usuario).filter(Usuario.id == i.reportado_por_id).first() if i.reportado_por_id else None

    # ── Alumno responsable ────────────────────────────────────────────────────
    # SESION     → alumno que tenía la PC en ESA sesión (testigo directo, certeza ALTA)
    # RECEPCION  → alumno de la sesión ANTERIOR a la que recibió el daño (certeza MEDIA)
    alumno_responsable = None
    certeza            = None
    if i.origen == "SESION" and i.origen_id and i.computadora_id:
        certeza = "ALTA"
        try:
            asig = db.query(AsignacionPC).filter(
                AsignacionPC.sesion_id      == i.origen_id,
                AsignacionPC.computadora_id == i.computadora_id,
            ).order_by(AsignacionPC.hora_asignacion.desc()).first()
            if asig:
                alumno_responsable = {
                    "nombre":    asig.alumno_nombre,
                    "matricula": asig.alumno_matricula,
                }
        except Exception:
            pass
    elif i.origen == "RECEPCION" and i.computadora_id:
        certeza = "MEDIA"
        try:
            # Buscar la última AsignacionPC de sesiones DISTINTAS a la actual
            asig = db.query(AsignacionPC).filter(
                AsignacionPC.computadora_id == i.computadora_id,
                AsignacionPC.sesion_id      != (i.origen_id or 0),
            ).order_by(AsignacionPC.hora_asignacion.desc()).first()
            if asig:
                from models.sesion import SesionClase as SC
                sesion_ant = db.query(SC).filter(SC.id == asig.sesion_id).first()
                alumno_responsable = {
                    "nombre":          asig.alumno_nombre,
                    "matricula":       asig.alumno_matricula,
                    "sesion_anterior": {
                        "id":     sesion_ant.id           if sesion_ant else None,
                        "codigo": sesion_ant.codigo_sesion if sesion_ant else None,
                        "materia": sesion_ant.materia      if sesion_ant else None,
                        "inicio": sesion_ant.inicio.isoformat() if sesion_ant and sesion_ant.inicio else None,
                    },
                }
        except Exception:
            pass

    # ── Adeudo vinculado a este incidente ────────────────────────────────────
    adeudo_vinculado = db.query(Adeudo).filter(Adeudo.incidente_id == i.id).first()
    seguimientos = db.query(SeguimientoIncidente).filter(
        SeguimientoIncidente.incidente_id == i.id
    ).order_by(
        SeguimientoIncidente.creado_en.asc(),
        SeguimientoIncidente.id.asc(),
    ).all()

    return {
        "id":               i.id,
        "activo_id":        i.activo_id,
        "activo_nombre":    activo.nombre if activo else None,
        "activo_codigo":    activo.codigo_inventario if activo else None,
        "activo_categoria": activo.categoria if activo else None,
        "computadora_id":   i.computadora_id,
        "pc_codigo":        pc_info["codigo"] if pc_info else None,
        "laboratorio_id":   i.laboratorio_id or (lab.id if lab else None),
        "laboratorio_nombre": lab.nombre if lab else None,
        "origen":           i.origen,
        "origen_id":        i.origen_id,
        "tipo":             i.tipo,
        "descripcion":      i.descripcion,
        "reportado_por":    reporter.nombre if reporter else None,
        "fecha_reporte":    i.fecha_reporte.isoformat() if i.fecha_reporte else None,
        "estado":           i.estado,
        "prioridad":        i.prioridad,
        "notas_seguimiento": i.notas_seguimiento,
        "fecha_resolucion": i.fecha_resolucion.isoformat() if i.fecha_resolucion else None,
        "costo_reparacion": i.costo_reparacion,
        "cerrado":           i.estado in ESTADOS_INCIDENTE_CERRADOS,
        "seguimientos": [
            {
                "id": s.id,
                "tipo": s.tipo,
                "texto": s.texto,
                "estado_anterior": s.estado_anterior,
                "estado_nuevo": s.estado_nuevo,
                "creado_en": s.creado_en.isoformat() if s.creado_en else None,
                "usuario_id": s.usuario_id,
                "usuario_nombre": s.usuario.nombre if s.usuario else "Sistema",
            }
            for s in seguimientos
        ],
        "alumno_responsable": alumno_responsable,
        "certeza":            certeza,
        # Adeudo vinculado (si existe)
        "adeudo_id":          adeudo_vinculado.id     if adeudo_vinculado else None,
        "adeudo_estado":      adeudo_vinculado.estado if adeudo_vinculado else None,
        "adeudo_persona":     adeudo_vinculado.persona_nombre if adeudo_vinculado else None,
    }


# ─── Activos ───────────────────────────────────────────────────────────────────

@router.get("/ubicaciones", summary="Listar ubicaciones de inventario")
def listar_ubicaciones(
    activo: Optional[bool] = True,
    tipo: Optional[str] = None,
    departamento_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    q = db.query(UbicacionInventario)
    if activo is not None:
        q = q.filter(UbicacionInventario.activo == activo)
    if tipo:
        q = q.filter(UbicacionInventario.tipo == tipo.upper())
    if departamento_id:
        q = q.filter(UbicacionInventario.departamento_id == departamento_id)
    return [_serializar_ubicacion(u, db) for u in q.order_by(UbicacionInventario.edificio, UbicacionInventario.nombre).all()]


@router.post("/ubicaciones", status_code=status.HTTP_201_CREATED, summary="Crear ubicacion de inventario")
def crear_ubicacion(
    data: UbicacionInventarioIn,
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN)),
):
    tipo = data.tipo.upper()
    tipos_ubicacion_validos = _catalogo_claves(db, CATALOGO_TIPO_UBICACION)
    if tipo not in tipos_ubicacion_validos:
        raise HTTPException(status_code=422, detail=f"Tipo de ubicacion invalido. Use: {tipos_ubicacion_validos}")
    if data.departamento_id and not db.query(Departamento).filter(Departamento.id == data.departamento_id).first():
        raise HTTPException(status_code=404, detail="Departamento no encontrado")
    ubicacion = UbicacionInventario(
        nombre=data.nombre.strip(),
        tipo=tipo,
        edificio=data.edificio.strip() if data.edificio else None,
        piso=data.piso.strip() if data.piso else None,
        referencia=data.referencia.strip() if data.referencia else None,
        departamento_id=data.departamento_id,
        activo=data.activo,
        creado_en=_utcnow(),
    )
    db.add(ubicacion)
    db.commit()
    db.refresh(ubicacion)
    return _serializar_ubicacion(ubicacion, db)


@router.get("/departamentos-opciones", summary="Departamentos disponibles para inventario")
def departamentos_opciones_inventario(
    modo: str = "lectura",
    activo: bool = True,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    modo = (modo or "lectura").lower()
    if modo not in ("lectura", "escritura"):
        raise HTTPException(status_code=422, detail="modo debe ser lectura o escritura")

    scope_global = _es_admin_inventario_global(current_user) or puede_validar_inventario(db, current_user)
    query = db.query(Departamento)
    if activo is not None:
        query = query.filter(Departamento.activo == activo)

    if not scope_global:
        if modo == "escritura":
            ids = departamentos_inventario(db, current_user)
        else:
            ids = _departamentos_visibles_inventario(db, current_user) or []
        if not ids:
            return {"items": [], "scope_global": False, "modo": modo}
        query = query.filter(Departamento.id.in_(ids))

    items = query.order_by(Departamento.nombre).all()
    return {
        "items": [_serializar_departamento_opcion(dep) for dep in items],
        "scope_global": scope_global,
        "modo": modo,
    }


@router.get("/resguardantes-opciones", summary="Usuarios disponibles como resguardantes de inventario")
def resguardantes_opciones(
    departamento_id: Optional[int] = None,
    laboratorio_id: Optional[int] = None,
    buscar: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    q = db.query(Usuario).filter(Usuario.activo == True)
    if _es_rol_laboratorio(current_user):
        if not current_user.laboratorio_id:
            return []
        q = q.filter(Usuario.laboratorio_id == current_user.laboratorio_id)
    else:
        departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
        if departamentos_visibles is not None:
            if not departamentos_visibles:
                return []
            q = q.filter(Usuario.departamento_id.in_(departamentos_visibles))
    if departamento_id:
        _asegurar_write_inventario_departamento(db, current_user, departamento_id)
        q = q.filter(Usuario.departamento_id == departamento_id)
    if laboratorio_id:
        assert_lab_write(laboratorio_id, current_user)
        q = q.filter(Usuario.laboratorio_id == laboratorio_id)
    if buscar:
        like = f"%{buscar.strip()}%"
        q = q.filter((Usuario.nombre.ilike(like)) | (Usuario.email.ilike(like)) | (Usuario.numero_empleado.ilike(like)))
    usuarios = q.order_by(Usuario.nombre).limit(100).all()
    return [_serializar_usuario_resguardo(u, db) for u in usuarios]


@router.get("/activos", summary="Listar activos")
def listar_activos(
    laboratorio_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    ubicacion_id: Optional[int] = None,
    alcance: Optional[str]         = None,
    tipo_inventario: Optional[str] = None,
    estado_admin: Optional[str]    = None,
    categoria: Optional[str]      = None,
    estado: Optional[str]         = None,
    solo_activos: bool             = True,
    solo_disponibles: bool         = False,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    _actualizar_estado_prestamos(db)
    q = _query_activos_filtrados(
        db, current_user, laboratorio_id, departamento_id, ubicacion_id,
        alcance, tipo_inventario, estado_admin, categoria, estado, solo_activos,
    )
    if q is None:
        return []

    activos = q.order_by(Activo.categoria, Activo.nombre).all()
    result  = [_serializar_activo(a, db) for a in activos]

    if solo_disponibles:
        result = [
            a for a in result
            if not a["prestado"] and (a.get("estado_admin") or "VALIDADO") == "VALIDADO"
        ]
    return result


@router.get("/activos/exportar", summary="Exportar corte de inventario a Excel")
def exportar_activos(
    request: Request,
    laboratorio_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    ubicacion_id: Optional[int] = None,
    alcance: Optional[str]         = None,
    tipo_inventario: Optional[str] = None,
    estado_admin: Optional[str]    = None,
    categoria: Optional[str]      = None,
    estado: Optional[str]         = None,
    buscar: Optional[str]         = None,
    solo_activos: bool             = True,
    solo_disponibles: bool         = False,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _actualizar_estado_prestamos(db)
    q = _query_activos_filtrados(
        db, current_user, laboratorio_id, departamento_id, ubicacion_id,
        alcance, tipo_inventario, estado_admin, categoria, estado, solo_activos,
    )
    if q is not None and not estado_admin:
        q = q.filter(or_(Activo.estado_admin == "VALIDADO", Activo.estado_admin.is_(None)))
    activos = [] if q is None else q.order_by(Activo.categoria, Activo.nombre).all()
    filas = [_serializar_activo(a, db) for a in activos]
    if solo_disponibles:
        filas = [a for a in filas if not a["prestado"]]
    if buscar:
        q_buscar = _normalizar(buscar)
        campos_busqueda = [
            "nombre", "codigo_inventario", "numero_oficial", "marca", "modelo",
            "numero_serie", "departamento_nombre", "laboratorio_nombre",
            "ubicacion_label", "responsable_nombre", "resguardante_externo_nombre",
        ]
        filas = [
            item for item in filas
            if q_buscar in _normalizar(" ".join(str(item.get(c) or "") for c in campos_busqueda))
        ]

    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws = wb.create_sheet("Inventario")

    green = "007A53"
    header_fill = PatternFill("solid", fgColor=green)
    soft_fill = PatternFill("solid", fgColor="E7F5EE")
    border = Border(bottom=Side(style="thin", color="D9E2EC"))

    ws_resumen["A1"] = "Corte de inventario"
    ws_resumen["A1"].font = Font(size=16, bold=True, color=green)
    ws_resumen["A2"] = "Este archivo es una copia de control para revision fisica y administrativa; no es un respaldo restaurable del sistema."
    ws_resumen["A2"].font = Font(italic=True, color="64748B")

    lab_filtro = db.query(Laboratorio).filter(Laboratorio.id == laboratorio_id).first() if laboratorio_id else None
    dep_filtro = db.query(Departamento).filter(Departamento.id == departamento_id).first() if departamento_id else None

    filtros = {
        "Generado por": current_user.nombre,
        "Fecha": _utcnow().strftime("%Y-%m-%d %H:%M"),
        "Laboratorio": lab_filtro.nombre if lab_filtro else "Todos los visibles",
        "Departamento": dep_filtro.nombre if dep_filtro else "Todos los visibles",
        "Alcance": alcance or "Todos",
        "Categoria": categoria or "Todas",
        "Estado operativo": estado or "Todos",
        "Estado administrativo": estado_admin or "Todos",
        "Busqueda": buscar or "Sin busqueda",
        "Total exportado": len(filas),
    }
    for row, (label, value) in enumerate(filtros.items(), start=4):
        ws_resumen.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2, value=value)
        ws_resumen.cell(row=row, column=1).fill = soft_fill
        ws_resumen.cell(row=row, column=1).border = border
        ws_resumen.cell(row=row, column=2).border = border
    ws_resumen.column_dimensions["A"].width = 26
    ws_resumen.column_dimensions["B"].width = 55

    columnas = [
        ("#", lambda a, i: i),
        ("Codigo SIGA", lambda a, i: a.get("codigo_inventario")),
        ("No. oficial", lambda a, i: a.get("numero_oficial")),
        ("Nombre", lambda a, i: a.get("nombre")),
        ("Categoria", lambda a, i: a.get("categoria")),
        ("Estado operativo", lambda a, i: a.get("estado")),
        ("Estado administrativo", lambda a, i: a.get("estado_admin")),
        ("Alcance", lambda a, i: a.get("alcance")),
        ("Departamento", lambda a, i: a.get("departamento_nombre")),
        ("Laboratorio", lambda a, i: a.get("laboratorio_nombre")),
        ("Ubicacion", lambda a, i: a.get("ubicacion_label") or a.get("ubicacion_nombre")),
        ("Resguardante", lambda a, i: a.get("responsable_nombre") or a.get("resguardante_externo_nombre")),
        ("Marca", lambda a, i: a.get("marca")),
        ("Modelo", lambda a, i: a.get("modelo")),
        ("Serie", lambda a, i: a.get("numero_serie")),
        ("Area", lambda a, i: a.get("area")),
        ("Valor", lambda a, i: a.get("valor")),
        ("Prestado", lambda a, i: "SI" if a.get("prestado") else "NO"),
        ("Observaciones", lambda a, i: a.get("observaciones")),
    ]
    ws.append([c[0] for c in columnas])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for idx, item in enumerate(filas, start=1):
        ws.append([getter(item, idx) for _, getter in columnas])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [8, 22, 18, 34, 24, 18, 22, 18, 32, 30, 34, 30, 18, 22, 22, 16, 14, 12, 42]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    registrar(
        db,
        Accion.EXPORTAR_INVENTARIO,
        Recurso.ACTIVO,
        usuario=current_user,
        detalle={
            "total": len(filas),
            "laboratorio_id": laboratorio_id,
            "departamento_id": departamento_id,
            "categoria": categoria,
            "estado": estado,
        },
        request=request,
    )

    filename = f"corte_inventario_{_utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/activos/etiquetas", summary="Generar etiquetas QR de inventario")
def generar_etiquetas_activos(
    request: Request,
    ids: Optional[str] = None,
    laboratorio_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    ubicacion_id: Optional[int] = None,
    alcance: Optional[str]         = None,
    tipo_inventario: Optional[str] = None,
    estado_admin: Optional[str]    = None,
    categoria: Optional[str]       = None,
    estado: Optional[str]          = None,
    buscar: Optional[str]          = None,
    solo_activos: bool             = True,
    solo_disponibles: bool         = False,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _actualizar_estado_prestamos(db)
    if ids:
        try:
            ids_int = [int(x.strip()) for x in ids.split(",") if x.strip()]
        except ValueError:
            raise HTTPException(status_code=422, detail="ids debe ser una lista de numeros separados por coma")
        if not ids_int:
            raise HTTPException(status_code=422, detail="Selecciona al menos un activo")
        activos = db.query(Activo).filter(Activo.id.in_(ids_int)).order_by(Activo.categoria, Activo.nombre).all()
        for activo in activos:
            _asegurar_acceso_activo_departamental(db, activo, current_user)
            _asegurar_activo_validado(activo)
    else:
        q = _query_activos_filtrados(
            db, current_user, laboratorio_id, departamento_id, ubicacion_id,
            alcance, tipo_inventario, estado_admin, categoria, estado, solo_activos,
        )
        if q is not None:
            q = q.filter(or_(Activo.estado_admin == "VALIDADO", Activo.estado_admin.is_(None)))
        activos = [] if q is None else q.order_by(Activo.categoria, Activo.nombre).limit(200).all()

    filas = [_serializar_activo(a, db) for a in activos]
    if solo_disponibles:
        filas = [a for a in filas if not a["prestado"]]
    if buscar:
        q_buscar = _normalizar(buscar)
        campos_busqueda = [
            "nombre", "codigo_inventario", "numero_oficial", "marca", "modelo",
            "numero_serie", "departamento_nombre", "laboratorio_nombre",
            "ubicacion_label", "responsable_nombre", "resguardante_externo_nombre",
        ]
        filas = [
            item for item in filas
            if q_buscar in _normalizar(" ".join(str(item.get(c) or "") for c in campos_busqueda))
        ]
    if len(filas) > 200:
        filas = filas[:200]
    if not filas:
        raise HTTPException(status_code=404, detail="No hay activos para generar etiquetas")

    output = _generar_pdf_etiquetas(filas)
    registrar(
        db,
        Accion.EXPORTAR_INVENTARIO,
        Recurso.ACTIVO,
        usuario=current_user,
        detalle={
            "tipo": "etiquetas_qr",
            "total": len(filas),
            "ids": ids,
            "laboratorio_id": laboratorio_id,
            "departamento_id": departamento_id,
        },
        request=request,
    )
    filename = f"etiquetas_inventario_{_utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/activos/{activo_id}/etiqueta", summary="Generar etiqueta QR del activo")
def generar_etiqueta_activo(
    request: Request,
    activo_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    activo = db.query(Activo).filter(Activo.id == activo_id).first()
    _asegurar_acceso_activo_departamental(db, activo, current_user)
    _asegurar_activo_validado(activo)
    data = _serializar_activo(activo, db)
    output = _generar_pdf_etiquetas([data])
    registrar(
        db,
        Accion.EXPORTAR_INVENTARIO,
        Recurso.ACTIVO,
        usuario=current_user,
        detalle={"tipo": "etiqueta_qr", "activo_id": activo_id, "codigo": activo.codigo_inventario},
        request=request,
    )
    filename = f"etiqueta_{activo.codigo_inventario}.pdf".replace("/", "-")
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/activos/{activo_id}/resguardo", summary="Descargar formato de resguardo del activo")
def descargar_resguardo_activo(
    request: Request,
    activo_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    activo = db.query(Activo).filter(Activo.id == activo_id).first()
    _asegurar_acceso_activo_departamental(db, activo, current_user)
    _asegurar_activo_validado(activo)

    data = _serializar_activo(activo, db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resguardo"

    green = "007A53"
    soft = "E7F5EE"
    border = Border(bottom=Side(style="thin", color="D9E2EC"))

    ws.merge_cells("A1:D1")
    ws["A1"] = "Formato de resguardo de activo"
    ws["A1"].font = Font(size=16, bold=True, color=green)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = "Control interno de entrega/recepcion de inventario institucional"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="64748B")

    filas = [
        ("Codigo SIGA", data.get("codigo_inventario")),
        ("No. oficial/patrimonial", data.get("numero_oficial")),
        ("Nombre del activo", data.get("nombre")),
        ("Categoria", data.get("categoria")),
        ("Marca / modelo", " / ".join([v for v in [data.get("marca"), data.get("modelo")] if v])),
        ("No. serie", data.get("numero_serie")),
        ("Estado operativo", data.get("estado")),
        ("Estado administrativo", data.get("estado_admin")),
        ("Departamento/Laboratorio", data.get("departamento_nombre") or data.get("laboratorio_nombre")),
        ("Ubicacion registrada", data.get("ubicacion_label") or data.get("ubicacion_nombre")),
        ("Resguardante", data.get("responsable_nombre") or data.get("resguardante_externo_nombre")),
        ("Valor", data.get("valor")),
        ("Fecha de emision", _utcnow().strftime("%Y-%m-%d %H:%M")),
        ("Emitido por", current_user.nombre),
    ]
    start = 4
    for idx, (label, value) in enumerate(filas, start=start):
        ws.cell(idx, 1, label).font = Font(bold=True)
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor=soft)
        ws.cell(idx, 2, value if value is not None else "")
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
        for col in range(1, 5):
            ws.cell(idx, col).border = border

    obs_row = start + len(filas) + 2
    ws.merge_cells(start_row=obs_row, start_column=1, end_row=obs_row, end_column=4)
    ws.cell(obs_row, 1, "Observaciones de entrega").font = Font(bold=True)
    ws.cell(obs_row + 1, 1, "")
    ws.merge_cells(start_row=obs_row + 1, start_column=1, end_row=obs_row + 4, end_column=4)
    ws.cell(obs_row + 1, 1).alignment = Alignment(vertical="top", wrap_text=True)

    sign_row = obs_row + 7
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=2)
    ws.merge_cells(start_row=sign_row, start_column=3, end_row=sign_row, end_column=4)
    ws.cell(sign_row, 1, "Entrega").alignment = Alignment(horizontal="center")
    ws.cell(sign_row, 3, "Recibe / resguardante").alignment = Alignment(horizontal="center")
    ws.cell(sign_row, 1).font = Font(bold=True)
    ws.cell(sign_row, 3).font = Font(bold=True)
    ws.merge_cells(start_row=sign_row + 4, start_column=1, end_row=sign_row + 4, end_column=2)
    ws.merge_cells(start_row=sign_row + 4, start_column=3, end_row=sign_row + 4, end_column=4)
    ws.cell(sign_row + 4, 1, "Nombre y firma").alignment = Alignment(horizontal="center")
    ws.cell(sign_row + 4, 3, "Nombre y firma").alignment = Alignment(horizontal="center")

    for col, width in {"A": 24, "B": 28, "C": 24, "D": 28}.items():
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    registrar(
        db,
        Accion.EXPORTAR_INVENTARIO,
        Recurso.ACTIVO,
        usuario=current_user,
        recurso_id=activo.id,
        detalle={"formato": "RESGUARDO_ACTIVO", "codigo": activo.codigo_inventario},
        request=request,
    )
    filename = f"resguardo_{activo.codigo_inventario}.xlsx".replace("/", "-")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/activos", status_code=status.HTTP_201_CREATED, summary="Registrar activo")
def crear_activo(
    request: Request,
    data: ActivoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    # RLS: LAB_ADMIN solo puede crear activos en su propio laboratorio
    alcance = (data.alcance or "LABORATORIO").upper()
    if alcance not in ALCANCES_ACTIVO:
        raise HTTPException(status_code=422, detail=f"Alcance invalido. Use: {ALCANCES_ACTIVO}")
    if _es_rol_laboratorio(current_user):
        if alcance != "LABORATORIO":
            raise HTTPException(status_code=403, detail="Solo Super Admin puede crear activos institucionales")
        if not current_user.laboratorio_id:
            raise HTTPException(status_code=403, detail="No tienes laboratorio asignado")
        if data.laboratorio_id and data.laboratorio_id != current_user.laboratorio_id:
            raise HTTPException(status_code=403, detail="Solo puedes crear activos en tu laboratorio")
        data.laboratorio_id = current_user.laboratorio_id
    if current_user.rol == RolUsuario.LAB_ADMIN and alcance != "LABORATORIO":
        raise HTTPException(status_code=403, detail="Solo Super Admin puede crear activos institucionales")
    # La adscripcion de un activo de laboratorio la define el laboratorio.
    # El departamento solo aplica al inventario institucional.
    departamento_operacion_id = None if alcance == "LABORATORIO" else data.departamento_id
    if not _es_admin_inventario_global(current_user):
        if not departamentos_inventario(db, current_user):
            raise HTTPException(status_code=403, detail="No tienes permiso para gestionar inventario departamental")
        departamento_operacion_id = _resolver_departamento_inventario(db, current_user, data.departamento_id)
        _asegurar_write_inventario_departamento(db, current_user, departamento_operacion_id)
        if alcance != "INSTITUCIONAL":
            raise HTTPException(status_code=403, detail="El inventario departamental debe registrarse como institucional")
    tipo_inventario = (data.tipo_inventario or "ACTIVO").upper()
    if tipo_inventario not in TIPOS_INVENTARIO:
        raise HTTPException(status_code=422, detail=f"Tipo de inventario invalido. Use: {TIPOS_INVENTARIO}")
    estado_admin = (data.estado_admin or "VALIDADO").upper()
    if estado_admin not in ESTADOS_ADMIN_ACTIVO:
        raise HTTPException(status_code=422, detail=f"Estado administrativo invalido. Use: {ESTADOS_ADMIN_ACTIVO}")
    if not puede_validar_inventario(db, current_user):
        estado_admin = "BORRADOR"
    unidad_medida = (data.unidad_medida or "PIEZA").upper()
    if unidad_medida not in UNIDADES_MEDIDA:
        raise HTTPException(status_code=422, detail=f"Unidad de medida invalida. Use: {UNIDADES_MEDIDA}")

    categorias_validas = _catalogo_claves(db, CATALOGO_TIPO_CATEGORIA)
    if data.categoria.upper() not in categorias_validas:
        raise HTTPException(status_code=422, detail=f"Categoria invalida. Use: {categorias_validas}")
    if alcance == "LABORATORIO":
        if not data.laboratorio_id:
            raise HTTPException(status_code=422, detail="laboratorio_id es requerido para activos de laboratorio")
        assert_lab_write(data.laboratorio_id, current_user)
        if not db.query(Laboratorio).filter(Laboratorio.id == data.laboratorio_id, Laboratorio.activo == True).first():
            raise HTTPException(status_code=404, detail="Laboratorio no encontrado")
    elif data.laboratorio_id:
        assert_lab_write(data.laboratorio_id, current_user)
    if departamento_operacion_id and not db.query(Departamento).filter(Departamento.id == departamento_operacion_id).first():
        raise HTTPException(status_code=404, detail="Departamento no encontrado")
    if data.ubicacion_id and not db.query(UbicacionInventario).filter(UbicacionInventario.id == data.ubicacion_id, UbicacionInventario.activo == True).first():
        raise HTTPException(status_code=404, detail="Ubicacion no encontrada")
    if data.responsable_id and not db.query(Usuario).filter(Usuario.id == data.responsable_id, Usuario.activo == True).first():
        raise HTTPException(status_code=404, detail="Responsable no encontrado")

    # Auto-generar código si no se proporcionó
    area_codigo = data.area
    if not area_codigo and departamento_operacion_id:
        dep = db.query(Departamento).filter(Departamento.id == departamento_operacion_id).first()
        area_codigo = dep.clave if dep else None
    codigo = data.codigo_inventario or _generar_codigo(db, data.categoria, area_codigo)

    if db.query(Activo).filter(Activo.codigo_inventario == codigo).first():
        raise HTTPException(status_code=409, detail=f"Ya existe un activo con código '{codigo}'")

    payload = data.model_dump(exclude={"codigo_inventario"})
    if payload.get("numero_oficial"):
        payload["numero_oficial"] = payload["numero_oficial"].strip()
        if db.query(Activo).filter(Activo.numero_oficial == payload["numero_oficial"]).first():
            raise HTTPException(status_code=409, detail=f"Ya existe un activo con numero oficial '{payload['numero_oficial']}'")
    else:
        payload["numero_oficial"] = None
    payload["categoria"] = payload["categoria"].upper()
    payload["alcance"] = alcance
    payload["departamento_id"] = departamento_operacion_id
    payload["tipo_inventario"] = tipo_inventario
    payload["estado_admin"] = estado_admin
    payload["unidad_medida"] = unidad_medida
    # Patrimonial siempre: cantidad=1, PIEZA, sin stock mínimo (blindado también en schema)
    payload["cantidad"] = 1.0
    payload["unidad_medida"] = "PIEZA"
    payload["stock_minimo"] = None
    # Resguardante: si viene responsable_id, limpiar texto libre y viceversa
    if payload.get("responsable_id"):
        payload["resguardante_externo_nombre"] = None
    elif payload.get("resguardante_externo_nombre"):
        payload["responsable_id"] = None
    if payload.get("ubicacion_tipo"):
        payload["ubicacion_tipo"] = payload["ubicacion_tipo"].upper()
        tipos_ubicacion_validos = _catalogo_claves(db, CATALOGO_TIPO_UBICACION)
        if payload["ubicacion_tipo"] not in tipos_ubicacion_validos:
            raise HTTPException(status_code=422, detail=f"Tipo de ubicacion invalido. Use: {tipos_ubicacion_validos}")
    a = Activo(**payload, codigo_inventario=codigo, fecha_adquisicion=_utcnow())
    db.add(a)
    db.commit()
    db.refresh(a)
    registrar(db, accion=Accion.CREAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=a.id,
              detalle={"codigo": a.codigo_inventario, "nombre": a.nombre, "categoria": a.categoria},
              request=request)
    if (a.estado_admin or "").upper() in ("BORRADOR", "EN_REVISION", "OBSERVADO"):
        from routers.notificaciones import crear_notificacion
        for usuario_id in _destinatarios_revision_institucional(db, current_user.id):
            crear_notificacion(
                db,
                usuario_id,
                "INVENTARIO_REVISION",
                "Activo pendiente de revision",
                f"{a.codigo_inventario} · {a.nombre} fue registrado y requiere validacion institucional.",
                f"/admin/inventario?tab=revision&estado_admin={a.estado_admin}&buscar={a.codigo_inventario}",
                enviar_email=False,
            )
        db.commit()
    return _serializar_activo(a, db)


@router.get("/activos/{activo_id}", summary="Detalle de activo")
def obtener_activo(
    activo_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    a = db.query(Activo).filter(Activo.id == activo_id).first()
    # RLS: LAB_ADMIN no puede ver activos de otros laboratorios (devuelve 404)
    assert_resource_access(a, current_user)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None and a.departamento_id not in departamentos_visibles:
        raise HTTPException(status_code=404, detail="Recurso no encontrado")
    historial = db.query(Prestamo).filter(
        Prestamo.activo_id == activo_id
    ).order_by(Prestamo.fecha_salida.desc()).limit(10).all()
    detalle = _serializar_activo(a, db)
    detalle["historial_prestamos"] = [_serializar_prestamo(p, db) for p in historial]
    return detalle


@router.put("/activos/{activo_id}", summary="Editar activo")
def editar_activo(
    request: Request,
    activo_id: int,
    data: ActivoUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    a = db.query(Activo).filter(Activo.id == activo_id).first()
    # RLS: LAB_ADMIN solo puede editar activos de su laboratorio
    _asegurar_acceso_activo_departamental(db, a, current_user)
    _asegurar_write_inventario_departamento(db, current_user, a.departamento_id)
    if (a.estado_admin or "VALIDADO").upper() == "RECHAZADO" and current_user.rol != RolUsuario.SUPER_ADMIN:
        raise HTTPException(
            status_code=409,
            detail=(
                "El activo no fue autorizado y ya no puede corregirse. "
                "Super Admin debe reabrirlo como borrador o en revision."
            ),
        )
    campos = data.model_dump(exclude_unset=True)
    campos_con_trazabilidad = {
        "alcance",
        "laboratorio_id",
        "departamento_id",
        "ubicacion_id",
        "ubicacion_tipo",
        "ubicacion_nombre",
        "responsable_id",
        "resguardante_externo_nombre",
    }
    cambios_trazables = [
        campo for campo in campos_con_trazabilidad
        if campo in campos and campos[campo] != getattr(a, campo)
    ]
    estado_admin_actual = (a.estado_admin or "VALIDADO").upper()
    permite_correccion_directa = estado_admin_actual in ("BORRADOR", "EN_REVISION", "OBSERVADO")
    if cambios_trazables and not permite_correccion_directa:
        raise HTTPException(
            status_code=409,
            detail=(
                "Los cambios de adscripcion, departamento, ubicacion o resguardante "
                "deben registrarse desde Movimiento para conservar el historial."
            ),
        )
    if "numero_oficial" in campos:
        campos["numero_oficial"] = (campos["numero_oficial"] or "").strip() or None
        if campos["numero_oficial"]:
            duplicado = db.query(Activo).filter(
                Activo.numero_oficial == campos["numero_oficial"],
                Activo.id != activo_id,
            ).first()
            if duplicado:
                raise HTTPException(status_code=409, detail=f"Ya existe un activo con numero oficial '{campos['numero_oficial']}'")
    if "categoria" in campos:
        campos["categoria"] = campos["categoria"].upper()
        categorias_validas = _catalogo_claves(db, CATALOGO_TIPO_CATEGORIA)
        if campos["categoria"] not in categorias_validas:
            raise HTTPException(status_code=422, detail=f"Categoria invalida. Use: {categorias_validas}")
    if "alcance" in campos:
        campos["alcance"] = campos["alcance"].upper()
        if campos["alcance"] not in ALCANCES_ACTIVO:
            raise HTTPException(status_code=422, detail=f"Alcance invalido. Use: {ALCANCES_ACTIVO}")
        if _es_rol_laboratorio(current_user) and campos["alcance"] != "LABORATORIO":
            raise HTTPException(status_code=403, detail="Solo Super Admin puede convertir activos a institucionales")
        if (
            not _es_admin_inventario_global(current_user)
            and not puede_validar_inventario(db, current_user)
            and campos["alcance"] != "INSTITUCIONAL"
        ):
            raise HTTPException(status_code=403, detail="El inventario departamental debe mantenerse como institucional")
    if "tipo_inventario" in campos:
        campos["tipo_inventario"] = campos["tipo_inventario"].upper()
        if campos["tipo_inventario"] not in TIPOS_INVENTARIO:
            raise HTTPException(status_code=422, detail=f"Tipo de inventario invalido. Use: {TIPOS_INVENTARIO}")
    if "estado_admin" in campos:
        if not puede_validar_inventario(db, current_user):
            raise HTTPException(
                status_code=403,
                detail="Solo Inventario Institucional puede cambiar el estado administrativo del activo",
            )
        campos["estado_admin"] = campos["estado_admin"].upper()
        if campos["estado_admin"] not in ESTADOS_ADMIN_ACTIVO:
            raise HTTPException(status_code=422, detail=f"Estado administrativo invalido. Use: {ESTADOS_ADMIN_ACTIVO}")
    if "unidad_medida" in campos:
        campos["unidad_medida"] = campos["unidad_medida"].upper()
        if campos["unidad_medida"] not in UNIDADES_MEDIDA:
            raise HTTPException(status_code=422, detail=f"Unidad de medida invalida. Use: {UNIDADES_MEDIDA}")
    destino_lab = campos.get("laboratorio_id", a.laboratorio_id)
    destino_alcance = campos.get("alcance", a.alcance or "LABORATORIO")
    if destino_alcance == "LABORATORIO":
        if not destino_lab:
            raise HTTPException(status_code=422, detail="laboratorio_id es requerido para activos de laboratorio")
        assert_lab_write(destino_lab, current_user)
        if not db.query(Laboratorio).filter(Laboratorio.id == destino_lab, Laboratorio.activo == True).first():
            raise HTTPException(status_code=404, detail="Laboratorio no encontrado")
    if "departamento_id" in campos:
        if campos["departamento_id"]:
            if not db.query(Departamento).filter(Departamento.id == campos["departamento_id"]).first():
                raise HTTPException(status_code=404, detail="Departamento no encontrado")
            _asegurar_write_inventario_departamento(db, current_user, campos["departamento_id"])
        elif not _es_admin_inventario_global(current_user):
            raise HTTPException(status_code=422, detail="El activo departamental requiere departamento")
    if "ubicacion_id" in campos and campos["ubicacion_id"] and not db.query(UbicacionInventario).filter(UbicacionInventario.id == campos["ubicacion_id"], UbicacionInventario.activo == True).first():
        raise HTTPException(status_code=404, detail="Ubicacion no encontrada")
    if "responsable_id" in campos and campos["responsable_id"] and not db.query(Usuario).filter(Usuario.id == campos["responsable_id"], Usuario.activo == True).first():
        raise HTTPException(status_code=404, detail="Responsable no encontrado")
    if "ubicacion_tipo" in campos and campos["ubicacion_tipo"]:
        campos["ubicacion_tipo"] = campos["ubicacion_tipo"].upper()
        tipos_ubicacion_validos = _catalogo_claves(db, CATALOGO_TIPO_UBICACION)
        if campos["ubicacion_tipo"] not in tipos_ubicacion_validos:
            raise HTTPException(status_code=422, detail=f"Tipo de ubicacion invalido. Use: {tipos_ubicacion_validos}")
    if "ubicacion_nombre" in campos:
        campos["ubicacion_nombre"] = (campos["ubicacion_nombre"] or "").strip() or None
    if "resguardante_externo_nombre" in campos:
        campos["resguardante_externo_nombre"] = (campos["resguardante_externo_nombre"] or "").strip() or None
    if campos.get("ubicacion_id"):
        campos["ubicacion_nombre"] = None
    elif campos.get("ubicacion_nombre"):
        campos["ubicacion_id"] = None
    if campos.get("responsable_id"):
        campos["resguardante_externo_nombre"] = None
    elif campos.get("resguardante_externo_nombre"):
        campos["responsable_id"] = None
    tipo_final = campos.get("tipo_inventario", a.tipo_inventario or "ACTIVO")
    if tipo_final == "ACTIVO":
        campos["cantidad"] = 1
        campos["stock_minimo"] = None
    for campo, valor in campos.items():
        setattr(a, campo, valor)
    db.commit()
    db.refresh(a)
    registrar(db, accion=Accion.EDITAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=a.id,
              detalle={"campos": list(campos.keys())},
              request=request)
    return _serializar_activo(a, db)


@router.post("/activos/{activo_id}/validacion", summary="Cambiar estado administrativo del activo")
def cambiar_validacion_activo(
    request: Request,
    activo_id: int,
    data: ActivoValidacionUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    if not puede_validar_inventario(db, current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo Super Admin o Inventario Institucional puede revisar y validar activos",
        )
    a = db.query(Activo).filter(Activo.id == activo_id).first()
    _asegurar_acceso_activo_departamental(db, a, current_user)
    _asegurar_write_inventario_departamento(db, current_user, a.departamento_id)

    estado = (data.estado_admin or "").upper()
    estados_validacion = ["BORRADOR", "EN_REVISION", "OBSERVADO", "VALIDADO", "RECHAZADO"]
    if estado not in estados_validacion:
        raise HTTPException(status_code=422, detail=f"Estado de validacion invalido. Use: {estados_validacion}")
    observaciones = (data.observaciones or "").strip()
    if estado in ("OBSERVADO", "RECHAZADO") and not observaciones:
        raise HTTPException(
            status_code=422,
            detail="Debes indicar el motivo para observar o rechazar el activo",
        )

    anterior = a.estado_admin or "VALIDADO"
    a.estado_admin = estado
    db.commit()
    db.refresh(a)
    registrar(
        db,
        accion=Accion.EDITAR_ACTIVO,
        recurso=Recurso.ACTIVO,
        usuario=current_user,
        recurso_id=a.id,
        detalle={
            "flujo": "VALIDACION_INVENTARIO",
            "estado_anterior": anterior,
            "estado_nuevo": estado,
            "observaciones": observaciones or None,
        },
        request=request,
    )
    from routers.notificaciones import crear_notificacion
    etiquetas = {
        "BORRADOR": ("Activo devuelto a borrador", "quedo nuevamente en borrador"),
        "EN_REVISION": ("Activo en revision", "esta siendo revisado"),
        "OBSERVADO": ("Activo observado", "requiere correcciones"),
        "VALIDADO": ("Activo validado", "fue validado oficialmente"),
        "RECHAZADO": ("Activo no autorizado", "no fue autorizado"),
    }
    titulo, resultado = etiquetas[estado]
    mensaje = f"{a.codigo_inventario} · {a.nombre} {resultado}."
    if observaciones:
        mensaje += f" Motivo: {observaciones}"
    for usuario_id in _destinatarios_validacion_activo(a, db):
        if usuario_id == current_user.id:
            continue
        crear_notificacion(
            db,
            usuario_id,
            "INVENTARIO_VALIDACION",
            titulo,
            mensaje,
            f"/admin/inventario?estado_admin={estado}&buscar={a.codigo_inventario}",
            enviar_email=False,
        )
    db.commit()
    return _serializar_activo(a, db)


@router.delete("/activos/{activo_id}", summary="Dar de baja activo")
def dar_baja_activo(
    request: Request,
    activo_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB))
):
    a = db.query(Activo).filter(Activo.id == activo_id).first()
    # RLS: LAB_ADMIN solo puede dar de baja activos de su laboratorio
    assert_resource_access(a, current_user)
    _asegurar_activo_validado(a)
    prestamo_activo = db.query(Prestamo).filter(
        Prestamo.activo_id == activo_id,
        Prestamo.estado.in_(["ACTIVO", "VENCIDO"])
    ).first()
    if prestamo_activo:
        raise HTTPException(status_code=409, detail="No se puede dar de baja: el activo tiene un préstamo activo")
    a.activo = False
    a.estado = "BAJA"
    db.commit()
    registrar(db, accion=Accion.ELIMINAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=a.id,
              detalle={"nombre": a.nombre, "codigo": a.codigo_inventario},
              request=request)
    return {"mensaje": f"Activo '{a.nombre}' dado de baja"}


# ─── Préstamos ─────────────────────────────────────────────────────────────────

@router.get("/movimientos", summary="Listar movimientos de inventario")
def listar_movimientos(
    activo_id: Optional[int] = None,
    estado: Optional[str] = None,
    tipo: Optional[str] = None,
    departamento_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(MovimientoInventario).join(Activo, MovimientoInventario.activo_id == Activo.id)
    if _es_rol_laboratorio(current_user):
        q = q.filter(Activo.laboratorio_id == current_user.laboratorio_id)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None:
        if not departamentos_visibles:
            return []
        q = q.filter(Activo.departamento_id.in_(departamentos_visibles))
    if activo_id:
        q = q.filter(MovimientoInventario.activo_id == activo_id)
    if estado:
        q = q.filter(MovimientoInventario.estado == estado.upper())
    if tipo:
        q = q.filter(MovimientoInventario.tipo == tipo.upper())
    if departamento_id:
        q = q.filter(
            (MovimientoInventario.departamento_origen_id == departamento_id) |
            (MovimientoInventario.departamento_destino_id == departamento_id)
        )
    movimientos = q.order_by(MovimientoInventario.fecha_solicitud.desc()).limit(250).all()
    return [_serializar_movimiento(m, db) for m in movimientos]


@router.post("/activos/{activo_id}/movimientos", status_code=status.HTTP_201_CREATED, summary="Solicitar movimiento de inventario")
def solicitar_movimiento(
    request: Request,
    activo_id: int,
    data: MovimientoInventarioCreate,
    aplicar_directo: bool = True,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    activo = db.query(Activo).filter(Activo.id == activo_id).first()
    _asegurar_acceso_activo_departamental(db, activo, current_user)
    _asegurar_activo_validado(activo)
    _asegurar_write_inventario_departamento(db, current_user, activo.departamento_id)
    _validar_destinos_movimiento(data, db)
    cruza_departamento = data.departamento_destino_id and data.departamento_destino_id != activo.departamento_id
    if cruza_departamento and current_user.rol != RolUsuario.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Los movimientos entre departamentos requieren administracion central")
    tipo = data.tipo.upper()
    cantidad = data.cantidad or (1 if (activo.tipo_inventario or "ACTIVO") == "ACTIVO" else activo.cantidad or 1)

    if (activo.tipo_inventario or "ACTIVO") == "ACTIVO" and cantidad != 1:
        raise HTTPException(status_code=422, detail="Los activos individuales solo pueden moverse con cantidad 1")

    mov = MovimientoInventario(
        activo_id=activo.id,
        tipo=tipo,
        estado="SOLICITADO",
        departamento_origen_id=activo.departamento_id,
        departamento_destino_id=data.departamento_destino_id,
        ubicacion_origen_id=activo.ubicacion_id,
        ubicacion_destino_id=data.ubicacion_destino_id,
        resguardante_origen_id=activo.responsable_id,
        resguardante_destino_id=data.resguardante_destino_id,
        ubicacion_origen_nombre=activo.ubicacion_nombre,
        ubicacion_destino_nombre=data.ubicacion_destino_nombre.strip() if data.ubicacion_destino_nombre else None,
        resguardante_origen_nombre=activo.resguardante_externo_nombre,
        resguardante_destino_nombre=data.resguardante_destino_nombre.strip() if data.resguardante_destino_nombre else None,
        cantidad=cantidad,
        solicitado_por_id=current_user.id,
        fecha_solicitud=_utcnow(),
        observaciones=data.observaciones,
    )
    if aplicar_directo and (current_user.rol == RolUsuario.SUPER_ADMIN or not cruza_departamento):
        ahora = _utcnow()
        mov.estado = "RECIBIDO"
        mov.autorizado_por_id = current_user.id
        mov.recibido_por_id = current_user.id
        mov.fecha_autorizacion = ahora
        mov.fecha_recepcion = ahora
        _aplicar_movimiento_recibido(activo, mov)
    db.add(mov)
    db.commit()
    db.refresh(mov)
    registrar(db, accion=Accion.EDITAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=activo.id,
              detalle={"movimiento_id": mov.id, "tipo": mov.tipo, "estado": mov.estado},
              request=request)
    return _serializar_movimiento(mov, db)


def _es_movimiento_institucional(mov: MovimientoInventario, activo: Activo) -> bool:
    """True si el movimiento cruza laboratorios, departamentos distintos o involucra activo institucional."""
    if activo.alcance == "INSTITUCIONAL":
        return True
    if mov.departamento_destino_id and mov.departamento_origen_id != mov.departamento_destino_id:
        return True
    return False


@router.post("/movimientos/{movimiento_id}/{accion}", summary="Actualizar estado de movimiento")
def actualizar_movimiento(
    request: Request,
    movimiento_id: int,
    accion: str,
    data: MovimientoInventarioEstado = MovimientoInventarioEstado(),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN))
):
    mov = db.query(MovimientoInventario).filter(MovimientoInventario.id == movimiento_id).first()
    if not mov:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    activo = db.query(Activo).filter(Activo.id == mov.activo_id).first()
    assert_resource_access(activo, current_user)
    accion = accion.lower()
    ahora = _utcnow()
    es_institucional = _es_movimiento_institucional(mov, activo)

    if accion == "autorizar":
        if mov.estado != "SOLICITADO":
            raise HTTPException(status_code=409, detail="Solo se pueden autorizar movimientos solicitados")
        # Movimientos institucionales/inter-departamentales solo los autoriza SUPER_ADMIN
        if es_institucional and current_user.rol != RolUsuario.SUPER_ADMIN:
            raise HTTPException(
                status_code=403,
                detail="Los movimientos institucionales o entre departamentos requieren autorización de administración central"
            )
        mov.estado = "AUTORIZADO"
        mov.autorizado_por_id = current_user.id
        mov.fecha_autorizacion = ahora
    elif accion == "rechazar":
        if mov.estado not in ("SOLICITADO", "AUTORIZADO"):
            raise HTTPException(status_code=409, detail="El movimiento ya no puede rechazarse")
        if es_institucional and current_user.rol != RolUsuario.SUPER_ADMIN:
            raise HTTPException(status_code=403, detail="Solo administración central puede rechazar movimientos institucionales")
        mov.estado = "RECHAZADO"
        mov.autorizado_por_id = current_user.id
        mov.fecha_autorizacion = ahora
    elif accion == "entregar":
        if mov.estado != "AUTORIZADO":
            raise HTTPException(status_code=409, detail="Solo se pueden entregar movimientos autorizados")
        # Entrega: el LAB_ADMIN del laboratorio de origen o SUPER_ADMIN
        if current_user.rol == RolUsuario.LAB_ADMIN:
            lab_origen = activo.laboratorio_id
            if lab_origen and current_user.laboratorio_id != lab_origen:
                raise HTTPException(status_code=403, detail="Solo el responsable del laboratorio de origen puede registrar la entrega")
        mov.estado = "ENTREGADO"
        mov.entregado_por_id = current_user.id
        mov.fecha_entrega = ahora
    elif accion == "recibir":
        if mov.estado not in ("AUTORIZADO", "ENTREGADO"):
            raise HTTPException(status_code=409, detail="Solo se pueden recibir movimientos autorizados o entregados")
        _aplicar_movimiento_recibido(activo, mov)
        mov.estado = "RECIBIDO"
        mov.recibido_por_id = current_user.id
        mov.fecha_recepcion = ahora
    elif accion == "cancelar":
        if mov.estado in ("RECIBIDO", "RECHAZADO"):
            raise HTTPException(status_code=409, detail="El movimiento ya no puede cancelarse")
        mov.estado = "CANCELADO"
    else:
        raise HTTPException(status_code=422, detail="Accion invalida. Use: autorizar, rechazar, entregar, recibir o cancelar")

    if data.observaciones:
        mov.observaciones = ((mov.observaciones or "") + "\n" + data.observaciones).strip()
    db.commit()
    db.refresh(mov)
    registrar(db, accion=Accion.EDITAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=activo.id,
              detalle={"movimiento_id": mov.id, "accion": accion, "estado": mov.estado},
              request=request)
    return _serializar_movimiento(mov, db)


@router.get("/bajas", summary="Listar solicitudes de baja patrimonial")
def listar_bajas(
    estado: Optional[str] = None,
    activo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(SolicitudBajaInventario).join(Activo, SolicitudBajaInventario.activo_id == Activo.id)
    if _es_rol_laboratorio(current_user):
        q = q.filter(Activo.laboratorio_id == current_user.laboratorio_id)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None:
        if not departamentos_visibles:
            return []
        q = q.filter(Activo.departamento_id.in_(departamentos_visibles))
    if estado:
        q = q.filter(SolicitudBajaInventario.estado == estado.upper())
    if activo_id:
        q = q.filter(SolicitudBajaInventario.activo_id == activo_id)
    bajas = q.order_by(SolicitudBajaInventario.fecha_solicitud.desc()).limit(250).all()
    return [_serializar_solicitud_baja(b, db) for b in bajas]


@router.post("/activos/{activo_id}/baja", status_code=status.HTTP_201_CREATED, summary="Solicitar baja patrimonial")
def solicitar_baja(
    request: Request,
    activo_id: int,
    data: SolicitudBajaCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    activo = db.query(Activo).filter(Activo.id == activo_id).first()
    _asegurar_acceso_activo_departamental(db, activo, current_user)
    _asegurar_activo_validado(activo)
    _asegurar_write_inventario_departamento(db, current_user, activo.departamento_id)
    abierta = db.query(SolicitudBajaInventario).filter(
        SolicitudBajaInventario.activo_id == activo_id,
        SolicitudBajaInventario.estado.in_(["SOLICITADA", "EN_REVISION", "VALIDADA_FISICAMENTE", "AUTORIZADA"]),
    ).first()
    if abierta:
        raise HTTPException(status_code=409, detail="El activo ya tiene una solicitud de baja abierta")
    baja = SolicitudBajaInventario(
        activo_id=activo.id,
        estado="SOLICITADA",
        motivo=data.motivo,
        diagnostico=data.diagnostico,
        evidencia_url=data.evidencia_url,
        destino_final=data.destino_final,
        observaciones=data.observaciones,
        solicitado_por_id=current_user.id,
        fecha_solicitud=_utcnow(),
    )
    activo.estado_admin = "BAJA_SOLICITADA"
    db.add(baja)
    db.commit()
    db.refresh(baja)
    registrar(db, accion=Accion.EDITAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=activo.id,
              detalle={"solicitud_baja_id": baja.id, "estado": baja.estado},
              request=request)
    return _serializar_solicitud_baja(baja, db)


@router.post("/bajas/{baja_id}/{accion}", summary="Actualizar tramite de baja patrimonial")
def actualizar_baja(
    request: Request,
    baja_id: int,
    accion: str,
    data: SolicitudBajaAccion = SolicitudBajaAccion(),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB))
):
    baja = db.query(SolicitudBajaInventario).filter(SolicitudBajaInventario.id == baja_id).first()
    if not baja:
        raise HTTPException(status_code=404, detail="Solicitud de baja no encontrada")
    activo = db.query(Activo).filter(Activo.id == baja.activo_id).first()
    assert_resource_access(activo, current_user)
    accion = accion.lower()
    ahora = _utcnow()

    if accion == "revisar":
        if baja.estado != "SOLICITADA":
            raise HTTPException(status_code=409, detail="Solo se pueden revisar bajas solicitadas")
        baja.estado = "EN_REVISION"
        baja.revisado_por_id = current_user.id
        baja.fecha_revision = ahora
    elif accion == "validar":
        if baja.estado not in ("SOLICITADA", "EN_REVISION"):
            raise HTTPException(status_code=409, detail="La baja no esta lista para validacion fisica")
        baja.estado = "VALIDADA_FISICAMENTE"
        baja.validado_por_id = current_user.id
        baja.fecha_validacion = ahora
    elif accion == "autorizar":
        if baja.estado != "VALIDADA_FISICAMENTE":
            raise HTTPException(status_code=409, detail="La baja requiere validacion fisica previa")
        # Solo SUPER_ADMIN puede autorizar bajas patrimoniales
        if current_user.rol != RolUsuario.SUPER_ADMIN:
            raise HTTPException(status_code=403, detail="Solo administracion central puede autorizar bajas patrimoniales")
        baja.estado = "AUTORIZADA"
        baja.autorizado_por_id = current_user.id   # campo correcto — no revisado_por_id
        baja.fecha_autorizacion = ahora
    elif accion == "rechazar":
        if baja.estado in ("EJECUTADA", "CANCELADA"):
            raise HTTPException(status_code=409, detail="La baja ya no puede rechazarse")
        baja.estado = "RECHAZADA"
        activo.estado_admin = "OBSERVADO"
    elif accion == "ejecutar":
        if baja.estado != "AUTORIZADA":
            raise HTTPException(status_code=409, detail="Solo se pueden ejecutar bajas autorizadas")
        baja.estado = "EJECUTADA"
        baja.ejecutado_por_id = current_user.id
        baja.fecha_ejecucion = ahora
        activo.estado_admin = "BAJA_EJECUTADA"
        activo.estado = "BAJA"
        activo.activo = False
    elif accion == "cancelar":
        if baja.estado == "EJECUTADA":
            raise HTTPException(status_code=409, detail="Una baja ejecutada no puede cancelarse")
        baja.estado = "CANCELADA"
        activo.estado_admin = "VALIDADO"
    else:
        raise HTTPException(status_code=422, detail="Accion invalida. Use: revisar, validar, autorizar, rechazar, ejecutar o cancelar")

    if data.destino_final:
        baja.destino_final = data.destino_final
    if data.observaciones:
        baja.observaciones = ((baja.observaciones or "") + "\n" + data.observaciones).strip()
    db.commit()
    db.refresh(baja)
    registrar(db, accion=Accion.EDITAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=activo.id,
              detalle={"solicitud_baja_id": baja.id, "accion": accion, "estado": baja.estado},
              request=request)
    return _serializar_solicitud_baja(baja, db)


@router.get("/levantamientos", summary="Listar levantamientos fisicos")
def listar_levantamientos(
    estado: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(LevantamientoInventario)
    if _es_rol_laboratorio(current_user):
        q = q.filter(LevantamientoInventario.laboratorio_id == current_user.laboratorio_id)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None:
        if not departamentos_visibles:
            return []
        q = q.filter(LevantamientoInventario.departamento_id.in_(departamentos_visibles))
    if estado:
        q = q.filter(LevantamientoInventario.estado == estado.upper())
    return [_serializar_levantamiento(l, db) for l in q.order_by(LevantamientoInventario.fecha_inicio.desc()).all()]


@router.get("/levantamientos/{levantamiento_id}/detalle", summary="Detalle operativo de levantamiento fisico")
def detalle_levantamiento(
    levantamiento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    l = db.query(LevantamientoInventario).filter(LevantamientoInventario.id == levantamiento_id).first()
    _asegurar_acceso_levantamiento(db, l, current_user)
    activos = _query_activos_levantamiento(db, l, current_user)
    revisiones = db.query(RevisionLevantamientoInventario).filter(
        RevisionLevantamientoInventario.levantamiento_id == levantamiento_id
    ).all()
    revisados_ids = {r.activo_id for r in revisiones}
    por_estado = {estado: 0 for estado in ESTADOS_REVISION_LEVANTAMIENTO}
    for r in revisiones:
        por_estado[r.estado] = por_estado.get(r.estado, 0) + 1
    pendientes = [a for a in activos if a.id not in revisados_ids]
    return {
        "levantamiento": _serializar_levantamiento(l, db),
        "activos": [_serializar_activo(a, db) for a in activos],
        "revisiones": [_serializar_revision(r, db) for r in revisiones],
        "pendientes": [_serializar_activo(a, db) for a in pendientes],
        "resumen": {
            "total_esperado": len(activos),
            "total_revisado": len(revisiones),
            "total_pendiente": len(pendientes),
            "por_estado": por_estado,
            "porcentaje": round((len(revisiones) / len(activos)) * 100) if activos else 0,
        },
    }


@router.post("/levantamientos", status_code=status.HTTP_201_CREATED, summary="Crear campana de levantamiento fisico")
def crear_levantamiento(
    data: LevantamientoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    departamento_operacion_id = data.departamento_id
    if not _es_admin_inventario_global(current_user):
        departamento_operacion_id = _resolver_departamento_inventario(db, current_user, data.departamento_id)
        _asegurar_write_inventario_departamento(db, current_user, departamento_operacion_id)
    if departamento_operacion_id and not db.query(Departamento).filter(Departamento.id == departamento_operacion_id).first():
        raise HTTPException(status_code=404, detail="Departamento no encontrado")
    lab_id = data.laboratorio_id
    if _es_rol_laboratorio(current_user):
        if not current_user.laboratorio_id:
            raise HTTPException(status_code=403, detail="No tienes laboratorio asignado")
        if lab_id and lab_id != current_user.laboratorio_id:
            raise HTTPException(status_code=403, detail="Solo puedes crear levantamientos de tu laboratorio")
        lab_id = current_user.laboratorio_id
    if not _es_admin_inventario_global(current_user):
        lab_id = None
    if lab_id:
        assert_lab_write(lab_id, current_user)
        if not db.query(Laboratorio).filter(Laboratorio.id == lab_id, Laboratorio.activo == True).first():
            raise HTTPException(status_code=404, detail="Laboratorio no encontrado")
    elif current_user.rol == RolUsuario.LAB_ADMIN:
        lab_id = current_user.laboratorio_id
    l = LevantamientoInventario(
        nombre=data.nombre,
        departamento_id=departamento_operacion_id,
        laboratorio_id=lab_id,
        observaciones=data.observaciones,
        creado_por_id=current_user.id,
        fecha_inicio=_utcnow(),
    )
    db.add(l)
    db.commit()
    db.refresh(l)
    return _serializar_levantamiento(l, db)


@router.post("/levantamientos/{levantamiento_id}/revisiones", status_code=status.HTTP_201_CREATED, summary="Registrar revision fisica de un bien")
def registrar_revision_levantamiento(
    request: Request,
    levantamiento_id: int,
    data: RevisionLevantamientoIn,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    l = db.query(LevantamientoInventario).filter(LevantamientoInventario.id == levantamiento_id).first()
    _asegurar_acceso_levantamiento(db, l, current_user)
    if l.estado != "ABIERTO":
        raise HTTPException(status_code=409, detail="El levantamiento no esta abierto")
    activo = db.query(Activo).filter(Activo.id == data.activo_id).first()
    _asegurar_acceso_activo_departamental(db, activo, current_user)
    _asegurar_write_inventario_departamento(db, current_user, activo.departamento_id)
    activos_scope = _query_activos_levantamiento(db, l, current_user)
    if activo.id not in {a.id for a in activos_scope}:
        raise HTTPException(status_code=422, detail="El activo no pertenece al alcance de este levantamiento")
    estado = data.estado.upper()
    if estado not in ESTADOS_REVISION_LEVANTAMIENTO:
        raise HTTPException(status_code=422, detail=f"Estado de revision invalido. Use: {ESTADOS_REVISION_LEVANTAMIENTO}")
    revision = db.query(RevisionLevantamientoInventario).filter(
        RevisionLevantamientoInventario.levantamiento_id == levantamiento_id,
        RevisionLevantamientoInventario.activo_id == data.activo_id,
    ).first()
    if not revision:
        revision = RevisionLevantamientoInventario(
            levantamiento_id=levantamiento_id,
            activo_id=data.activo_id,
            revisado_por_id=current_user.id,
        )
        db.add(revision)
    revision.estado = estado
    revision.ubicacion_reportada = data.ubicacion_reportada
    revision.resguardante_reportado = data.resguardante_reportado
    revision.observaciones = data.observaciones
    revision.evidencia_url = data.evidencia_url
    revision.fecha_revision = _utcnow()
    if estado == "NO_LOCALIZADO":
        activo.estado_admin = "OBSERVADO"
    elif estado == "PROPUESTO_BAJA":
        activo.estado_admin = "BAJA_SOLICITADA"
    db.commit()
    db.refresh(revision)
    registrar(db, accion=Accion.EDITAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user, recurso_id=activo.id,
              detalle={"levantamiento_id": l.id, "revision_id": revision.id, "estado_revision": revision.estado},
              request=request)
    return _serializar_revision(revision, db)


@router.post("/levantamientos/{levantamiento_id}/cerrar", summary="Cerrar levantamiento fisico")
def cerrar_levantamiento(
    request: Request,
    levantamiento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    l = db.query(LevantamientoInventario).filter(LevantamientoInventario.id == levantamiento_id).first()
    _asegurar_acceso_levantamiento(db, l, current_user)
    activos = _query_activos_levantamiento(db, l, current_user)
    revisiones = db.query(RevisionLevantamientoInventario).filter(
        RevisionLevantamientoInventario.levantamiento_id == levantamiento_id
    ).all()
    pendientes = max(0, len(activos) - len({r.activo_id for r in revisiones}))
    l.estado = "CERRADO"
    l.fecha_cierre = _utcnow()
    db.commit()
    db.refresh(l)
    registrar(db, accion=Accion.EDITAR_ACTIVO, recurso=Recurso.ACTIVO,
              usuario=current_user,
              detalle={"levantamiento_id": l.id, "accion": "CERRAR_LEVANTAMIENTO", "total_esperado": len(activos), "total_revisado": len(revisiones), "pendientes": pendientes},
              request=request)
    return _serializar_levantamiento(l, db)


@router.get("/activos/{activo_id}/expediente", summary="Expediente digital del bien")
def expediente_activo(
    activo_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    activo = db.query(Activo).filter(Activo.id == activo_id).first()
    assert_resource_access(activo, current_user)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None and activo.departamento_id not in departamentos_visibles:
        raise HTTPException(status_code=404, detail="Recurso no encontrado")
    movimientos = db.query(MovimientoInventario).filter(MovimientoInventario.activo_id == activo_id).order_by(MovimientoInventario.fecha_solicitud.desc()).all()
    bajas = db.query(SolicitudBajaInventario).filter(SolicitudBajaInventario.activo_id == activo_id).order_by(SolicitudBajaInventario.fecha_solicitud.desc()).all()
    revisiones = db.query(RevisionLevantamientoInventario).filter(RevisionLevantamientoInventario.activo_id == activo_id).order_by(RevisionLevantamientoInventario.fecha_revision.desc()).all()
    prestamos = db.query(Prestamo).filter(Prestamo.activo_id == activo_id).order_by(Prestamo.fecha_salida.desc()).all()
    incidentes = db.query(Incidente).filter(Incidente.activo_id == activo_id).order_by(Incidente.fecha_reporte.desc()).all()
    auditoria = db.query(AuditLog).filter(
        AuditLog.recurso == Recurso.ACTIVO,
        AuditLog.recurso_id == activo_id,
    ).order_by(AuditLog.timestamp.desc()).limit(80).all()

    timeline = []
    for m in movimientos:
        mov_data = _serializar_movimiento(m, db)
        origen = (
            mov_data.get("departamento_origen_nombre") or
            mov_data.get("ubicacion_origen_nombre") or
            mov_data.get("resguardante_origen_nombre")
        )
        destino = (
            mov_data.get("departamento_destino_nombre") or
            mov_data.get("ubicacion_destino_nombre") or
            mov_data.get("resguardante_destino_nombre")
        )
        timeline.append(_evento_expediente(
            "MOVIMIENTO",
            m.fecha_recepcion or m.fecha_entrega or m.fecha_autorizacion or m.fecha_solicitud,
            (m.tipo or "MOVIMIENTO").replace("_", " "),
            f"{origen or 'Origen sin dato'} -> {destino or 'Destino sin dato'}",
            m.estado,
            None,
            {"movimiento_id": m.id},
        ))
    for b in bajas:
        timeline.append(_evento_expediente(
            "BAJA",
            b.fecha_ejecucion or b.fecha_autorizacion or b.fecha_solicitud,
            f"Baja patrimonial: {b.estado}",
            b.motivo,
            b.estado,
            None,
            {"baja_id": b.id},
        ))
    for r in revisiones:
        timeline.append(_evento_expediente(
            "LEVANTAMIENTO",
            r.fecha_revision,
            f"Revision fisica: {r.estado}",
            r.observaciones or r.ubicacion_reportada,
            r.estado,
            None,
            {"revision_id": r.id, "levantamiento_id": r.levantamiento_id},
        ))
    for p in prestamos:
        timeline.append(_evento_expediente(
            "PRESTAMO",
            p.fecha_retorno_real or p.fecha_salida,
            f"Prestamo: {p.estado}",
            f"{p.solicitante_nombre} ({p.solicitante_id_escolar})",
            p.estado,
            None,
            {"prestamo_id": p.id},
        ))
    for i in incidentes:
        timeline.append(_evento_expediente(
            "INCIDENTE",
            i.fecha_reporte,
            f"Incidente: {i.tipo}",
            i.descripcion,
            i.estado,
            None,
            {"incidente_id": i.id, "prioridad": i.prioridad},
        ))
    for log in auditoria:
        detalle = log.detalle or {}
        if detalle.get("flujo") == "VALIDACION_INVENTARIO":
            titulo = f"Validacion: {detalle.get('estado_anterior', '—')} -> {detalle.get('estado_nuevo', '—')}"
            descripcion = detalle.get("observaciones")
            estado = detalle.get("estado_nuevo")
        else:
            titulo = log.accion.replace("_", " ")
            descripcion = ", ".join(detalle.get("campos", [])) if isinstance(detalle.get("campos"), list) else None
            estado = None
        timeline.append(_evento_expediente(
            "AUDITORIA",
            log.timestamp,
            titulo,
            descripcion,
            estado,
            log.usuario_nombre,
            {"audit_id": log.id, "accion": log.accion, "detalle": detalle},
        ))
    timeline = sorted([e for e in timeline if e["fecha"]], key=lambda e: e["fecha"], reverse=True)

    activo_data = _serializar_activo(activo, db)
    alertas = []
    if not activo.numero_oficial:
        alertas.append("Sin numero oficial/patrimonial")
    if not activo.departamento_id and (activo.alcance or "").upper() == "INSTITUCIONAL":
        alertas.append("Sin departamento responsable")
    if not activo.ubicacion_id and not activo.ubicacion_nombre:
        alertas.append("Sin ubicacion registrada")
    if not activo.responsable_id and not activo.resguardante_externo_nombre:
        alertas.append("Sin resguardante")
    if (activo.estado_admin or "").upper() != "VALIDADO":
        alertas.append(f"Validacion pendiente: {activo.estado_admin}")

    return {
        "activo": activo_data,
        "resumen": {
            "alertas": alertas,
            "total_eventos": len(timeline),
            "estado_admin": activo.estado_admin or "VALIDADO",
            "estado_operativo": activo.estado,
            "prestado": bool(activo_data.get("prestado")),
            "ultima_actualizacion": timeline[0]["fecha"] if timeline else None,
        },
        "movimientos": [_serializar_movimiento(m, db) for m in movimientos],
        "bajas": [_serializar_solicitud_baja(b, db) for b in bajas],
        "levantamientos": [_serializar_revision(r, db) for r in revisiones],
        "prestamos": [_serializar_prestamo(p, db) for p in prestamos],
        "incidentes": [_serializar_incidente(i, db) for i in incidentes],
        "auditoria": [_serializar_audit_log(log) for log in auditoria],
        "timeline": timeline,
    }


@router.get("/prestamos", summary="Listar préstamos")
def listar_prestamos(
    estado: Optional[str]         = None,
    laboratorio_id: Optional[int] = None,
    activo_id: Optional[int]      = None,
    vencidos_solo: bool            = False,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    _actualizar_estado_prestamos(db)
    q = db.query(Prestamo)
    if estado:
        q = q.filter(Prestamo.estado == estado.upper())
    if vencidos_solo:
        q = q.filter(Prestamo.estado == "VENCIDO")
    if activo_id:
        q = q.filter(Prestamo.activo_id == activo_id)
    if laboratorio_id or _es_rol_laboratorio(current_user):
        activos_query = db.query(Activo.id)
        activos_query = _filtrar_lab_asignado(activos_query, Activo, current_user, laboratorio_id)
        activo_ids = [row[0] for row in activos_query.all()]
        if not activo_ids:
            return []
        q = q.filter(Prestamo.activo_id.in_(activo_ids))

    prestamos = q.order_by(Prestamo.fecha_salida.desc()).all()
    return [_serializar_prestamo(p, db) for p in prestamos]


@router.get("/prestamos/vencidos", summary="Préstamos vencidos (alertas)")
def prestamos_vencidos(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    _actualizar_estado_prestamos(db)
    q = db.query(Prestamo).filter(Prestamo.estado == "VENCIDO")
    if _es_rol_laboratorio(current_user):
        activo_ids = [
            row[0]
            for row in db.query(Activo.id)
            .filter(Activo.laboratorio_id == current_user.laboratorio_id)
            .all()
        ]
        if not activo_ids:
            return []
        q = q.filter(Prestamo.activo_id.in_(activo_ids))
    return [_serializar_prestamo(p, db) for p in q.order_by(Prestamo.fecha_retorno_esperada).all()]


@router.post("/prestamos", status_code=status.HTTP_201_CREATED, summary="Registrar préstamo")
def crear_prestamo(
    request: Request,
    data: PrestamoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    if current_user.rol == RolUsuario.ALUMNO:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    _actualizar_estado_prestamos(db)

    activo_ids = list(dict.fromkeys(data.activo_ids or ([data.activo_id] if data.activo_id else [])))
    if not activo_ids:
        raise HTTPException(status_code=422, detail="Selecciona al menos un activo")

    activos = db.query(Activo).filter(Activo.id.in_(activo_ids), Activo.activo == True).all()
    activos_por_id = {activo.id: activo for activo in activos}
    faltantes = [activo_id for activo_id in activo_ids if activo_id not in activos_por_id]
    if faltantes:
        raise HTTPException(
            status_code=404,
            detail=f"Activos no encontrados o dados de baja: {', '.join(map(str, faltantes))}",
        )

    prestamos_activos = {
        row[0]
        for row in db.query(Prestamo.activo_id).filter(
            Prestamo.activo_id.in_(activo_ids),
            Prestamo.estado.in_(["ACTIVO", "VENCIDO"]),
        ).all()
    }
    if prestamos_activos:
        nombres = [
            activos_por_id[activo_id].nombre
            for activo_id in activo_ids
            if activo_id in prestamos_activos
        ]
        raise HTTPException(
            status_code=409,
            detail=f"Ya tienen un préstamo activo: {', '.join(nombres)}",
        )

    for activo_id in activo_ids:
        activo = activos_por_id[activo_id]
        assert_resource_access(activo, current_user)
        _asegurar_activo_validado(activo)
        if activo.estado != "OPERATIVO":
            raise HTTPException(
                status_code=400,
                detail=f"El activo '{activo.nombre}' está en estado '{activo.estado}' y no se puede prestar",
            )

    # Parsear fecha de devolución esperada
    fecha_retorno = None
    if data.fecha_devolucion_esperada:
        try:
            # Parsear solo la fecha (YYYY-MM-DD) y ponerla al final del día
            # para evitar falsos positivos por diferencia UTC vs hora local
            fecha_date = datetime.date.fromisoformat(data.fecha_devolucion_esperada[:10])
        except ValueError:
            raise HTTPException(status_code=422, detail="Formato de fecha inválido, use YYYY-MM-DD")
        # Rechazar solo si la fecha es estrictamente anterior a hoy
        if fecha_date < datetime.date.today():
            raise HTTPException(status_code=422, detail="La fecha de devolución no puede ser en el pasado")
        # Guardar como datetime al final del día
        fecha_retorno = datetime.datetime.combine(fecha_date, datetime.time(23, 59, 59))
    else:
        # Por defecto 7 días
        fecha_retorno = _utcnow() + datetime.timedelta(days=7)

    # Guardar los datos compartidos de la solicitud como metadato.
    import json
    meta = json.dumps({
        "receptor_tipo": data.receptor_tipo,
        "proposito": data.proposito,
        "notas": data.notas,
    })
    obs_salida = f"__meta__{meta}"
    folio = f"PRE-{datetime.date.today():%Y%m%d}-{uuid.uuid4().hex[:8].upper()}"
    fecha_salida = _utcnow()
    prestamos = []
    for activo_id in activo_ids:
        prestamo = Prestamo(
            folio=folio,
            activo_id=activo_id,
            solicitante_nombre=data.receptor_nombre,
            solicitante_id_escolar=data.receptor_matricula or "",
            autorizado_por=current_user.id,
            fecha_salida=fecha_salida,
            fecha_retorno_esperada=fecha_retorno,
            estado="ACTIVO",
            condicion_salida="BUENO",
            observaciones_salida=obs_salida,
        )
        db.add(prestamo)
        prestamos.append(prestamo)

    db.commit()
    for prestamo in prestamos:
        db.refresh(prestamo)
        registrar(
            db,
            accion=Accion.CREAR_PRESTAMO,
            recurso=Recurso.PRESTAMO,
            usuario=current_user,
            recurso_id=prestamo.id,
            detalle={
                "folio": folio,
                "activo_id": prestamo.activo_id,
                "receptor": prestamo.solicitante_nombre,
                "total_activos": len(prestamos),
            },
            request=request,
        )

    serializados = [_serializar_prestamo(prestamo, db) for prestamo in prestamos]
    if len(serializados) == 1:
        return serializados[0]
    return {
        "folio": folio,
        "total_activos": len(serializados),
        "estado": "ACTIVO",
        "prestamos": serializados,
    }


@router.post("/prestamos/grupos/{folio}/devolver", summary="Registrar devolución total o parcial")
def devolver_grupo_prestamos(
    request: Request,
    folio: str,
    data: PrestamoGrupoDevolver,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    prestamos_pendientes = db.query(Prestamo).filter(
        Prestamo.folio == folio,
        Prestamo.estado.in_(["ACTIVO", "VENCIDO"]),
    ).order_by(Prestamo.id).all()
    if not prestamos_pendientes:
        raise HTTPException(status_code=404, detail="No hay activos pendientes para este folio")

    if data.prestamo_ids:
        ids_solicitados = set(data.prestamo_ids)
        ids_disponibles = {prestamo.id for prestamo in prestamos_pendientes}
        ids_invalidos = ids_solicitados - ids_disponibles
        if ids_invalidos:
            raise HTTPException(
                status_code=400,
                detail="La selección incluye préstamos que no pertenecen al folio o ya fueron devueltos",
            )
        prestamos_pendientes = [
            prestamo for prestamo in prestamos_pendientes if prestamo.id in ids_solicitados
        ]

    for prestamo in prestamos_pendientes:
        activo = db.query(Activo).filter(Activo.id == prestamo.activo_id).first()
        assert_resource_access(activo, current_user)

    resultados = [
        devolver_prestamo(request, prestamo.id, data, db, current_user)
        for prestamo in prestamos_pendientes
    ]
    return {
        "folio": folio,
        "devueltos": len(resultados),
        "prestamos": resultados,
    }


@router.post("/prestamos/{prestamo_id}/devolver", summary="Registrar devolución")
def devolver_prestamo(
    request: Request,
    prestamo_id: int,
    data: PrestamoDevolver,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    if data.condicion_devolucion not in ("BUENO", "REGULAR", "MALO", "DAÑADO"):
        raise HTTPException(status_code=422, detail="Condición de devolución no válida")

    p = db.query(Prestamo).filter(Prestamo.id == prestamo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    if p.estado == "DEVUELTO":
        raise HTTPException(status_code=400, detail="Este préstamo ya fue devuelto")

    activo = db.query(Activo).filter(Activo.id == p.activo_id).first()
    assert_resource_access(activo, current_user)

    p.estado                = "DEVUELTO"
    p.fecha_retorno_real    = _utcnow()
    p.condicion_retorno     = data.condicion_devolucion
    p.observaciones_retorno = data.notas_devolucion

    # ── Auto-crear incidente y actualizar estado del activo si regresa dañado ──
    activo = db.query(Activo).filter(Activo.id == p.activo_id).first()
    if data.condicion_devolucion in ("MALO", "DAÑADO") and activo:
        nuevo_estado = "DAÑADO" if data.condicion_devolucion == "DAÑADO" else "MANTENIMIENTO"
        activo.estado = nuevo_estado

        incidente = Incidente(
            activo_id        = activo.id,
            laboratorio_id   = activo.laboratorio_id,
            origen           = "PRESTAMO",
            origen_id        = p.id,
            tipo             = "DAÑO" if data.condicion_devolucion == "DAÑADO" else "MANTENIMIENTO",
            descripcion      = data.notas_devolucion or f"Equipo devuelto en condición {data.condicion_devolucion} por {p.solicitante_nombre}",
            reportado_por_id = current_user.id,
            prioridad        = "ALTA" if data.condicion_devolucion == "DAÑADO" else "MEDIA",
            estado           = "PENDIENTE",
        )
        db.add(incidente)

        # ── Auto-crear adeudo por daño al equipo ──────────────────────────────
        # Obtener tipo de persona desde metadato guardado en observaciones_salida
        import json as _json
        obs_salida = p.observaciones_salida or ""
        try:
            meta = _json.loads(obs_salida[8:] if obs_salida.startswith("__meta__") else (obs_salida or "{}"))
            persona_tipo = meta.get("receptor_tipo", "ALUMNO")
        except Exception:
            persona_tipo = "ALUMNO"

        nombre_equipo = f"{activo.nombre} (cód. {activo.codigo_inventario})" if activo else "Equipo"
        adeudo_danio = Adeudo(
            persona_nombre        = p.solicitante_nombre,
            persona_identificador = p.solicitante_id_escolar or "",
            persona_tipo          = persona_tipo,
            origen_tipo           = "PRESTAMO",
            tipo                  = "DAÑO",
            descripcion           = (
                f"Equipo devuelto dañado — Préstamo #{p.id}. "
                f"{nombre_equipo}. "
                f"Condición al devolver: {data.condicion_devolucion}. "
                + (f"Nota: {data.notas_devolucion}" if data.notas_devolucion else "")
            ),
            prestamo_id           = p.id,
            laboratorio_id        = activo.laboratorio_id if activo else None,
            estado                = "PENDIENTE",
            reportado_por_id      = current_user.id,
            fecha_reporte         = _utcnow(),
        )
        db.add(adeudo_danio)

        # Notificar a admins del laboratorio
        try:
            from routers.notificaciones import crear_notificacion
            admins = db.query(Usuario).filter(
                Usuario.activo == True,
                Usuario.rol.in_(["SUPER_ADMIN", "LAB_ADMIN", "RESPONSABLE_LAB"]),
            ).all()
            for admin in admins:
                if admin.rol == "SUPER_ADMIN" or admin.laboratorio_id == activo.laboratorio_id:
                    crear_notificacion(
                        db, admin.id,
                        "PRESTAMO_VENCIDO",
                        f"⚠️ Equipo devuelto dañado — {p.solicitante_nombre}",
                        f"{nombre_equipo} devuelto en condición {data.condicion_devolucion}. "
                        f"Se generó adeudo automáticamente.",
                    )
        except Exception:
            pass

    # ── Auto-resolver adeudo por vencimiento si existía (el equipo fue devuelto) ──
    adeudo_vinculado = db.query(Adeudo).filter(
        Adeudo.prestamo_id == prestamo_id,
        Adeudo.tipo.in_(["PRESTAMO_VENCIDO", "PRESTAMO_NO_DEVUELTO"]),
        Adeudo.estado.in_(["PENDIENTE", "EN_REVISION"]),
    ).first()
    if adeudo_vinculado:
        adeudo_vinculado.estado           = "RESUELTO"
        adeudo_vinculado.fecha_resolucion = _utcnow()
        adeudo_vinculado.resuelto_por_id  = current_user.id
        adeudo_vinculado.notas_resolucion = (
            f"Préstamo devuelto el {_utcnow().strftime('%d/%m/%Y')} "
            f"en condición {data.condicion_devolucion}. "
            + (data.notas_devolucion or "")
        )

    db.commit()
    db.refresh(p)
    registrar(db, accion=Accion.DEVOLVER_PRESTAMO, recurso=Recurso.PRESTAMO,
              usuario=current_user, recurso_id=p.id,
              detalle={"activo_id": p.activo_id, "condicion": data.condicion_devolucion},
              request=request)
    return _serializar_prestamo(p, db)


# --- Incidentes ────────────────────────────────────────────────────────────────

@router.get("/incidentes", summary="Listar incidentes de mantenimiento")
def listar_incidentes(
    estado:         Optional[str] = None,
    prioridad:      Optional[str] = None,
    laboratorio_id: Optional[int] = None,
    tipo:           Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(Incidente)
    if estado:
        q = q.filter(Incidente.estado == estado.upper())
    if prioridad:
        q = q.filter(Incidente.prioridad == prioridad.upper())
    if tipo:
        q = q.filter(Incidente.tipo == tipo.upper())
    q = _filtrar_lab_asignado(q, Incidente, current_user, laboratorio_id)

    incidentes = q.order_by(Incidente.fecha_reporte.desc()).all()
    return [_serializar_incidente(i, db) for i in incidentes]


@router.post("/incidentes", status_code=status.HTTP_201_CREATED, summary="Crear incidente manualmente")
def crear_incidente(
    request: Request,
    data: IncidenteCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    if current_user.rol == RolUsuario.ALUMNO:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    # Los reportes de limpieza, iluminacion o seguridad pueden ser generales
    # del laboratorio. El responsable asociara un equipo despues si aplica.
    if not data.activo_id and not data.computadora_id and not data.laboratorio_id:
        raise HTTPException(
            status_code=422,
            detail="El incidente debe indicar un laboratorio, un activo o una computadora"
        )
    if data.activo_id and data.computadora_id:
        raise HTTPException(
            status_code=422,
            detail="El incidente no puede asociarse simultaneamente a un activo y a una computadora"
        )

    # Si viene activo_id, validar el alta oficial y extraer laboratorio_id.
    lab_id = data.laboratorio_id
    activo = None
    if data.activo_id:
        activo = db.query(Activo).filter(Activo.id == data.activo_id).first()
        _asegurar_acceso_activo_departamental(db, activo, current_user)
        _asegurar_activo_validado(activo)
        if lab_id and activo.laboratorio_id and lab_id != activo.laboratorio_id:
            raise HTTPException(status_code=422, detail="El activo no pertenece al laboratorio indicado")
        lab_id = activo.laboratorio_id or lab_id

    # Si viene computadora_id, validar y extraer laboratorio_id.
    if data.computadora_id:
        from models.laboratorio import Computadora
        pc = db.query(Computadora).filter(Computadora.id == data.computadora_id).first()
        if not pc:
            raise HTTPException(status_code=404, detail="Computadora no encontrada")
        if lab_id and lab_id != pc.laboratorio_id:
            raise HTTPException(status_code=422, detail="La computadora no pertenece al laboratorio indicado")
        lab_id = pc.laboratorio_id

    if lab_id:
        lab = db.query(Laboratorio).filter(Laboratorio.id == lab_id).first()
        if not lab:
            raise HTTPException(status_code=404, detail="Laboratorio no encontrado")
        if _es_rol_laboratorio(current_user) and current_user.laboratorio_id != lab.id:
            raise HTTPException(status_code=403, detail="Solo puedes reportar incidentes de tu laboratorio")

    i = Incidente(
        activo_id        = data.activo_id,
        computadora_id   = data.computadora_id,
        laboratorio_id   = lab_id,
        origen           = data.origen.upper(),
        origen_id        = data.origen_id,
        tipo             = data.tipo.upper(),
        descripcion      = data.descripcion,
        reportado_por_id = current_user.id,
        prioridad        = data.prioridad.upper(),
        estado           = "PENDIENTE",
    )
    db.add(i)

    # Si el incidente es sobre un activo de inventario → actualizar su estado
    if data.activo_id and data.tipo.upper() in ("DAÑO", "MANTENIMIENTO"):
        if activo:
            activo.estado = "DAÑADO" if data.tipo.upper() == "DAÑO" else "MANTENIMIENTO"

    # ── Si es una PC de cómputo → ponerla en MANTENIMIENTO inmediatamente ──
    # Esto la bloquea en el mapa de sesiones hasta que el admin la repare
    if data.computadora_id and data.tipo.upper() in ("DAÑO", "MANTENIMIENTO", "OTRO"):
        from models.laboratorio import Computadora
        pc = db.query(Computadora).filter(Computadora.id == data.computadora_id).first()
        if pc:
            pc.estado = "MANTENIMIENTO"

    db.commit()
    db.refresh(i)
    registrar(db, accion=Accion.CREAR_MANTENIMIENTO, recurso=Recurso.INCIDENTE,
              usuario=current_user, recurso_id=i.id,
              detalle={"tipo": i.tipo, "prioridad": i.prioridad,
                       "laboratorio_id": i.laboratorio_id, "origen": i.origen,
                       "activo_id": i.activo_id, "computadora_id": i.computadora_id},
              request=request)
    return _serializar_incidente(i, db)


@router.put("/incidentes/{incidente_id}", summary="Actualizar estado/seguimiento del incidente")
def actualizar_incidente(
    request: Request,
    incidente_id: int,
    data: IncidenteUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    if current_user.rol == RolUsuario.ALUMNO:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    i = db.query(Incidente).filter(Incidente.id == incidente_id).first()
    if not i:
        raise HTTPException(status_code=404, detail="Incidente no encontrado")
    _asegurar_acceso_incidente(i, current_user)
    if i.estado in ESTADOS_INCIDENTE_CERRADOS:
        raise HTTPException(
            status_code=409,
            detail=(
                "La incidencia ya esta cerrada y su expediente es de solo lectura. "
                "Puedes agregar un seguimiento o reabrirla indicando el motivo."
            ),
        )

    campos = data.model_dump(exclude_unset=True)
    motivo_vinculacion = (campos.pop("motivo_vinculacion", None) or "").strip()
    nota_legacy = (campos.pop("notas_seguimiento", None) or "").strip()
    estado_anterior = i.estado
    vinculo_creado = None

    if "activo_id" in campos or "computadora_id" in campos:
        if current_user.rol not in (
            RolUsuario.SUPER_ADMIN,
            RolUsuario.LAB_ADMIN,
            RolUsuario.RESPONSABLE_LAB,
        ):
            raise HTTPException(
                status_code=403,
                detail="Solo el responsable del laboratorio puede asociar el reporte con un equipo"
            )

        activo_id = campos.get("activo_id")
        computadora_id = campos.get("computadora_id")
        if i.activo_id or i.computadora_id:
            mismo_activo = activo_id == i.activo_id and computadora_id in (None, i.computadora_id)
            misma_pc = computadora_id == i.computadora_id and activo_id in (None, i.activo_id)
            if not (mismo_activo or misma_pc):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "El equipo relacionado ya forma parte de la trazabilidad del incidente. "
                        "No puede cambiarse desde la actualización normal."
                    )
                )
            campos.pop("activo_id", None)
            campos.pop("computadora_id", None)
            activo_id = None
            computadora_id = None
        if activo_id and computadora_id:
            raise HTTPException(
                status_code=422,
                detail="El incidente no puede asociarse simultaneamente a un activo y a una computadora"
            )
        if (activo_id or computadora_id) and len(motivo_vinculacion) < 5:
            raise HTTPException(
                status_code=422,
                detail="Debes indicar cómo se identificó el equipo relacionado"
            )

        if activo_id:
            activo_asignado = db.query(Activo).filter(Activo.id == activo_id).first()
            _asegurar_acceso_activo_departamental(db, activo_asignado, current_user)
            _asegurar_activo_validado(activo_asignado)
            if i.laboratorio_id and activo_asignado.laboratorio_id != i.laboratorio_id:
                raise HTTPException(status_code=422, detail="El activo no pertenece al laboratorio del reporte")
            campos["computadora_id"] = None
            vinculo_creado = {
                "tipo": "ACTIVO",
                "id": activo_asignado.id,
                "codigo": activo_asignado.codigo_inventario,
                "motivo": motivo_vinculacion,
            }

        if computadora_id:
            from models.laboratorio import Computadora
            pc_asignada = db.query(Computadora).filter(Computadora.id == computadora_id).first()
            if not pc_asignada:
                raise HTTPException(status_code=404, detail="Computadora no encontrada")
            if i.laboratorio_id and pc_asignada.laboratorio_id != i.laboratorio_id:
                raise HTTPException(status_code=422, detail="La computadora no pertenece al laboratorio del reporte")
            if _es_rol_laboratorio(current_user) and current_user.laboratorio_id != pc_asignada.laboratorio_id:
                raise HTTPException(status_code=403, detail="Solo puedes gestionar equipos de tu laboratorio")
            campos["activo_id"] = None
            vinculo_creado = {
                "tipo": "COMPUTADORA",
                "id": pc_asignada.id,
                "codigo": pc_asignada.codigo,
                "motivo": motivo_vinculacion,
            }
            if i.tipo.upper() in ("DAÑO", "MANTENIMIENTO", "OTRO"):
                pc_asignada.estado = "MANTENIMIENTO"

    # ── Bloquear reabrir un incidente que ya generó un adeudo ────────────────
    nuevo_estado = campos.get("estado", "").upper() if "estado" in campos else ""
    if nuevo_estado == "DADO_DE_BAJA":
        activo_destino = campos.get("activo_id", i.activo_id)
        pc_destino = campos.get("computadora_id", i.computadora_id)
        if not activo_destino and not pc_destino:
            raise HTTPException(
                status_code=422,
                detail="Una observacion general no puede marcarse como dada de baja sin un equipo relacionado",
            )
    if nuevo_estado in ("PENDIENTE", "EN_REVISION"):
        adeudo_vinculado = db.query(Adeudo).filter(Adeudo.incidente_id == incidente_id).first()
        if adeudo_vinculado and adeudo_vinculado.estado not in ("RESUELTO", "CANCELADO"):
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Este incidente tiene un adeudo vinculado en estado '{adeudo_vinculado.estado}' "
                    f"(ID #{adeudo_vinculado.id}). No se puede reabrir hasta que el adeudo sea "
                    f"resuelto o cancelado. Si necesitas revisar el equipo, crea un nuevo incidente de inspección."
                )
            )

    for campo, valor in campos.items():
        if isinstance(valor, str):
            valor = valor.upper() if campo in ("estado", "prioridad") else valor
        setattr(i, campo, valor)

    # Si se marca como REPARADO → registrar fecha y reactivar el equipo
    if "estado" in campos and campos["estado"].upper() == "REPARADO":
        i.fecha_resolucion = _utcnow()
        # Activo de inventario → vuelve a OPERATIVO
        if i.activo_id:
            activo = db.query(Activo).filter(Activo.id == i.activo_id).first()
            if activo:
                activo.estado = "OPERATIVO"
        # PC de cómputo → vuelve a OPERATIVO (disponible en mapa de sesiones)
        if i.computadora_id:
            from models.laboratorio import Computadora
            pc = db.query(Computadora).filter(Computadora.id == i.computadora_id).first()
            if pc:
                pc.estado = "OPERATIVO"

    # Si se cierra sin adeudo → solo registrar fecha de resolución
    if "estado" in campos and campos["estado"].upper() == "CERRADO_SIN_ADEUDO":
        i.fecha_resolucion = _utcnow()

    # Si se marca como DADO_DE_BAJA → dar de baja el equipo
    if "estado" in campos and campos["estado"].upper() == "DADO_DE_BAJA":
        i.fecha_resolucion = _utcnow()
        # Activo de inventario → baja definitiva
        if i.activo_id:
            activo = db.query(Activo).filter(Activo.id == i.activo_id).first()
            if activo:
                activo.estado = "BAJA"
                activo.activo = False
        # PC de cómputo → estado BAJA (no aparece en mapa)
        if i.computadora_id:
            from models.laboratorio import Computadora
            pc = db.query(Computadora).filter(Computadora.id == i.computadora_id).first()
            if pc:
                pc.estado = "BAJA"
                pc.activa = False

    if nota_legacy:
        db.add(SeguimientoIncidente(
            incidente_id=i.id,
            usuario_id=current_user.id,
            tipo="NOTA",
            texto=nota_legacy,
        ))
    if vinculo_creado:
        db.add(SeguimientoIncidente(
            incidente_id=i.id,
            usuario_id=current_user.id,
            tipo="VINCULACION",
            texto=(
                f"Se relaciono {vinculo_creado['tipo'].lower()} "
                f"{vinculo_creado['codigo']}: {vinculo_creado['motivo']}"
            ),
        ))
    if nuevo_estado and nuevo_estado != estado_anterior:
        db.add(SeguimientoIncidente(
            incidente_id=i.id,
            usuario_id=current_user.id,
            tipo="CAMBIO_ESTADO",
            texto=f"Estado cambiado de {estado_anterior} a {nuevo_estado}",
            estado_anterior=estado_anterior,
            estado_nuevo=nuevo_estado,
        ))

    db.commit()
    db.refresh(i)
    if vinculo_creado:
        registrar(
            db,
            accion=Accion.VINCULAR_EQUIPO_INCIDENTE,
            recurso=Recurso.INCIDENTE,
            usuario=current_user,
            recurso_id=i.id,
            detalle={
                **vinculo_creado,
                "laboratorio_id": i.laboratorio_id,
            },
            request=request,
        )
    # Audit log when closing/resolving
    if "estado" in campos and campos["estado"].upper() in ("REPARADO", "CERRADO_SIN_ADEUDO", "DADO_DE_BAJA"):
        registrar(db, accion=Accion.CERRAR_MANTENIMIENTO, recurso=Recurso.INCIDENTE,
                  usuario=current_user, recurso_id=i.id,
                  detalle={"estado_anterior": estado_anterior, "estado_nuevo": i.estado,
                           "laboratorio_id": i.laboratorio_id},
                  request=request)
    return _serializar_incidente(i, db)


@router.post("/incidentes/{incidente_id}/seguimientos", summary="Agregar nota al historial del incidente")
def agregar_seguimiento_incidente(
    request: Request,
    incidente_id: int,
    data: IncidenteSeguimientoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if current_user.rol == RolUsuario.ALUMNO:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    incidente = db.query(Incidente).filter(Incidente.id == incidente_id).first()
    if not incidente:
        raise HTTPException(status_code=404, detail="Incidente no encontrado")
    _asegurar_acceso_incidente(incidente, current_user)

    texto = data.texto.strip()
    if len(texto) < 2:
        raise HTTPException(status_code=422, detail="Escribe el seguimiento realizado")

    db.add(SeguimientoIncidente(
        incidente_id=incidente.id,
        usuario_id=current_user.id,
        tipo="NOTA",
        texto=texto,
    ))
    db.commit()
    db.refresh(incidente)
    registrar(
        db,
        accion=Accion.AGREGAR_SEGUIMIENTO_INCIDENTE,
        recurso=Recurso.INCIDENTE,
        usuario=current_user,
        recurso_id=incidente.id,
        detalle={"estado": incidente.estado, "laboratorio_id": incidente.laboratorio_id},
        request=request,
    )
    return _serializar_incidente(incidente, db)


@router.post("/incidentes/{incidente_id}/reabrir", summary="Reabrir un incidente cerrado")
def reabrir_incidente(
    request: Request,
    incidente_id: int,
    data: IncidenteReabrir,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if current_user.rol not in (
        RolUsuario.SUPER_ADMIN,
        RolUsuario.LAB_ADMIN,
        RolUsuario.RESPONSABLE_LAB,
    ):
        raise HTTPException(status_code=403, detail="Solo un responsable puede reabrir incidencias")

    incidente = db.query(Incidente).filter(Incidente.id == incidente_id).first()
    if not incidente:
        raise HTTPException(status_code=404, detail="Incidente no encontrado")
    _asegurar_acceso_incidente(incidente, current_user)
    if incidente.estado not in ESTADOS_INCIDENTE_CERRADOS:
        raise HTTPException(status_code=409, detail="La incidencia no esta cerrada")

    adeudo_vinculado = db.query(Adeudo).filter(Adeudo.incidente_id == incidente_id).first()
    if adeudo_vinculado and adeudo_vinculado.estado not in ("RESUELTO", "CANCELADO"):
        raise HTTPException(
            status_code=409,
            detail="No se puede reabrir mientras tenga un adeudo pendiente",
        )

    motivo = data.motivo.strip()
    estado_anterior = incidente.estado
    incidente.estado = "EN_REVISION"
    incidente.fecha_resolucion = None

    if incidente.activo_id:
        activo = db.query(Activo).filter(Activo.id == incidente.activo_id).first()
        if activo:
            activo.activo = True
            activo.estado = "MANTENIMIENTO"
    if incidente.computadora_id:
        pc = db.query(Computadora).filter(Computadora.id == incidente.computadora_id).first()
        if pc:
            pc.activa = True
            pc.estado = "MANTENIMIENTO"

    db.add(SeguimientoIncidente(
        incidente_id=incidente.id,
        usuario_id=current_user.id,
        tipo="REAPERTURA",
        texto=f"Incidencia reabierta: {motivo}",
        estado_anterior=estado_anterior,
        estado_nuevo="EN_REVISION",
    ))
    db.commit()
    db.refresh(incidente)
    registrar(
        db,
        accion=Accion.REABRIR_INCIDENTE,
        recurso=Recurso.INCIDENTE,
        usuario=current_user,
        recurso_id=incidente.id,
        detalle={
            "estado_anterior": estado_anterior,
            "estado_nuevo": incidente.estado,
            "motivo": motivo,
            "laboratorio_id": incidente.laboratorio_id,
        },
        request=request,
    )
    return _serializar_incidente(incidente, db)


@router.get("/incidentes/estadisticas", summary="Resumen de incidentes")
def estadisticas_incidentes(
    laboratorio_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(Incidente)
    if laboratorio_id:
        q = q.filter(Incidente.laboratorio_id == laboratorio_id)
    elif current_user.rol == RolUsuario.LAB_ADMIN:
        q = q.filter(Incidente.laboratorio_id == current_user.laboratorio_id)

    todos = q.all()
    return {
        "total":       len(todos),
        "pendientes":  sum(1 for i in todos if i.estado == "PENDIENTE"),
        "en_revision": sum(1 for i in todos if i.estado == "EN_REVISION"),
        "reparados":   sum(1 for i in todos if i.estado == "REPARADO"),
        "dados_de_baja": sum(1 for i in todos if i.estado == "DADO_DE_BAJA"),
        "alta_prioridad": sum(1 for i in todos if i.prioridad == "ALTA" and i.estado not in ("REPARADO", "DADO_DE_BAJA")),
    }


@router.get("/catalogo", summary="Catalogo configurable de inventario")
def listar_catalogo_inventario(
    tipo: Optional[str] = None,
    solo_activos: bool = False,
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    tipos = [tipo.upper()] if tipo else CATALOGO_TIPOS
    for t in tipos:
        if t not in CATALOGO_TIPOS:
            raise HTTPException(status_code=422, detail=f"Tipo de catalogo invalido. Use: {CATALOGO_TIPOS}")
    return {
        "items": [
            item
            for t in tipos
            for item in _catalogo_items(db, t, solo_activos=solo_activos)
        ]
    }


@router.post("/catalogo", status_code=status.HTTP_201_CREATED, summary="Crear item del catalogo de inventario")
def crear_catalogo_inventario(
    data: CatalogoInventarioIn,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _asegurar_permiso_catalogo_inventario(db, current_user)
    tipo = data.tipo.upper()
    if tipo not in CATALOGO_TIPOS:
        raise HTTPException(status_code=422, detail=f"Tipo de catalogo invalido. Use: {CATALOGO_TIPOS}")
    alcance = (data.alcance or "AMBOS").upper()
    if alcance not in CATALOGO_ALCANCES:
        raise HTTPException(status_code=422, detail=f"Alcance invalido. Use: {CATALOGO_ALCANCES}")
    clave = _normalizar_clave_catalogo(data.clave or data.nombre)
    base_claves = {i["clave"] for i in _catalogo_base_items(tipo)}
    if clave in base_claves:
        raise HTTPException(status_code=409, detail="Ese elemento ya existe como catalogo base")
    if db.query(CatalogoInventarioItem).filter(
        CatalogoInventarioItem.tipo == tipo,
        CatalogoInventarioItem.clave == clave,
    ).first():
        raise HTTPException(status_code=409, detail=f"Ya existe un elemento con clave '{clave}'")
    prefijo = None
    if tipo == CATALOGO_TIPO_CATEGORIA:
        prefijo = _normalizar_clave_catalogo(data.prefijo_codigo or clave)[:12]
    item = CatalogoInventarioItem(
        tipo=tipo,
        clave=clave,
        nombre=data.nombre.strip(),
        prefijo_codigo=prefijo,
        alcance=alcance,
        activo=data.activo,
        protegido=False,
        creado_por_id=current_user.id,
        creado_en=_utcnow(),
        actualizado_en=_utcnow(),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _serializar_catalogo_inventario(item)


@router.put("/catalogo/{item_id}", summary="Actualizar item del catalogo de inventario")
def actualizar_catalogo_inventario(
    item_id: int,
    data: CatalogoInventarioUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _asegurar_permiso_catalogo_inventario(db, current_user)
    item = db.query(CatalogoInventarioItem).filter(CatalogoInventarioItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Elemento de catalogo no encontrado")
    if item.protegido:
        raise HTTPException(status_code=409, detail="Los elementos base no se pueden editar")
    cambios = data.model_dump(exclude_unset=True)
    if "nombre" in cambios and cambios["nombre"] is not None:
        item.nombre = cambios["nombre"].strip()
    if "prefijo_codigo" in cambios and item.tipo == CATALOGO_TIPO_CATEGORIA:
        item.prefijo_codigo = _normalizar_clave_catalogo(cambios["prefijo_codigo"] or item.clave)[:12]
    if "alcance" in cambios and cambios["alcance"] is not None:
        alcance = cambios["alcance"].upper()
        if alcance not in CATALOGO_ALCANCES:
            raise HTTPException(status_code=422, detail=f"Alcance invalido. Use: {CATALOGO_ALCANCES}")
        item.alcance = alcance
    if "activo" in cambios and cambios["activo"] is not None:
        item.activo = cambios["activo"]
    item.actualizado_en = _utcnow()
    db.commit()
    db.refresh(item)
    return _serializar_catalogo_inventario(item)


@router.get("/categorias", summary="Categorías disponibles")
def listar_categorias(
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    categorias_items = _catalogo_items(db, CATALOGO_TIPO_CATEGORIA, solo_activos=True)
    tipos_ubicacion_items = _catalogo_items(db, CATALOGO_TIPO_UBICACION, solo_activos=True)
    return {
        "categorias": [i["clave"] for i in categorias_items],
        "categorias_items": categorias_items,
        "estados": ESTADOS_ACTIVO,
        "condiciones": CONDICIONES,
        "alcances": ALCANCES_ACTIVO,
        "tipos_inventario": TIPOS_INVENTARIO,
        "estados_admin": ESTADOS_ADMIN_ACTIVO,
        "tipos_ubicacion": [i["clave"] for i in tipos_ubicacion_items],
        "tipos_ubicacion_items": tipos_ubicacion_items,
        "tipos_movimiento": TIPOS_MOVIMIENTO,
        "estados_movimiento": ESTADOS_MOVIMIENTO,
        "estados_baja": ESTADOS_BAJA,
        "estados_levantamiento": ESTADOS_LEVANTAMIENTO,
        "estados_revision_levantamiento": ESTADOS_REVISION_LEVANTAMIENTO,
    }


@router.get("/labs-nombres", summary="Nombres exactos de laboratorios activos")
def labs_nombres(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Devuelve los nombres tal como están registrados, para usarlos en la plantilla."""
    query = db.query(Laboratorio).filter(Laboratorio.activo == True)
    if _es_rol_laboratorio(current_user):
        query = query.filter(Laboratorio.id == current_user.laboratorio_id)
    labs = query.order_by(Laboratorio.nombre).all()
    return [{"id": l.id, "nombre": l.nombre} for l in labs]


@router.get("/estadisticas", summary="Resumen del inventario")
def estadisticas(
    laboratorio_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    ubicacion_id: Optional[int] = None,
    alcance: Optional[str] = None,
    tipo_inventario: Optional[str] = None,
    estado_admin: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    _actualizar_estado_prestamos(db)
    q_a = db.query(Activo).filter(Activo.activo == True)
    q_p = db.query(Prestamo)

    q_a = _filtrar_lab_asignado(q_a, Activo, current_user, laboratorio_id)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None:
        if not departamentos_visibles:
            activos = []
            ids = []
            return {
                "total_activos": 0, "operativos": 0, "en_mantenimiento": 0,
                "bajas": 0, "prestamos_activos": 0, "prestamos_vencidos": 0,
                "institucionales": 0, "bajas_pendientes": 0, "no_localizados": 0,
                "por_categoria": {}, "por_departamento": {}, "por_estado_admin": {},
            }
        if departamento_id:
            if departamento_id not in departamentos_visibles:
                raise HTTPException(status_code=403, detail="Solo puedes consultar estadisticas de tu departamento")
        else:
            q_a = q_a.filter(Activo.departamento_id.in_(departamentos_visibles))
    if departamento_id:
        q_a = q_a.filter(Activo.departamento_id == departamento_id)
    if ubicacion_id:
        q_a = q_a.filter(Activo.ubicacion_id == ubicacion_id)
    if alcance:
        q_a = q_a.filter(Activo.alcance == alcance.upper())
    if tipo_inventario:
        q_a = q_a.filter(Activo.tipo_inventario == tipo_inventario.upper())

    activos_scope = q_a.all()
    por_estado_admin = {}
    for activo in activos_scope:
        estado_validacion = activo.estado_admin or "VALIDADO"
        por_estado_admin[estado_validacion] = por_estado_admin.get(estado_validacion, 0) + 1

    if estado_admin:
        estado_filtro = estado_admin.upper()
        activos = [
            activo for activo in activos_scope
            if (activo.estado_admin or "VALIDADO") == estado_filtro
        ]
    else:
        activos = [
            activo for activo in activos_scope
            if (activo.estado_admin or "VALIDADO") == "VALIDADO"
        ]
    ids = [a.id for a in activos]
    prestamos = db.query(Prestamo).filter(Prestamo.activo_id.in_(ids)).all() if ids else []
    bajas = db.query(SolicitudBajaInventario).filter(SolicitudBajaInventario.activo_id.in_(ids)).all() if ids else []
    revisiones = db.query(RevisionLevantamientoInventario).filter(RevisionLevantamientoInventario.activo_id.in_(ids)).all() if ids else []
    ahora = _utcnow()
    en_7 = ahora + datetime.timedelta(days=7)
    mantenimientos = db.query(MantenimientoPreventivo).filter(
        MantenimientoPreventivo.activo_id.in_(ids),
        MantenimientoPreventivo.estado.in_(["PENDIENTE", "EN_PROCESO"]),
    ).all() if ids else []

    por_categoria = {}
    por_departamento = {}
    for a in activos:
        por_categoria[a.categoria] = por_categoria.get(a.categoria, 0) + 1
        key = a.departamento_id or 0
        por_departamento[key] = por_departamento.get(key, 0) + 1

    solicitudes_prestamo = {}
    for prestamo in prestamos:
        clave = prestamo.folio or f"PRE-{prestamo.id}"
        solicitudes_prestamo.setdefault(clave, []).append(prestamo)

    solicitudes_activas = 0
    solicitudes_vencidas = 0
    solicitudes_devueltas = 0
    for items in solicitudes_prestamo.values():
        if any(item.estado == "VENCIDO" for item in items):
            solicitudes_vencidas += 1
        elif any(item.estado == "ACTIVO" for item in items):
            solicitudes_activas += 1
        else:
            solicitudes_devueltas += 1

    return {
        "total_activos": len(activos),
        "operativos": sum(1 for a in activos if a.estado == "OPERATIVO"),
        "en_mantenimiento": sum(1 for a in activos if a.estado == "MANTENIMIENTO"),
        "dañados": sum(1 for a in activos if a.estado == "DAÑADO"),
        "prestamos_totales":  len(prestamos),
        "prestamos_activos":  sum(1 for p in prestamos if p.estado == "ACTIVO"),
        "prestamos_vencidos": sum(1 for p in prestamos if p.estado == "VENCIDO"),
        "prestamos_devueltos": sum(1 for p in prestamos if p.estado == "DEVUELTO"),
        "solicitudes_prestamo_totales": len(solicitudes_prestamo),
        "solicitudes_prestamo_activas": solicitudes_activas,
        "solicitudes_prestamo_vencidas": solicitudes_vencidas,
        "solicitudes_prestamo_devueltas": solicitudes_devueltas,
        "por_categoria": por_categoria,
        "por_estado_admin": por_estado_admin,
        "institucionales": sum(1 for a in activos if (a.alcance or "").upper() == "INSTITUCIONAL"),
        "de_laboratorio": sum(1 for a in activos if (a.alcance or "LABORATORIO").upper() == "LABORATORIO"),
        "activos_individuales": sum(1 for a in activos if (a.tipo_inventario or "ACTIVO").upper() == "ACTIVO"),
        "bajas_pendientes": sum(1 for b in bajas if b.estado in ("SOLICITADA", "EN_REVISION", "VALIDADA_FISICAMENTE", "AUTORIZADA")),
        "bajas_ejecutadas": sum(1 for b in bajas if b.estado == "EJECUTADA"),
        "no_localizados": sum(1 for r in revisiones if r.estado == "NO_LOCALIZADO"),
        "propuestos_baja": sum(1 for r in revisiones if r.estado == "PROPUESTO_BAJA"),
        "mantenimientos_pendientes": len(mantenimientos),
        "mantenimientos_vencidos": sum(1 for m in mantenimientos if m.fecha_limite and m.fecha_limite < ahora),
        "mantenimientos_proximos": sum(1 for m in mantenimientos if m.fecha_limite and ahora <= m.fecha_limite <= en_7),
        "por_departamento": por_departamento,
    }


@router.get("/mantenimiento-alertas", summary="Alertas de mantenimiento preventivo por inventario")
def alertas_mantenimiento_inventario(
    laboratorio_id: Optional[int] = None,
    departamento_id: Optional[int] = None,
    ubicacion_id: Optional[int] = None,
    alcance: Optional[str] = None,
    estado_admin: Optional[str] = None,
    categoria: Optional[str] = None,
    estado: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    q = _query_activos_filtrados(
        db,
        current_user,
        laboratorio_id=laboratorio_id,
        departamento_id=departamento_id,
        ubicacion_id=ubicacion_id,
        alcance=alcance,
        estado_admin=estado_admin,
        categoria=categoria,
        estado=estado,
        solo_activos=True,
    )
    activos = [] if q is None else q.all()
    activos_map = {a.id: a for a in activos}
    ids = list(activos_map.keys())
    ahora = _utcnow()
    en_7 = ahora + datetime.timedelta(days=7)
    mantenimientos = db.query(MantenimientoPreventivo).filter(
        MantenimientoPreventivo.activo_id.in_(ids),
        MantenimientoPreventivo.estado.in_(["PENDIENTE", "EN_PROCESO"]),
    ).order_by(
        MantenimientoPreventivo.fecha_limite.is_(None),
        MantenimientoPreventivo.fecha_limite,
        MantenimientoPreventivo.fecha_programada,
    ).all() if ids else []

    items = []
    for m in mantenimientos[:30]:
        activo = activos_map.get(m.activo_id)
        if not activo:
            continue
        dep = db.query(Departamento).filter(Departamento.id == activo.departamento_id).first() if activo.departamento_id else None
        lab = db.query(Laboratorio).filter(Laboratorio.id == activo.laboratorio_id).first() if activo.laboratorio_id else None
        estado_alerta = "PROGRAMADO"
        if m.fecha_limite and m.fecha_limite < ahora:
            estado_alerta = "VENCIDO"
        elif m.fecha_limite and ahora <= m.fecha_limite <= en_7:
            estado_alerta = "PROXIMO"
        items.append({
            "id": m.id,
            "activo_id": activo.id,
            "codigo_inventario": activo.codigo_inventario,
            "activo_nombre": activo.nombre,
            "departamento_nombre": dep.nombre if dep else None,
            "laboratorio_nombre": lab.nombre if lab else None,
            "tipo": m.tipo,
            "estado": m.estado,
            "estado_alerta": estado_alerta,
            "fecha_programada": m.fecha_programada.isoformat() if m.fecha_programada else None,
            "fecha_limite": m.fecha_limite.isoformat() if m.fecha_limite else None,
        })

    return {
        "total_pendientes": len(mantenimientos),
        "vencidos": sum(1 for m in mantenimientos if m.fecha_limite and m.fecha_limite < ahora),
        "proximos_7": sum(1 for m in mantenimientos if m.fecha_limite and ahora <= m.fecha_limite <= en_7),
        "programados": sum(1 for m in mantenimientos if not m.fecha_limite or m.fecha_limite > en_7),
        "items": items,
    }


# ─── Importación masiva desde Plantilla_Inventario_UTC.xlsx ────────────────────
#
# Columnas de la plantilla simplificada (0-indexed):
#   A(0)=# (fila)   B(1)=NOMBRE*   C(2)=LABORATORIO*   D(3)=CATEGORÍA*
#   E(4)=ÁREA        F(5)=MARCA      G(6)=MODELO          H(7)=NUM_SERIE
#   I(8)=ESPECIFICACIONES   J(9)=VALOR   K(10)=ESTADO
#   L(11)=RESGUARDO   M(12)=OBSERVACIONES
#
# El sistema genera el número de inventario automáticamente (UTC-[ÁREA]-[TIPO]-[SEQ])

@router.post("/activos/importar", summary="Importar activos desde Plantilla_Inventario_UTC.xlsx")
async def importar_activos(
    request: Request,
    file: UploadFile = File(...),
    estado_admin_default: str = "BORRADOR",
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")
    if not _es_admin_inventario_global(current_user) and not departamentos_inventario(db, current_user):
        raise HTTPException(status_code=403, detail="No tienes permiso para importar inventario departamental")
    estado_admin_default = (estado_admin_default or "BORRADOR").upper()
    if estado_admin_default not in ESTADOS_ADMIN_ACTIVO:
        raise HTTPException(status_code=422, detail=f"Estado administrativo invalido. Use: {ESTADOS_ADMIN_ACTIVO}")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(400, "Archivo Excel inválido o dañado")

    ws = wb["Inventario"] if "Inventario" in wb.sheetnames else wb.active

    def _v(row, idx):
        v = row[idx] if idx < len(row) else None
        return str(v).strip() if v is not None else ""

    # Cache de laboratorios (nombre normalizado sin acentos → objeto)
    labs_query = db.query(Laboratorio).filter(Laboratorio.activo == True)
    if _es_rol_laboratorio(current_user):
        labs_query = labs_query.filter(Laboratorio.id == current_user.laboratorio_id)
    labs_list = labs_query.all()
    labs = {_normalizar(lab.nombre): lab for lab in labs_list}
    departamentos_list = db.query(Departamento).filter(Departamento.activo == True).all()
    departamentos_por_nombre = {_normalizar(d.nombre): d for d in departamentos_list}
    departamentos_por_clave = {_normalizar(d.clave): d for d in departamentos_list if d.clave}

    def _resolver_departamento_excel(texto: str) -> int | None:
        if not texto:
            return current_user.departamento_id
        key = _normalizar(texto)
        dep = departamentos_por_clave.get(key) or departamentos_por_nombre.get(key)
        return dep.id if dep else None

    def _normalizar_numero_oficial(valor: str) -> str | None:
        valor = (valor or "").strip()
        return valor or None

    def _buscar_existente(codigo_siga: str | None, numero_oficial: str | None) -> Activo | None:
        if codigo_siga:
            existente = db.query(Activo).filter(Activo.codigo_inventario == codigo_siga).first()
            if existente:
                return existente
        if numero_oficial:
            return db.query(Activo).filter(Activo.numero_oficial == numero_oficial).first()
        return None

    creados = 0
    actualizados = 0
    duplicados_posibles = 0
    errores = []
    activos_creados_ids = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        # Fila completamente vacía → fin del listado
        if all(v is None for v in row):
            break
        # Fila de ejemplo (primera celda es "→")
        if str(row[0] or "").strip() == "→":
            continue

        nombre     = _v(row, 1)
        lab_nombre = _v(row, 2).upper()
        categoria  = _v(row, 3).upper() or "OTRO"
        area       = _v(row, 4) or None          # prefijo opcional (ej. "LTI")
        marca      = _v(row, 5) or None
        modelo     = _v(row, 6) or None
        num_serie  = _v(row, 7) or None
        specs      = _v(row, 8) or None
        valor_raw  = row[9]  if 9  < len(row) else None
        estado     = _v(row, 10).upper() or "OPERATIVO"
        resguardo  = _v(row, 11) or None
        obs        = _v(row, 12) or None
        codigo_siga = _v(row, 13).upper() or None
        numero_oficial = _normalizar_numero_oficial(_v(row, 14))
        departamento_txt = _v(row, 15)
        alcance_excel = _v(row, 16).upper()
        estado_admin = _v(row, 17).upper() or estado_admin_default
        if not puede_validar_inventario(db, current_user):
            estado_admin = "BORRADOR"

        fila_errs = []
        if not nombre:
            fila_errs.append("Nombre/descripción requerido (columna B)")

        # Normalizar categoría
        if categoria not in _catalogo_claves(db, CATALOGO_TIPO_CATEGORIA):
            categoria = "OTRO"

        # Normalizar estado
        if estado not in ESTADOS_ACTIVO:
            estado = "OPERATIVO"
        if estado_admin not in ESTADOS_ADMIN_ACTIVO:
            fila_errs.append(f"Estado administrativo invalido: {estado_admin}")

        # Parsear valor monetario
        try:
            valor = float(valor_raw) if valor_raw is not None else None
        except (ValueError, TypeError):
            valor = None

        alcance = "LABORATORIO" if (
            alcance_excel == "LABORATORIO" or
            (_es_admin_inventario_global(current_user) and not alcance_excel and _v(row, 2))
        ) else "INSTITUCIONAL"
        lab = None
        departamento_id = None

        if _es_rol_laboratorio(current_user):
            alcance = "LABORATORIO"
            lab = _buscar_lab(_v(row, 2), labs, db, current_user)
            if not lab and current_user.laboratorio_id:
                lab = db.query(Laboratorio).filter(
                    Laboratorio.id == current_user.laboratorio_id,
                    Laboratorio.activo == True,
                ).first()
            if not lab:
                fila_errs.append("Tu usuario no tiene un laboratorio asignado para importar activos")
        elif alcance == "LABORATORIO":
            lab = _buscar_lab(_v(row, 2), labs, db, current_user)
            if not lab:
                nombres_disponibles = ", ".join(f"'{l.nombre}'" for l in labs_list)
                fila_errs.append(
                    f"Laboratorio '{_v(row, 2)}' no encontrado. "
                    f"Nombres disponibles: {nombres_disponibles}"
                )
        else:
            departamento_id = _resolver_departamento_excel(departamento_txt)
            if not departamento_id:
                fila_errs.append(f"Departamento '{departamento_txt}' no encontrado")
            else:
                try:
                    departamento_id = _resolver_departamento_inventario(db, current_user, departamento_id)
                    _asegurar_write_inventario_departamento(db, current_user, departamento_id)
                except HTTPException as exc:
                    fila_errs.append(str(exc.detail))

        if fila_errs:
            errores.append({"fila": row_idx, "codigo": "—", "nombre": nombre or "—", "errores": fila_errs})
            continue

        # Generar número de inventario automáticamente
        existente = _buscar_existente(codigo_siga, numero_oficial)
        if existente:
            try:
                if _es_rol_laboratorio(current_user):
                    assert_lab_write(existente.laboratorio_id, current_user)
                else:
                    _asegurar_write_inventario_departamento(db, current_user, existente.departamento_id)
            except HTTPException as exc:
                errores.append({"fila": row_idx, "codigo": existente.codigo_inventario, "nombre": nombre, "errores": [str(exc.detail)]})
                continue

            if numero_oficial:
                otro = db.query(Activo).filter(
                    Activo.numero_oficial == numero_oficial,
                    Activo.id != existente.id,
                ).first()
                if otro:
                    errores.append({
                        "fila": row_idx,
                        "codigo": existente.codigo_inventario,
                        "nombre": nombre,
                        "errores": [f"El numero oficial '{numero_oficial}' ya pertenece a {otro.codigo_inventario}"],
                    })
                    continue

            existente.nombre = nombre
            existente.categoria = categoria
            existente.area = area
            existente.marca = marca
            existente.modelo = modelo
            existente.numero_serie = num_serie
            existente.especificaciones = specs
            existente.valor = valor
            existente.estado = estado
            existente.estado_admin = estado_admin
            existente.numero_oficial = numero_oficial
            existente.resguardante_externo_nombre = resguardo
            existente.observaciones = obs
            existente.tipo_inventario = "ACTIVO"
            existente.cantidad = 1.0
            existente.unidad_medida = "PIEZA"
            existente.stock_minimo = None
            existente.alcance = alcance
            existente.laboratorio_id = lab.id if lab else None
            existente.departamento_id = departamento_id
            db.flush()
            actualizados += 1
            continue

        if not codigo_siga and num_serie:
            posible = db.query(Activo).filter(
                Activo.numero_serie == num_serie,
                Activo.nombre == nombre,
            ).first()
            if posible:
                duplicados_posibles += 1
                errores.append({
                    "fila": row_idx,
                    "codigo": posible.codigo_inventario,
                    "nombre": nombre,
                    "errores": [
                        "Posible duplicado por nombre y numero de serie. Agrega Codigo SIGA o No. oficial para actualizarlo.",
                    ],
                })
                continue

        area_codigo = area
        if not area_codigo and departamento_id:
            dep = db.query(Departamento).filter(Departamento.id == departamento_id).first()
            area_codigo = dep.clave if dep else None
        codigo = codigo_siga or _generar_codigo(db, categoria, area_codigo)

        if db.query(Activo).filter(Activo.codigo_inventario == codigo).first():
            errores.append({"fila": row_idx, "codigo": codigo, "nombre": nombre, "errores": [f"Ya existe un activo con codigo '{codigo}'"]})
            continue
        if numero_oficial and db.query(Activo).filter(Activo.numero_oficial == numero_oficial).first():
            errores.append({"fila": row_idx, "codigo": codigo, "nombre": nombre, "errores": [f"Ya existe un activo con numero oficial '{numero_oficial}'"]})
            continue

        nuevo_activo = Activo(
            codigo_inventario = codigo,
            numero_oficial    = numero_oficial,
            nombre            = nombre,
            alcance           = alcance,
            laboratorio_id    = lab.id if lab else None,
            departamento_id   = departamento_id,
            tipo_inventario   = "ACTIVO",
            estado_admin      = estado_admin,
            categoria         = categoria,
            area              = area,
            marca             = marca,
            modelo            = modelo,
            numero_serie      = num_serie,
            especificaciones  = specs,
            valor             = valor,
            cantidad          = 1.0,
            unidad_medida     = "PIEZA",
            stock_minimo      = None,
            estado            = estado,
            resguardante_externo_nombre = resguardo,
            observaciones     = obs,
            fecha_adquisicion = _utcnow(),
        )
        db.add(nuevo_activo)
        # Flush para que _generar_codigo en la siguiente fila vea este código
        db.flush()
        activos_creados_ids.append(nuevo_activo.id)
        creados += 1

    db.commit()
    for activo_id in activos_creados_ids:
        registrar(
            db,
            Accion.CREAR_ACTIVO,
            Recurso.ACTIVO,
            usuario=current_user,
            recurso_id=activo_id,
            detalle={"origen": "IMPORTACION_EXCEL"},
            request=request,
        )
    registrar(
        db,
        Accion.IMPORTAR_ACTIVOS,
        Recurso.ACTIVO,
        usuario=current_user,
        detalle={
            "creados": creados,
            "actualizados": actualizados,
            "duplicados_posibles": duplicados_posibles,
            "errores": len(errores),
            "estado_admin_default": estado_admin_default,
        },
        request=request,
    )
    return {
        "creados":       creados,
        "actualizados":  actualizados,
        "duplicados_posibles": duplicados_posibles,
        "estado_admin_default": estado_admin_default,
        "total_errores": len(errores),
        "errores":       errores,
    }


# ══════════════════════════════════════════════════════════════════════════════
# MANTENIMIENTO PREVENTIVO
# ══════════════════════════════════════════════════════════════════════════════

TIPOS_MANT = [
    "LIMPIEZA_FISICA","REVISION_SOFTWARE","ACTUALIZACION","REVISION_HARDWARE",
    "FORMATEO","RESPALDO","INSPECCION","REPARACION_MOBILIARIO","AJUSTE_MOBILIARIO",
    "SUSTITUCION_PARTES","LIMPIEZA_GENERAL","ELECTRICO","CLIMATIZACION","PLOMERIA",
    "PINTURA","SEGURIDAD","CALIBRACION","DESINFECCION","OTRO",
]
PERIODOS   = ["SEMANAL","MENSUAL","TRIMESTRAL","SEMESTRAL","ANUAL","UNICO"]
ESTADOS_MP = ["PENDIENTE","EN_PROCESO","COMPLETADO","OMITIDO"]

class MantPrevCreate(BaseModel):
    activo_id:        Optional[int]  = None
    computadora_id:   Optional[int]  = None
    laboratorio_id:   Optional[int]  = None
    tipo:             str
    periodicidad:     str             = "TRIMESTRAL"
    fecha_programada: str             # ISO string
    fecha_limite:     Optional[str]   = None
    descripcion:      Optional[str]   = None
    checklist:        Optional[str]   = None   # JSON string

class MantPrevUpdate(BaseModel):
    estado:           Optional[str]  = None
    fecha_inicio:     Optional[str]  = None
    fecha_completado: Optional[str]  = None
    notas_result:     Optional[str]  = None
    costo:            Optional[float]= None
    duracion_min:     Optional[int]  = None
    checklist:        Optional[str]  = None
    descripcion:      Optional[str]  = None
    fecha_programada: Optional[str]  = None
    fecha_limite:     Optional[str]  = None


def _serializar_mp(mp: MantenimientoPreventivo, db: Session) -> dict:
    activo_nombre = None
    activo_categoria = None
    activo_alcance = None
    activo_responsable = None
    activo_ubicacion = None
    if mp.activo_id:
        a = db.query(Activo).filter(Activo.id == mp.activo_id).first()
        if a:
            activo_nombre = f"{a.nombre} ({a.codigo_inventario})"
            activo_categoria = a.categoria
            activo_alcance = a.alcance
            activo_responsable = a.resguardante_externo_nombre
            activo_ubicacion = a.ubicacion_nombre
    from models.laboratorio import Computadora
    pc_codigo = None
    if mp.computadora_id:
        pc = db.query(Computadora).filter(Computadora.id == mp.computadora_id).first()
        if pc:
            pc_codigo = pc.codigo
    completado_por = None
    if mp.completado_por_id:
        u = db.query(Usuario).filter(Usuario.id == mp.completado_por_id).first()
        if u:
            completado_por = u.nombre
    lab_nombre = None
    if mp.laboratorio_id:
        l = db.query(Laboratorio).filter(Laboratorio.id == mp.laboratorio_id).first()
        if l:
            lab_nombre = l.nombre
    return {
        "id":               mp.id,
        "activo_id":        mp.activo_id,
        "activo_nombre":    activo_nombre,
        "activo_categoria": activo_categoria,
        "activo_alcance":   activo_alcance,
        "activo_responsable": activo_responsable,
        "activo_ubicacion": activo_ubicacion,
        "computadora_id":   mp.computadora_id,
        "pc_codigo":        pc_codigo,
        "laboratorio_id":   mp.laboratorio_id,
        "laboratorio_nombre": lab_nombre,
        "tipo":             mp.tipo,
        "periodicidad":     mp.periodicidad,
        "fecha_programada": mp.fecha_programada.isoformat() if mp.fecha_programada else None,
        "fecha_limite":     mp.fecha_limite.isoformat() if mp.fecha_limite else None,
        "estado":           mp.estado,
        "fecha_inicio":     mp.fecha_inicio.isoformat() if mp.fecha_inicio else None,
        "fecha_completado": mp.fecha_completado.isoformat() if mp.fecha_completado else None,
        "completado_por":   completado_por,
        "descripcion":      mp.descripcion,
        "checklist":        mp.checklist,
        "notas_result":     mp.notas_result,
        "costo":            mp.costo,
        "duracion_min":     mp.duracion_min,
        "fecha_creacion":   mp.fecha_creacion.isoformat() if mp.fecha_creacion else None,
    }


@router.get("/mantenimientos-preventivos", summary="Listar mantenimientos preventivos")
def listar_mantenimientos(
    laboratorio_id: Optional[int] = None,
    estado:         Optional[str] = None,
    activo_id:      Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(MantenimientoPreventivo)
    if laboratorio_id:
        q = q.filter(MantenimientoPreventivo.laboratorio_id == laboratorio_id)
    elif _es_rol_laboratorio(current_user) and current_user.laboratorio_id:
        q = q.filter(MantenimientoPreventivo.laboratorio_id == current_user.laboratorio_id)
    elif not _es_admin_inventario_global(current_user):
        departamentos_visibles = _departamentos_visibles_inventario(db, current_user) or []
        if not departamentos_visibles:
            return []
        activos_ids = [
            row[0] for row in db.query(Activo.id)
            .filter(Activo.departamento_id.in_(departamentos_visibles))
            .all()
        ]
        if not activos_ids:
            return []
        q = q.filter(MantenimientoPreventivo.activo_id.in_(activos_ids))
    if estado:
        q = q.filter(MantenimientoPreventivo.estado == estado.upper())
    if activo_id:
        activo = db.query(Activo).filter(Activo.id == activo_id).first()
        _asegurar_acceso_activo_departamental(db, activo, current_user)
        q = q.filter(MantenimientoPreventivo.activo_id == activo_id)
    items = q.order_by(MantenimientoPreventivo.fecha_programada).all()
    return [_serializar_mp(m, db) for m in items]


@router.post("/mantenimientos-preventivos", summary="Programar mantenimiento preventivo")
def crear_mantenimiento(
    data: MantPrevCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    activo = db.query(Activo).filter(Activo.id == data.activo_id).first() if data.activo_id else None
    if activo:
        _asegurar_acceso_activo_departamental(db, activo, current_user)
        _asegurar_activo_validado(activo)
        _asegurar_write_inventario_departamento(db, current_user, activo.departamento_id)
    elif data.laboratorio_id:
        assert_lab_write(data.laboratorio_id, current_user)
    elif current_user.rol not in (RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB):
        raise HTTPException(status_code=403, detail="Selecciona un activo de tu departamento para programar mantenimiento")

    tipo = data.tipo.upper()
    periodicidad = data.periodicidad.upper()
    if tipo not in TIPOS_MANT:
        raise HTTPException(status_code=422, detail="Tipo de mantenimiento no valido")
    if periodicidad not in PERIODOS:
        raise HTTPException(status_code=422, detail="Periodicidad no valida")

    mp = MantenimientoPreventivo(
        activo_id        = data.activo_id,
        computadora_id   = data.computadora_id,
        laboratorio_id   = data.laboratorio_id or (activo.laboratorio_id if activo else None),
        tipo             = tipo,
        periodicidad     = periodicidad,
        fecha_programada = datetime.datetime.fromisoformat(data.fecha_programada),
        fecha_limite     = datetime.datetime.fromisoformat(data.fecha_limite) if data.fecha_limite else None,
        descripcion      = data.descripcion,
        checklist        = data.checklist,
        estado           = "PENDIENTE",
    )
    db.add(mp)
    db.commit()
    db.refresh(mp)
    return _serializar_mp(mp, db)


@router.put("/mantenimientos-preventivos/{mp_id}", summary="Actualizar mantenimiento preventivo")
def actualizar_mantenimiento(
    mp_id: int,
    data: MantPrevUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    mp = db.query(MantenimientoPreventivo).filter(MantenimientoPreventivo.id == mp_id).first()
    if not mp:
        raise HTTPException(status_code=404, detail="Mantenimiento no encontrado")
    activo = db.query(Activo).filter(Activo.id == mp.activo_id).first() if mp.activo_id else None
    if activo:
        _asegurar_acceso_activo_departamental(db, activo, current_user)
        _asegurar_write_inventario_departamento(db, current_user, activo.departamento_id)
    elif mp.laboratorio_id:
        assert_lab_write(mp.laboratorio_id, current_user)
    elif current_user.rol not in (RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB):
        raise HTTPException(status_code=403, detail="No puedes actualizar este mantenimiento")

    if data.estado:
        mp.estado = data.estado.upper()
        if activo and data.estado.upper() == "EN_PROCESO":
            activo.estado = "MANTENIMIENTO"
        if data.estado.upper() == "COMPLETADO":
            mp.completado_por_id = current_user.id
            mp.fecha_completado  = _utcnow()
            if activo and activo.estado == "MANTENIMIENTO":
                activo.estado = "OPERATIVO"
            # Auto-generar el siguiente según periodicidad
            delta_map = {
                "SEMANAL":    datetime.timedelta(weeks=1),
                "MENSUAL":    datetime.timedelta(days=30),
                "TRIMESTRAL": datetime.timedelta(days=90),
                "SEMESTRAL":  datetime.timedelta(days=180),
                "ANUAL":      datetime.timedelta(days=365),
            }
            delta = delta_map.get(mp.periodicidad)
            if delta:
                siguiente = MantenimientoPreventivo(
                    activo_id        = mp.activo_id,
                    computadora_id   = mp.computadora_id,
                    laboratorio_id   = mp.laboratorio_id,
                    tipo             = mp.tipo,
                    periodicidad     = mp.periodicidad,
                    fecha_programada = mp.fecha_completado + delta,
                    fecha_limite     = (mp.fecha_completado + delta + datetime.timedelta(days=14)),
                    descripcion      = mp.descripcion,
                    checklist        = mp.checklist,
                    estado           = "PENDIENTE",
                )
                db.add(siguiente)

    if data.fecha_inicio:
        mp.fecha_inicio = datetime.datetime.fromisoformat(data.fecha_inicio)
    if data.notas_result is not None:
        mp.notas_result = data.notas_result
    if data.costo is not None:
        mp.costo = data.costo
    if data.duracion_min is not None:
        mp.duracion_min = data.duracion_min
    if data.checklist is not None:
        mp.checklist = data.checklist
    if data.descripcion is not None:
        mp.descripcion = data.descripcion
    if data.fecha_programada:
        mp.fecha_programada = datetime.datetime.fromisoformat(data.fecha_programada)
    if data.fecha_limite:
        mp.fecha_limite = datetime.datetime.fromisoformat(data.fecha_limite)

    db.commit()
    db.refresh(mp)
    return _serializar_mp(mp, db)


@router.delete("/mantenimientos-preventivos/{mp_id}", summary="Eliminar mantenimiento preventivo")
def eliminar_mantenimiento(
    mp_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    mp = db.query(MantenimientoPreventivo).filter(MantenimientoPreventivo.id == mp_id).first()
    if not mp:
        raise HTTPException(status_code=404, detail="No encontrado")
    activo = db.query(Activo).filter(Activo.id == mp.activo_id).first() if mp.activo_id else None
    if activo:
        _asegurar_acceso_activo_departamental(db, activo, current_user)
        _asegurar_write_inventario_departamento(db, current_user, activo.departamento_id)
    elif mp.laboratorio_id:
        assert_lab_write(mp.laboratorio_id, current_user)
    elif current_user.rol not in (RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN, RolUsuario.RESPONSABLE_LAB):
        raise HTTPException(status_code=403, detail="No puedes eliminar este mantenimiento")
    db.delete(mp)
    db.commit()
    return {"ok": True}


# ── Historial unificado por activo ────────────────────────────────────────────

@router.get("/activos/{activo_id}/historial", summary="Historial completo de un activo")
def historial_activo(
    activo_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    activo = db.query(Activo).filter(Activo.id == activo_id).first()
    if not activo:
        raise HTTPException(status_code=404, detail="Activo no encontrado")
    assert_resource_access(activo, current_user)
    departamentos_visibles = _departamentos_visibles_inventario(db, current_user)
    if departamentos_visibles is not None and activo.departamento_id not in departamentos_visibles:
        raise HTTPException(status_code=404, detail="Activo no encontrado")

    eventos = []

    # ── Incidentes ─────────────────────────────────────────────────────────
    for i in db.query(Incidente).filter(Incidente.activo_id == activo_id).all():
        rep_nombre = None
        if i.reportado_por_id:
            u = db.query(Usuario).filter(Usuario.id == i.reportado_por_id).first()
            if u: rep_nombre = u.nombre
        eventos.append({
            "tipo_evento":  "INCIDENTE",
            "fecha":        i.fecha_reporte.isoformat() if i.fecha_reporte else None,
            "fecha_fin":    i.fecha_resolucion.isoformat() if i.fecha_resolucion else None,
            "titulo":       f"{i.tipo.capitalize()} — {i.estado}",
            "descripcion":  i.descripcion,
            "estado":       i.estado,
            "prioridad":    i.prioridad,
            "notas":        i.notas_seguimiento,
            "costo":        i.costo_reparacion,
            "usuario":      rep_nombre,
            "id_ref":       i.id,
        })

    # ── Préstamos ──────────────────────────────────────────────────────────
    for p in db.query(Prestamo).filter(Prestamo.activo_id == activo_id).all():
        meta_prestamo = _meta_prestamo(p)
        eventos.append({
            "tipo_evento":  "PRESTAMO",
            "fecha":        p.fecha_salida.isoformat() if p.fecha_salida else None,
            "fecha_fin":    p.fecha_retorno_real.isoformat() if p.fecha_retorno_real else None,
            "titulo":       f"Prestamo - {p.solicitante_nombre}",
            "descripcion":  meta_prestamo["descripcion"],
            "receptor_tipo": meta_prestamo["receptor_tipo"],
            "proposito":    meta_prestamo["proposito"],
            "estado":       p.estado,
            "condicion_salida":  p.condicion_salida,
            "condicion_retorno": p.condicion_retorno,
            "notas":        p.observaciones_retorno,
            "usuario":      p.solicitante_nombre,
            "id_ref":       p.id,
        })

    # Mantenimientos preventivos
    for mp in db.query(MantenimientoPreventivo).filter(MantenimientoPreventivo.activo_id == activo_id).all():
        comp_nombre = None
        if mp.completado_por_id:
            u = db.query(Usuario).filter(Usuario.id == mp.completado_por_id).first()
            if u: comp_nombre = u.nombre
        eventos.append({
            "tipo_evento":  "MANTENIMIENTO_PREVENTIVO",
            "fecha":        mp.fecha_programada.isoformat() if mp.fecha_programada else None,
            "fecha_fin":    mp.fecha_completado.isoformat() if mp.fecha_completado else None,
            "titulo":       "Preventivo - " + mp.tipo.replace("_", " ").title(),
            "descripcion":  mp.descripcion,
            "estado":       mp.estado,
            "notas":        mp.notas_result,
            "costo":        mp.costo,
            "duracion_min": mp.duracion_min,
            "usuario":      comp_nombre,
            "id_ref":       mp.id,
        })

    eventos.sort(key=lambda e: e["fecha"] or "", reverse=True)
    return {
        "activo_id": activo_id,
        "activo": {
            "id": activo.id,
            "nombre": activo.nombre,
            "codigo": activo.codigo_inventario,
            "codigo_inventario": activo.codigo_inventario,
            "categoria": activo.categoria,
            "marca": activo.marca,
            "modelo": activo.modelo,
            "estado": activo.estado,
            "resguardante_externo_nombre": activo.resguardante_externo_nombre,
            "responsable_nombre": activo.responsable.nombre if activo.responsable else None,
            "laboratorio_id": activo.laboratorio_id,
        },
        "total_eventos": len(eventos),
        "eventos": eventos,
    }


@router.get("/computadoras/{pc_id}/historial", summary="Historial completo de una computadora")
def historial_computadora(
    pc_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    pc = db.query(Computadora).filter(Computadora.id == pc_id).first()
    if not pc:
        raise HTTPException(status_code=404, detail="Computadora no encontrada")
    if _es_rol_laboratorio(current_user) and current_user.laboratorio_id != pc.laboratorio_id:
        raise HTTPException(status_code=404, detail="Computadora no encontrada")

    lab = db.query(Laboratorio).filter(Laboratorio.id == pc.laboratorio_id).first()
    eventos = []

    for i in db.query(Incidente).filter(Incidente.computadora_id == pc_id).all():
        rep_nombre = None
        if i.reportado_por_id:
            u = db.query(Usuario).filter(Usuario.id == i.reportado_por_id).first()
            if u:
                rep_nombre = u.nombre
        eventos.append({
            "tipo_evento": "INCIDENTE",
            "fecha": i.fecha_reporte.isoformat() if i.fecha_reporte else None,
            "fecha_fin": i.fecha_resolucion.isoformat() if i.fecha_resolucion else None,
            "titulo": f"{i.tipo.capitalize()} - {i.estado}",
            "descripcion": i.descripcion,
            "estado": i.estado,
            "prioridad": i.prioridad,
            "notas": i.notas_seguimiento,
            "costo": i.costo_reparacion,
            "usuario": rep_nombre,
            "id_ref": i.id,
        })

    for mp in db.query(MantenimientoPreventivo).filter(MantenimientoPreventivo.computadora_id == pc_id).all():
        comp_nombre = None
        if mp.completado_por_id:
            u = db.query(Usuario).filter(Usuario.id == mp.completado_por_id).first()
            if u:
                comp_nombre = u.nombre
        eventos.append({
            "tipo_evento": "MANTENIMIENTO_PREVENTIVO",
            "fecha": mp.fecha_programada.isoformat() if mp.fecha_programada else None,
            "fecha_fin": mp.fecha_completado.isoformat() if mp.fecha_completado else None,
            "titulo": "Preventivo - " + mp.tipo.replace("_", " ").title(),
            "descripcion": mp.descripcion,
            "estado": mp.estado,
            "notas": mp.notas_result,
            "costo": mp.costo,
            "duracion_min": mp.duracion_min,
            "usuario": comp_nombre,
            "id_ref": mp.id,
        })

    eventos.sort(key=lambda e: e["fecha"] or "", reverse=True)
    return {
        "computadora_id": pc_id,
        "activo": {
            "id": pc.id,
            "nombre": pc.codigo,
            "codigo": pc.codigo,
            "codigo_inventario": pc.activo.codigo_inventario if pc.activo else None,
            "categoria": "Puesto de laboratorio",
            "marca": pc.activo.marca if pc.activo else None,
            "modelo": pc.activo.modelo if pc.activo else None,
            "estado": pc.estado,
            "laboratorio_id": pc.laboratorio_id,
            "laboratorio_nombre": lab.nombre if lab else None,
            "fila": pc.fila,
        },
        "total_eventos": len(eventos),
        "eventos": eventos,
    }
