from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import datetime, io, calendar, collections


def _utcnow() -> datetime.datetime:
    return datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None)


from database import get_db
from models.laboratorio import Laboratorio, Computadora
from models.sesion import SesionClase, AsignacionPC
from models.inventario import Activo, Prestamo, Incidente
from models.usuario import Usuario, RolUsuario
from models.adeudo import Adeudo
from dependencies import get_current_user
from rls import assert_report_access

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/reportes", tags=["Reportes"])

MESES_ES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
DIAS_ES  = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

# ─── helpers de estilo ────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

AZUL_OSC  = "1E3A5F"
AZUL_MED  = "2D6A9F"
AZUL_CLR  = "BDD7EE"
VERDE_OSC = "1B5E20"
VERDE_CLR = "C8E6C9"
ROJO_CLR  = "FFCDD2"
AMBAR_CLR = "FFF9C4"
GRIS_CLR  = "F5F5F5"
GRIS_ROW  = "EEF2F7"

def _set_hdr(ws, row, col, txt, bg=AZUL_OSC, fg="FFFFFF", bold=True, size=10,
             width=None, wrap=False):
    c = ws.cell(row=row, column=col, value=txt)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold, size=size, color=fg)
    c.alignment = _center(wrap)
    c.border    = _border("3A5A8A")
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def _data_row(ws, row, values, bg=None, fmts=None, aligns=None):
    fill_obj = _fill(bg) if bg else None
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        if fill_obj:
            c.fill = fill_obj
        c.font      = _font(size=10)
        c.border    = _border()
        c.alignment = _left()
        if fmts and fmts[i-1]:
            c.number_format = fmts[i-1]
        if aligns and aligns[i-1]:
            c.alignment = aligns[i-1]

# ─── Agregación de datos ──────────────────────────────────────────────────────

def _periodo(mes: int, anio: int):
    ini = datetime.datetime(anio, mes, 1)
    fin = datetime.datetime(anio, mes, calendar.monthrange(anio, mes)[1], 23, 59, 59)
    return ini, fin

def _datos_mes(db: Session, lab_id: int, mes: int, anio: int) -> dict:
    ini, fin = _periodo(mes, anio)

    lab = db.query(Laboratorio).filter(Laboratorio.id == lab_id).first()
    if not lab:
        return None

    # ── Sesiones ──────────────────────────────────────────────────────────────
    sesiones = db.query(SesionClase).filter(
        SesionClase.laboratorio_id == lab_id,
        SesionClase.inicio >= ini,
        SesionClase.inicio <= fin,
    ).all()

    # Docentes activos
    docentes_ids  = list({s.docente_id for s in sesiones})
    docentes_obj  = {u.id: u for u in db.query(Usuario).filter(Usuario.id.in_(docentes_ids)).all()}

    # Sesiones por docente
    ses_por_docente = collections.defaultdict(list)
    for s in sesiones:
        ses_por_docente[s.docente_id].append(s)

    # Alumnos (asignaciones únicas por matrícula)
    asig_ids  = [s.id for s in sesiones]
    asignaciones = db.query(AsignacionPC).filter(AsignacionPC.sesion_id.in_(asig_ids)).all() if asig_ids else []
    alumnos_unicos = {}
    for a in asignaciones:
        if a.alumno_matricula not in alumnos_unicos:
            alumnos_unicos[a.alumno_matricula] = {"nombre": a.alumno_nombre, "matricula": a.alumno_matricula, "sesiones": 0}
        alumnos_unicos[a.alumno_matricula]["sesiones"] += 1

    # Horas de uso
    horas_totales = 0.0
    for s in sesiones:
        fin_s = s.fin_real or s.fin_estimado
        if fin_s and s.inicio:
            horas_totales += (fin_s - s.inicio).total_seconds() / 3600

    # Horas pico (día semana × hora)
    horas_pico = collections.defaultdict(int)    # (dia_semana 0-6, hora) → count
    for s in sesiones:
        if s.inicio:
            horas_pico[(s.inicio.weekday(), s.inicio.hour)] += 1

    # ── Computadoras ─────────────────────────────────────────────────────────
    pcs = db.query(Computadora).filter(
        Computadora.laboratorio_id == lab_id,
        Computadora.activa == True
    ).all()
    pcs_operativas   = sum(1 for p in pcs if p.estado == "OPERATIVO")
    pcs_mant         = sum(1 for p in pcs if p.estado == "MANTENIMIENTO")
    pcs_danadas      = sum(1 for p in pcs if p.estado not in ("OPERATIVO","MANTENIMIENTO"))

    # ── Activos de inventario ─────────────────────────────────────────────────
    activos = db.query(Activo).filter(
        Activo.laboratorio_id == lab_id,
        Activo.activo == True
    ).all()
    activos_operativos = sum(1 for a in activos if a.estado == "OPERATIVO")
    activos_mant       = sum(1 for a in activos if a.estado == "MANTENIMIENTO")
    activos_danados    = sum(1 for a in activos if a.estado == "DAÑADO")

    # ── Préstamos del periodo ─────────────────────────────────────────────────
    activo_ids = [a.id for a in activos]
    prestamos = db.query(Prestamo).filter(
        Prestamo.activo_id.in_(activo_ids),
        Prestamo.fecha_salida >= ini,
        Prestamo.fecha_salida <= fin,
    ).all() if activo_ids else []
    activos_map = {a.id: a for a in activos}

    prestamos_activos  = sum(1 for p in prestamos if p.estado == "ACTIVO")
    prestamos_vencidos = sum(1 for p in prestamos if p.estado == "VENCIDO")
    prestamos_devueltos= sum(1 for p in prestamos if p.estado == "DEVUELTO")

    # ── Incidentes del periodo ────────────────────────────────────────────────
    incidentes = db.query(Incidente).filter(
        Incidente.laboratorio_id == lab_id,
        Incidente.fecha_reporte >= ini,
        Incidente.fecha_reporte <= fin,
    ).all()
    inc_pendientes = sum(1 for i in incidentes if i.estado == "PENDIENTE")
    inc_reparados  = sum(1 for i in incidentes if i.estado == "REPARADO")
    inc_baja       = sum(1 for i in incidentes if i.estado == "DADO_DE_BAJA")

    # ── Adeudos vinculados a los incidentes del periodo ───────────────────────
    inc_ids = [i.id for i in incidentes]
    adeudos_inc = db.query(Adeudo).filter(
        Adeudo.incidente_id.in_(inc_ids)
    ).all() if inc_ids else []
    # dict: incidente_id → adeudo (el más reciente si hubiera varios)
    adeudo_por_incidente: dict = {}
    for a in adeudos_inc:
        if a.incidente_id not in adeudo_por_incidente:
            adeudo_por_incidente[a.incidente_id] = a

    # ── Mes anterior para comparativa ─────────────────────────────────────────
    mes_ant  = mes - 1 if mes > 1 else 12
    anio_ant = anio if mes > 1 else anio - 1
    ini_a, fin_a = _periodo(mes_ant, anio_ant)
    ses_ant = db.query(SesionClase).filter(
        SesionClase.laboratorio_id == lab_id,
        SesionClase.inicio >= ini_a,
        SesionClase.inicio <= fin_a,
    ).count()
    asig_ant_ids = [s.id for s in db.query(SesionClase).filter(
        SesionClase.laboratorio_id == lab_id,
        SesionClase.inicio >= ini_a,
        SesionClase.inicio <= fin_a,
    ).all()]
    alumnos_ant = len({a.alumno_matricula for a in db.query(AsignacionPC).filter(
        AsignacionPC.sesion_id.in_(asig_ant_ids)
    ).all()}) if asig_ant_ids else 0

    return {
        "laboratorio": {"id": lab.id, "nombre": lab.nombre, "capacidad": lab.capacidad},
        "periodo": {"mes": mes, "anio": anio, "mes_nombre": MESES_ES[mes]},
        "sesiones": {
            "total": len(sesiones),
            "horas_total": round(horas_totales, 1),
            "detalle": sesiones,
            "por_docente": ses_por_docente,
            "horas_pico": horas_pico,
            "asignaciones": asignaciones,
        },
        "docentes": {
            "total": len(docentes_ids),
            "obj": docentes_obj,
        },
        "alumnos": {
            "total_unicos": len(alumnos_unicos),
            "detalle": list(alumnos_unicos.values()),
        },
        "pcs": {
            "total": len(pcs), "operativas": pcs_operativas,
            "mantenimiento": pcs_mant, "danadas": pcs_danadas,
        },
        "activos": {
            "total": len(activos), "operativos": activos_operativos,
            "mantenimiento": activos_mant, "danados": activos_danados,
            "detalle": activos,
        },
        "prestamos": {
            "total": len(prestamos), "activos": prestamos_activos,
            "vencidos": prestamos_vencidos, "devueltos": prestamos_devueltos,
            "detalle": prestamos,
        },
        "incidentes": {
            "total": len(incidentes), "pendientes": inc_pendientes,
            "reparados": inc_reparados, "baja": inc_baja,
            "detalle": incidentes,
        },
        "adeudo_por_incidente": adeudo_por_incidente,
        "activos_map": activos_map,
        "comparativa": {
            "sesiones_mes_ant": ses_ant,
            "alumnos_mes_ant": alumnos_ant,
            "mes_ant_nombre": MESES_ES[mes_ant],
        },
    }


# ─── Endpoint JSON ────────────────────────────────────────────────────────────

@router.get("/mensual", summary="Datos del reporte mensual (JSON)")
def reporte_mensual_json(
    laboratorio_id: int,
    mes:  int = None,
    anio: int = None,
    db:   Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    # RLS: LAB_ADMIN solo puede ver reportes de su propio laboratorio
    assert_report_access(laboratorio_id, current_user)

    hoy = datetime.date.today()
    mes  = mes  or hoy.month
    anio = anio or hoy.year
    if not (1 <= mes <= 12):
        raise HTTPException(400, "Mes inválido (1-12)")

    d = _datos_mes(db, laboratorio_id, mes, anio)
    if d is None:
        raise HTTPException(404, "Laboratorio no encontrado")

    return {
        "laboratorio":   d["laboratorio"],
        "periodo":       d["periodo"],
        "sesiones":      {k: v for k, v in d["sesiones"].items() if k not in ("detalle","por_docente","horas_pico")},
        "docentes":      {"total": d["docentes"]["total"]},
        "alumnos":       {"total_unicos": d["alumnos"]["total_unicos"]},
        "pcs":           d["pcs"],
        "activos":       {k: v for k, v in d["activos"].items() if k != "detalle"},
        "prestamos":     {k: v for k, v in d["prestamos"].items() if k != "detalle"},
        "incidentes":    {k: v for k, v in d["incidentes"].items() if k != "detalle"},
        "comparativa":   d["comparativa"],
    }


# ─── Generador Excel ──────────────────────────────────────────────────────────

def _build_excel(d: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    lab    = d["laboratorio"]
    per    = d["periodo"]
    titulo = f"Reporte Mensual — {lab['nombre']} — {per['mes_nombre']} {per['anio']}"
    fecha_gen = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── HOJA 1: RESUMEN ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 10

    # Portada
    ws.merge_cells("B2:H2")
    t = ws.cell(row=2, column=2, value=f"📊  {titulo}")
    t.font      = Font(bold=True, size=16, color="FFFFFF")
    t.fill      = _fill(AZUL_OSC)
    t.alignment = _left(wrap=False)

    ws.merge_cells("B3:H3")
    s = ws.cell(row=3, column=2, value=f"Universidad Tecnológica de Candelaria  |  Generado: {fecha_gen}")
    s.font      = Font(italic=True, size=10, color="A8C8F0")
    s.fill      = _fill(AZUL_MED)
    s.alignment = _left(wrap=False)

    # Tarjetas de métricas (fila 5-8)
    METRICAS = [
        ("🗓️ Sesiones", d["sesiones"]["total"], AZUL_OSC, "FFFFFF"),
        ("👩‍🏫 Docentes", d["docentes"]["total"], "1B4F72", "FFFFFF"),
        ("🎓 Alumnos", d["alumnos"]["total_unicos"], VERDE_OSC, "FFFFFF"),
        ("⏱️ Horas de uso", d["sesiones"]["horas_total"], "4A235A", "FFFFFF"),
        ("💻 PCs operativas", f"{d['pcs']['operativas']}/{d['pcs']['total']}", "1B6B3A", "FFFFFF"),
        ("🔧 En mantenimiento", d["pcs"]["mantenimiento"] + d["activos"]["mantenimiento"], "7D6608", "FFFFFF"),
        ("📤 Préstamos activos", d["prestamos"]["total"], "154360", "FFFFFF"),
        ("⚠️ Incidentes", d["incidentes"]["total"], "78281F", "FFFFFF"),
    ]
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 36
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 10

    for i, (etiq, val, bg, fg) in enumerate(METRICAS, 2):
        col = i
        ws.column_dimensions[get_column_letter(col)].width = 16
        lbl = ws.cell(row=5, column=col, value=etiq)
        lbl.font = Font(bold=True, size=9, color="AAAAAA")
        lbl.alignment = _center()

        num = ws.cell(row=6, column=col, value=val)
        num.font      = Font(bold=True, size=22, color=fg)
        num.fill      = _fill(bg)
        num.alignment = _center()
        num.border    = _border("FFFFFF")

    # Comparativa mes anterior
    ws.row_dimensions[9].height  = 10
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 20
    ws.row_dimensions[12].height = 20
    ws.row_dimensions[13].height = 20

    ws.merge_cells("B10:H10")
    h = ws.cell(row=10, column=2, value="COMPARATIVA VS MES ANTERIOR")
    h.font = Font(bold=True, size=10, color="FFFFFF")
    h.fill = _fill(AZUL_MED)
    h.alignment = _center()
    h.border = _border("3A5A8A")

    comp = d["comparativa"]
    ses_act   = d["sesiones"]["total"]
    ses_ant   = comp["sesiones_mes_ant"]
    alum_act  = d["alumnos"]["total_unicos"]
    alum_ant  = comp["alumnos_mes_ant"]

    def _delta(act, ant):
        if ant == 0:
            return "N/D"
        pct = ((act - ant) / ant) * 100
        return f"{'▲' if pct >= 0 else '▼'} {abs(pct):.0f}%"

    for r, (lbl, act, ant, mn) in enumerate([
        ("Sesiones realizadas",   ses_act,  ses_ant,  comp["mes_ant_nombre"]),
        ("Alumnos atendidos",     alum_act, alum_ant, comp["mes_ant_nombre"]),
    ], 11):
        ws.cell(row=r, column=2, value=lbl).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2).alignment = _left()
        ws.cell(row=r, column=3, value=f"{per['mes_nombre']}: {act}").alignment = _center()
        ws.cell(row=r, column=4, value=f"{mn}: {ant}").alignment = _center()
        d_cell = ws.cell(row=r, column=5, value=_delta(act, ant))
        d_cell.alignment = _center()
        d_cell.font = Font(bold=True, size=10,
            color="1B6B3A" if isinstance(_delta(act, ant), str) and "▲" in str(_delta(act, ant))
            else "78281F")

    # ── HOJA 2: SESIONES POR DOCENTE ─────────────────────────────────────────
    ws2 = wb.create_sheet("Sesiones por Docente")
    ws2.sheet_view.showGridLines = False

    HDRS2 = [("Docente",30),("Materia",28),("Grupo",10),
             ("Fecha",14),("Hora Inicio",13),("Hora Fin",13),
             ("Duración (h)",14),("Alumnos",10),("Estado",14)]
    ws2.merge_cells(f"A1:{get_column_letter(len(HDRS2))}1")
    t2 = ws2.cell(row=1, column=1, value=f"SESIONES — {per['mes_nombre']} {per['anio']}  |  {lab['nombre']}")
    t2.font = Font(bold=True, size=13, color="FFFFFF"); t2.fill = _fill(AZUL_OSC)
    t2.alignment = _left(); ws2.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(HDRS2, 1):
        _set_hdr(ws2, 2, col, h, width=w)

    # Índice de asignaciones por sesión para conteo rápido
    _asig_por_sesion: dict = {}
    for _a in d["sesiones"].get("asignaciones", []):
        _asig_por_sesion.setdefault(_a.sesion_id, []).append(_a)

    sesiones_ord = sorted(d["sesiones"]["detalle"], key=lambda s: s.inicio or datetime.datetime.min)
    for r, s in enumerate(sesiones_ord, 3):
        doc  = d["docentes"]["obj"].get(s.docente_id)
        nombre_doc = doc.nombre if doc else f"ID {s.docente_id}"
        fin_s = s.fin_real or s.fin_estimado
        dur   = round((fin_s - s.inicio).total_seconds() / 3600, 2) if fin_s and s.inicio else ""
        n_alumnos = len(_asig_por_sesion.get(s.id, []))
        fill_bg = GRIS_ROW if r % 2 == 0 else None
        _data_row(ws2, r, [
            nombre_doc, s.materia, s.grupo,
            s.inicio.strftime("%d/%m/%Y") if s.inicio else "",
            s.inicio.strftime("%H:%M") if s.inicio else "",
            fin_s.strftime("%H:%M") if fin_s else "—",
            dur, n_alumnos or "", s.estado
        ], bg=fill_bg)

    # ── HOJA 3: ALUMNOS ATENDIDOS ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Alumnos Atendidos")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:D1")
    t3 = ws3.cell(row=1, column=1, value=f"ALUMNOS ATENDIDOS — {per['mes_nombre']} {per['anio']}  |  {lab['nombre']}")
    t3.font = Font(bold=True, size=13, color="FFFFFF"); t3.fill = _fill(VERDE_OSC)
    t3.alignment = _left(); ws3.row_dimensions[1].height = 28

    for col, (h, w) in enumerate([("Matrícula",14),("Nombre",34),("Sesiones",12),("Lab",24)], 1):
        _set_hdr(ws3, 2, col, h, bg=VERDE_OSC, width=w)

    alumnos_ord = sorted(d["alumnos"]["detalle"], key=lambda a: a["nombre"])
    for r, al in enumerate(alumnos_ord, 3):
        fill_bg = GRIS_ROW if r % 2 == 0 else None
        _data_row(ws3, r, [al["matricula"], al["nombre"], al["sesiones"], lab["nombre"]], bg=fill_bg)

    ws3.cell(row=len(alumnos_ord)+4, column=1, value="TOTAL ALUMNOS ÚNICOS:").font = Font(bold=True)
    ws3.cell(row=len(alumnos_ord)+4, column=2, value=d["alumnos"]["total_unicos"]).font = Font(bold=True)

    # ── HOJA 4: HORAS PICO ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Horas Pico")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:H1")
    t4 = ws4.cell(row=1, column=1, value=f"DISTRIBUCIÓN DE USO — {per['mes_nombre']} {per['anio']}  |  {lab['nombre']}")
    t4.font = Font(bold=True, size=13, color="FFFFFF"); t4.fill = _fill("4A235A")
    t4.alignment = _left(); ws4.row_dimensions[1].height = 28

    ws4.cell(row=2, column=1, value="Hora / Día").font = Font(bold=True, size=10, color="FFFFFF")
    ws4.cell(row=2, column=1).fill = _fill("4A235A")
    ws4.cell(row=2, column=1).alignment = _center()
    ws4.column_dimensions["A"].width = 12

    for ci, dia in enumerate(DIAS_ES, 2):
        c = ws4.cell(row=2, column=ci, value=dia)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = _fill("4A235A"); c.alignment = _center()
        ws4.column_dimensions[get_column_letter(ci)].width = 12

    HORAS = list(range(7, 22))  # 7:00 a 21:00
    for ri, hora in enumerate(HORAS, 3):
        c = ws4.cell(row=ri, column=1, value=f"{hora:02d}:00 - {hora+1:02d}:00")
        c.font = Font(bold=True, size=9); c.fill = _fill("EEF2F7"); c.alignment = _center()
        ws4.row_dimensions[ri].height = 18
        for di in range(7):
            cnt = d["sesiones"]["horas_pico"].get((di, hora), 0)
            cell = ws4.cell(row=ri, column=di+2, value=cnt if cnt else "")
            cell.alignment = _center()
            if cnt >= 3:
                cell.fill = _fill("1B5E20"); cell.font = Font(bold=True, color="FFFFFF")
            elif cnt == 2:
                cell.fill = _fill("A5D6A7"); cell.font = Font(bold=True)
            elif cnt == 1:
                cell.fill = _fill("E8F5E9")
            else:
                cell.fill = _fill("FAFAFA")
            cell.border = _border()

    # ── HOJA 5: INVENTARIO ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Inventario")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:G1")
    t5 = ws5.cell(row=1, column=1, value=f"ESTADO DEL INVENTARIO — {per['mes_nombre']} {per['anio']}  |  {lab['nombre']}")
    t5.font = Font(bold=True, size=13, color="FFFFFF"); t5.fill = _fill("1B6B3A")
    t5.alignment = _left(); ws5.row_dimensions[1].height = 28

    # Resumen PCs
    ws5.cell(row=2, column=1, value="COMPUTADORAS").font = Font(bold=True, size=11)
    for col, (h, w) in enumerate([("Código",14),("Estado",14),("Especificaciones",40),("Lab",24)],2):
        _set_hdr(ws5, 3, col, h, bg="1B6B3A", width=w)
    _set_hdr(ws5, 3, 1, "#", bg="1B6B3A", width=5)

    pcs = d.get("pcs_detalle", [])  # se llenará si disponible
    # Resumen activos de inventario
    act_inicio = 5 + 3  # se ajusta según número de PCs

    ws5.cell(row=3+1, column=1, value="Resumen PCs:").font = Font(bold=True)
    for col, (lbl, val) in enumerate([
        ("Total", d["pcs"]["total"]),
        ("Operativas", d["pcs"]["operativas"]),
        ("Mantenimiento", d["pcs"]["mantenimiento"]),
        ("Con fallas", d["pcs"]["danadas"]),
    ], 1):
        ws5.cell(row=4, column=col).value = lbl
        ws5.cell(row=4, column=col).font  = Font(bold=True, size=9)
        ws5.cell(row=5, column=col).value = val
        ws5.cell(row=5, column=col).font  = Font(bold=True, size=14)
        ws5.cell(row=5, column=col).fill  = _fill(VERDE_CLR if lbl == "Operativas" else (AMBAR_CLR if lbl == "Mantenimiento" else (ROJO_CLR if lbl == "Con fallas" else GRIS_CLR)))
        ws5.cell(row=5, column=col).alignment = _center()

    # Activos de inventario
    ws5.row_dimensions[7].height = 8
    ws5.cell(row=8, column=1, value="ACTIVOS DE INVENTARIO").font = Font(bold=True, size=11)
    HDRS5 = [("#",5),("Código",18),("Nombre",30),("Categoría",16),
             ("Marca/Modelo",22),("Estado",14),("Resguardo",24)]
    for col, (h, w) in enumerate(HDRS5, 1):
        _set_hdr(ws5, 9, col, h, bg="1B6B3A", width=w)

    activos_ord = sorted(d["activos"]["detalle"], key=lambda a: (a.categoria, a.nombre))
    for r, a in enumerate(activos_ord, 10):
        fill_bg = GRIS_ROW if r % 2 == 0 else None
        _data_row(ws5, r, [
            r-9, a.codigo_inventario, a.nombre, a.categoria.replace("_"," "),
            f"{a.marca or ''} {a.modelo or ''}".strip(),
            a.estado,
            a.resguardo_nombre or "—"
        ], bg=fill_bg)
        # Color por estado
        estado_colors = {"OPERATIVO": "E8F5E9", "MANTENIMIENTO": "FFF9C4",
                         "DAÑADO": "FFCDD2", "BAJA": "EEEEEE"}
        ws5.cell(row=r, column=6).fill = _fill(estado_colors.get(a.estado, "FFFFFF"))

    # ── HOJA 6: PRÉSTAMOS ─────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Préstamos")
    ws6.sheet_view.showGridLines = False
    ws6.merge_cells("A1:H1")
    t6 = ws6.cell(row=1, column=1, value=f"PRÉSTAMOS — {per['mes_nombre']} {per['anio']}  |  {lab['nombre']}")
    t6.font = Font(bold=True, size=13, color="FFFFFF"); t6.fill = _fill("154360")
    t6.alignment = _left(); ws6.row_dimensions[1].height = 28

    HDRS6 = [("#",5),("Equipo",30),("Código Inv.",16),
             ("Solicitante",28),("Matricula",14),
             ("Fecha Salida",14),("Fecha Retorno",14),("Estado",12)]
    for col, (h, w) in enumerate(HDRS6, 1):
        _set_hdr(ws6, 2, col, h, bg="154360", width=w)

    prest_ord = sorted(d["prestamos"]["detalle"], key=lambda p: p.fecha_salida or datetime.datetime.min, reverse=True)
    for r, p in enumerate(prest_ord, 3):
        act  = d["activos_map"].get(p.activo_id)
        nombre_act = act.nombre if act else f"ID {p.activo_id}"
        codigo_act = act.codigo_inventario if act else "—"
        fill_bg = GRIS_ROW if r % 2 == 0 else None
        if p.estado == "VENCIDO":
            fill_bg = "FFCDD2"
        elif p.estado == "DEVUELTO":
            fill_bg = "E8F5E9"
        _data_row(ws6, r, [
            r-2, nombre_act, codigo_act,
            p.solicitante_nombre, p.solicitante_id_escolar or "—",
            p.fecha_salida.strftime("%d/%m/%Y") if p.fecha_salida else "—",
            p.fecha_retorno_esperada.strftime("%d/%m/%Y") if p.fecha_retorno_esperada else "—",
            p.estado
        ], bg=fill_bg)

    if not prest_ord:
        ws6.cell(row=3, column=1, value="No hubo préstamos en este periodo.").font = Font(italic=True, color="888888")

    # ── HOJA 7: INCIDENTES ────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Incidentes y Mantenimiento")
    ws7.sheet_view.showGridLines = False
    ws7.merge_cells("A1:H1")
    t7 = ws7.cell(row=1, column=1, value=f"INCIDENTES Y MANTENIMIENTO — {per['mes_nombre']} {per['anio']}  |  {lab['nombre']}")
    t7.font = Font(bold=True, size=13, color="FFFFFF"); t7.fill = _fill("78281F")
    t7.alignment = _left(); ws7.row_dimensions[1].height = 28

    HDRS7 = [("#",5),("Equipo",30),("Tipo",14),("Descripción",40),
             ("Prioridad",12),("Estado",16),("Fecha Reporte",14),("Fecha Resolución",16),
             ("Adeudo",18)]
    for col, (h, w) in enumerate(HDRS7, 1):
        _set_hdr(ws7, 2, col, h, bg="78281F", width=w)

    # Colores semánticos para la columna Adeudo
    ADEUDO_COLORS = {
        "PENDIENTE":   ("FFF3CD", "856404"),   # amarillo — requiere atención
        "EN_REVISION": ("CCE5FF", "004085"),   # azul — en proceso
        "RESUELTO":    ("D4EDDA", "155724"),   # verde — cerrado
        "EXONERADO":   ("E2E3E5", "383D41"),   # gris — exonerado
    }

    adeudo_map = d.get("adeudo_por_incidente", {})

    inc_ord = sorted(d["incidentes"]["detalle"], key=lambda i: i.fecha_reporte or datetime.datetime.min, reverse=True)
    for r, i in enumerate(inc_ord, 3):
        act = d["activos_map"].get(i.activo_id) if i.activo_id else None
        nombre_eq = act.nombre if act else (f"PC ID {i.computadora_id}" if i.computadora_id else "—")
        fill_bg = GRIS_ROW if r % 2 == 0 else None
        prio_colors = {"ALTA": "FFCDD2", "MEDIA": "FFF9C4", "BAJA": "E8F5E9"}
        if i.prioridad:
            fill_bg = prio_colors.get(i.prioridad.upper(), fill_bg)

        # Datos del adeudo vinculado (si existe)
        adeudo = adeudo_map.get(i.id)
        adeudo_label = "—"
        adeudo_fill  = None
        adeudo_font_color = "000000"
        if adeudo:
            lbl_map = {"PENDIENTE":"Pendiente","EN_REVISION":"En revisión",
                       "RESUELTO":"Resuelto","EXONERADO":"Exonerado"}
            adeudo_label = lbl_map.get(adeudo.estado, adeudo.estado)
            if adeudo.monto_estimado:
                adeudo_label += f" (${adeudo.monto_estimado:.2f})"
            bg_hex, fg_hex = ADEUDO_COLORS.get(adeudo.estado, ("FFFFFF","000000"))
            adeudo_fill       = bg_hex
            adeudo_font_color = fg_hex

        _data_row(ws7, r, [
            r-2, nombre_eq, i.tipo, i.descripcion or "—",
            i.prioridad, i.estado,
            i.fecha_reporte.strftime("%d/%m/%Y") if i.fecha_reporte else "—",
            i.fecha_resolucion.strftime("%d/%m/%Y") if i.fecha_resolucion else "Pendiente",
            adeudo_label,
        ], bg=fill_bg)

        # Sobreescribir color de la celda Adeudo si hay dato
        if adeudo and adeudo_fill:
            col_adeudo = len(HDRS7)  # última columna
            cell = ws7.cell(row=r, column=col_adeudo)
            cell.fill = PatternFill("solid", fgColor=adeudo_fill)
            cell.font = Font(size=9, bold=(adeudo.estado == "PENDIENTE"),
                             color=adeudo_font_color)

    if not inc_ord:
        ws7.cell(row=3, column=1, value="No se reportaron incidentes en este periodo.").font = Font(italic=True, color="888888")

    # ── Color de pestañas ─────────────────────────────────────────────────────
    tabs = {"Resumen": "1E3A5F", "Sesiones por Docente": "2D6A9F",
            "Alumnos Atendidos": "1B6B3A", "Horas Pico": "4A235A",
            "Inventario": "1B6B3A", "Préstamos": "154360",
            "Incidentes y Mantenimiento": "78281F"}
    for ws_obj in wb.worksheets:
        color = tabs.get(ws_obj.title, "2D6A9F")
        ws_obj.sheet_properties.tabColor = color

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Endpoint Excel ───────────────────────────────────────────────────────────

@router.get("/mensual/excel", summary="Descargar reporte mensual en Excel")
def reporte_mensual_excel(
    laboratorio_id: int,
    mes:  int = None,
    anio: int = None,
    db:   Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    # RLS: LAB_ADMIN solo puede exportar reportes de su propio laboratorio
    assert_report_access(laboratorio_id, current_user)

    hoy = datetime.date.today()
    mes  = mes  or hoy.month
    anio = anio or hoy.year
    if not (1 <= mes <= 12):
        raise HTTPException(400, "Mes inválido (1-12)")

    d = _datos_mes(db, laboratorio_id, mes, anio)
    if d is None:
        raise HTTPException(404, "Laboratorio no encontrado")

    buf = _build_excel(d)
    nombre_lab  = d["laboratorio"]["nombre"].replace(" ","_")[:20]
    mes_nombre  = MESES_ES[mes]
    filename    = f"Reporte_{nombre_lab}_{mes_nombre}_{anio}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Endpoint: Stats para Dashboard ──────────────────────────────────────────

@router.get("/dashboard", summary="Estadísticas en tiempo real para el dashboard")
def dashboard_stats(
    laboratorio_id: Optional[int] = None,
    db:   Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    ahora  = _utcnow()
    hoy_ini = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    hoy_fin = ahora.replace(hour=23, minute=59, second=59)

    # Si es LAB_ADMIN, restringir a su lab
    lab_filter_id = laboratorio_id
    if current_user.rol == RolUsuario.LAB_ADMIN and current_user.laboratorio_id:
        lab_filter_id = current_user.laboratorio_id

    # ── Sesiones ──────────────────────────────────────────────────────────────
    q_ses = db.query(SesionClase)
    if lab_filter_id:
        q_ses = q_ses.filter(SesionClase.laboratorio_id == lab_filter_id)

    sesiones_hoy    = q_ses.filter(SesionClase.inicio >= hoy_ini, SesionClase.inicio <= hoy_fin).all()
    sesiones_activas = q_ses.filter(SesionClase.estado == "ABIERTA").all()

    # Semana actual (lunes a hoy)
    lunes = (ahora - datetime.timedelta(days=ahora.weekday())).replace(hour=0, minute=0, second=0)
    sesiones_semana = q_ses.filter(SesionClase.inicio >= lunes).count()

    # Alumnos únicos hoy
    ids_hoy = [s.id for s in sesiones_hoy]
    alumnos_hoy = len({a.alumno_matricula for a in db.query(AsignacionPC).filter(
        AsignacionPC.sesion_id.in_(ids_hoy)
    ).all()}) if ids_hoy else 0

    # Sesiones últimos 7 días (para la gráfica)
    DIAS_ES = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]
    sesiones_7d = []
    for i in range(6, -1, -1):
        dia  = ahora - datetime.timedelta(days=i)
        ini  = dia.replace(hour=0, minute=0, second=0, microsecond=0)
        fin  = dia.replace(hour=23, minute=59, second=59)
        q_d  = db.query(SesionClase).filter(SesionClase.inicio >= ini, SesionClase.inicio <= fin)
        if lab_filter_id:
            q_d = q_d.filter(SesionClase.laboratorio_id == lab_filter_id)
        cnt  = q_d.count()
        sesiones_7d.append({
            "fecha": dia.strftime("%Y-%m-%d"),
            "dia":   DIAS_ES[dia.weekday()],
            "count": cnt,
            "es_hoy": i == 0,
        })

    # ── PCs / Computadoras ────────────────────────────────────────────────────
    q_pc = db.query(Computadora).filter(Computadora.activa == True)
    if lab_filter_id:
        q_pc = q_pc.filter(Computadora.laboratorio_id == lab_filter_id)
    pcs_all  = q_pc.all()
    pcs_stat = {
        "total":        len(pcs_all),
        "operativas":   sum(1 for p in pcs_all if p.estado == "OPERATIVO"),
        "mantenimiento":sum(1 for p in pcs_all if p.estado == "MANTENIMIENTO"),
        "danadas":      sum(1 for p in pcs_all if p.estado not in ("OPERATIVO","MANTENIMIENTO")),
    }

    # ── Activos de inventario ─────────────────────────────────────────────────
    q_act = db.query(Activo).filter(Activo.activo == True)
    if lab_filter_id:
        q_act = q_act.filter(Activo.laboratorio_id == lab_filter_id)
    activos_all = q_act.all()

    # ── Préstamos ─────────────────────────────────────────────────────────────
    # Actualizar vencidos
    vencidos_upd = db.query(Prestamo).filter(
        Prestamo.estado == "ACTIVO",
        Prestamo.fecha_retorno_esperada < ahora,
    ).all()
    for p in vencidos_upd:
        p.estado = "VENCIDO"
    if vencidos_upd:
        db.commit()

    activo_ids = [a.id for a in activos_all]
    if activo_ids:
        prestamos_activos  = db.query(Prestamo).filter(
            Prestamo.activo_id.in_(activo_ids), Prestamo.estado == "ACTIVO").count()
        prestamos_vencidos = db.query(Prestamo).filter(
            Prestamo.activo_id.in_(activo_ids), Prestamo.estado == "VENCIDO").count()
    else:
        prestamos_activos = prestamos_vencidos = 0

    # ── Incidentes abiertos ───────────────────────────────────────────────────
    q_inc = db.query(Incidente).filter(Incidente.estado.in_(["PENDIENTE","EN_REVISION"]))
    if lab_filter_id:
        q_inc = q_inc.filter(Incidente.laboratorio_id == lab_filter_id)
    incidentes_abiertos = q_inc.count()

    # ── Próximas reservaciones (hoy en adelante) ──────────────────────────────
    from models.horario import Reservacion, HorarioDisponible
    q_res = (
        db.query(Reservacion, HorarioDisponible)
        .join(HorarioDisponible, Reservacion.horario_id == HorarioDisponible.id)
        .filter(Reservacion.estado.in_(["APROBADA", "PROGRAMADA"]))
    )
    if lab_filter_id:
        q_res = q_res.filter(Reservacion.laboratorio_id == lab_filter_id)
    reservas_raw = q_res.all()

    # Calcular la próxima fecha real para cada reservación (basada en dia_semana)
    hoy_date = ahora.date()
    proximas_con_fecha = []
    for r, h in reservas_raw:
        dias_hasta = (h.dia_semana - hoy_date.weekday()) % 7
        proxima_fecha = hoy_date + datetime.timedelta(days=dias_hasta)
        proximas_con_fecha.append((proxima_fecha, h.hora_inicio, r, h))

    proximas_con_fecha.sort(key=lambda x: (x[0], x[1]))
    proximas_con_fecha = proximas_con_fecha[:5]

    proximas_data = []
    for fecha, _, r, h in proximas_con_fecha:
        doc = db.query(Usuario).filter(Usuario.id == r.docente_id).first()
        proximas_data.append({
            "id":        r.id,
            "materia":   r.materia,
            "docente":   doc.nombre if doc else "—",
            "fecha":     fecha.isoformat(),
            "hora_ini":  h.hora_inicio,
            "hora_fin":  h.hora_fin,
            "grupo":     r.grupo,
        })

    # ── Alertas accionables ───────────────────────────────────────────────────
    # 1. Incidentes sin atender (PENDIENTE)
    q_inc_pend = db.query(Incidente).filter(Incidente.estado == "PENDIENTE")
    if lab_filter_id:
        q_inc_pend = q_inc_pend.filter(Incidente.laboratorio_id == lab_filter_id)
    inc_pend_total = q_inc_pend.count()
    inc_pend_list  = q_inc_pend.order_by(Incidente.fecha_reporte).limit(5).all()

    alerta_incidentes = [
        {
            "id":          i.id,
            "tipo":        i.tipo,
            "descripcion": (i.descripcion or "Sin descripción")[:80],
            "dias":        (ahora - i.fecha_reporte).days if i.fecha_reporte else 0,
            "prioridad":   getattr(i, "prioridad", "MEDIA"),
        }
        for i in inc_pend_list
    ]

    # 2. Préstamos vencidos (detalle)
    if activo_ids:
        prest_venc_rows = (
            db.query(Prestamo, Activo)
            .join(Activo, Prestamo.activo_id == Activo.id)
            .filter(Prestamo.estado == "VENCIDO", Prestamo.activo_id.in_(activo_ids))
            .order_by(Prestamo.fecha_retorno_esperada)
            .limit(5)
            .all()
        )
    else:
        prest_venc_rows = []

    alerta_prestamos = [
        {
            "id":          p.id,
            "persona":     p.solicitante_nombre,
            "id_escolar":  p.solicitante_id_escolar,
            "activo":      a.nombre,
            "dias_vencido": max(0, (ahora - p.fecha_retorno_esperada).days) if p.fecha_retorno_esperada else 0,
        }
        for p, a in prest_venc_rows
    ]

    # 3. Adeudos pendientes más de 7 días
    umbral_adeudo = ahora - datetime.timedelta(days=7)
    q_adeudos_crit = db.query(Adeudo).filter(
        Adeudo.estado == "PENDIENTE",
        Adeudo.fecha_reporte < umbral_adeudo,
    )
    if lab_filter_id:
        q_adeudos_crit = q_adeudos_crit.filter(Adeudo.laboratorio_id == lab_filter_id)
    adeudos_crit_total = q_adeudos_crit.count()
    adeudos_crit_list  = q_adeudos_crit.order_by(Adeudo.fecha_reporte).limit(5).all()

    alerta_adeudos = [
        {
            "id":      d.id,
            "persona": d.persona_nombre,
            "tipo":    d.tipo,
            "dias":    (ahora - d.fecha_reporte).days if d.fecha_reporte else 0,
        }
        for d in adeudos_crit_list
    ]

    total_alertas = inc_pend_total + prestamos_vencidos + adeudos_crit_total

    # ── Labs disponibles (para selector del dashboard) ────────────────────────
    labs = db.query(Laboratorio).filter(Laboratorio.activo == True).all()

    return {
        "sesiones": {
            "activas":   len(sesiones_activas),
            "hoy":       len(sesiones_hoy),
            "semana":    sesiones_semana,
        },
        "alumnos_hoy":          alumnos_hoy,
        "pcs":                  pcs_stat,
        "activos_inventario":   len(activos_all),
        "prestamos": {
            "activos":   prestamos_activos,
            "vencidos":  prestamos_vencidos,
        },
        "incidentes_abiertos":  incidentes_abiertos,
        "sesiones_7d":          sesiones_7d,
        "proximas_reservaciones": proximas_data,
        "labs": [{"id": l.id, "nombre": l.nombre} for l in labs],
        "alertas": {
            "total":                total_alertas,
            "incidentes_pendientes": alerta_incidentes,
            "incidentes_total":      inc_pend_total,
            "prestamos_vencidos":    alerta_prestamos,
            "prestamos_total":       prestamos_vencidos,
            "adeudos_criticos":      alerta_adeudos,
            "adeudos_total":         adeudos_crit_total,
        },
    }


# ─── Helpers cuatrimestre ─────────────────────────────────────────────────────

def _cuatrimestre_actual(ahora: datetime.datetime):
    """Devuelve (num, anio) del cuatrimestre en curso."""
    m = ahora.month
    if m <= 4:   return 1, ahora.year
    elif m <= 8: return 2, ahora.year
    else:        return 3, ahora.year

def _cuatrimestre_anterior(num: int, anio: int):
    if num == 1: return 3, anio - 1
    elif num == 2: return 1, anio
    else:          return 2, anio

def _rango_cuatrimestre(num: int, anio: int):
    rangos = {
        1: (datetime.datetime(anio, 1, 1), datetime.datetime(anio, 4, 30, 23, 59, 59)),
        2: (datetime.datetime(anio, 5, 1), datetime.datetime(anio, 8, 31, 23, 59, 59)),
        3: (datetime.datetime(anio, 9, 1), datetime.datetime(anio, 12, 31, 23, 59, 59)),
    }
    return rangos[num]

def _meses_cuatrimestre(num: int):
    return {1: [1,2,3,4], 2: [5,6,7,8], 3: [9,10,11,12]}[num]

def _nombre_cuatrimestre(num: int, anio: int):
    nombres = {1: f"ENE-ABR {anio}", 2: f"MAY-AGO {anio}", 3: f"SEP-DIC {anio}"}
    return nombres[num]


# ─── Endpoint: Análisis comparativo ──────────────────────────────────────────

@router.get("/comparativo", summary="Análisis comparativo cuatrimestral")
def reporte_comparativo(
    laboratorio_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    ahora = _utcnow()

    # RLS
    lab_filter = laboratorio_id
    if current_user.rol == RolUsuario.LAB_ADMIN and current_user.laboratorio_id:
        lab_filter = current_user.laboratorio_id

    # Cuatrimestres
    num_act, anio_act = _cuatrimestre_actual(ahora)
    num_ant, anio_ant = _cuatrimestre_anterior(num_act, anio_act)
    ini_act, fin_act  = _rango_cuatrimestre(num_act, anio_act)
    ini_ant, fin_ant  = _rango_cuatrimestre(num_ant, anio_ant)

    def _q_ses(ini, fin):
        q = db.query(SesionClase).filter(
            SesionClase.inicio >= ini,
            SesionClase.inicio <= fin,
        )
        if lab_filter:
            q = q.filter(SesionClase.laboratorio_id == lab_filter)
        return q

    # ── 1. Tendencia: sesiones mes a mes ─────────────────────────────────────
    def _ses_por_mes(meses_list: list, anio: int):
        resultado = []
        for m in meses_list:
            ini_m, fin_m = _periodo(m, anio)
            q = db.query(SesionClase).filter(
                SesionClase.inicio >= ini_m,
                SesionClase.inicio <= fin_m,
            )
            if lab_filter:
                q = q.filter(SesionClase.laboratorio_id == lab_filter)
            resultado.append({
                "mes":    m,
                "nombre": MESES_ES[m],
                "count":  q.count(),
            })
        return resultado

    tendencia_actual   = _ses_por_mes(_meses_cuatrimestre(num_act), anio_act)
    tendencia_anterior = _ses_por_mes(_meses_cuatrimestre(num_ant), anio_ant)

    # Totales cuatrimestrales
    total_act = sum(d["count"] for d in tendencia_actual)
    total_ant = sum(d["count"] for d in tendencia_anterior)

    # ── 2. Horas pico — cuatrimestre actual ───────────────────────────────────
    DIAS_ES_SHORT = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    HORAS_LAB = list(range(7, 22))   # 07:00 – 21:00

    ses_act = _q_ses(ini_act, fin_act).all()
    horas_pico_raw = collections.defaultdict(int)
    for s in ses_act:
        if s.inicio:
            horas_pico_raw[(s.inicio.weekday(), s.inicio.hour)] += 1

    horas_pico = []
    for dia in range(7):
        for hora in HORAS_LAB:
            horas_pico.append({
                "dia":   dia,
                "dia_nombre": DIAS_ES_SHORT[dia],
                "hora":  hora,
                "count": horas_pico_raw.get((dia, hora), 0),
            })

    # ── 3. Top docentes — cuatrimestre actual ─────────────────────────────────
    doc_count = collections.defaultdict(lambda: {"sesiones": 0, "horas": 0.0, "nombre": ""})
    for s in ses_act:
        if s.docente_id:
            doc_count[s.docente_id]["sesiones"] += 1
            fin_s = s.fin_real or s.fin_estimado
            if fin_s and s.inicio:
                doc_count[s.docente_id]["horas"] += (fin_s - s.inicio).total_seconds() / 3600

    doc_ids = list(doc_count.keys())
    doc_objs = {u.id: u for u in db.query(Usuario).filter(Usuario.id.in_(doc_ids)).all()}
    for did, v in doc_count.items():
        u = doc_objs.get(did)
        v["nombre"] = u.nombre if u else f"#{did}"
        v["horas"]  = round(v["horas"], 1)

    top_docentes = sorted(
        [{"docente_id": k, **v} for k, v in doc_count.items()],
        key=lambda x: x["sesiones"], reverse=True
    )[:10]

    # ── 4. Computadoras con más incidentes recurrentes ─────────────────────────
    # Último año calendario
    hace_un_anio = ahora - datetime.timedelta(days=365)
    q_inc = db.query(Incidente).filter(Incidente.fecha_reporte >= hace_un_anio)
    if lab_filter:
        q_inc = q_inc.filter(Incidente.laboratorio_id == lab_filter)
    incidentes_ano = q_inc.all()

    pc_inc_count = collections.defaultdict(lambda: {"total": 0, "pendientes": 0, "nombre": "", "numero": ""})
    for inc in incidentes_ano:
        if inc.computadora_id:
            pc_inc_count[inc.computadora_id]["total"] += 1
            if inc.estado == "PENDIENTE":
                pc_inc_count[inc.computadora_id]["pendientes"] += 1

    pc_ids = list(pc_inc_count.keys())
    pc_objs = {p.id: p for p in db.query(Computadora).filter(Computadora.id.in_(pc_ids)).all()}
    for pid, v in pc_inc_count.items():
        pc = pc_objs.get(pid)
        v["codigo"]  = pc.codigo  if pc else f"PC #{pid}"
        v["numero"]  = pc.numero  if pc else ""
        v["estado"]  = pc.estado  if pc else ""

    computadoras_criticas = sorted(
        [{"computadora_id": k, **v} for k, v in pc_inc_count.items()],
        key=lambda x: x["total"], reverse=True
    )[:10]
    # quitar incidentes sin computadora_id (activos de inventario)
    computadoras_criticas = [c for c in computadoras_criticas if c["codigo"] != ""]

    # ── 5. Alumnos únicos por cuatrimestre ────────────────────────────────────
    def _alumnos_unicos(ini, fin):
        ids = [s.id for s in _q_ses(ini, fin).all()]
        if not ids:
            return 0
        return len({a.alumno_matricula for a in db.query(AsignacionPC).filter(
            AsignacionPC.sesion_id.in_(ids)
        ).all()})

    alumnos_act = _alumnos_unicos(ini_act, fin_act)
    alumnos_ant = _alumnos_unicos(ini_ant, fin_ant)

    # ── 6. Horas totales por cuatrimestre ─────────────────────────────────────
    def _horas(sesiones_list):
        h = 0.0
        for s in sesiones_list:
            fin_s = s.fin_real or s.fin_estimado
            if fin_s and s.inicio:
                h += (fin_s - s.inicio).total_seconds() / 3600
        return round(h, 1)

    ses_ant_all = _q_ses(ini_ant, fin_ant).all()
    horas_act = _horas(ses_act)
    horas_ant = _horas(ses_ant_all)

    # ── Labs disponibles ──────────────────────────────────────────────────────
    labs = db.query(Laboratorio).filter(Laboratorio.activo == True).all()

    return {
        "cuatrimestre_actual":   {"num": num_act, "anio": anio_act, "nombre": _nombre_cuatrimestre(num_act, anio_act)},
        "cuatrimestre_anterior": {"num": num_ant, "anio": anio_ant, "nombre": _nombre_cuatrimestre(num_ant, anio_ant)},
        "tendencia": {
            "actual":   tendencia_actual,
            "anterior": tendencia_anterior,
            "total_actual":   total_act,
            "total_anterior": total_ant,
        },
        "horas_pico":            horas_pico,
        "top_docentes":          top_docentes,
        "computadoras_criticas": computadoras_criticas,
        "resumen": {
            "sesiones_actual":  total_act,
            "sesiones_anterior": total_ant,
            "alumnos_actual":   alumnos_act,
            "alumnos_anterior": alumnos_ant,
            "horas_actual":     horas_act,
            "horas_anterior":   horas_ant,
               },
        "labs": [{"id": l.id, "nombre": l.nombre} for l in labs],
    }


# ─── Reporte de Cumplimiento Docente ─────────────────────────────────────────

def _contar_ocurrencias(dia_semana: int, fecha_inicio: datetime.date, fecha_fin: datetime.date) -> int:
    """Cuenta cuántas veces cae ese día de la semana en el rango [inicio, fin]."""
    count = 0
    d = fecha_inicio
    while d <= fecha_fin:
        if d.weekday() == dia_semana:
            count += 1
        d += datetime.timedelta(days=1)
    return count


DIAS_NOMBRES = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"}


@router.get("/cumplimiento", summary="Cumplimiento docente por cuatrimestre")
def reporte_cumplimiento(
    laboratorio_id: Optional[int] = None,
    cuatrimestre: Optional[str] = None,
    docente_id: Optional[int] = None,
    fecha_inicio: Optional[str] = None,  # YYYY-MM-DD
    fecha_fin: Optional[str] = None,     # YYYY-MM-DD
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from models.horario import Reservacion, HorarioDisponible
    from models.cumplimiento import EventoCumplimiento

    # Parse fechas
    f_ini: Optional[datetime.date] = None
    f_fin: Optional[datetime.date] = None
    if fecha_inicio:
        try:
            f_ini = datetime.date.fromisoformat(fecha_inicio)
        except ValueError:
            pass
    if fecha_fin:
        try:
            f_fin = datetime.date.fromisoformat(fecha_fin)
        except ValueError:
            pass

    # Fallback: mes actual si no se proveen fechas
    if not f_ini or not f_fin:
        hoy = datetime.date.today()
        f_ini = hoy.replace(day=1)
        # último día del mes
        last_day = calendar.monthrange(hoy.year, hoy.month)[1]
        f_fin = hoy.replace(day=last_day)

    # Consulta base de reservaciones
    q = db.query(Reservacion).filter(Reservacion.estado != "CANCELADA")
    if laboratorio_id:
        q = q.filter(Reservacion.laboratorio_id == laboratorio_id)
    if cuatrimestre:
        q = q.filter(Reservacion.cuatrimestre == cuatrimestre)
    if docente_id:
        q = q.filter(Reservacion.docente_id == docente_id)

    reservaciones = q.all()

    # Agrupar por docente
    docentes_dict: dict = {}  # docente_id → {datos}

    for res in reservaciones:
        did = res.docente_id
        if did not in docentes_dict:
            doc = db.query(Usuario).filter(Usuario.id == did).first()
            docentes_dict[did] = {
                "docente_id": did,
                "docente_nombre": doc.nombre if doc else f"#{did}",
                "reservaciones_activas": 0,
                "clases_esperadas": 0,
                "impartidas": 0,
                "no_asistio": 0,
                "cancelacion_tardia": 0,
                "detalle": [],
            }

        docentes_dict[did]["reservaciones_activas"] += 1

        # Calcular clases esperadas para esta reservación
        horario = db.query(HorarioDisponible).filter(HorarioDisponible.id == res.horario_id).first()
        clases_esp = 0
        dia_nombre = ""
        hora_ini_str = ""
        if horario:
            dia_nombre = DIAS_NOMBRES.get(horario.dia_semana, str(horario.dia_semana))
            hora_ini_str = horario.hora_inicio
            clases_esp = _contar_ocurrencias(horario.dia_semana, f_ini, f_fin)

        docentes_dict[did]["clases_esperadas"] += clases_esp

        # Contar eventos de cumplimiento para esta reservación en el rango
        eventos_q = db.query(EventoCumplimiento).filter(
            EventoCumplimiento.reservacion_id == res.id,
            EventoCumplimiento.fecha >= f_ini,
            EventoCumplimiento.fecha <= f_fin,
        ).all()

        impartidas = sum(1 for e in eventos_q if e.tipo == "IMPARTIDA")
        no_asistio = sum(1 for e in eventos_q if e.tipo == "NO_ASISTIO")
        cancelacion = sum(1 for e in eventos_q if e.tipo == "CANCELADA_TARDIA")

        docentes_dict[did]["impartidas"] += impartidas
        docentes_dict[did]["no_asistio"] += no_asistio
        docentes_dict[did]["cancelacion_tardia"] += cancelacion

        docentes_dict[did]["detalle"].append({
            "reservacion_id": res.id,
            "materia": res.materia,
            "grupo": res.grupo,
            "dia_nombre": dia_nombre,
            "hora_inicio": hora_ini_str,
            "clases_esperadas": clases_esp,
            "impartidas": impartidas,
            "no_asistio": no_asistio,
            "cancelacion_tardia": cancelacion,
            "eventos": [
                {
                    "tipo": e.tipo,
                    "fecha": e.fecha.isoformat() if e.fecha else None,
                    "motivo": e.motivo,
                }
                for e in sorted(eventos_q, key=lambda x: x.fecha)
            ],
        })

    # Calcular métricas finales por docente
    result = []
    for did, d in docentes_dict.items():
        total_registrados = d["impartidas"] + d["no_asistio"] + d["cancelacion_tardia"]
        sin_registro = max(0, d["clases_esperadas"] - total_registrados)
        total_para_pct = max(d["impartidas"] + d["no_asistio"] + d["cancelacion_tardia"], 1)
        porcentaje = round((d["impartidas"] / total_para_pct) * 100, 1)

        if porcentaje >= 85:
            semaforo = "verde"
        elif porcentaje >= 70:
            semaforo = "amarillo"
        else:
            semaforo = "rojo"

        result.append({
            **d,
            "sin_registro": sin_registro,
            "porcentaje_cumplimiento": porcentaje,
            "semaforo": semaforo,
        })

    # Ordenar por porcentaje ascendente (peores primero)
    result.sort(key=lambda x: x["porcentaje_cumplimiento"])

    return {
        "filtros": {
            "laboratorio_id": laboratorio_id,
            "cuatrimestre": cuatrimestre,
            "docente_id": docente_id,
            "fecha_inicio": f_ini.isoformat(),
            "fecha_fin": f_fin.isoformat(),
        },
        "docentes": result,
    }
