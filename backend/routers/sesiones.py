from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from typing import Optional, List
from database import get_db
from models.sesion import SesionClase, AsignacionPC, ObservacionPC
from models.horario import Reservacion
from models.laboratorio import Laboratorio, Computadora
from models.usuario import Usuario, RolUsuario
from models.catalogo import CatalogoAlumno
from dependencies import get_current_user, require_roles
from permissions import require_permission
from ws.mapa import manager, _snapshot_lab
from routers.notificaciones import crear_notificacion
import datetime
import uuid
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/sesiones", tags=["Sesiones de Clase"])


# ─── Schemas ───────────────────────────────────────────────────────────────────

class SesionCreate(BaseModel):
    laboratorio_id: int
    materia: str        = Field(..., min_length=2, max_length=100)
    grupo: str          = Field(..., min_length=1, max_length=20)
    reservacion_id: Optional[int] = None
    fin_estimado_min: Optional[int] = Field(None, ge=30, le=300, description="Duración estimada en minutos")

class SesionCerrar(BaseModel):
    observacion_general: Optional[str] = None

class AsignacionCreate(BaseModel):
    computadora_id: int
    alumno_nombre: str   = Field(..., min_length=2, max_length=100)
    alumno_matricula: str = Field(..., min_length=2, max_length=20)

class ObservacionCreate(BaseModel):
    computadora_id: Optional[int] = None
    tipo: str       = Field(default="SIN_NOVEDAD", description="SIN_NOVEDAD, FALLA_HARDWARE, FALLA_SOFTWARE, LIMPIEZA, OTRO")
    descripcion: Optional[str] = None
    prioridad: str  = Field(default="BAJA", description="BAJA, MEDIA, ALTA")

class ReportePCCreate(BaseModel):
    computadora_id: int
    nota: str        = Field(..., min_length=3, max_length=500)
    bloquear: bool   = Field(default=False, description="Si True, cambia la PC a MANTENIMIENTO")


# ─── Helpers ───────────────────────────────────────────────────────────────────

def _serializar_sesion(s: SesionClase, db: Session) -> dict:
    lab  = db.query(Laboratorio).filter(Laboratorio.id == s.laboratorio_id).first()
    doc  = db.query(Usuario).filter(Usuario.id == s.docente_id).first()
    asigs = db.query(AsignacionPC).filter(
        AsignacionPC.sesion_id == s.id,
        AsignacionPC.hora_liberacion == None  # noqa: E711
    ).count()

    # Calcular tiempo restante (en segundos) o exceso si está abierta
    ahora = datetime.datetime.utcnow()
    tiempo_restante_seg = None
    en_overtime = False
    if s.estado == "ABIERTA" and s.fin_estimado:
        diff = (s.fin_estimado - ahora).total_seconds()
        tiempo_restante_seg = int(diff)   # negativo si está en overtime
        en_overtime = diff < 0

    return {
        "id": s.id,
        "codigo_sesion": s.codigo_sesion,
        "laboratorio_id": s.laboratorio_id,
        "laboratorio_nombre": lab.nombre if lab else None,
        "docente_id": s.docente_id,
        "docente_nombre": doc.nombre if doc else None,
        "materia": s.materia,
        "grupo": s.grupo,
        "inicio": s.inicio.isoformat() if s.inicio else None,
        "fin_estimado": s.fin_estimado.isoformat() if s.fin_estimado else None,
        "fin_real": s.fin_real.isoformat() if s.fin_real else None,
        "estado": s.estado,
        "pcs_ocupadas": asigs,
        "observacion_general": s.observacion_general,
        "reservacion_id": s.reservacion_id,
        "overtime_min": s.overtime_min or 0,
        "tiempo_restante_seg": tiempo_restante_seg,
        "en_overtime": en_overtime,
    }

def _get_sesion_activa_docente(docente_id: int, db: Session) -> Optional[SesionClase]:
    return db.query(SesionClase).filter(
        SesionClase.docente_id == docente_id,
        SesionClase.estado == "ABIERTA"
    ).first()


# ─── Sesiones ──────────────────────────────────────────────────────────────────

@router.get("", summary="Listar sesiones")
def listar_sesiones(
    laboratorio_id: Optional[int] = None,
    estado: Optional[str]         = None,
    docente_id: Optional[int]     = None,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(SesionClase)
    if laboratorio_id:
        q = q.filter(SesionClase.laboratorio_id == laboratorio_id)
    elif current_user.rol == RolUsuario.LAB_ADMIN:
        q = q.filter(SesionClase.laboratorio_id == current_user.laboratorio_id)
    if estado:
        q = q.filter(SesionClase.estado == estado)
    if docente_id:
        q = q.filter(SesionClase.docente_id == docente_id)
    elif current_user.rol == RolUsuario.DOCENTE:
        q = q.filter(SesionClase.docente_id == current_user.id)

    sesiones = q.order_by(SesionClase.inicio.desc()).limit(limit).all()
    return [_serializar_sesion(s, db) for s in sesiones]


@router.get("/activas", summary="Sesiones abiertas ahora mismo")
def sesiones_activas(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    q = db.query(SesionClase).filter(SesionClase.estado == "ABIERTA")
    if current_user.rol == RolUsuario.DOCENTE:
        q = q.filter(SesionClase.docente_id == current_user.id)
    elif current_user.rol == RolUsuario.LAB_ADMIN:
        q = q.filter(SesionClase.laboratorio_id == current_user.laboratorio_id)
    return [_serializar_sesion(s, db) for s in q.all()]


@router.post("", status_code=status.HTTP_201_CREATED, summary="Abrir sesión de clase")
async def abrir_sesion(
    data: SesionCreate,
    db: Session = Depends(get_db),
    # RBAC: solo SUPER_ADMIN, LAB_ADMIN y DOCENTE pueden abrir sesiones
    current_user: Usuario = Depends(require_permission("sesiones:write"))
):

    # Solo un docente puede tener una sesión abierta a la vez
    if current_user.rol == RolUsuario.DOCENTE:
        ya_abierta = _get_sesion_activa_docente(current_user.id, db)
        if ya_abierta:
            raise HTTPException(
                status_code=409,
                detail=f"Ya tienes una sesión abierta: {ya_abierta.codigo_sesion}"
            )

    # Verificar que el laboratorio exista y esté activo
    lab = db.query(Laboratorio).filter(
        Laboratorio.id == data.laboratorio_id, Laboratorio.activo == True
    ).first()
    if not lab:
        raise HTTPException(status_code=404, detail="Laboratorio no encontrado")

    # No puede haber otra sesión abierta en el mismo lab
    otra = db.query(SesionClase).filter(
        SesionClase.laboratorio_id == data.laboratorio_id,
        SesionClase.estado == "ABIERTA"
    ).first()
    if otra:
        raise HTTPException(
            status_code=409,
            detail=f"Ya hay una sesión abierta en este laboratorio: {otra.codigo_sesion}"
        )

    ahora = datetime.datetime.utcnow()
    fin_est = ahora + datetime.timedelta(minutes=data.fin_estimado_min) if data.fin_estimado_min else None
    codigo  = f"SES-{ahora.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

    sesion = SesionClase(
        reservacion_id=data.reservacion_id,
        laboratorio_id=data.laboratorio_id,
        docente_id=current_user.id,
        materia=data.materia,
        grupo=data.grupo,
        codigo_sesion=codigo,
        inicio=ahora,
        fin_estimado=fin_est,
        estado="ABIERTA",
    )
    db.add(sesion)
    db.commit()
    db.refresh(sesion)

    # Broadcast WebSocket
    await manager.broadcast(data.laboratorio_id, {
        "tipo": "sesion_abierta",
        "sesion": {
            "id": sesion.id,
            "codigo": codigo,
            "materia": data.materia,
            "grupo": data.grupo,
            "docente": current_user.nombre,
        }
    })

    return _serializar_sesion(sesion, db)


@router.get("/{sesion_id}", summary="Detalle de sesión")
def obtener_sesion(
    sesion_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta sesión")
    return _serializar_sesion(s, db)


@router.get("/{sesion_id}/mapa", summary="Snapshot HTTP del mapa de PCs (polling fallback)")
def mapa_sesion(
    sesion_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Devuelve el estado actual de todas las PCs del laboratorio de la sesión.
    Usado como fallback cuando WebSocket no está disponible."""
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    pcs = _snapshot_lab(s.laboratorio_id, db)
    return {"pcs": pcs, "sesion_id": sesion_id, "laboratorio_id": s.laboratorio_id}


@router.post("/{sesion_id}/cerrar", summary="Cerrar sesión de clase")
async def cerrar_sesion(
    sesion_id: int,
    data: SesionCerrar,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    if s.estado != "ABIERTA":
        raise HTTPException(status_code=400, detail="La sesión no está abierta")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(status_code=403, detail="No puedes cerrar una sesión que no es tuya")

    # Liberar todas las PCs que quedaron asignadas
    asigs_abiertas = db.query(AsignacionPC).filter(
        AsignacionPC.sesion_id == sesion_id,
        AsignacionPC.hora_liberacion == None  # noqa: E711
    ).all()
    ahora = datetime.datetime.utcnow()
    for a in asigs_abiertas:
        a.hora_liberacion = ahora

    s.estado = "CERRADA"
    s.fin_real = ahora
    s.observacion_general = data.observacion_general
    # Calcular overtime
    overtime_min = 0
    if s.fin_estimado and ahora > s.fin_estimado:
        overtime_min = int((ahora - s.fin_estimado).total_seconds() / 60)
    s.overtime_min = overtime_min
    db.commit()
    db.refresh(s)

    # ── Notificación de overtime al cerrar ───────────────────────────────────
    if overtime_min > 0:
        try:
            lab = db.query(Laboratorio).filter(Laboratorio.id == s.laboratorio_id).first()
            lab_nombre = lab.nombre if lab else f"Lab {s.laboratorio_id}"
            admins = db.query(Usuario).filter(
                Usuario.laboratorio_id == s.laboratorio_id
            ).all()
            if not admins:
                admins = db.query(Usuario).filter(Usuario.rol == RolUsuario.SUPER_ADMIN).all()
            docente_nombre = current_user.nombre
            for adm in admins:
                crear_notificacion(
                    db, adm.id,
                    tipo="OVERTIME",
                    titulo=f"Sesión cerrada con overtime — {lab_nombre}",
                    mensaje=(
                        f"La sesión #{s.id} de {docente_nombre} en {lab_nombre} "
                        f"tuvo {overtime_min} min de overtime. "
                        f"Fin estimado: {s.fin_estimado.strftime('%H:%M') if s.fin_estimado else 'N/D'}, "
                        f"fin real: {ahora.strftime('%H:%M')}."
                    ),
                    url="/admin/sesiones",
                )
            db.commit()
        except Exception:
            pass

    # Broadcast WebSocket
    await manager.broadcast(s.laboratorio_id, {
        "tipo": "sesion_cerrada",
        "sesion_id": sesion_id,
    })

    return _serializar_sesion(s, db)


# ─── Asignaciones de PC ────────────────────────────────────────────────────────

@router.get("/{sesion_id}/asignaciones", summary="Lista de asignaciones")
def listar_asignaciones(
    sesion_id: int,
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user)
):
    asigs = db.query(AsignacionPC).filter(AsignacionPC.sesion_id == sesion_id).all()
    result = []
    for a in asigs:
        pc = db.query(Computadora).filter(Computadora.id == a.computadora_id).first()
        result.append({
            "id": a.id,
            "sesion_id": a.sesion_id,
            "computadora_id": a.computadora_id,
            "pc_codigo": pc.codigo if pc else None,
            "pc_fila": pc.fila if pc else None,
            "alumno_nombre": a.alumno_nombre,
            "alumno_matricula": a.alumno_matricula,
            "hora_asignacion": a.hora_asignacion.isoformat() if a.hora_asignacion else None,
            "hora_liberacion": a.hora_liberacion.isoformat() if a.hora_liberacion else None,
            "activa": a.hora_liberacion is None,
        })
    return result


@router.post("/{sesion_id}/asignaciones", status_code=status.HTTP_201_CREATED, summary="Asignar PC a alumno")
async def asignar_pc(
    sesion_id: int,
    data: AsignacionCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id, SesionClase.estado == "ABIERTA").first()
    if not s:
        raise HTTPException(status_code=404, detail="Sesión no encontrada o no está abierta")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta sesión")

    # Verificar que la PC exista y pertenezca al lab
    pc = db.query(Computadora).filter(
        Computadora.id == data.computadora_id,
        Computadora.laboratorio_id == s.laboratorio_id,
        Computadora.activa == True,
    ).first()
    if not pc:
        raise HTTPException(status_code=404, detail="Computadora no encontrada en este laboratorio")
    if pc.estado not in ("OPERATIVO", "EN_CLASE"):
        raise HTTPException(status_code=400, detail=f"La PC está en estado {pc.estado}, no se puede asignar")

    # Verificar que la PC no esté ya asignada en esta sesión
    ya_asignada = db.query(AsignacionPC).filter(
        AsignacionPC.sesion_id == sesion_id,
        AsignacionPC.computadora_id == data.computadora_id,
        AsignacionPC.hora_liberacion == None  # noqa: E711
    ).first()
    if ya_asignada:
        raise HTTPException(status_code=409, detail="Esta PC ya está asignada en la sesión")

    asig = AsignacionPC(
        sesion_id=sesion_id,
        computadora_id=data.computadora_id,
        alumno_nombre=data.alumno_nombre,
        alumno_matricula=data.alumno_matricula,
        hora_asignacion=datetime.datetime.utcnow(),
    )
    db.add(asig)
    db.commit()
    db.refresh(asig)

    # Broadcast WebSocket: PC ocupada
    await manager.broadcast(s.laboratorio_id, {
        "tipo": "pc_actualizada",
        "pc": {
            "pc_id": pc.id,
            "codigo": pc.codigo,
            "fila": pc.fila,
            "estado": "OCUPADA",
            "alumno": {
                "nombre": data.alumno_nombre,
                "matricula": data.alumno_matricula,
                "asignacion_id": asig.id,
            },
            "sesion_id": sesion_id,
        }
    })

    return {
        "id": asig.id,
        "computadora_id": asig.computadora_id,
        "pc_codigo": pc.codigo,
        "alumno_nombre": asig.alumno_nombre,
        "alumno_matricula": asig.alumno_matricula,
        "hora_asignacion": asig.hora_asignacion.isoformat(),
    }


@router.delete("/{sesion_id}/asignaciones/{asig_id}", summary="Liberar PC")
async def liberar_pc(
    sesion_id: int,
    asig_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    asig = db.query(AsignacionPC).filter(
        AsignacionPC.id == asig_id,
        AsignacionPC.sesion_id == sesion_id,
        AsignacionPC.hora_liberacion == None  # noqa: E711
    ).first()
    if not asig:
        raise HTTPException(status_code=404, detail="Asignación no encontrada o ya liberada")

    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta sesión")

    pc = db.query(Computadora).filter(Computadora.id == asig.computadora_id).first()
    asig.hora_liberacion = datetime.datetime.utcnow()
    db.commit()

    # Broadcast WebSocket: PC libre (en sesión)
    await manager.broadcast(s.laboratorio_id, {
        "tipo": "pc_actualizada",
        "pc": {
            "pc_id": pc.id,
            "codigo": pc.codigo,
            "fila": pc.fila,
            "estado": "EN_CLASE",
            "alumno": None,
            "sesion_id": sesion_id,
        }
    })

    return {"mensaje": f"PC {pc.codigo} liberada"}


# ─── Observaciones ─────────────────────────────────────────────────────────────

@router.get("/{sesion_id}/observaciones", summary="Observaciones de la sesión")
def listar_observaciones(
    sesion_id: int,
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user)
):
    obs = db.query(ObservacionPC).filter(ObservacionPC.sesion_id == sesion_id).all()
    result = []
    for o in obs:
        pc = db.query(Computadora).filter(Computadora.id == o.computadora_id).first() if o.computadora_id else None
        result.append({
            "id": o.id,
            "sesion_id": o.sesion_id,
            "computadora_id": o.computadora_id,
            "pc_codigo": pc.codigo if pc else None,
            "tipo": o.tipo,
            "descripcion": o.descripcion,
            "prioridad": o.prioridad,
            "atendida": o.atendida,
        })
    return result


@router.post("/{sesion_id}/observaciones", status_code=status.HTTP_201_CREATED, summary="Registrar observación")
def crear_observacion(
    sesion_id: int,
    data: ObservacionCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    obs = ObservacionPC(
        sesion_id=sesion_id,
        computadora_id=data.computadora_id,
        tipo=data.tipo,
        descripcion=data.descripcion,
        prioridad=data.prioridad,
        atendida=False,
    )
    db.add(obs)
    db.commit()
    db.refresh(obs)

    pc = db.query(Computadora).filter(Computadora.id == obs.computadora_id).first() if obs.computadora_id else None
    return {
        "id": obs.id,
        "sesion_id": obs.sesion_id,
        "computadora_id": obs.computadora_id,
        "pc_codigo": pc.codigo if pc else None,
        "tipo": obs.tipo,
        "descripcion": obs.descripcion,
        "prioridad": obs.prioridad,
        "atendida": obs.atendida,
    }


# ─── Reporte de PC al cierre de sesión ────────────────────────────────────────

@router.post("/{sesion_id}/reportar-pc", status_code=status.HTTP_201_CREATED,
             summary="Reportar PC con problema al cerrar sesión")
async def reportar_pc_cierre(
    sesion_id: int,
    data: ReportePCCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    # Verificar que la PC pertenece al laboratorio de la sesión
    pc = db.query(Computadora).filter(
        Computadora.id == data.computadora_id,
        Computadora.laboratorio_id == s.laboratorio_id
    ).first()
    if not pc:
        raise HTTPException(status_code=404, detail="PC no encontrada en este laboratorio")

    # Crear incidente en Mantenimiento
    from models.inventario import Incidente
    descripcion = f"[PC {pc.codigo}] {data.nota} — reportado al cierre de sesión {s.codigo_sesion} ({s.materia})"
    incidente = Incidente(
        computadora_id=pc.id,
        laboratorio_id=s.laboratorio_id,
        origen="SESION",
        origen_id=s.id,
        tipo="MANTENIMIENTO",
        prioridad="MEDIA",
        descripcion=descripcion,
        reportado_por_id=current_user.id,
        estado="PENDIENTE",
    )
    db.add(incidente)

    # Si se pidió bloquear: cambiar estado de la PC a MANTENIMIENTO
    if data.bloquear:
        pc.estado = "MANTENIMIENTO"

    db.commit()
    db.refresh(incidente)

    # Broadcast WebSocket si se bloqueó la PC
    if data.bloquear:
        await manager.broadcast(s.laboratorio_id, {
            "tipo": "pc_actualizada",
            "pc": {
                "pc_id": pc.id,
                "codigo": pc.codigo,
                "fila": pc.fila,
                "estado": "MANTENIMIENTO",
                "alumno": None,
                "sesion_id": sesion_id,
                "bloqueada": True,
            }
        })

    return {
        "incidente_id": incidente.id,
        "pc_codigo": pc.codigo,
        "bloqueada": data.bloquear,
        "estado_pc": pc.estado,
        "mensaje": (
            f"PC {pc.codigo} reportada"
            + (" y bloqueada temporalmente" if data.bloquear else "")
            + ". El administrador recibirá el aviso."
        ),
    }


# ─── Lista de Asistencia ───────────────────────────────────────────────────────

# ── Estilos Excel (locales para no duplicar los de reportes.py) ───────────────
def _as_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _as_font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color)

def _as_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _as_center():
    return Alignment(horizontal="center", vertical="center")

def _as_left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _get_asistencia_data(sesion_id: int, db: Session) -> dict | None:
    """Recopila datos enriquecidos de asistencia para una sesión."""
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        return None

    lab = db.query(Laboratorio).filter(Laboratorio.id == s.laboratorio_id).first()
    doc = db.query(Usuario).filter(Usuario.id == s.docente_id).first()

    asigs = db.query(AsignacionPC).filter(AsignacionPC.sesion_id == sesion_id).all()

    # Enriquecer con catálogo: buscar la entrada más reciente por matrícula
    matriculas = list({a.alumno_matricula for a in asigs})
    cat_map: dict = {}
    if matriculas:
        for ca in db.query(CatalogoAlumno).filter(CatalogoAlumno.matricula.in_(matriculas)).all():
            if ca.matricula not in cat_map:
                cat_map[ca.matricula] = ca

    pcs_ids = list({a.computadora_id for a in asigs})
    pcs_map = {p.id: p for p in db.query(Computadora).filter(Computadora.id.in_(pcs_ids)).all()} if pcs_ids else {}

    alumnos = []
    for a in asigs:
        pc  = pcs_map.get(a.computadora_id)
        cat = cat_map.get(a.alumno_matricula)
        fin = a.hora_liberacion or s.fin_real
        duracion_min = (
            int((fin - a.hora_asignacion).total_seconds() / 60)
            if a.hora_asignacion and fin else None
        )
        alumnos.append({
            "asignacion_id":    a.id,
            "alumno_nombre":    a.alumno_nombre,
            "alumno_matricula": a.alumno_matricula,
            "pc_codigo":        pc.codigo if pc else None,
            "pc_fila":          pc.fila   if pc else None,
            "hora_entrada":     a.hora_asignacion.isoformat() if a.hora_asignacion else None,
            "hora_salida":      a.hora_liberacion.isoformat() if a.hora_liberacion else None,
            "duracion_min":     duracion_min,
            "activa":           a.hora_liberacion is None,
            # Campos del catálogo (None si el alumno no está registrado)
            "carrera":          cat.carrera      if cat else None,
            "cuatrimestre":     cat.cuatrimestre if cat else None,
            "grupo_catalogo":   cat.grupo        if cat else None,
            "periodo":          cat.periodo      if cat else None,
        })

    fin_sesion = s.fin_real or s.fin_estimado
    dur_sesion_min = (
        int((fin_sesion - s.inicio).total_seconds() / 60)
        if s.inicio and fin_sesion else None
    )

    return {
        "sesion": {
            "id":            s.id,
            "codigo_sesion": s.codigo_sesion,
            "materia":       s.materia,
            "grupo":         s.grupo,
            "estado":        s.estado,
            "inicio":        s.inicio.isoformat()       if s.inicio       else None,
            "fin_real":      s.fin_real.isoformat()     if s.fin_real     else None,
            "fin_estimado":  s.fin_estimado.isoformat() if s.fin_estimado else None,
            "duracion_min":  dur_sesion_min,
        },
        "laboratorio": {
            "id":     lab.id     if lab else None,
            "nombre": lab.nombre if lab else "—",
        },
        "docente": {
            "id":     doc.id     if doc else None,
            "nombre": doc.nombre if doc else "—",
            "email":  doc.email  if doc else None,
        },
        "alumnos":       alumnos,
        "total_alumnos": len(alumnos),
    }


def _build_asistencia_excel(data: dict) -> io.BytesIO:
    """Genera el Excel de lista de asistencia."""
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Asistencia"
    ws.sheet_view.showGridLines = False

    ses     = data["sesion"]
    lab     = data["laboratorio"]
    doc     = data["docente"]
    alumnos = sorted(data["alumnos"], key=lambda x: x["alumno_nombre"])

    AZUL_OSC  = "1E3A5F"
    AZUL_MED  = "2D6A9F"
    GRIS_ROW  = "EEF2F7"

    # ── Fila 1: título ────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1,
                value=f"LISTA DE ASISTENCIA — {ses['materia']}  [{ses['codigo_sesion']}]")
    t.font      = Font(bold=True, size=14, color="FFFFFF")
    t.fill      = _as_fill(AZUL_OSC)
    t.alignment = _as_left()
    ws.row_dimensions[1].height = 32
    ws.column_dimensions["A"].width = 5

    # ── Filas 2-9: metadata ───────────────────────────────────────────────────
    meta = [
        ("Laboratorio:",    lab["nombre"]),
        ("Docente:",        doc["nombre"]),
        ("Grupo:",          ses["grupo"]),
        ("Fecha / Inicio:", (ses["inicio"] or "—")[:16].replace("T", "  ")),
        ("Fin real:",       (ses["fin_real"] or "—")[:16].replace("T", "  ") if ses["fin_real"] else "—"),
        ("Duración:",       f"{ses['duracion_min']} min" if ses["duracion_min"] else "—"),
        ("Estado:",         ses["estado"]),
        ("Total alumnos:",  data["total_alumnos"]),
    ]
    for ri, (lbl, val) in enumerate(meta, 2):
        c_lbl = ws.cell(row=ri, column=2, value=lbl)
        c_lbl.font = Font(bold=True, size=9, color="444444")
        c_val = ws.cell(row=ri, column=3, value=str(val))
        c_val.font = Font(size=9)
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28

    # ── Fila cabecera de tabla ────────────────────────────────────────────────
    HDR_ROW = len(meta) + 3
    HDRS = [
        ("#",             5),  ("Matrícula",     14), ("Nombre completo",  30),
        ("Carrera",      22),  ("Cuatrimestre",  14), ("Grupo",             8),
        ("PC",           10),  ("Entrada",       16), ("Salida",           16),
        ("Duración\n(min)", 13),
    ]
    for col, (h, w) in enumerate(HDRS, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.fill      = _as_fill(AZUL_OSC)
        c.font      = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _as_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HDR_ROW].height = 28

    # ── Filas de datos ────────────────────────────────────────────────────────
    for r, al in enumerate(alumnos, HDR_ROW + 1):
        fill_obj = _as_fill(GRIS_ROW) if r % 2 == 0 else None
        entrada = (al["hora_entrada"] or "—")[:16].replace("T", " ") if al["hora_entrada"] else "—"
        salida  = (
            (al["hora_salida"] or "")[:16].replace("T", " ") if al["hora_salida"]
            else ("En sesión" if al["activa"] else "—")
        )
        vals = [
            r - HDR_ROW,
            al["alumno_matricula"],
            al["alumno_nombre"],
            al.get("carrera")        or "—",
            al.get("cuatrimestre")   or "—",
            al.get("grupo_catalogo") or "—",
            al.get("pc_codigo")      or "—",
            entrada, salida,
            al.get("duracion_min")   or "—",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            if fill_obj:
                c.fill = fill_obj
            c.font      = Font(size=9)
            c.border    = _as_border()
            c.alignment = _as_center() if ci in (1, 5, 6, 7, 8, 9, 10) else _as_left()
        ws.row_dimensions[r].height = 16

    # ── Fila de total ─────────────────────────────────────────────────────────
    tot = HDR_ROW + len(alumnos) + 1
    ws.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True, size=10)
    c_tot = ws.cell(row=tot, column=2, value=data["total_alumnos"])
    c_tot.font = Font(bold=True, size=13, color="FFFFFF")
    c_tot.fill = _as_fill(AZUL_MED)
    c_tot.alignment = _as_center()

    ws.sheet_properties.tabColor = "1E3A5F"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/{sesion_id}/asistencia", summary="Lista de asistencia enriquecida de la sesión")
def lista_asistencia(
    sesion_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Retorna la lista de asistencia de una sesión con datos del catálogo de alumnos."""
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(404, "Sesión no encontrada")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(403, "Acceso denegado")

    data = _get_asistencia_data(sesion_id, db)
    return data


@router.get("/{sesion_id}/asistencia/excel", summary="Exportar lista de asistencia en Excel")
def asistencia_excel(
    sesion_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Descarga la lista de asistencia de la sesión como archivo Excel."""
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(404, "Sesión no encontrada")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(403, "Acceso denegado")

    data     = _get_asistencia_data(sesion_id, db)
    buf      = _build_asistencia_excel(data)
    mat_slug = s.materia[:18].replace(" ", "_")
    filename = f"Asistencia_{s.codigo_sesion}_{mat_slug}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
