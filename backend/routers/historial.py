"""
Router: Historial de Alumno
Endpoints:
  GET /reportes/historial-alumno        → JSON historial completo por matrícula
  GET /reportes/historial-alumno/excel  → Excel descargable
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import datetime, io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models.sesion import SesionClase, AsignacionPC
from models.laboratorio import Laboratorio, Computadora
from models.usuario import Usuario, RolUsuario
from models.catalogo import CatalogoAlumno
from dependencies import get_current_user

router = APIRouter(prefix="/reportes", tags=["Historial Alumno"])

# ─── Helper de datos ─────────────────────────────────────────────────────────

def _get_historial(matricula: str, db: Session) -> dict:
    asigs = db.query(AsignacionPC).filter(
        AsignacionPC.alumno_matricula == matricula
    ).all()

    cat = db.query(CatalogoAlumno).filter(
        CatalogoAlumno.matricula == matricula
    ).order_by(CatalogoAlumno.id.desc()).first()

    nombre_base = asigs[0].alumno_nombre if asigs else "—"
    nombre_cat  = (
        f"{cat.nombres} {cat.apellido_paterno} {cat.apellido_materno}".strip()
        if cat else None
    )

    alumno_info = {
        "matricula":    matricula,
        "nombre":       nombre_cat or nombre_base,
        "carrera":      cat.carrera      if cat else None,
        "cuatrimestre": cat.cuatrimestre if cat else None,
        "grupo":        cat.grupo        if cat else None,
        "periodo":      cat.periodo      if cat else None,
    }

    if not asigs:
        return {"alumno": alumno_info, "total_sesiones": 0, "total_horas": 0.0, "historial": []}

    sesion_ids = list({a.sesion_id for a in asigs})
    sesiones_map = {
        s.id: s for s in db.query(SesionClase).filter(
            SesionClase.id.in_(sesion_ids)
        ).all()
    }
    lab_ids  = list({s.laboratorio_id for s in sesiones_map.values()})
    labs_map = {
        l.id: l for l in db.query(Laboratorio).filter(
            Laboratorio.id.in_(lab_ids)
        ).all()
    }
    doc_ids  = list({s.docente_id for s in sesiones_map.values()})
    docs_map = {
        u.id: u for u in db.query(Usuario).filter(
            Usuario.id.in_(doc_ids)
        ).all()
    }
    pc_ids   = list({a.computadora_id for a in asigs})
    pcs_map  = {
        p.id: p for p in db.query(Computadora).filter(
            Computadora.id.in_(pc_ids)
        ).all()
    } if pc_ids else {}

    horas_totales = 0.0
    historial = []

    for a in asigs:
        ses = sesiones_map.get(a.sesion_id)
        if not ses:
            continue
        lab = labs_map.get(ses.laboratorio_id)
        doc = docs_map.get(ses.docente_id)
        pc  = pcs_map.get(a.computadora_id)

        fin   = a.hora_liberacion or ses.fin_real
        horas = None
        if a.hora_asignacion and fin:
            horas = round((fin - a.hora_asignacion).total_seconds() / 3600, 2)
            horas_totales += horas

        historial.append({
            "sesion_id":      ses.id,
            "codigo_sesion":  ses.codigo_sesion,
            "materia":        ses.materia,
            "grupo":          ses.grupo,
            "docente":        doc.nombre if doc else "—",
            "laboratorio":    lab.nombre if lab else "—",
            "laboratorio_id": lab.id     if lab else None,
            "fecha":          ses.inicio.strftime("%Y-%m-%d") if ses.inicio else None,
            "hora_entrada":   a.hora_asignacion.isoformat() if a.hora_asignacion else None,
            "hora_salida":    a.hora_liberacion.isoformat() if a.hora_liberacion else None,
            "horas":          horas,
            "pc_codigo":      pc.codigo if pc else None,
            "estado_sesion":  ses.estado,
        })

    historial.sort(key=lambda x: x["fecha"] or "", reverse=True)

    return {
        "alumno":         alumno_info,
        "total_sesiones": len(historial),
        "total_horas":    round(horas_totales, 2),
        "historial":      historial,
    }


# ─── Excel ────────────────────────────────────────────────────────────────────

def _bd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _build_historial_excel(data: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"
    ws.sheet_view.showGridLines = False

    al       = data["alumno"]
    historial = data["historial"]

    AZUL_OSC = "1E3A5F"
    AZUL_MED = "2D6A9F"
    GRIS_ROW = "EEF2F7"

    def _fill(h):
        return PatternFill("solid", fgColor=h)

    # Título
    ws.merge_cells("A1:I1")
    t = ws.cell(row=1, column=1,
                value=f"HISTORIAL DE ASISTENCIA — {al['nombre']}  [{al['matricula']}]")
    t.font      = Font(bold=True, size=14, color="FFFFFF")
    t.fill      = _fill(AZUL_OSC)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Metadata
    meta = [
        ("Matrícula:",    al["matricula"]),
        ("Nombre:",       al["nombre"]),
        ("Carrera:",      al.get("carrera")      or "—"),
        ("Cuatrimestre:", al.get("cuatrimestre") or "—"),
        ("Grupo:",        al.get("grupo")        or "—"),
        ("Período:",      al.get("periodo")      or "—"),
        ("Total sesiones:", data["total_sesiones"]),
        ("Horas totales:",  f"{data['total_horas']} h"),
    ]
    for ri, (lbl, val) in enumerate(meta, 2):
        ws.cell(row=ri, column=1, value=lbl).font = Font(bold=True, size=9, color="444444")
        ws.cell(row=ri, column=2, value=str(val)).font = Font(size=9)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32

    # Cabecera
    HDR_ROW = len(meta) + 3
    HDRS = [
        ("#", 5), ("Fecha", 12), ("Materia", 30), ("Grupo", 8),
        ("Docente", 24), ("Laboratorio", 22), ("PC", 10), ("Horas", 10), ("Estado", 14),
    ]
    for col, (h, w) in enumerate(HDRS, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.fill      = _fill(AZUL_OSC)
        c.font      = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _bd()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HDR_ROW].height = 22

    # Datos
    for r, h in enumerate(historial, HDR_ROW + 1):
        fill_obj = _fill(GRIS_ROW) if r % 2 == 0 else None
        vals = [
            r - HDR_ROW, h.get("fecha") or "—", h.get("materia") or "—",
            h.get("grupo") or "—", h.get("docente") or "—",
            h.get("laboratorio") or "—", h.get("pc_codigo") or "—",
            h.get("horas") if h.get("horas") is not None else "—",
            h.get("estado_sesion") or "—",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            if fill_obj:
                c.fill = fill_obj
            c.font      = Font(size=9)
            c.border    = _bd()
            c.alignment = Alignment(
                horizontal="center" if ci in (1, 2, 4, 7, 8, 9) else "left",
                vertical="center"
            )
        ws.row_dimensions[r].height = 16

    # Totales
    tot = HDR_ROW + len(historial) + 1
    ws.cell(row=tot, column=1, value="TOTAL SESIONES").font = Font(bold=True)
    c1 = ws.cell(row=tot, column=2, value=data["total_sesiones"])
    c1.font = Font(bold=True, size=12, color="FFFFFF")
    c1.fill = _fill(AZUL_MED)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=tot, column=7, value="TOTAL HORAS").font = Font(bold=True)
    c2 = ws.cell(row=tot, column=8, value=f"{data['total_horas']} h")
    c2.font = Font(bold=True, size=12, color="FFFFFF")
    c2.fill = _fill("1B5E20")
    c2.alignment = Alignment(horizontal="center", vertical="center")

    ws.sheet_properties.tabColor = "1E3A5F"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/historial-alumno", summary="Historial de asistencia por alumno (JSON)")
def historial_alumno_json(
    matricula: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if current_user.rol == RolUsuario.DOCENTE:
        raise HTTPException(403, "Acceso restringido a administradores")
    return _get_historial(matricula, db)


@router.get("/historial-alumno/excel", summary="Historial de asistencia por alumno (Excel)")
def historial_alumno_excel(
    matricula: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if current_user.rol == RolUsuario.DOCENTE:
        raise HTTPException(403, "Acceso restringido a administradores")
    data     = _get_historial(matricula, db)
    buf      = _build_historial_excel(data)
    slug     = matricula.replace("/", "-")
    filename = f"Historial_{slug}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
