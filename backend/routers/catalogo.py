from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from typing import Optional, List
from database import get_db
from models.catalogo import CatalogoAlumno, CatalogoMateria
from models.usuario import Usuario, RolUsuario
from dependencies import get_current_user, require_roles
import openpyxl
import io
import datetime

router = APIRouter(prefix="/catalogo", tags=["Catálogo"])


# ─── Schemas ───────────────────────────────────────────────────────────────────

class AlumnoCreate(BaseModel):
    matricula:        str = Field(..., min_length=1, max_length=30)
    apellido_paterno: str = Field(..., min_length=1, max_length=80)
    apellido_materno: str = Field(..., min_length=1, max_length=80)
    nombres:          str = Field(..., min_length=1, max_length=100)
    carrera:          str = Field(..., min_length=2, max_length=120)
    cuatrimestre:     int = Field(..., ge=1, le=12)
    grupo:            str = Field(..., min_length=1, max_length=5)
    periodo:          str = Field(..., min_length=4, max_length=20)

class AlumnoUpdate(BaseModel):
    matricula:        Optional[str]  = None
    apellido_paterno: Optional[str]  = None
    apellido_materno: Optional[str]  = None
    nombres:          Optional[str]  = None
    carrera:          Optional[str]  = None
    cuatrimestre:     Optional[int]  = None
    grupo:            Optional[str]  = None
    periodo:          Optional[str]  = None
    activo:           Optional[bool] = None

class MateriaCreate(BaseModel):
    nombre:               str          = Field(..., min_length=2, max_length=200)
    carrera:              Optional[str] = None
    cuatrimestre_oficial: Optional[int] = None
    periodo:              Optional[str] = None

class MateriaUpdate(BaseModel):
    nombre:               Optional[str]  = None
    carrera:              Optional[str]  = None
    cuatrimestre_oficial: Optional[int]  = None
    periodo:              Optional[str]  = None
    activo:               Optional[bool] = None


# ─── Helpers ───────────────────────────────────────────────────────────────────

def _norm(val) -> str:
    return str(val).strip() if val is not None else ""

def _serializar_alumno(a: CatalogoAlumno) -> dict:
    return {
        "id":               a.id,
        "matricula":        a.matricula,
        "apellido_paterno": a.apellido_paterno,
        "apellido_materno": a.apellido_materno,
        "nombres":          a.nombres,
        "nombre_completo":  f"{a.apellido_paterno} {a.apellido_materno} {a.nombres}".strip(),
        "carrera":          a.carrera,
        "cuatrimestre":     a.cuatrimestre,
        "grupo":            a.grupo,
        "periodo":          a.periodo,
        "activo":           a.activo,
    }

def _serializar_materia(m: CatalogoMateria) -> dict:
    return {
        "id":                   m.id,
        "nombre":               m.nombre,
        "carrera":              m.carrera,
        "cuatrimestre_oficial": m.cuatrimestre_oficial,
        "periodo":              m.periodo,
        "activo":               m.activo,
    }

_admin_roles = require_roles(
    RolUsuario.SUPER_ADMIN,
    RolUsuario.SERVICIOS_ESCOLARES,
)


# ═══════════════════════════════════════════════════════════════════════════════
#  ALUMNOS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/alumnos", summary="Listar alumnos del catálogo")
def listar_alumnos(
    periodo:      Optional[str]  = None,
    carrera:      Optional[str]  = None,
    grupo:        Optional[str]  = None,
    cuatrimestre: Optional[int]  = None,
    activo:       Optional[bool] = None,
    q:            Optional[str]  = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    query = db.query(CatalogoAlumno)
    if periodo      is not None: query = query.filter(CatalogoAlumno.periodo      == periodo)
    if carrera      is not None: query = query.filter(CatalogoAlumno.carrera      == carrera)
    if grupo        is not None: query = query.filter(CatalogoAlumno.grupo        == grupo)
    if cuatrimestre is not None: query = query.filter(CatalogoAlumno.cuatrimestre == cuatrimestre)
    if activo       is not None: query = query.filter(CatalogoAlumno.activo       == activo)
    if q:
        term = f"%{q}%"
        query = query.filter(
            CatalogoAlumno.nombres.ilike(term) |
            CatalogoAlumno.apellido_paterno.ilike(term) |
            CatalogoAlumno.apellido_materno.ilike(term) |
            CatalogoAlumno.matricula.ilike(term)
        )
    return [_serializar_alumno(a) for a in
            query.order_by(CatalogoAlumno.apellido_paterno, CatalogoAlumno.nombres).all()]


@router.get("/alumnos/buscar", summary="Autocomplete de alumnos para SesionActiva")
def buscar_alumnos(
    q:       str           = "",
    periodo: Optional[str] = None,
    grupo:   Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    if len(q.strip()) < 2:
        return []
    term  = f"%{q.strip()}%"
    query = db.query(CatalogoAlumno).filter(
        CatalogoAlumno.activo == True,
        (CatalogoAlumno.nombres.ilike(term) |
         CatalogoAlumno.apellido_paterno.ilike(term) |
         CatalogoAlumno.apellido_materno.ilike(term) |
         CatalogoAlumno.matricula.ilike(term))
    )
    if periodo: query = query.filter(CatalogoAlumno.periodo == periodo)
    if grupo:   query = query.filter(CatalogoAlumno.grupo   == grupo)
    return [
        {
            "id":               a.id,
            "nombres":          a.nombres,
            "apellido_paterno": a.apellido_paterno,
            "apellido_materno": a.apellido_materno,
            "matricula":        a.matricula,
            "grupo":            a.grupo,
            "carrera":          a.carrera,
            "periodo":          a.periodo,
        }
        for a in query.order_by(CatalogoAlumno.apellido_paterno).limit(10).all()
    ]


@router.get("/buscar-personas", summary="Búsqueda combinada: alumnos + personal")
def buscar_personas(
    q:  str = "",
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Autocomplete unificado para Consulta de Persona.
    Devuelve alumnos del catálogo Y usuarios del sistema (docentes, administrativos, etc.).
    """
    if len(q.strip()) < 2:
        return []
    term = f"%{q.strip()}%"

    # ── Alumnos ──────────────────────────────────────────────────────────────
    alumnos = db.query(CatalogoAlumno).filter(
        CatalogoAlumno.activo.isnot(False),
        (CatalogoAlumno.nombres.ilike(term) |
         CatalogoAlumno.apellido_paterno.ilike(term) |
         CatalogoAlumno.apellido_materno.ilike(term) |
         CatalogoAlumno.matricula.ilike(term))
    ).order_by(CatalogoAlumno.apellido_paterno).limit(10).all()

    # ── Personal (usuarios que no son ALUMNO) ─────────────────────────────
    personal = db.query(Usuario).filter(
        Usuario.activo == True,
        Usuario.rol != RolUsuario.ALUMNO,
        (Usuario.nombre.ilike(term) |
         Usuario.numero_empleado.ilike(term))
    ).order_by(Usuario.nombre).limit(10).all()

    results = []

    for a in alumnos:
        nombre = f"{a.apellido_paterno} {a.apellido_materno or ''} {a.nombres}".strip()
        results.append({
            "tipo":           "ALUMNO",
            "identificador":  a.matricula,
            "nombre":         nombre,
            "subtitulo":      f"Matrícula: {a.matricula}",
            "extra":          a.carrera,
            # campos para compatibilidad con AutocompleteInput legacy
            "id":               a.id,
            "nombres":          a.nombres,
            "apellido_paterno": a.apellido_paterno,
            "apellido_materno": a.apellido_materno,
            "matricula":        a.matricula,
            "grupo":            a.grupo,
            "carrera":          a.carrera,
        })

    for u in personal:
        results.append({
            "tipo":          "PERSONAL",
            "identificador": u.numero_empleado or f"USR{u.id}",
            "nombre":        u.nombre,
            "subtitulo":     f"No. Emp: {u.numero_empleado or 'N/A'} · {u.rol.value}",
            "extra":         None,
            "id":            u.id,
        })

    return results


@router.post("/alumnos", status_code=status.HTTP_201_CREATED, summary="Crear alumno manualmente")
def crear_alumno(
    data: AlumnoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    matricula = _norm(data.matricula)
    periodo   = _norm(data.periodo)

    existe = db.query(CatalogoAlumno).filter(
        CatalogoAlumno.matricula == matricula,
    ).first()
    if existe:
        raise HTTPException(status_code=409,
            detail="Ya existe un alumno con esa matrícula. Edítalo para actualizar sus datos.")

    alumno = CatalogoAlumno(
        matricula        = matricula,
        apellido_paterno = _norm(data.apellido_paterno),
        apellido_materno = _norm(data.apellido_materno),
        nombres          = _norm(data.nombres),
        carrera          = _norm(data.carrera),
        cuatrimestre     = data.cuatrimestre,
        grupo            = _norm(data.grupo).upper(),
        periodo          = periodo,
    )
    db.add(alumno)
    db.commit()
    db.refresh(alumno)
    return _serializar_alumno(alumno)


@router.put("/alumnos/{alumno_id}", summary="Actualizar alumno")
def actualizar_alumno(
    alumno_id: int,
    data: AlumnoUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    a = db.query(CatalogoAlumno).filter(CatalogoAlumno.id == alumno_id).first()
    if not a:
        raise HTTPException(404, "Alumno no encontrado")
    for field, val in data.dict(exclude_none=True).items():
        if field == "grupo" and val:
            val = val.upper()
        setattr(a, field, val)
    db.commit()
    db.refresh(a)
    return _serializar_alumno(a)


@router.delete("/alumnos/{alumno_id}", summary="Desactivar alumno")
def eliminar_alumno(
    alumno_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    a = db.query(CatalogoAlumno).filter(CatalogoAlumno.id == alumno_id).first()
    if not a:
        raise HTTPException(404, "Alumno no encontrado")
    a.activo = False
    db.commit()
    return {"mensaje": "Alumno desactivado"}


@router.post("/alumnos/importar", summary="Importar alumnos desde Excel (Plantilla_Alumnos_UTECAN.xlsx)")
async def importar_alumnos(
    file: UploadFile = File(...),
    preview: bool = False,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    """
    Importa (o previsualiza) alumnos desde un Excel con la Plantilla_Alumnos_UTECAN.

    - **preview=false** (por defecto): aplica los cambios en la base de datos.
    - **preview=true**: analiza el archivo y devuelve el desglose sin guardar nada.
    """
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(400, "Archivo Excel inválido o dañado")

    ws = wb["Alumnos"] if "Alumnos" in wb.sheetnames else wb.active

    creados           = 0
    actualizados      = 0
    sin_cambios       = 0
    errores           = []
    cambios_sensibles = []   # [{matricula, nombre, campo, antes, despues}]

    # Campos considerados sensibles: cambios que el admin debe revisar antes de confirmar
    CAMPOS_SENSIBLES = ("cuatrimestre", "carrera", "grupo", "periodo")

    # Plantilla: fila 1=título, 2=leyenda, 3=cabeceras, 4=ejemplo → datos desde fila 5
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        # Detener en la primera fila completamente vacía
        if all(v is None or _norm(v) == "" for v in row):
            break

        matricula        = _norm(row[0]) if row[0] is not None else ""
        apellido_paterno = _norm(row[1]) if row[1] is not None else ""
        apellido_materno = _norm(row[2]) if row[2] is not None else ""
        nombres          = _norm(row[3]) if row[3] is not None else ""
        carrera          = _norm(row[4]) if row[4] is not None else ""
        cuat_raw         = row[5]
        grupo            = _norm(row[6]).upper() if row[6] is not None else ""
        periodo          = _norm(row[7]) if row[7] is not None else ""

        # Validar campos requeridos
        fila_errores = []
        if not matricula:        fila_errores.append("matrícula vacía")
        if not apellido_paterno: fila_errores.append("apellido paterno vacío")
        if not nombres:          fila_errores.append("nombre(s) vacío")
        if not carrera:          fila_errores.append("carrera vacía")
        if not grupo:            fila_errores.append("grupo vacío")
        if not periodo:          fila_errores.append("periodo vacío")

        try:
            cuatrimestre = int(cuat_raw) if cuat_raw is not None else 0
            if not (1 <= cuatrimestre <= 12):
                fila_errores.append("cuatrimestre fuera de rango (1–12)")
        except (ValueError, TypeError):
            fila_errores.append("cuatrimestre inválido")
            cuatrimestre = 0

        if fila_errores:
            errores.append({
                "fila":    row_idx,
                "datos":   f"{matricula or '?'} — {apellido_paterno} {nombres}".strip(" —"),
                "errores": fila_errores,
            })
            continue

        existente = db.query(CatalogoAlumno).filter(
            CatalogoAlumno.matricula == matricula,
        ).first()

        nombre_completo = f"{apellido_paterno} {apellido_materno} {nombres}".strip()

        if existente:
            incoming = {
                "apellido_paterno": apellido_paterno,
                "apellido_materno": apellido_materno,
                "nombres":          nombres,
                "carrera":          carrera,
                "cuatrimestre":     cuatrimestre,
                "grupo":            grupo,
                "periodo":          periodo,
            }
            hay_cambio = False
            for campo, nuevo_val in incoming.items():
                actual_val = getattr(existente, campo)
                if str(actual_val).strip() != str(nuevo_val).strip():
                    hay_cambio = True
                    if campo in CAMPOS_SENSIBLES:
                        cambios_sensibles.append({
                            "matricula": matricula,
                            "nombre":    nombre_completo,
                            "campo":     campo,
                            "antes":     actual_val,
                            "despues":   nuevo_val,
                        })

            if hay_cambio:
                actualizados += 1
                if not preview:
                    existente.apellido_paterno = apellido_paterno
                    existente.apellido_materno = apellido_materno
                    existente.nombres          = nombres
                    existente.carrera          = carrera
                    existente.cuatrimestre     = cuatrimestre
                    existente.grupo            = grupo
                    existente.periodo          = periodo
                    existente.activo           = True
            else:
                sin_cambios += 1
                if not preview:
                    # Reactivar si estaba inactivo (sin contar como actualización)
                    if not existente.activo:
                        existente.activo = True
        else:
            creados += 1
            if not preview:
                db.add(CatalogoAlumno(
                    matricula        = matricula,
                    apellido_paterno = apellido_paterno,
                    apellido_materno = apellido_materno,
                    nombres          = nombres,
                    carrera          = carrera,
                    cuatrimestre     = cuatrimestre,
                    grupo            = grupo,
                    periodo          = periodo,
                ))

    if not preview:
        db.commit()

    return {
        "preview":           preview,
        "creados":           creados,
        "actualizados":      actualizados,
        "sin_cambios":       sin_cambios,
        "total_errores":     len(errores),
        "errores":           errores,
        "cambios_sensibles": cambios_sensibles,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  MATERIAS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/materias", summary="Listar materias del catálogo")
def listar_materias(
    periodo: Optional[str]  = None,
    carrera: Optional[str]  = None,
    activo:  Optional[bool] = None,
    q:       Optional[str]  = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    query = db.query(CatalogoMateria)
    if periodo is not None: query = query.filter(CatalogoMateria.periodo == periodo)
    if carrera is not None: query = query.filter(CatalogoMateria.carrera == carrera)
    if activo  is not None: query = query.filter(CatalogoMateria.activo  == activo)
    if q:       query = query.filter(CatalogoMateria.nombre.ilike(f"%{q}%"))
    return [_serializar_materia(m) for m in
            query.order_by(CatalogoMateria.nombre).all()]


@router.get("/materias/buscar", summary="Autocomplete de materias para Reservaciones")
def buscar_materias(
    q:  str = "",
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Devuelve todas las combinaciones materia+carrera+cuatrimestre que coinciden con la búsqueda.
    NO deduplica por nombre: si "Inglés" existe en 3 carreras, devuelve 3 opciones distintas.
    Incluye un campo `label` listo para mostrar en el selector del frontend.
    """
    if len(q.strip()) < 2:
        return []

    resultados = db.query(CatalogoMateria).filter(
        CatalogoMateria.activo == True,
        CatalogoMateria.nombre.ilike(f"%{q.strip()}%")
    ).order_by(CatalogoMateria.nombre, CatalogoMateria.carrera).limit(20).all()

    def _ordinal(n) -> str:
        if n is None:
            return "?"
        sufijos = {1:"er", 2:"do", 3:"er", 4:"to", 5:"to", 6:"to",
                   7:"mo", 8:"vo", 9:"no", 10:"mo", 11:"vo", 12:"vo"}
        return f"{n}{sufijos.get(int(n), 'o')}"

    items = []
    for m in resultados:
        tiene_carrera = bool(m.carrera and m.carrera.strip())
        tiene_cuat    = m.cuatrimestre_oficial is not None

        if tiene_carrera and tiene_cuat:
            label = f"{m.nombre} · {m.carrera} · {_ordinal(m.cuatrimestre_oficial)} cuatrimestre"
        elif tiene_carrera:
            label = f"{m.nombre} · {m.carrera}"
        elif tiene_cuat:
            label = f"{m.nombre} · {_ordinal(m.cuatrimestre_oficial)} cuatrimestre"
        else:
            label = m.nombre   # materia sin contexto académico registrado

        items.append({
            "id":                   m.id,
            "nombre":               m.nombre,
            "carrera":              m.carrera,
            "cuatrimestre_oficial": m.cuatrimestre_oficial,
            "periodo":              m.periodo,
            "label":                label,
            "tiene_contexto":       tiene_carrera or tiene_cuat,
        })

    return items


@router.post("/materias", status_code=status.HTTP_201_CREATED, summary="Crear materia manualmente")
def crear_materia(
    data: MateriaCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    m = CatalogoMateria(
        nombre               = _norm(data.nombre),
        carrera              = _norm(data.carrera) if data.carrera else None,
        cuatrimestre_oficial = data.cuatrimestre_oficial,
        periodo              = _norm(data.periodo) if data.periodo else None,
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return _serializar_materia(m)


@router.put("/materias/{materia_id}", summary="Actualizar materia")
def actualizar_materia(
    materia_id: int,
    data: MateriaUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    m = db.query(CatalogoMateria).filter(CatalogoMateria.id == materia_id).first()
    if not m:
        raise HTTPException(404, "Materia no encontrada")
    for field, val in data.dict(exclude_none=True).items():
        setattr(m, field, val)
    db.commit()
    db.refresh(m)
    return _serializar_materia(m)


@router.delete("/materias/{materia_id}", summary="Desactivar materia")
def eliminar_materia(
    materia_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    m = db.query(CatalogoMateria).filter(CatalogoMateria.id == materia_id).first()
    if not m:
        raise HTTPException(404, "Materia no encontrada")
    m.activo = False
    db.commit()
    return {"mensaje": "Materia desactivada"}


@router.post("/materias/importar", summary="Importar materias desde Excel (hoja concentrado)")
async def importar_materias(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_admin_roles),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(400, "Archivo Excel inválido o dañado")

    # Detectar formato según nombre de hoja
    if "concentrado" in wb.sheetnames:
        # Formato concentrado UTECAN: PERIODO(0), NIVEL(1), CARR.(2), CUAT.(3), GPO.(4), ASIGNATURA(5)
        ws = wb["concentrado"]
        min_row = 2
        col_nombre  = 5
        col_carrera = 2
        col_cuat    = 3
        col_periodo = 0
    elif "Materias" in wb.sheetnames:
        # Formato Plantilla_Materias_UTECAN.xlsx:
        # Fila 1=título, 2=leyenda, 3=vacía, 4=encabezados, 5=ejemplo, 6+=datos
        # Col A(0)=#, B(1)=nombre, C(2)=carrera, D(3)=cuatrimestre, E(4)=periodo
        ws = wb["Materias"]
        min_row = 6
        col_nombre  = 1
        col_carrera = 2
        col_cuat    = 3
        col_periodo = 4
    else:
        # Formato simple genérico: nombre(0), carrera(1), cuat(2), periodo(3)
        ws = wb.active
        min_row = 2
        col_nombre  = 0
        col_carrera = 1
        col_cuat    = 2
        col_periodo = 3

    creados     = 0
    actualizados = 0
    errores     = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), start=min_row):
        if all(v is None for v in row):
            break
        # Saltar fila de ejemplo (marcada con "→" en primera celda)
        if str(row[0] or "").strip() == "→":
            continue
        nombre  = _norm(row[col_nombre])  if col_nombre < len(row) and row[col_nombre]  is not None else ""
        carrera = _norm(row[col_carrera]) if col_carrera < len(row) and row[col_carrera] is not None else None
        periodo = _norm(row[col_periodo]) if col_periodo < len(row) and row[col_periodo] is not None else None

        try:
            cuat = int(row[col_cuat]) if col_cuat < len(row) and row[col_cuat] is not None else None
        except (ValueError, TypeError):
            cuat = None

        if not nombre:
            continue

        existente = db.query(CatalogoMateria).filter(
            CatalogoMateria.nombre  == nombre,
            CatalogoMateria.periodo == periodo,
        ).first()

        if existente:
            existente.carrera              = carrera
            existente.cuatrimestre_oficial = cuat
            existente.activo               = True
            actualizados += 1
        else:
            db.add(CatalogoMateria(
                nombre               = nombre,
                carrera              = carrera,
                cuatrimestre_oficial = cuat,
                periodo              = periodo,
            ))
            creados += 1

    db.commit()
    return {
        "creados":       creados,
        "actualizados":  actualizados,
        "total_errores": len(errores),
        "errores":       errores,
    }


# ─── Periodos disponibles (helper para selects en frontend) ────────────────────

@router.get("/periodos", summary="Periodos únicos en el catálogo")
def listar_periodos(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from sqlalchemy import distinct, union_all
    periodos_a = db.query(CatalogoAlumno.periodo).distinct()
    periodos_m = db.query(CatalogoMateria.periodo).distinct()
    todos = set()
    for (p,) in periodos_a.all():
        if p: todos.add(p)
    for (p,) in periodos_m.all():
        if p: todos.add(p)
    return sorted(todos, reverse=True)


@router.get("/carreras", summary="Carreras únicas en el catálogo")
def listar_carreras(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carreras = set()
    for (c,) in db.query(CatalogoAlumno.carrera).distinct().all():
        if c: carreras.add(c)
    for (c,) in db.query(CatalogoMateria.carrera).distinct().all():
        if c: carreras.add(c)
    return sorted(carreras)


# ─── Periodo actual calculado por fecha ───────────────────────────────────────

def _calcular_periodo(fecha: datetime.date) -> str:
    """
    Calcula el período escolar UTECAN según la fecha:
      1 ene – 30 abr  →  ENE-ABR YYYY
      1 may – 31 ago  →  MAY-AGO YYYY
      1 sep – 31 dic  →  SEP-DIC YYYY
    """
    mes  = fecha.month
    anio = fecha.year
    if mes <= 4:
        return f"ENE-ABR {anio}"
    elif mes <= 8:
        return f"MAY-AGO {anio}"
    else:
        return f"SEP-DIC {anio}"


@router.get("/periodo-actual", summary="Período escolar actual calculado por fecha")
def periodo_actual(
    current_user: Usuario = Depends(get_current_user),
):
    hoy    = datetime.date.today()
    periodo = _calcular_periodo(hoy)
    return {
        "periodo": periodo,
        "fecha":   hoy.isoformat(),
        "mes":     hoy.month,
    }
