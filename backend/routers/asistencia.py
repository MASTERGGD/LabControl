"""
Router: Lista de Asistencia por Sesión
Endpoints:
  GET /sesiones/{sesion_id}/asistencia        → JSON enriquecido con CatalogoAlumno
  GET /sesiones/{sesion_id}/asistencia/excel  → Excel descargable
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import datetime, io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models.sesion import SesionClase, AsignacionPC
from models.laboratorio import Laboratorio, Computadora
from models.usuario import Usuario, RolUsuario
from models.catalogo import CatalogoAlumno
from dependencies import get_current_user

router = APIRouter(prefix="/sesiones", tags=["Asistencia"])

# ─── Estilos ──────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── Helper de datos ─────────────────────────────────────────────────────────

def _get_asistencia_data(sesion_id: int, db: Session):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        return None

    lab = db.query(Laboratorio).filter(Laboratorio.id == s.laboratorio_id).first()
    doc = db.query(Usuario).filter(Usuario.id == s.docente_id).first()
    asigs = db.query(AsignacionPC).filter(AsignacionPC.sesion_id == sesion_id).all()

    # Enriquecer con catálogo
    matriculas = list({a.alumno_matricula for a in asigs})
    cat_map = {}
    if matriculas:
        for ca in db.query(CatalogoAlumno).filter(
            CatalogoAlumno.matricula.in_(matriculas)
        ).all():
            if ca.matricula not in cat_map:
                cat_map[ca.matricula] = ca

    pcs_ids = list({a.computadora_id for a in asigs})
    pcs_map = {
        p.id: p for p in db.query(Computadora).filter(
            Computadora.id.in_(pcs_ids)
        ).all()
    } if pcs_ids else {}

    alumnos = []
    for a in asigs:
        pc  = pcs_map.get(a.computadora_id)
        cat = cat_map.get(a.alumno_matricula)
        fin = a.hora_liberacion or s.fin_real
        duracion_min = (
            int((fin - a.hora_asignacion).total_seconds() / 60)
            if a.hora_asignacion and fin else None
        )
        alumnos.append({
            "asignacion_id":    a.id,
            "alumno_nombre":    a.alumno_nombre,
            "alumno_matricula": a.alumno_matricula,
            "pc_codigo":        pc.codigo if pc else None,
            "pc_fila":          pc.fila   if pc else None,
            "hora_entrada":     a.hora_asignacion.isoformat() if a.hora_asignacion else None,
            "hora_salida":      a.hora_liberacion.isoformat() if a.hora_liberacion else None,
            "duracion_min":     duracion_min,
            "activa":           a.hora_liberacion is None,
            "carrera":          cat.carrera      if cat else None,
            "cuatrimestre":     cat.cuatrimestre if cat else None,
            "grupo_catalogo":   cat.grupo        if cat else None,
            "periodo":          cat.periodo      if cat else None,
        })

    fin_sesion = s.fin_real or s.fin_estimado
    dur_sesion_min = (
        int((fin_sesion - s.inicio).total_seconds() / 60)
        if s.inicio and fin_sesion else None
    )

    return {
        "sesion": {
            "id":            s.id,
            "codigo_sesion": s.codigo_sesion,
            "materia":       s.materia,
            "grupo":         s.grupo,
            "estado":        s.estado,
            "inicio":        s.inicio.isoformat()       if s.inicio       else None,
            "fin_real":      s.fin_real.isoformat()     if s.fin_real     else None,
            "fin_estimado":  s.fin_estimado.isoformat() if s.fin_estimado else None,
            "duracion_min":  dur_sesion_min,
        },
        "laboratorio": {
            "id":     lab.id     if lab else None,
            "nombre": lab.nombre if lab else "—",
        },
        "docente": {
            "id":     doc.id     if doc else None,
            "nombre": doc.nombre if doc else "—",
            "email":  doc.email  if doc else None,
        },
        "alumnos":       alumnos,
        "total_alumnos": len(alumnos),
    }


# ─── Excel ────────────────────────────────────────────────────────────────────

def _build_asistencia_excel(data: dict) -> io.BytesIO:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Asistencia"
    ws.sheet_view.showGridLines = False

    ses     = data["sesion"]
    lab     = data["laboratorio"]
    doc     = data["docente"]
    alumnos = sorted(data["alumnos"], key=lambda x: x["alumno_nombre"])

    AZUL_OSC = "1E3A5F"
    AZUL_MED = "2D6A9F"
    GRIS_ROW = "EEF2F7"

    # Título
    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1,
                value=f"LISTA DE ASISTENCIA — {ses['materia']}  [{ses['codigo_sesion']}]")
    t.font      = Font(bold=True, size=14, color="FFFFFF")
    t.fill      = _fill(AZUL_OSC)
    t.alignment = _left()
    ws.row_dimensions[1].height = 32
    ws.column_dimensions["A"].width = 5

    # Metadata
    meta = [
        ("Laboratorio:",    lab["nombre"]),
        ("Docente:",        doc["nombre"]),
        ("Grupo:",          ses["grupo"]),
        ("Fecha/Inicio:",   (ses["inicio"] or "—")[:16].replace("T", "  ")),
        ("Fin real:",       (ses["fin_real"] or "—")[:16].replace("T", "  ") if ses["fin_real"] else "—"),
        ("Duración:",       f"{ses['duracion_min']} min" if ses["duracion_min"] else "—"),
        ("Estado:",         ses["estado"]),
        ("Total alumnos:",  data["total_alumnos"]),
    ]
    for ri, (lbl, val) in enumerate(meta, 2):
        ws.cell(row=ri, column=2, value=lbl).font  = Font(bold=True, size=9, color="444444")
        ws.cell(row=ri, column=3, value=str(val)).font = Font(size=9)
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28

    # Cabecera
    HDR_ROW = len(meta) + 3
    HDRS = [
        ("#", 5), ("Matrícula", 14), ("Nombre completo", 30),
        ("Carrera", 22), ("Cuatrimestre", 14), ("Grupo", 8),
        ("PC", 10), ("Entrada", 16), ("Salida", 16), ("Duración\n(min)", 13),
    ]
    for col, (h, w) in enumerate(HDRS, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.fill      = _fill(AZUL_OSC)
        c.font      = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HDR_ROW].height = 28

    # Datos
    for r, al in enumerate(alumnos, HDR_ROW + 1):
        fill_obj = _fill(GRIS_ROW) if r % 2 == 0 else None
        entrada = (al["hora_entrada"] or "")[:16].replace("T", " ") if al["hora_entrada"] else "—"
        salida  = (
            al["hora_salida"][:16].replace("T", " ") if al["hora_salida"]
            else ("En sesión" if al["activa"] else "—")
        )
        vals = [
            r - HDR_ROW, al["alumno_matricula"], al["alumno_nombre"],
            al.get("carrera") or "—", al.get("cuatrimestre") or "—",
            al.get("grupo_catalogo") or "—", al.get("pc_codigo") or "—",
            entrada, salida, al.get("duracion_min") or "—",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            if fill_obj:
                c.fill = fill_obj
            c.font      = Font(size=9)
            c.border    = _border()
            c.alignment = _center() if ci in (1, 5, 6, 7, 8, 9, 10) else _left()
        ws.row_dimensions[r].height = 16

    # Total
    tot = HDR_ROW + len(alumnos) + 1
    ws.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True, size=10)
    c_tot = ws.cell(row=tot, column=2, value=data["total_alumnos"])
    c_tot.font = Font(bold=True, size=13, color="FFFFFF")
    c_tot.fill = _fill(AZUL_MED)
    c_tot.alignment = _center()

    ws.sheet_properties.tabColor = "1E3A5F"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/{sesion_id}/asistencia", summary="Lista de asistencia enriquecida")
def lista_asistencia(
    sesion_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(404, "Sesión no encontrada")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(403, "Acceso denegado")
    return _get_asistencia_data(sesion_id, db)


@router.get("/{sesion_id}/asistencia/excel", summary="Exportar asistencia en Excel")
def asistencia_excel(
    sesion_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    s = db.query(SesionClase).filter(SesionClase.id == sesion_id).first()
    if not s:
        raise HTTPException(404, "Sesión no encontrada")
    if current_user.rol == RolUsuario.DOCENTE and s.docente_id != current_user.id:
        raise HTTPException(403, "Acceso denegado")

    data     = _get_asistencia_data(sesion_id, db)
    buf      = _build_asistencia_excel(data)
    mat_slug = s.materia[:18].replace(" ", "_")
    filename = f"Asistencia_{s.codigo_sesion}_{mat_slug}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
