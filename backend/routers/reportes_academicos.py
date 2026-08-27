import datetime
import hashlib
import io
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Optional
from xml.sax.saxutils import escape

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import Image, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from database import get_db
from dependencies import get_current_user
from models.catalogo import GrupoAcademico, InscripcionAlumno, PeriodoEscolar
from models.docencia import AsistenciaDocente, CargaDocente, ClaseDocente, CorreccionAsistenciaDocente, SeguimientoAlumnoDocente
from models.reporte_academico import EmisionReporteAcademico
from models.usuario import Usuario
from services.calendario_academico import estado_fecha_academica
from services.timezone import as_mx, format_fecha_corta_mx, now_mx, today_mx
from services.user_permissions import puede_gestionar_materias


router = APIRouter(prefix="/reportes-academicos", tags=["Reportes académicos"])
ESTADOS_CERRADOS = {"ATENDIDO", "CUMPLIDO", "CUMPLIDO_PARCIAL", "NO_CUMPLIDO", "CERRADO"}
MIN_SESIONES_PORCENTAJE = 3
MIN_SESIONES_PRIORIDAD = 5
LOGO_UTECAN = Path(__file__).resolve().parent.parent / "assets" / "tutoria" / "utecan_logo.jpg"


def _autorizar(db: Session, user: Usuario):
    if not puede_gestionar_materias(db, user):
        raise HTTPException(403, "Solo Dirección de División de Carrera puede consultar este reporte")


def _ids_grupos(valor: str) -> list[int]:
    try:
        ids = list(dict.fromkeys(int(item) for item in valor.split(",") if item.strip()))
    except ValueError:
        raise HTTPException(422, "La selección de grupos no es válida")
    if not ids or len(ids) > 50:
        raise HTTPException(422, "Selecciona entre 1 y 50 grupos")
    return ids


def _nombre_alumno(alumno) -> str:
    texto = f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombres}".strip()
    return " ".join(parte.capitalize() for parte in texto.split())


def _nombre_persona(valor: str) -> str:
    limpio = re.sub(r"\s*\(demo\)\s*$", "", valor or "", flags=re.IGNORECASE).strip()
    articulos = {"de", "del", "la", "las", "los", "y"}
    resultado = []
    for indice, parte in enumerate(limpio.split()):
        minuscula = parte.lower()
        if indice and minuscula in articulos:
            resultado.append(minuscula)
        elif parte.isupper() and len(parte) <= 4 and parte not in {"DE", "DEL", "LA", "LAS", "LOS"}:
            resultado.append(parte)
        else:
            resultado.append(parte.capitalize())
    return " ".join(resultado)


def _texto_presentacion(valor: Optional[str]) -> Optional[str]:
    if not valor:
        return None
    texto = re.sub(r"\s+", " ", valor.strip())
    return texto[:1].upper() + texto[1:] if texto else None


def _cantidad(cantidad: int, singular: str, plural: Optional[str] = None) -> str:
    return f"{cantidad} {singular if cantidad == 1 else (plural or singular + 's')}"


def _porcentaje(valor: Optional[float]) -> str:
    if valor is None:
        return "—"
    numero = float(valor)
    return f"{int(numero)}%" if numero.is_integer() else f"{numero:.1f}%"


def _fecha_iso(valor: str) -> datetime.date:
    return datetime.date.fromisoformat(valor)


def _rango_fecha_es(desde: str, hasta: str) -> str:
    inicio, fin = _fecha_iso(desde), _fecha_iso(hasta)
    meses = ("", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre")
    if inicio.year == fin.year:
        return f"{inicio.day} de {meses[inicio.month]} al {fin.day} de {meses[fin.month]} de {fin.year}"
    return f"{inicio.day} de {meses[inicio.month]} de {inicio.year} al {fin.day} de {meses[fin.month]} de {fin.year}"


def _nivel_presentacion(valor: str) -> str:
    return (valor or "").lower().capitalize()


def _hora_excel(valor: Optional[str]) -> Optional[datetime.time]:
    if not valor:
        return None
    return datetime.datetime.strptime(valor, "%H:%M").time()


def _carrera_corta(valor: str) -> str:
    alto = (valor or "").upper()
    conocidos = {
        "INTELIGENCIA ARTIFICIAL": "TSU IA",
        "DESARROLLO Y GESTIÓN DE SOFTWARE": "ING. DGS",
        "DESARROLLO Y GESTION DE SOFTWARE": "ING. DGS",
        "CONTADURÍA": "CONTADURÍA",
        "CONTADURIA": "CONTADURÍA",
        "AGRICULTURA SUSTENTABLE": "AGRICULTURA",
        "PARAMÉDICO": "TSU PARAMÉDICO",
        "PARAMEDICO": "TSU PARAMÉDICO",
    }
    for clave, corto in conocidos.items():
        if clave in alto:
            return corto
    palabras = [p for p in re.findall(r"[A-ZÁÉÍÓÚÑ]+", alto) if p not in {"DE", "DEL", "EN", "Y", "LA", "EL", "SUPERIOR", "UNIVERSITARIO"}]
    return " ".join(palabras[:4]) or valor


def _limites_periodo(clave: str) -> tuple[datetime.date, datetime.date]:
    match = re.search(r"(ENE-ABR|MAY-AGO|SEP-DIC)\s+(\d{4})", (clave or "").upper())
    if not match:
        hoy = datetime.date.today()
        return hoy.replace(month=1, day=1), hoy
    bloque, anio_txt = match.groups(); anio = int(anio_txt)
    if bloque == "ENE-ABR": return datetime.date(anio, 1, 1), datetime.date(anio, 4, 30)
    if bloque == "MAY-AGO": return datetime.date(anio, 5, 1), datetime.date(anio, 8, 31)
    return datetime.date(anio, 9, 1), datetime.date(anio, 12, 31)


def _fechas_programadas(
    db: Session,
    periodo_id: int,
    carga: CargaDocente,
    inicio: datetime.date,
    fin: datetime.date,
    lectividad: dict[datetime.date, bool],
) -> int:
    if fin < inicio:
        return 0
    total = 0; fecha = inicio
    while fecha <= fin:
        if fecha not in lectividad:
            lectividad[fecha] = estado_fecha_academica(db, periodo_id, fecha)["es_lectiva"]
        if fecha.weekday() == carga.dia_semana and lectividad[fecha]:
            total += 1
        fecha += datetime.timedelta(days=1)
    return total


def _mostrar_asistencia(asistio: int, registros: int, sesiones: int) -> dict:
    porcentaje = round(asistio * 100 / registros, 1) if registros else None
    publicable = sesiones >= MIN_SESIONES_PORCENTAJE and porcentaje is not None
    return {"porcentaje": porcentaje if publicable else None, "porcentaje_observado": porcentaje,
            "asistio": asistio, "registros": registros,
            "sesiones_base": sesiones, "publicable": publicable,
            "texto": f"{asistio} de {registros}" if registros else "Sin registros"}


def _slug(valor: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", valor).strip("_") or "reporte_academico"


def _datos_reporte(db: Session, periodo_id: int, grupo_ids: list[int], desde: Optional[datetime.date], hasta: Optional[datetime.date]):
    periodo = db.query(PeriodoEscolar).filter(PeriodoEscolar.id == periodo_id).first()
    if not periodo:
        raise HTTPException(404, "Periodo escolar no encontrado")
    grupos = db.query(GrupoAcademico).filter(
        GrupoAcademico.id.in_(grupo_ids), GrupoAcademico.periodo_id == periodo_id,
    ).order_by(GrupoAcademico.carrera, GrupoAcademico.cuatrimestre, GrupoAcademico.grupo).all()
    if len(grupos) != len(grupo_ids):
        raise HTTPException(422, "Uno o más grupos no pertenecen al periodo seleccionado")
    periodo_inicio, periodo_fin = _limites_periodo(periodo.clave)
    inicio_reporte = max(desde or periodo_inicio, periodo_inicio)
    fin_reporte = min(hasta or periodo_fin, periodo_fin, datetime.date.today())

    cargas = db.query(CargaDocente).filter(
        CargaDocente.periodo_id == periodo_id,
        CargaDocente.grupo_academico_id.in_(grupo_ids),
        CargaDocente.tipo_actividad == "CLASE",
        CargaDocente.activo == True,
    ).order_by(CargaDocente.grupo_academico_id, CargaDocente.id).all()
    carga_ids = [c.id for c in cargas]
    clases_q = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id.in_(carga_ids), ClaseDocente.estado == "CERRADA",
    )
    clases_q = clases_q.filter(ClaseDocente.fecha >= inicio_reporte, ClaseDocente.fecha <= fin_reporte)
    clases = clases_q.order_by(ClaseDocente.fecha.desc()).all()
    clase_ids = [c.id for c in clases]
    correcciones = db.query(CorreccionAsistenciaDocente).filter(
        CorreccionAsistenciaDocente.clase_docente_id.in_(clase_ids),
    ).order_by(CorreccionAsistenciaDocente.clase_docente_id, CorreccionAsistenciaDocente.creado_en, CorreccionAsistenciaDocente.id).all() if clase_ids else []
    asistencias = db.query(AsistenciaDocente).filter(
        AsistenciaDocente.clase_docente_id.in_(clase_ids),
    ).order_by(AsistenciaDocente.clase_docente_id, AsistenciaDocente.alumno_id).all() if clase_ids else []
    seguimientos = db.query(SeguimientoAlumnoDocente).filter(
        SeguimientoAlumnoDocente.carga_docente_id.in_(carga_ids),
        SeguimientoAlumnoDocente.tipo != "CALIFICACION",
    ).all() if carga_ids else []
    inscripciones = db.query(InscripcionAlumno).filter(
        InscripcionAlumno.grupo_academico_id.in_(grupo_ids), InscripcionAlumno.estado == "ACTIVO",
    ).all()

    carga_por_id = {c.id: c for c in cargas}
    alumno_grupo = {i.alumno_id: i.grupo_academico_id for i in inscripciones}
    alumnos = {i.alumno_id: i.alumno for i in inscripciones}
    asistencias_alumno = defaultdict(list)
    for item in asistencias:
        asistencias_alumno[item.alumno_id].append(item.estado)
    seguimientos_alumno = defaultdict(list)
    for item in seguimientos:
        seguimientos_alumno[item.alumno_id].append(item)

    materias_map = {}
    for carga in cargas:
        key = (carga.grupo_academico_id, carga.materia_id or carga.actividad_nombre, carga.docente_id)
        materias_map.setdefault(key, {"cargas": [], "clases": []})["cargas"].append(carga)
    for clase in clases:
        carga = carga_por_id[clase.carga_docente_id]
        key = (carga.grupo_academico_id, carga.materia_id or carga.actividad_nombre, carga.docente_id)
        materias_map.setdefault(key, {"cargas": [carga], "clases": []})["clases"].append(clase)

    materias = []
    lectividad: dict[datetime.date, bool] = {}
    for bloque in materias_map.values():
        carga = bloque["cargas"][0]
        sesiones = bloque["clases"]
        ids = {c.id for c in sesiones}
        estados = [a.estado for a in asistencias if a.clase_docente_id in ids]
        asistio = sum(1 for e in estados if e in {"PRESENTE", "RETARDO", "JUSTIFICADA"})
        programadas = sum(
            _fechas_programadas(db, periodo_id, item, inicio_reporte, fin_reporte, lectividad)
            for item in bloque["cargas"]
        )
        cobertura = round(len(sesiones) * 100 / programadas, 1) if programadas else None
        ultima = max(sesiones, key=lambda c: c.fecha) if sesiones else None
        avances = [c.avance_planeacion for c in sesiones if c.avance_planeacion is not None]
        asistencia_info = _mostrar_asistencia(asistio, len(estados), len(sesiones))
        materias.append({
            "grupo_id": carga.grupo_academico_id,
            "materia": carga.actividad_nombre,
            "docente": _nombre_persona(carga.docente.nombre) if carga.docente else "Sin docente",
            "sesiones": len(sesiones), "sesiones_programadas": programadas, "cobertura": cobertura,
            "asistencia": asistencia_info["porcentaje"], "asistencia_detalle": asistencia_info,
            "avance_sesion": round(sum(avances) / len(avances), 1) if avances else None,
            "ultimo_tema": _texto_presentacion(ultima.tema_impartido) if ultima else None,
            "pendiente": _texto_presentacion(ultima.tema_pendiente) if ultima else None,
            "extemporaneas": sum(1 for c in sesiones if c.es_extemporanea),
            "corregidas": len({r.clase_docente_id for r in correcciones if r.clase_docente_id in ids}),
        })

    alumnos_atencion = []
    for alumno_id, alumno in alumnos.items():
        estados = asistencias_alumno[alumno_id]
        asistio = sum(1 for e in estados if e in {"PRESENTE", "RETARDO", "JUSTIFICADA"})
        porcentaje = round(asistio * 100 / len(estados), 1) if estados else None
        abiertos = [s for s in seguimientos_alumno[alumno_id] if s.estado not in ESTADOS_CERRADOS]
        faltas = estados.count("FALTA")
        if faltas or abiertos:
            sesiones_base = len(estados)
            if sesiones_base < MIN_SESIONES_PORCENTAJE and not abiertos:
                nivel = "DATOS INSUFICIENTES"
            elif len(abiertos) >= 2 or (sesiones_base >= MIN_SESIONES_PRIORIDAD and porcentaje is not None and porcentaje < 80):
                nivel = "PRIORITARIO"
            elif abiertos or (sesiones_base >= MIN_SESIONES_PORCENTAJE and porcentaje is not None and porcentaje < 85):
                nivel = "ATENCIÓN"
            else:
                nivel = "OBSERVACIÓN"
            alumnos_atencion.append({
                "grupo_id": alumno_grupo.get(alumno_id), "alumno_id": alumno_id,
                "matricula": alumno.matricula, "nombre": _nombre_alumno(alumno),
                "asistencia": porcentaje if sesiones_base >= MIN_SESIONES_PORCENTAJE else None,
                "asistencias_registradas": asistio, "registros": sesiones_base,
                "faltas": faltas, "seguimientos_abiertos": len(abiertos), "nivel": nivel,
            })

    incidencias = [{
        "grupo_id": carga_por_id[c.carga_docente_id].grupo_academico_id,
        "fecha": c.fecha.isoformat(), "materia": carga_por_id[c.carga_docente_id].actividad_nombre,
        "docente": _nombre_persona(carga_por_id[c.carga_docente_id].docente.nombre) if carga_por_id[c.carga_docente_id].docente else "",
        "tipo": (c.incidencia_tipo or "OTRA").replace("_", " ").title(),
        "descripcion": _texto_presentacion(c.incidencias), "requiere_seguimiento": c.incidencia_requiere_seguimiento,
    } for c in clases if c.incidencias]

    observaciones = [{
        "grupo_id": carga_por_id[c.carga_docente_id].grupo_academico_id,
        "fecha": c.fecha.isoformat(), "materia": carga_por_id[c.carga_docente_id].actividad_nombre,
        "docente": _nombre_persona(carga_por_id[c.carga_docente_id].docente.nombre) if carga_por_id[c.carga_docente_id].docente else "",
        "tema": _texto_presentacion(c.tema_impartido), "actividades": _texto_presentacion(c.actividades_realizadas), "pendiente": _texto_presentacion(c.tema_pendiente),
    } for c in clases if c.tema_impartido or c.actividades_realizadas or c.tema_pendiente]

    grupos_json = []
    for grupo in grupos:
        inscritos = [i for i in inscripciones if i.grupo_academico_id == grupo.id]
        mats = [m for m in materias if m["grupo_id"] == grupo.id]
        estados = [a.estado for a in asistencias if alumno_grupo.get(a.alumno_id) == grupo.id]
        asistio = sum(1 for e in estados if e in {"PRESENTE", "RETARDO", "JUSTIFICADA"})
        sesiones_registradas = sum(m["sesiones"] for m in mats)
        sesiones_programadas = sum(m["sesiones_programadas"] for m in mats)
        asistencia_info = _mostrar_asistencia(asistio, len(estados), sesiones_registradas)
        grupos_json.append({
            "id": grupo.id, "nombre": f"{grupo.cuatrimestre}° {grupo.grupo}", "carrera": grupo.carrera,
            "carrera_corta": _carrera_corta(grupo.carrera),
            "alumnos": len(inscritos), "materias": len(mats),
            "sesiones": sesiones_registradas, "sesiones_programadas": sesiones_programadas,
            "cobertura": round(sesiones_registradas * 100 / sesiones_programadas, 1) if sesiones_programadas else None,
            "asistencia": asistencia_info["porcentaje"], "asistencia_detalle": asistencia_info,
            "incidencias": sum(1 for i in incidencias if i["grupo_id"] == grupo.id),
            "alumnos_atencion": sum(1 for a in alumnos_atencion if a["grupo_id"] == grupo.id),
        })

    total_estados = [a.estado for a in asistencias]
    total_asistio = sum(1 for e in total_estados if e in {"PRESENTE", "RETARDO", "JUSTIFICADA"})
    programadas_total = sum(m["sesiones_programadas"] for m in materias)
    asistencia_general = _mostrar_asistencia(total_asistio, len(total_estados), len(clases))
    correcciones_por_clase = defaultdict(list)
    correcciones_por_asistencia = defaultdict(list)
    for correccion in correcciones:
        correcciones_por_clase[correccion.clase_docente_id].append(correccion)
        if correccion.asistencia_id:
            correcciones_por_asistencia[correccion.asistencia_id].append(correccion)
    sesiones_detalle = []
    for clase in clases:
        carga = carga_por_id[clase.carga_docente_id]
        estados_clase = [a.estado for a in asistencias if a.clase_docente_id == clase.id]
        movimientos = correcciones_por_clase[clase.id]
        sesiones_detalle.append({
            "sesion_id": clase.id, "grupo_id": carga.grupo_academico_id,
            "fecha": clase.fecha.isoformat(), "materia": carga.actividad_nombre,
            "docente": _nombre_persona(carga.docente.nombre) if carga.docente else "Sin docente",
            "hora_inicio": carga.hora_inicio, "hora_fin": carga.hora_fin, "estado": clase.estado,
            "presentes": estados_clase.count("PRESENTE"), "faltas": estados_clase.count("FALTA"),
            "retardos": estados_clase.count("RETARDO"), "justificadas": estados_clase.count("JUSTIFICADA"),
            "total_registros": len(estados_clase), "tema": _texto_presentacion(clase.tema_impartido),
            "cumplimiento_declarado": clase.avance_planeacion,
            "extemporanea": bool(clase.es_extemporanea), "motivo_extemporaneo": clase.motivo_extemporaneo,
            "movimientos_correccion": len(movimientos),
            "ultima_correccion": max((r.creado_en for r in movimientos), default=None).isoformat() if movimientos else None,
        })
    asistencia_sesion = []
    for asistencia in asistencias:
        clase = next(c for c in clases if c.id == asistencia.clase_docente_id)
        carga = carga_por_id[clase.carga_docente_id]
        alumno = alumnos.get(asistencia.alumno_id) or asistencia.alumno
        movimientos = sorted(correcciones_por_asistencia[asistencia.id], key=lambda r: (r.creado_en, r.id))
        estado_original = movimientos[0].estado_anterior if movimientos and movimientos[0].estado_anterior else asistencia.estado
        asistencia_sesion.append({
            "sesion_id": clase.id, "grupo_id": carga.grupo_academico_id,
            "fecha": clase.fecha.isoformat(), "materia": carga.actividad_nombre,
            "docente": _nombre_persona(carga.docente.nombre) if carga.docente else "Sin docente",
            "matricula": alumno.matricula, "alumno": _nombre_alumno(alumno),
            "estado_original": estado_original, "estado_actual": asistencia.estado,
            "observacion": asistencia.observacion, "movimientos_correccion": len(movimientos),
        })
    sesiones_especiales = [{
        "grupo_id": carga_por_id[c.carga_docente_id].grupo_academico_id,
        "fecha": c.fecha.isoformat(), "materia": carga_por_id[c.carga_docente_id].actividad_nombre,
        "docente": _nombre_persona(carga_por_id[c.carga_docente_id].docente.nombre) if carga_por_id[c.carga_docente_id].docente else "",
        "extemporanea": c.es_extemporanea, "motivo_extemporaneo": c.motivo_extemporaneo,
        "correcciones": len(correcciones_por_clase[c.id]),
        "ultima_correccion": max((r.creado_en for r in correcciones_por_clase[c.id]), default=None).isoformat() if correcciones_por_clase[c.id] else None,
    } for c in clases if c.es_extemporanea or any(r.clase_docente_id == c.id for r in correcciones)]
    return {
        "periodo": {"id": periodo.id, "clave": periodo.clave},
        "filtros": {"desde": inicio_reporte.isoformat(), "hasta": fin_reporte.isoformat()},
        "resumen": {"grupos": len(grupos), "alumnos": len(alumnos), "materias": len(materias),
                    "sesiones": len(clases), "sesiones_programadas": programadas_total,
                    "cobertura": round(len(clases) * 100 / programadas_total, 1) if programadas_total else None,
                    "asistencia": asistencia_general["porcentaje"], "asistencia_detalle": asistencia_general,
                    "incidencias": len(incidencias), "alumnos_atencion": len(alumnos_atencion)},
        "grupos": grupos_json, "materias": materias,
        "alumnos_atencion": sorted(alumnos_atencion, key=lambda a: (a["nivel"] != "PRIORITARIO", a["nombre"])),
        "incidencias": incidencias, "observaciones_academicas": observaciones,
        "sesiones_especiales": sesiones_especiales,
        "sesiones_detalle": sesiones_detalle,
        "asistencia_sesion": sorted(asistencia_sesion, key=lambda a: (a["fecha"], a["sesion_id"], a["matricula"])),
        "criterios": {"min_sesiones_porcentaje": MIN_SESIONES_PORCENTAJE, "min_sesiones_prioridad": MIN_SESIONES_PRIORIDAD,
                      "meta_institucional": None,
                      "niveles": "Datos insuficientes: menos de 3 registros sin seguimiento; Atención: seguimiento abierto o asistencia menor a 85% con 3 sesiones; Prioritario: dos seguimientos abiertos o asistencia menor a 80% con 5 sesiones."},
        "privacidad": "Documento de uso interno. Contiene datos personales. Incluye únicamente información académica y de asistencia; los seguimientos de carácter personal se consultan en el módulo de Tutoría con los permisos correspondientes.",
    }


def _obtener_emision(db: Session, data: dict, usuario: Usuario) -> EmisionReporteAcademico:
    contenido = json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    content_hash = hashlib.sha256(contenido.encode("utf-8")).hexdigest()
    existente = db.query(EmisionReporteAcademico).filter(
        EmisionReporteAcademico.content_hash == content_hash,
    ).first()
    if existente:
        return existente
    alcance = json.dumps(sorted(g["id"] for g in data["grupos"]))
    emision = EmisionReporteAcademico(
        periodo_id=data["periodo"]["id"], generado_por_id=usuario.id,
        alcance=alcance, fecha_desde=data["filtros"]["desde"], fecha_hasta=data["filtros"]["hasta"],
        content_hash=content_hash,
    )
    db.add(emision)
    try:
        db.flush()
        periodo_corto = re.sub(r"[^A-Z0-9]", "", data["periodo"]["clave"].upper())
        emision.folio = f"RA-{periodo_corto}-{emision.id:06d}"
        db.commit()
    except IntegrityError:
        db.rollback()
        existente = db.query(EmisionReporteAcademico).filter(
            EmisionReporteAcademico.content_hash == content_hash,
        ).first()
        if existente:
            return existente
        raise
    db.refresh(emision)
    return emision


@router.get("/catalogos")
def catalogos(db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    _autorizar(db, current_user)
    periodos = db.query(PeriodoEscolar).order_by(PeriodoEscolar.id.desc()).all()
    grupos = db.query(GrupoAcademico).filter(GrupoAcademico.activo == True).order_by(
        GrupoAcademico.periodo_id.desc(), GrupoAcademico.carrera, GrupoAcademico.cuatrimestre, GrupoAcademico.grupo,
    ).all()
    return {"periodos": [{"id": p.id, "clave": p.clave, "es_actual": p.es_actual} for p in periodos],
            "grupos": [{"id": g.id, "periodo_id": g.periodo_id, "carrera": g.carrera,
                        "nombre": f"{g.cuatrimestre}° {g.grupo}", "turno": g.turno} for g in grupos]}


@router.get("")
def consultar(periodo_id: int, grupos: str, desde: Optional[datetime.date] = None, hasta: Optional[datetime.date] = None,
              db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    _autorizar(db, current_user)
    if desde and hasta and hasta < desde:
        raise HTTPException(422, "La fecha final debe ser igual o posterior a la inicial")
    return _datos_reporte(db, periodo_id, _ids_grupos(grupos), desde, hasta)


def _excel(data: dict, emision: EmisionReporteAcademico, usuario: Usuario) -> io.BytesIO:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    encabezado = PatternFill("solid", fgColor="0F766E")
    borde = "D9E2E8"

    def hoja_datos(nombre, columnas, filas, formatos=None, anchos=None, mensaje_vacio=None):
        ws = wb.create_sheet(nombre); ws.sheet_view.showGridLines = False; ws.append(columnas)
        for fila in filas: ws.append(fila)
        for celda in ws[1]:
            celda.font = Font(bold=True, color="FFFFFF"); celda.fill = encabezado
            celda.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 32; ws.freeze_panes = "A2"
        if filas:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{len(filas)+1}"
        elif mensaje_vacio:
            ws.cell(2, 1, mensaje_vacio); ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas))
            ws.cell(2, 1).font = Font(italic=True, color="64748B")
        formatos = formatos or {}; anchos = anchos or {}
        for indice, columna in enumerate(columnas, 1):
            ws.column_dimensions[get_column_letter(indice)].width = anchos.get(columna, min(42, max(14, len(columna) + 3)))
            if columna in formatos:
                for fila in range(2, len(filas) + 2): ws.cell(fila, indice).number_format = formatos[columna]
        for row in ws.iter_rows(min_row=2, max_row=max(2, len(filas)+1)):
            for celda in row:
                if isinstance(celda.value, (int, float)) and celda.number_format == "General":
                    celda.number_format = "#,##0"
                celda.alignment = Alignment(vertical="top", wrap_text=celda.column <= len(columnas) and columnas[celda.column-1] in {"Descripción", "Tema", "Actividades", "Pendiente", "Motivo", "Observación"})
        return ws

    autor = getattr(emision, "generado_por", None) or usuario
    r = data["resumen"]; local = as_mx(emision.generado_en)
    resumen = wb.create_sheet("Resumen"); resumen.sheet_view.showGridLines = False
    resumen.merge_cells("A1:D1"); resumen["A1"] = "REPORTE ACADÉMICO DE GRUPOS"; resumen["A1"].font = Font(size=16, bold=True, color="0F172A")
    metadatos = [
        ("Folio", emision.folio), ("Periodo", data["periodo"]["clave"]),
        ("Fecha inicial", _fecha_iso(data["filtros"]["desde"])), ("Fecha final", _fecha_iso(data["filtros"]["hasta"])),
        ("Generado por", _nombre_persona(autor.nombre)), ("Cargo", autor.departamento.nombre if getattr(autor, "departamento", None) else "Dirección de División de Carrera"),
        ("Fecha y hora de emisión", local.replace(tzinfo=None)), ("Zona horaria", "America/Mexico_City (Campeche)"),
        ("Alcance", f'{_cantidad(r["grupos"], "grupo")}, {_cantidad(r["alumnos"], "alumno")} y {_cantidad(r["materias"], "materia")}'),
    ]
    for fila, (etiqueta, valor) in enumerate(metadatos, 3):
        resumen.cell(fila, 1, etiqueta).font = Font(bold=True, color="0F766E"); resumen.cell(fila, 2, valor)
    resumen["B5"].number_format = "dd/mm/yyyy"; resumen["B6"].number_format = "dd/mm/yyyy"; resumen["B9"].number_format = "dd/mm/yyyy hh:mm"
    inicio_indicadores = 14
    resumen.cell(inicio_indicadores, 1, "Indicador"); resumen.cell(inicio_indicadores, 2, "Valor")
    for celda in resumen[inicio_indicadores]: celda.font = Font(bold=True, color="FFFFFF"); celda.fill = encabezado
    indicadores = [
        ("Sesiones registradas", r["sesiones"]), ("Sesiones programadas", r["sesiones_programadas"]),
        ("Cobertura", r["cobertura"] / 100 if r["cobertura"] is not None else None),
        ("Asistencias", r["asistencia_detalle"]["asistio"]), ("Registros de asistencia", r["asistencia_detalle"]["registros"]),
        ("Porcentaje de asistencia observado", r["asistencia_detalle"]["porcentaje_observado"] / 100 if r["asistencia_detalle"]["porcentaje_observado"] is not None else None),
        ("Muestra suficiente para publicar porcentaje", "Sí" if r["asistencia_detalle"]["publicable"] else "No"),
        ("Incidencias", r["incidencias"]), ("Alumnos con indicador", r["alumnos_atencion"]),
    ]
    for fila, (etiqueta, valor) in enumerate(indicadores, inicio_indicadores + 1): resumen.cell(fila, 1, etiqueta); resumen.cell(fila, 2, valor)
    resumen.cell(inicio_indicadores + 3, 2).number_format = "0.0%"; resumen.cell(inicio_indicadores + 6, 2).number_format = "0.0%"
    for fila in range(inicio_indicadores + 1, inicio_indicadores + len(indicadores) + 1):
        if isinstance(resumen.cell(fila, 2).value, int): resumen.cell(fila, 2).number_format = "#,##0"
    notas_fila = inicio_indicadores + len(indicadores) + 3
    resumen.cell(notas_fila, 1, "Criterios de clasificación").font = Font(bold=True, color="0F766E")
    resumen.merge_cells(start_row=notas_fila+1, start_column=1, end_row=notas_fila+2, end_column=4); resumen.cell(notas_fila+1, 1, data["criterios"]["niveles"]); resumen.cell(notas_fila+1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    resumen.cell(notas_fila+4, 1, "Privacidad").font = Font(bold=True, color="0F766E")
    resumen.merge_cells(start_row=notas_fila+5, start_column=1, end_row=notas_fila+6, end_column=4); resumen.cell(notas_fila+5, 1, data["privacidad"]); resumen.cell(notas_fila+5, 1).alignment = Alignment(wrap_text=True, vertical="top")
    resumen.column_dimensions["A"].width = 42; resumen.column_dimensions["B"].width = 52; resumen.column_dimensions["C"].width = 20; resumen.column_dimensions["D"].width = 20

    grupos = {g["id"]: f'{g["nombre"]} · {g["carrera_corta"]}' for g in data["grupos"]}
    pct = {"Cobertura": "0.0%", "Asistencia observada": "0.0%", "Cumplimiento declarado": "0.0%", "% asistencia observado": "0.0%"}
    fecha = {"Fecha": "dd/mm/yyyy", "Última corrección": "dd/mm/yyyy hh:mm", "Hora inicio": "hh:mm", "Hora fin": "hh:mm"}
    hoja_datos("Grupos", ["Grupo", "Alumnos", "Materias", "Sesiones registradas", "Sesiones programadas", "Cobertura", "Asistencias", "Registros", "Asistencia observada", "Nota de muestra", "Incidencias", "Alumnos con indicador"], [[grupos[g["id"]],g["alumnos"],g["materias"],g["sesiones"],g["sesiones_programadas"],g["cobertura"]/100 if g["cobertura"] is not None else None,g["asistencia_detalle"]["asistio"],g["asistencia_detalle"]["registros"],g["asistencia_detalle"]["porcentaje_observado"]/100 if g["asistencia_detalle"]["porcentaje_observado"] is not None else None,"" if g["asistencia_detalle"]["publicable"] else "Muestra insuficiente",g["incidencias"],g["alumnos_atencion"]] for g in data["grupos"]], pct, {"Grupo":34})
    hoja_datos("Materias", ["Grupo","Materia","Docente","Sesiones registradas","Sesiones programadas","Cobertura","Asistencias","Registros","Asistencia observada","Nota de muestra","Cumplimiento declarado","Último tema","Pendiente","Extemporáneas","Sesiones corregidas"], [[grupos[m["grupo_id"]],m["materia"],m["docente"],m["sesiones"],m["sesiones_programadas"],m["cobertura"]/100 if m["cobertura"] is not None else None,m["asistencia_detalle"]["asistio"],m["asistencia_detalle"]["registros"],m["asistencia_detalle"]["porcentaje_observado"]/100 if m["asistencia_detalle"]["porcentaje_observado"] is not None else None,"" if m["asistencia_detalle"]["publicable"] else "Muestra insuficiente",m["avance_sesion"]/100 if m["avance_sesion"] is not None else None,m["ultimo_tema"],m["pendiente"],m["extemporaneas"],m["corregidas"]] for m in data["materias"]], pct, {"Grupo":34,"Materia":28,"Docente":30,"Último tema":40,"Pendiente":40})
    hoja_datos("Alumnos con indicador", ["Grupo","Matrícula","Alumno","Asistencias","Registros","% asistencia observado","Nota de muestra","Faltas","Seguimientos abiertos","Clasificación"], [[grupos.get(a["grupo_id"],""),a["matricula"],a["nombre"],a["asistencias_registradas"],a["registros"],a["asistencias_registradas"]/a["registros"] if a["registros"] else None,"" if a["registros"] >= MIN_SESIONES_PORCENTAJE else "Muestra insuficiente",a["faltas"],a["seguimientos_abiertos"],_nivel_presentacion(a["nivel"])] for a in data["alumnos_atencion"]], pct, {"Grupo":34,"Alumno":38,"Clasificación":22})
    hoja_datos("Sesiones", ["ID sesión","Grupo","Fecha","Materia","Docente","Hora inicio","Hora fin","Estado","Presentes","Faltas","Retardos","Justificadas","Total registros","Tema","Cumplimiento declarado","Extemporánea","Motivo","Movimientos de corrección","Última corrección"], [[s["sesion_id"],grupos.get(s["grupo_id"],""),_fecha_iso(s["fecha"]),s["materia"],s["docente"],_hora_excel(s["hora_inicio"]),_hora_excel(s["hora_fin"]),s["estado"],s["presentes"],s["faltas"],s["retardos"],s["justificadas"],s["total_registros"],s["tema"],s["cumplimiento_declarado"]/100 if s["cumplimiento_declarado"] is not None else None,"Sí" if s["extemporanea"] else "No",s["motivo_extemporaneo"],s["movimientos_correccion"],as_mx(datetime.datetime.fromisoformat(s["ultima_correccion"])).replace(tzinfo=None) if s["ultima_correccion"] else None] for s in data["sesiones_detalle"]], {**pct, **fecha}, {"Grupo":34,"Materia":28,"Docente":30,"Tema":45,"Motivo":45})
    hoja_datos("Asistencia alumno-sesión", ["ID sesión","Grupo","Fecha","Materia","Docente","Matrícula","Alumno","Estado original","Estado actual","Movimientos de corrección","Observación"], [[a["sesion_id"],grupos.get(a["grupo_id"],""),_fecha_iso(a["fecha"]),a["materia"],a["docente"],a["matricula"],a["alumno"],a["estado_original"],a["estado_actual"],a["movimientos_correccion"],a["observacion"]] for a in data["asistencia_sesion"]], fecha, {"Grupo":34,"Materia":28,"Docente":30,"Alumno":38,"Observación":45})
    hoja_datos("Incidencias", ["Grupo","Fecha","Materia","Docente","Tipo","Descripción","Seguimiento"], [[grupos.get(i["grupo_id"],""),_fecha_iso(i["fecha"]),i["materia"],i["docente"],i["tipo"],i["descripcion"],"Sí" if i["requiere_seguimiento"] else "No"] for i in data["incidencias"]], fecha, {"Grupo":34,"Materia":28,"Docente":30,"Descripción":60}, "Sin incidencias registradas en el periodo")
    hoja_datos("Observaciones académicas", ["Grupo","Fecha","Materia","Docente","Tema","Actividades","Pendiente"], [[grupos.get(o["grupo_id"],""),_fecha_iso(o["fecha"]),o["materia"],o["docente"],o["tema"],o["actividades"],o["pendiente"]] for o in data["observaciones_academicas"]], fecha, {"Grupo":34,"Materia":28,"Docente":30,"Tema":45,"Actividades":50,"Pendiente":45}, "Sin observaciones académicas registradas en el periodo")
    hoja_datos("Registros especiales", ["Grupo","Fecha","Materia","Docente","Extemporánea","Motivo","Movimientos de corrección","Última corrección"], [[grupos.get(s["grupo_id"],""),_fecha_iso(s["fecha"]),s["materia"],s["docente"],"Sí" if s["extemporanea"] else "No",s["motivo_extemporaneo"],s["correcciones"],as_mx(datetime.datetime.fromisoformat(s["ultima_correccion"])).replace(tzinfo=None) if s["ultima_correccion"] else None] for s in data["sesiones_especiales"]], fecha, {"Grupo":34,"Materia":28,"Docente":30,"Motivo":50}, "Sin capturas extemporáneas ni correcciones en el periodo")
    salida = io.BytesIO(); wb.save(salida); salida.seek(0); return salida


@router.get("/exportar.xlsx")
def exportar_excel(periodo_id: int, grupos: str, desde: Optional[datetime.date] = None, hasta: Optional[datetime.date] = None,
                   db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    _autorizar(db, current_user); data = _datos_reporte(db, periodo_id, _ids_grupos(grupos), desde, hasta)
    emision = _obtener_emision(db, data, current_user)
    nombre = _slug(f'Reporte_academico_{data["periodo"]["clave"]}_{today_mx().isoformat()}')
    return StreamingResponse(_excel(data, emision, current_user), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{nombre}.xlsx"', "X-Reporte-Folio": emision.folio})


def _canvas_numerado(folio: str, generado: str, responsable: str, cargo: str):
    class CanvasNumerado(Canvas):
        def __init__(self, *args, **kwargs):
            Canvas.__init__(self, *args, **kwargs); self._paginas = []
        def showPage(self):
            self._paginas.append(dict(self.__dict__)); self._startPage()
        def save(self):
            total = len(self._paginas)
            for estado in self._paginas:
                self.__dict__.update(estado); self.setStrokeColor(colors.HexColor("#0F766E")); self.setLineWidth(1.2)
                self.line(1.2*cm, 1.02*cm, landscape(letter)[0]-1.2*cm, 1.02*cm)
                self.setFillColor(colors.HexColor("#334155")); self.setFont("Helvetica", 6.3)
                self.drawString(1.2*cm, .70*cm, f"UTECAN · SIGA · Folio {folio}")
                self.drawCentredString(landscape(letter)[0]/2, .70*cm, generado)
                self.drawRightString(landscape(letter)[0]-1.2*cm, .70*cm, f"Página {self._pageNumber} de {total}")
                self.setFont("Helvetica", 5.8); self.drawString(1.2*cm, .43*cm, "Documento de uso interno. Contiene datos personales. Información académica y de asistencia; los seguimientos personales se consultan en Tutoría con los permisos correspondientes.")
                if self._pageNumber == total:
                    self.setStrokeColor(colors.HexColor("#94A3B8")); self.setLineWidth(.6)
                    self.line(6.0*cm, 2.05*cm, 10.0*cm, 2.05*cm); self.line(18.0*cm, 2.05*cm, 22.0*cm, 2.05*cm)
                    self.setFillColor(colors.HexColor("#475569")); self.setFont("Helvetica", 6.4)
                    self.drawCentredString(8.0*cm, 1.78*cm, "Elaboró"); self.drawCentredString(20.0*cm, 1.78*cm, "Vo.Bo.")
                    self.drawCentredString(8.0*cm, 1.53*cm, responsable[:55]); self.drawCentredString(20.0*cm, 1.53*cm, "Dirección Académica")
                    self.setFont("Helvetica", 5.8); self.drawCentredString(8.0*cm, 1.30*cm, cargo[:70])
                Canvas.showPage(self)
            Canvas.save(self)
    return CanvasNumerado


def _pdf(data: dict, usuario: Usuario, emision: EmisionReporteAcademico) -> tuple[io.BytesIO, str]:
    ahora = as_mx(emision.generado_en)
    folio = emision.folio
    autor = getattr(emision, "generado_por", None) or usuario
    cargo = autor.departamento.nombre if getattr(autor, "departamento", None) else "Dirección de División de Carrera"
    responsable = _nombre_persona(autor.nombre)
    generado = f"Generado {ahora.strftime('%d/%m/%Y %H:%M')} (hora de Campeche) por {responsable}"
    salida = io.BytesIO(); styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Institucion", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=15, leading=18, textColor=colors.HexColor("#0F172A"), alignment=TA_LEFT, spaceAfter=2))
    styles.add(ParagraphStyle(name="Division", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10, leading=12, textColor=colors.HexColor("#0F766E"), spaceAfter=7))
    styles.add(ParagraphStyle(name="TituloReporte", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=colors.HexColor("#0F172A"), spaceAfter=7))
    styles.add(ParagraphStyle(name="Seccion", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=9.5, leading=12, textColor=colors.HexColor("#0F172A"), spaceBefore=8, spaceAfter=4, borderColor=colors.HexColor("#0F766E"), borderWidth=0, borderPadding=0))
    styles.add(ParagraphStyle(name="Celda", parent=styles["BodyText"], fontName="Helvetica", fontSize=6.7, leading=8.2, textColor=colors.HexColor("#1E293B")))
    styles.add(ParagraphStyle(name="CeldaBold", parent=styles["Celda"], fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="Nota", parent=styles["BodyText"], fontSize=7, leading=9, textColor=colors.HexColor("#475569")))
    doc = SimpleDocTemplate(salida, pagesize=landscape(letter), rightMargin=1.2*cm, leftMargin=1.2*cm, topMargin=1.05*cm, bottomMargin=2.45*cm, title="Reporte académico de grupos", author=responsable, subject="Seguimiento académico de Dirección de División de Carrera")
    P=lambda valor, estilo="Celda": Paragraph(escape(str(valor if valor not in (None, "") else "—")), styles[estilo])
    def tabla(filas, anchos=None, alinear_numeros=False):
        contenido=[[P(c,"CeldaBold" if ri==0 else "Celda") for c in fila] for ri,fila in enumerate(filas)]
        t=Table(contenido, repeatRows=1, colWidths=anchos, hAlign="LEFT")
        comandos=[("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#0F172A")),("LINEBELOW",(0,0),(-1,0),1.5,colors.HexColor("#0F766E")),("LINEBELOW",(0,1),(-1,-1),.25,colors.HexColor("#E2E8F0")),("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]
        for fila in range(2,len(filas),2): comandos.append(("BACKGROUND",(0,fila),(-1,fila),colors.HexColor("#F8FAFC")))
        if alinear_numeros: comandos.append(("ALIGN",(2,1),(-1,-1),"CENTER"))
        t.setStyle(TableStyle(comandos)); return t

    grupos_n={g["id"]:f'{g["nombre"]} · {g["carrera_corta"]}' for g in data["grupos"]}
    alcance=", ".join(grupos_n[g["id"]] for g in data["grupos"])
    r=data["resumen"]; asistencia_general = _porcentaje(r["asistencia"]) if r["asistencia_detalle"]["publicable"] else f'{r["asistencia_detalle"]["texto"]} asistencias registradas (muestra insuficiente para porcentaje)'
    membrete_texto = [Paragraph("UNIVERSIDAD TECNOLÓGICA DE CANDELARIA",styles["Institucion"]),Paragraph("Dirección de División de Carrera",styles["Division"]),Paragraph("Reporte académico de grupos",styles["TituloReporte"])]
    if LOGO_UTECAN.exists():
        logo = Image(str(LOGO_UTECAN), width=5.4*cm, height=1.71*cm)
        membrete = Table([[logo, membrete_texto]], colWidths=[5.8*cm, 20.2*cm], style=TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
    else:
        membrete = Table([[membrete_texto]], colWidths=[26*cm])
    story=[membrete,Spacer(1,5),Table([[P("Folio","CeldaBold"),P(folio),P("Periodo","CeldaBold"),P(data["periodo"]["clave"]),P("Rango","CeldaBold"),P(_rango_fecha_es(data["filtros"]["desde"], data["filtros"]["hasta"]))],[P("Generó","CeldaBold"),P(responsable),P("Cargo","CeldaBold"),P(cargo),P("Alcance","CeldaBold"),P(alcance)]],colWidths=[1.6*cm,5.0*cm,1.6*cm,4.0*cm,1.6*cm,11.2*cm],hAlign="LEFT"),Spacer(1,7)]
    cobertura = (f'{r["sesiones"]} de {r["sesiones_programadas"]} sesiones programadas ({_porcentaje(r["cobertura"])})' if r["sesiones_programadas"] else "Programación no disponible; no se calcula cobertura")
    story += [Paragraph("Resumen ejecutivo",styles["Seccion"]), tabla([["Indicador","Resultado"],["Cobertura del registro",cobertura],["Asistencia observada",asistencia_general],["Alcance académico",f'{_cantidad(r["grupos"], "grupo")}, {_cantidad(r["alumnos"], "alumno")} y {_cantidad(r["materias"], "materia")}'],["Situaciones registradas",f'{_cantidad(r["incidencias"], "incidencia")} y {_cantidad(r["alumnos_atencion"], "alumno")} con indicador de seguimiento']], [5.2*cm,20.8*cm]),Paragraph("Referencia institucional: no se encuentra configurada; el reporte no califica el resultado contra una meta oficial.",styles["Nota"])]
    for g in data["grupos"]:
        filas=[["Materia","Docente","Sesiones reg./prog.","Cobertura","Asistencia","Cumplimiento declarado (%)","Último tema","Pendiente"]]
        for m in data["materias"]:
            if m["grupo_id"] != g["id"]: continue
            asistencia = _porcentaje(m["asistencia"]) if m["asistencia_detalle"]["publicable"] else f'{m["asistencia_detalle"]["texto"]} (muestra insuficiente)'
            sesiones = f'{m["sesiones"]}/{m["sesiones_programadas"]}' if m["sesiones_programadas"] else f'{m["sesiones"]}/—'
            filas.append([m["materia"],m["docente"],sesiones,_porcentaje(m["cobertura"]),asistencia,_porcentaje(m["avance_sesion"]),m["ultimo_tema"],m["pendiente"]])
        bloque=[Paragraph(f'{escape(g["nombre"])} · {escape(g["carrera_corta"])}',styles["Seccion"]),tabla(filas,[3.5*cm,3.5*cm,2.0*cm,1.8*cm,3.3*cm,2.4*cm,4.8*cm,4.8*cm])]
        story.append(KeepTogether(bloque) if len(filas)<=3 else bloque[0]);
        if len(filas)>3: story.append(bloque[1])
    if data["alumnos_atencion"]:
        story.append(Paragraph("Alumnos con indicador académico",styles["Seccion"])); filas=[["Grupo","Matrícula","Alumno","Asistencias / registros","Faltas","Seguimientos","Clasificación"]]
        for a in data["alumnos_atencion"]: filas.append([grupos_n.get(a["grupo_id"],""),a["matricula"],a["nombre"],f'{a["asistencias_registradas"]}/{a["registros"]}',a["faltas"],a["seguimientos_abiertos"],_nivel_presentacion(a["nivel"])])
        criterios = ["1. Datos insuficientes: menos de 3 registros y sin seguimiento abierto.", "2. Atención: seguimiento abierto o asistencia menor a 85% con al menos 3 sesiones.", "3. Prioritario: dos seguimientos abiertos o asistencia menor a 80% con al menos 5 sesiones."]
        notas_criterios = Table([[Paragraph(texto,styles["Nota"])] for texto in criterios],colWidths=[22*cm],style=TableStyle([("TOPPADDING",(0,0),(-1,-1),1),("BOTTOMPADDING",(0,0),(-1,-1),1),("LEFTPADDING",(0,0),(-1,-1),0)]))
        story += [tabla(filas,[3.2*cm,2.3*cm,5.2*cm,3.0*cm,1.4*cm,2.2*cm,3.4*cm]),KeepTogether([notas_criterios])]
    if data["sesiones_especiales"]:
        story.append(Paragraph("Sesiones con registro extemporáneo o corregido",styles["Seccion"])); filas=[["Grupo","Fecha","Materia","Docente","Registro","Motivo / movimientos"]]
        for s in data["sesiones_especiales"]:
            registro = " · ".join(x for x in ["Extemporáneo" if s["extemporanea"] else None, f'{_cantidad(s["correcciones"], "corrección", "correcciones")}' if s["correcciones"] else None] if x)
            detalle = s["motivo_extemporaneo"] or "Sin motivo extemporáneo"
            if s["correcciones"]: detalle += f' · {_cantidad(s["correcciones"], "movimiento")} de auditoría'
            filas.append([grupos_n.get(s["grupo_id"],""),format_fecha_corta_mx(_fecha_iso(s["fecha"])),s["materia"],s["docente"],registro,detalle])
        story.append(tabla(filas,[3.2*cm,2.1*cm,4.0*cm,4.0*cm,2.3*cm,9.0*cm]))
    if data["observaciones_academicas"]:
        story.append(Paragraph("Registro académico de las sesiones",styles["Seccion"])); filas=[["Grupo / fecha","Materia / docente","Tema y actividades","Pendiente"]]
        for o in data["observaciones_academicas"]: filas.append([f'{grupos_n.get(o["grupo_id"],"")} · {format_fecha_corta_mx(_fecha_iso(o["fecha"]))}',f'{o["materia"]} · {o["docente"]}'," · ".join(x for x in [o["tema"],o["actividades"]] if x),o["pendiente"]])
        story.append(tabla(filas,[4.6*cm,5.2*cm,9.0*cm,7.2*cm]))
    if data["incidencias"]:
        story.append(Paragraph("Incidencias generales",styles["Seccion"])); filas=[["Grupo / fecha","Materia / docente","Tipo","Descripción","Seguimiento"]]
        for i in data["incidencias"]: filas.append([f'{grupos_n.get(i["grupo_id"],"")} · {format_fecha_corta_mx(_fecha_iso(i["fecha"]))}',f'{i["materia"]} · {i["docente"]}',_nivel_presentacion(i["tipo"]),i["descripcion"],"Canalizada" if i["requiere_seguimiento"] else "Solo registro"])
        story.append(tabla(filas,[4.6*cm,5.2*cm,2.8*cm,10.2*cm,2.5*cm]))
    story.append(Paragraph("Nota: el cumplimiento es el porcentaje declarado por el docente al cerrar la sesión; no constituye una evaluación institucional del avance.", styles["Nota"]))
    doc.build(story,canvasmaker=_canvas_numerado(folio, generado, responsable, cargo)); salida.seek(0); return salida, folio


@router.get("/exportar.pdf")
def exportar_pdf(periodo_id: int, grupos: str, desde: Optional[datetime.date] = None, hasta: Optional[datetime.date] = None,
                 db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    _autorizar(db, current_user); data = _datos_reporte(db, periodo_id, _ids_grupos(grupos), desde, hasta)
    emision = _obtener_emision(db, data, current_user)
    salida, folio = _pdf(data, current_user, emision)
    nombre = _slug(f'Reporte_academico_{data["periodo"]["clave"]}_{today_mx().isoformat()}')
    return StreamingResponse(salida,media_type="application/pdf",headers={"Content-Disposition":f'attachment; filename="{nombre}.pdf"', "X-Reporte-Folio": folio})
