import datetime
import io
import re
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from pydantic import BaseModel, Field, model_validator
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from database import get_db
from dependencies import get_current_user
from models.usuario import Usuario, RolUsuario
from models.catalogo import (
    CatalogoAlumno, CatalogoMateria, GrupoAcademico, InscripcionAlumno, PeriodoEscolar,
)
from models.laboratorio import Laboratorio
from models.espacio import EspacioInstitucional
from models.horario import BloqueoSlot, HorarioDisponible, Reservacion, SolicitudConflicto
from models.docencia import (
    AsistenciaDocente, CargaDocente, ClaseDocente,
    CorreccionAsistenciaDocente,
    DetalleJustificacionAsistencia, JustificacionAsistenciaDocente,
    SeguimientoAlumnoDocente,
)
from models.tutoria import AsignacionTutoria, Canalizacion, GrupoTutorado, ReporteTutor
from models.cierre_academico import CierreAcademicoPeriodo, ConfirmacionCargaDocente
from routers.notificaciones import crear_notificacion
from services.tutoria_sync import grupo_tutoria_para_academico
from services.calendario_academico import estado_fecha_academica


router = APIRouter(prefix="/docencia", tags=["Módulo docente"])
MX = ZoneInfo("America/Mexico_City")
TIPOS = {"CLASE", "TUTORIA", "DESCARGA", "RECESO", "OTRA"}
ESTADOS_ASISTENCIA = {"PRESENTE", "FALTA", "RETARDO", "JUSTIFICADA"}


def _nombre_archivo(valor: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", valor or "reporte").strip("_") or "reporte"


def _ahora_mx():
    return datetime.datetime.now(MX)


def _solo_docente(user: Usuario):
    if user.rol not in {RolUsuario.DOCENTE, RolUsuario.SUPER_ADMIN}:
        raise HTTPException(403, "Acceso exclusivo del personal docente")


class CargaInput(BaseModel):
    periodo_id: int
    grupo_academico_id: Optional[int] = None
    materia_id: Optional[int] = None
    grupo_tutorado_id: Optional[int] = None
    tipo_actividad: str = "CLASE"
    actividad_nombre: str = Field(..., min_length=2, max_length=200)
    dia_semana: int = Field(..., ge=0, le=5)
    hora_inicio: str = Field(..., pattern=r"^\d{2}:\d{2}$")
    hora_fin: str = Field(..., pattern=r"^\d{2}:\d{2}$")
    espacio_nombre: Optional[str] = Field(None, max_length=180)
    laboratorio_id: Optional[int] = None
    uso_laboratorio: str = "EQUIPOS"
    observaciones: Optional[str] = Field(None, max_length=1000)

    @model_validator(mode="after")
    def validar(self):
        self.tipo_actividad = self.tipo_actividad.upper()
        if self.tipo_actividad not in TIPOS:
            raise ValueError("Tipo de actividad no válido")
        if self.hora_inicio >= self.hora_fin:
            raise ValueError("La hora final debe ser posterior a la inicial")
        if self.tipo_actividad == "CLASE" and not self.grupo_academico_id:
            raise ValueError("Las clases deben tener un grupo")
        if self.tipo_actividad == "TUTORIA" and not self.grupo_tutorado_id:
            raise ValueError("Selecciona uno de tus grupos tutorados")
        self.uso_laboratorio = self.uso_laboratorio.upper()
        if self.uso_laboratorio not in {"SOLO_AULA", "EQUIPOS", "USO_PARCIAL"}:
            raise ValueError("Tipo de uso del laboratorio no válido")
        return self


class CapturaExtemporaneaInput(BaseModel):
    fecha: datetime.date
    motivo: str = Field(..., min_length=5, max_length=500)


class ClaseNoImpartidaInput(BaseModel):
    fecha: datetime.date
    motivo: str = Field(..., min_length=5, max_length=500)
    programar_reposicion: bool = False
    requiere_reposicion: bool = True
    fecha_reposicion: Optional[datetime.date] = None
    hora_inicio: Optional[str] = Field(None, pattern=r"^\d{2}:\d{2}$")
    hora_fin: Optional[str] = Field(None, pattern=r"^\d{2}:\d{2}$")
    tema: Optional[str] = Field(None, max_length=300)

    @model_validator(mode="after")
    def validar(self):
        if self.programar_reposicion and not self.requiere_reposicion:
            raise ValueError("No se puede programar una reposicion cuando la clase no la requiere")
        if self.programar_reposicion:
            if not self.fecha_reposicion or not self.hora_inicio or not self.hora_fin:
                raise ValueError("Indica fecha y horario de la reposicion")
            if self.fecha_reposicion <= self.fecha:
                raise ValueError("La reposicion debe ser posterior a la clase original")
            if self.hora_inicio >= self.hora_fin:
                raise ValueError("La hora final debe ser posterior a la inicial")
        return self


class ReposicionInput(BaseModel):
    fecha: datetime.date
    fecha_original: datetime.date
    hora_inicio: str = Field(..., pattern=r"^\d{2}:\d{2}$")
    hora_fin: str = Field(..., pattern=r"^\d{2}:\d{2}$")
    motivo: str = Field(..., min_length=5, max_length=500)
    tema: Optional[str] = Field(None, max_length=300)

    @model_validator(mode="after")
    def validar(self):
        if self.hora_inicio >= self.hora_fin:
            raise ValueError("La hora final debe ser posterior a la inicial")
        if self.fecha <= self.fecha_original:
            raise ValueError("La reposición debe programarse después de la clase original")
        return self


class AsistenciaInput(BaseModel):
    estado: str
    observacion: Optional[str] = Field(None, max_length=500)


class CierreInput(BaseModel):
    observacion_general: Optional[str] = Field(None, max_length=1000)
    tema_impartido: Optional[str] = Field(None, max_length=300)
    avance_planeacion: Optional[int] = Field(None, ge=0, le=100)
    actividades_realizadas: Optional[str] = Field(None, max_length=2000)
    tarea_asignada: Optional[str] = Field(None, max_length=1500)
    incidencias: Optional[str] = Field(None, max_length=1500)
    incidencia_tipo: Optional[str] = Field(None, max_length=30)
    incidencia_requiere_seguimiento: Optional[bool] = None
    incidencia_solicita_justificacion: Optional[bool] = None
    tema_pendiente: Optional[str] = Field(None, max_length=1000)

    @model_validator(mode="after")
    def validar_incidencia(self):
        if self.incidencias:
            self.incidencia_tipo = (self.incidencia_tipo or "OTRA").upper()
            if self.incidencia_tipo not in {
                "ACADEMICA", "DISCIPLINA", "INFRAESTRUCTURA",
                "SUSPENSION_INSTITUCIONAL", "CONTINGENCIA", "SEGURIDAD", "OTRA",
            }:
                raise ValueError("Tipo de incidencia no válido")
        elif self.incidencias is not None:
            self.incidencia_tipo = None
            self.incidencia_requiere_seguimiento = False
            self.incidencia_solicita_justificacion = False
        return self


class CorreccionInput(BaseModel):
    motivo: str = Field(..., min_length=5, max_length=500)


class ReclasificarNoImpartidaInput(BaseModel):
    motivo: str = Field(..., min_length=5, max_length=500)
    requiere_reposicion: bool = False


class JustificacionMultipleInput(BaseModel):
    fecha_inicio: datetime.date
    fecha_fin: datetime.date
    asistencia_ids: list[int] = Field(..., min_length=1, max_length=100)
    motivo: str = Field(..., min_length=5, max_length=1000)
    folio: Optional[str] = Field(None, max_length=100)

    @model_validator(mode="after")
    def validar(self):
        if self.fecha_fin < self.fecha_inicio:
            raise ValueError("La fecha final debe ser igual o posterior a la inicial")
        if len(set(self.asistencia_ids)) != len(self.asistencia_ids):
            raise ValueError("Hay faltas repetidas en la selección")
        self.motivo = self.motivo.strip()
        self.folio = self.folio.strip() if self.folio else None
        return self


class SeguimientoInput(BaseModel):
    tipo: str
    titulo: str = Field(..., min_length=2, max_length=180)
    detalle: Optional[str] = Field(None, max_length=2000)
    calificacion: Optional[float] = Field(None, ge=0, le=10)
    estado: str = Field("REGISTRADO", max_length=20)
    fecha_limite: Optional[datetime.date] = None
    fecha_revision: Optional[datetime.date] = None
    categoria_reporte: str = Field("ACADEMICO", max_length=30)
    prioridad_reporte: str = Field("MEDIA", max_length=15)
    confidencial: bool = False

    @model_validator(mode="after")
    def validar(self):
        self.tipo = self.tipo.upper()
        self.estado = self.estado.upper()
        if self.tipo not in {"OBSERVACION", "ACUERDO", "CALIFICACION", "TUTORIA"}:
            raise ValueError("Tipo de seguimiento no válido")
        if self.tipo == "CALIFICACION" and self.calificacion is None:
            raise ValueError("La calificación es obligatoria")
        if self.tipo == "ACUERDO":
            if not self.fecha_limite or not self.fecha_revision:
                raise ValueError("El acuerdo requiere fecha limite y fecha de revision")
            if self.fecha_revision < self.fecha_limite:
                raise ValueError("La fecha de revision debe ser igual o posterior a la fecha limite")
        self.categoria_reporte = self.categoria_reporte.upper()
        self.prioridad_reporte = self.prioridad_reporte.upper()
        if self.categoria_reporte not in {"ACADEMICO", "ASISTENCIA", "CONDUCTA", "PERSONAL", "OTRO"}:
            raise ValueError("Categoría de reporte no válida")
        if self.prioridad_reporte not in {"BAJA", "MEDIA", "ALTA"}:
            raise ValueError("Prioridad de reporte no válida")
        return self


class EstadoSeguimientoInput(BaseModel):
    estado: str
    resultado_atencion: Optional[str] = Field(None, max_length=2000)
    fecha_limite: Optional[datetime.date] = None
    fecha_revision: Optional[datetime.date] = None

    @model_validator(mode="after")
    def validar(self):
        self.estado = self.estado.upper()
        estados_resultado = {"ATENDIDO", "CUMPLIDO", "CUMPLIDO_PARCIAL", "NO_CUMPLIDO", "CERRADO", "REPROGRAMADO"}
        if self.estado not in {"PENDIENTE", *estados_resultado}:
            raise ValueError("Estado de seguimiento no válido")
        if self.estado in estados_resultado and not (self.resultado_atencion or "").strip():
            raise ValueError("El resultado de la atención es obligatorio")
        if self.estado == "REPROGRAMADO":
            if not self.fecha_limite or not self.fecha_revision:
                raise ValueError("La reprogramacion requiere nuevas fechas")
            if self.fecha_revision < self.fecha_limite:
                raise ValueError("La fecha de revision debe ser igual o posterior a la fecha limite")
        return self


class AlertaTempranaInput(BaseModel):
    senal: str
    nivel: str = "ATENCION"
    comentario: Optional[str] = Field(None, max_length=1000)

    @model_validator(mode="after")
    def validar(self):
        self.senal = self.senal.upper()
        self.nivel = self.nivel.upper()
        senales = {
            "INASISTENCIA", "BAJO_DESEMPENO", "CAMBIO_CONDUCTA",
            "FALTA_PARTICIPACION", "SITUACION_PERSONAL", "OTRO",
        }
        if self.senal not in senales:
            raise ValueError("Señal de alerta no válida")
        if self.nivel not in {"OBSERVACION", "ATENCION", "URGENTE"}:
            raise ValueError("Nivel de alerta no válido")
        self.comentario = self.comentario.strip() if self.comentario else None
        if (self.senal == "OTRO" or self.nivel == "URGENTE") and not self.comentario:
            raise ValueError("Describe brevemente la situación")
        return self


class IncidenciaClaseInput(BaseModel):
    tipo: str
    descripcion: str = Field(..., min_length=5, max_length=2000)
    requiere_seguimiento: bool = False
    solicita_justificacion: bool = False

    @model_validator(mode="after")
    def validar(self):
        self.tipo = self.tipo.upper()
        if self.tipo not in {"ACADEMICA", "DISCIPLINA", "INFRAESTRUCTURA", "SUSPENSION_INSTITUCIONAL", "CONTINGENCIA", "SEGURIDAD", "OTRA"}:
            raise ValueError("Tipo de incidencia no válido")
        if self.solicita_justificacion and self.tipo != "SUSPENSION_INSTITUCIONAL":
            raise ValueError("La justificación colectiva solo aplica a actividades o suspensiones institucionales")
        self.descripcion = self.descripcion.strip()
        return self


def _docente_objetivo(user: Usuario) -> int:
    return user.id


def _periodo_actual(db: Session):
    periodos = db.query(PeriodoEscolar).filter(
        PeriodoEscolar.activo == True,
    ).order_by(PeriodoEscolar.id.desc()).all()
    hoy = _ahora_mx()
    bloque = "ENE-ABR" if hoy.month <= 4 else "MAY-AGO" if hoy.month <= 8 else "SEP-DIC"
    esperado = _normalizar_periodo(f"{bloque} {hoy.year}")
    por_fecha = next(
        (periodo for periodo in periodos if _normalizar_periodo(periodo.clave) == esperado),
        None,
    )
    return por_fecha or next((periodo for periodo in periodos if periodo.es_actual), None)


def _validar_periodo_actual(db: Session, periodo_id: int, *, permitir_revision_carga=False):
    cierre = db.query(CierreAcademicoPeriodo).filter_by(periodo_id=periodo_id).first()
    if cierre and cierre.estado == "CERRADO" and not permitir_revision_carga:
        raise HTTPException(409, "El cuatrimestre está cerrado y disponible solo para consulta; no se pueden crear ni activar materias.")
    periodo = _periodo_actual(db)
    if not periodo:
        raise HTTPException(
            409,
            "No hay un periodo escolar actual configurado. Solicita a Servicios Escolares que active el cuatrimestre.",
        )
    if periodo.id != periodo_id:
        raise HTTPException(
            409,
            f"{periodo.clave} es el periodo actual. El periodo seleccionado permanece disponible solo para consulta.",
        )
    return periodo


def _validar_carga_actual(db: Session, carga: CargaDocente):
    periodo = _validar_periodo_actual(db, carga.periodo_id, permitir_revision_carga=True)
    cierre = db.query(CierreAcademicoPeriodo).filter(CierreAcademicoPeriodo.periodo_id == carga.periodo_id).first()
    if cierre:
        confirmacion = db.query(ConfirmacionCargaDocente).filter(
            ConfirmacionCargaDocente.cierre_id == cierre.id,
            ConfirmacionCargaDocente.carga_docente_id == carga.id,
        ).first()
        reapertura_vigente = bool(
            confirmacion
            and confirmacion.estado == "REABIERTA"
            and confirmacion.reabierta_hasta
            and confirmacion.reabierta_hasta >= datetime.datetime.utcnow()
        )
        if cierre.estado == "CERRADO" and not reapertura_vigente:
            raise HTTPException(409, "El cuatrimestre está cerrado; solicita una reapertura a División de Carrera")
        if confirmacion and confirmacion.estado == "CONFIRMADA_DOCENTE":
            raise HTTPException(409, "Esta materia ya fue confirmada para el cierre")
        if confirmacion and confirmacion.estado == "REABIERTA" and confirmacion.reabierta_hasta and confirmacion.reabierta_hasta < datetime.datetime.utcnow():
            raise HTTPException(409, "La reapertura de esta materia ya venció")
    return periodo


def _normalizar_periodo(clave: str | None) -> str:
    return "-".join((clave or "").strip().upper().replace("_", "-").split()).replace("--", "-")


def _slots_para_carga(db: Session, data: CargaInput):
    if not data.laboratorio_id:
        return [], "SIN_LABORATORIO"
    periodo = db.query(PeriodoEscolar).filter(PeriodoEscolar.id == data.periodo_id).first()
    candidatos = db.query(HorarioDisponible).filter(
        HorarioDisponible.laboratorio_id == data.laboratorio_id,
        HorarioDisponible.dia_semana == data.dia_semana,
        HorarioDisponible.activo == True,
    ).order_by(HorarioDisponible.hora_inicio).all()
    candidatos = [
        h for h in candidatos
        if _normalizar_periodo(h.cuatrimestre) == _normalizar_periodo(periodo.clave if periodo else "")
        and h.hora_inicio < data.hora_fin and h.hora_fin > data.hora_inicio
    ]
    if not candidatos:
        return [], "SIN_HORARIOS"
    cursor = data.hora_inicio
    seleccionados = []
    for slot in candidatos:
        if slot.hora_fin <= cursor:
            continue
        if slot.hora_inicio > cursor:
            if not (cursor == "09:45" and slot.hora_inicio == "10:15"):
                return candidatos, "COBERTURA_INCOMPLETA"
        seleccionados.append(slot)
        cursor = max(cursor, slot.hora_fin)
        if cursor >= data.hora_fin:
            break
    if cursor < data.hora_fin:
        return seleccionados, "COBERTURA_INCOMPLETA"
    return seleccionados, None


def _estado_laboratorio(db: Session, data: CargaInput, current_user: Usuario, carga_id: int | None = None):
    slots, problema = _slots_para_carga(db, data)
    lab = db.query(Laboratorio).filter(Laboratorio.id == data.laboratorio_id).first() if data.laboratorio_id else None
    respuesta = {
        "estado": problema or "DISPONIBLE",
        "laboratorio_id": data.laboratorio_id,
        "laboratorio_nombre": lab.nombre if lab else None,
        "slots": [],
        "ocupaciones": [],
    }
    if problema:
        return respuesta
    todos_vinculados = True
    for slot in slots:
        bloqueo = db.query(BloqueoSlot).filter(
            BloqueoSlot.horario_id == slot.id, BloqueoSlot.activo == True,
        ).first()
        reserva = db.query(Reservacion).filter(
            Reservacion.horario_id == slot.id,
            Reservacion.estado.in_(["PROGRAMADA", "EN_DISPUTA", "EN_CURSO"]),
        ).first()
        respuesta["slots"].append({"id": slot.id, "hora_inicio": slot.hora_inicio, "hora_fin": slot.hora_fin})
        if bloqueo:
            respuesta["estado"] = "BLOQUEADO"
            respuesta["ocupaciones"].append({"hora": f"{slot.hora_inicio}–{slot.hora_fin}", "motivo": bloqueo.motivo})
            todos_vinculados = False
        elif reserva and not (carga_id and reserva.carga_docente_id == carga_id):
            docente = db.query(Usuario).filter(Usuario.id == reserva.docente_id).first()
            mi_solicitud = db.query(SolicitudConflicto).filter(
                SolicitudConflicto.reservacion_id == reserva.id,
                SolicitudConflicto.solicitante_id == current_user.id,
                SolicitudConflicto.estado == "PENDIENTE",
            ).first()
            respuesta["estado"] = "SOLICITADO" if mi_solicitud else "OCUPADO"
            respuesta["ocupaciones"].append({
                "hora": f"{slot.hora_inicio}–{slot.hora_fin}",
                "reservacion_id": reserva.id,
                "docente": docente.nombre if docente else "Docente",
                "materia": reserva.materia,
                "grupo": reserva.grupo,
            })
            todos_vinculados = False
        elif not reserva or not (carga_id and reserva.carga_docente_id == carga_id):
            todos_vinculados = False
    if todos_vinculados and slots:
        respuesta["estado"] = "RESERVADO"
    return respuesta


def _cancelar_reservas_carga(db: Session, carga_id: int):
    reservas = db.query(Reservacion).filter(
        Reservacion.carga_docente_id == carga_id,
        Reservacion.estado.in_(["PROGRAMADA", "EN_DISPUTA"]),
    ).all()
    for reserva in reservas:
        reserva.estado = "CANCELADA"


def _traslapa(inicio_a, fin_a, inicio_b, fin_b):
    return inicio_a < fin_b and fin_a > inicio_b


def _advertencias(db: Session, carga: CargaDocente, excluir_id=None):
    avisos = []
    q = db.query(CargaDocente).filter(
        CargaDocente.activo == True,
        CargaDocente.periodo_id == carga.periodo_id,
        CargaDocente.dia_semana == carga.dia_semana,
    )
    if excluir_id:
        q = q.filter(CargaDocente.id != excluir_id)
    for otra in q.all():
        if not _traslapa(carga.hora_inicio, carga.hora_fin, otra.hora_inicio, otra.hora_fin):
            continue
        if otra.docente_id == carga.docente_id:
            raise HTTPException(
                409,
                f"Ya tienes otra actividad de {otra.hora_inicio} a {otra.hora_fin}. "
                "Revisa el horario oficial antes de continuar.",
            )
        if carga.grupo_academico_id and otra.grupo_academico_id == carga.grupo_academico_id:
            avisos.append("El grupo aparece en otra actividad a la misma hora.")
        if carga.espacio_nombre and otra.espacio_nombre and (
            carga.espacio_nombre.strip().upper() == otra.espacio_nombre.strip().upper()
        ):
            avisos.append("El salón o espacio aparece ocupado a la misma hora.")
    return list(dict.fromkeys(avisos))


def _normalizar_identidad(valor):
    return " ".join(str(valor or "").strip().upper().split())


def _identidad_materia_catalogo(materia):
    return (
        _normalizar_identidad(materia.nombre),
        _normalizar_identidad(materia.carrera),
        materia.cuatrimestre_oficial,
    )


def _catalogo_materias_unico(materias):
    canonicas, alias = {}, {}
    for materia in sorted(materias, key=lambda item: item.id):
        clave = _identidad_materia_catalogo(materia)
        canonica = canonicas.setdefault(clave, materia)
        alias[materia.id] = canonica.id
    return list(canonicas.values()), alias


def _validar_identidad_academica(db: Session, data: CargaInput):
    if data.tipo_actividad != "CLASE":
        return
    grupo = db.query(GrupoAcademico).filter(
        GrupoAcademico.id == data.grupo_academico_id,
        GrupoAcademico.periodo_id == data.periodo_id,
        GrupoAcademico.activo == True,
    ).first()
    materia = db.query(CatalogoMateria).filter(
        CatalogoMateria.id == data.materia_id,
        CatalogoMateria.activo == True,
    ).first()
    if not grupo:
        raise HTTPException(422, "El grupo no pertenece al periodo seleccionado o ya no está activo")
    if not materia:
        raise HTTPException(422, "Selecciona una materia activa del catálogo académico")
    if materia.carrera and _normalizar_identidad(materia.carrera) != _normalizar_identidad(grupo.carrera):
        raise HTTPException(422, "La materia y el grupo pertenecen a carreras diferentes")
    if materia.cuatrimestre_oficial is not None and materia.cuatrimestre_oficial != grupo.cuatrimestre:
        raise HTTPException(422, "La materia y el grupo pertenecen a cuatrimestres diferentes")
    # La materia pertenece al plan de estudios permanente. El periodo se define
    # en CargaDocente junto con el grupo, docente y horario de cada oferta.


def _validar_asignacion_materia(db: Session, data: CargaInput, docente_id: int, excluir_id=None):
    """Una materia-grupo pertenece a un solo docente, aunque tenga varios bloques."""
    if data.tipo_actividad != "CLASE":
        return
    materia = db.get(CatalogoMateria, data.materia_id)
    ids_equivalentes = [m.id for m in db.query(CatalogoMateria).filter(CatalogoMateria.activo == True).all()
                        if materia and _identidad_materia_catalogo(m) == _identidad_materia_catalogo(materia)]
    consulta = db.query(CargaDocente).filter(
        CargaDocente.activo == True,
        CargaDocente.tipo_actividad == "CLASE",
        CargaDocente.periodo_id == data.periodo_id,
        CargaDocente.grupo_academico_id == data.grupo_academico_id,
        CargaDocente.materia_id.in_(ids_equivalentes or [data.materia_id]),
        CargaDocente.docente_id != docente_id,
    )
    if excluir_id:
        consulta = consulta.filter(CargaDocente.id != excluir_id)
    asignada = consulta.first()
    if asignada:
        docente = asignada.docente.nombre if asignada.docente else "otro docente"
        raise HTTPException(409, f"Esta materia ya está asignada a {docente} para el grupo seleccionado")


def _commit_asignacion(db: Session):
    try:
        db.commit()
    except IntegrityError as error:
        db.rollback()
        if "uq_docente_materia_grupo" in str(error.orig):
            raise HTTPException(409, "La materia y el grupo fueron asignados a otro docente mientras guardabas") from error
        raise


def _validar_asignacion_tutoria(db: Session, data: CargaInput, docente_id: int, periodo: PeriodoEscolar):
    if data.tipo_actividad != "TUTORIA":
        return None
    grupo = db.query(GrupoTutorado).filter(
        GrupoTutorado.id == data.grupo_tutorado_id,
        GrupoTutorado.tutor_id == docente_id,
        GrupoTutorado.activo == True,
        GrupoTutorado.estado.in_(["ACTIVO", "PREPARACION"]),
        GrupoTutorado.periodo == periodo.clave,
    ).first()
    if not grupo:
        raise HTTPException(409, "El grupo tutorado no está asignado a tu cuenta en este periodo")
    return grupo


def _serializar_carga(c: CargaDocente, db: Session):
    grupo = c.grupo_academico
    lab = c.laboratorio
    reservas_lab = db.query(Reservacion).filter(
        Reservacion.carga_docente_id == c.id,
        Reservacion.estado.in_(["PROGRAMADA", "EN_DISPUTA", "EN_CURSO"]),
    ).all()
    estado_reserva_lab = (
        "EN_DISPUTA" if any(r.estado == "EN_DISPUTA" for r in reservas_lab)
        else "RESERVADO" if reservas_lab else "SIN_RESERVACION"
    )
    ahora_local = _ahora_mx()
    minuto_actual = ahora_local.hour * 60 + ahora_local.minute
    reservas_programadas = [r for r in reservas_lab if r.estado == "PROGRAMADA"]
    reserva_en_ventana = next((
        r for r in reservas_programadas
        if r.horario
        and r.horario.dia_semana == ahora_local.weekday()
        and (
            (int(r.horario.hora_inicio[:2]) * 60 + int(r.horario.hora_inicio[3:5]) - 15)
            <= minuto_actual
            <= (int(r.horario.hora_fin[:2]) * 60 + int(r.horario.hora_fin[3:5]) + 15)
        )
    ), None)
    reserva_operable = (
        next((r for r in reservas_lab if r.estado == "EN_CURSO"), None)
        or reserva_en_ventana
        or next(iter(reservas_programadas), None)
    )
    return {
        "id": c.id,
        "periodo_id": c.periodo_id,
        "periodo": c.periodo.clave if c.periodo else None,
        "grupo_academico_id": c.grupo_academico_id,
        "grupo": (
            f"{grupo.cuatrimestre}° {grupo.grupo}" if grupo
            else f"{c.grupo_tutorado.cuatrimestre}° {c.grupo_tutorado.grupo}" if c.grupo_tutorado
            else None
        ),
        "carrera": grupo.carrera if grupo else c.grupo_tutorado.carrera if c.grupo_tutorado else None,
        "materia_id": c.materia_id,
        "grupo_tutorado_id": c.grupo_tutorado_id,
        "grupo_tutorado": (
            f"{c.grupo_tutorado.cuatrimestre}° {c.grupo_tutorado.grupo} · {c.grupo_tutorado.carrera}"
            if c.grupo_tutorado else None
        ),
        "actividad_nombre": c.actividad_nombre,
        "tipo_actividad": c.tipo_actividad,
        "dia_semana": c.dia_semana,
        "hora_inicio": c.hora_inicio,
        "hora_fin": c.hora_fin,
        "espacio_nombre": lab.nombre if lab else c.espacio_nombre,
        "laboratorio_id": c.laboratorio_id,
        "uso_laboratorio": c.uso_laboratorio or "EQUIPOS",
        "estado_reserva_laboratorio": estado_reserva_lab if c.laboratorio_id else None,
        "reservacion_laboratorio_id": reserva_operable.id if reserva_operable else None,
        "estado": c.estado,
        "observaciones": c.observaciones,
        "puede_iniciar_hoy": (
            c.tipo_actividad == "CLASE"
            and c.estado == "ACTIVO"
            and c.dia_semana == _ahora_mx().weekday()
        ),
    }


def _serializar_clase(clase: ClaseDocente):
    carga = clase.carga
    asistencias = clase.asistencias
    conteos = {estado: 0 for estado in ESTADOS_ASISTENCIA}
    for asistencia in asistencias:
        conteos[asistencia.estado] = conteos.get(asistencia.estado, 0) + 1
    observaciones = (clase.observacion_general or "").splitlines()
    correcciones_legacy = [linea for linea in observaciones if linea.startswith("[Correcci")]
    observacion_limpia = "\n".join(
        linea for linea in observaciones if not linea.startswith("[Correcci")
    ).strip() or None
    if (
        clase.estado == "NO_IMPARTIDA"
        and observacion_limpia
        and clase.motivo_no_impartida
        and observacion_limpia.strip().lower() == f"clase no impartida: {clase.motivo_no_impartida}".lower()
    ):
        observacion_limpia = None
    correcciones = [{
        "id": correccion.id,
        "tipo": correccion.tipo,
        "alumno_id": correccion.alumno_id,
        "alumno": (
            f"{correccion.alumno.apellido_paterno} {correccion.alumno.apellido_materno} "
            f"{correccion.alumno.nombres}"
        ).strip() if correccion.alumno else None,
        "estado_anterior": correccion.estado_anterior,
        "estado_nuevo": correccion.estado_nuevo,
        "motivo": correccion.motivo,
        "docente": correccion.docente.nombre if correccion.docente else None,
        "creado_en": correccion.creado_en.isoformat(),
    } for correccion in sorted(
        clase.correcciones_asistencia, key=lambda item: item.creado_en, reverse=True,
    )]
    correcciones.extend({
        "id": f"legacy-{indice}", "tipo": "APERTURA", "alumno_id": None,
        "alumno": None, "estado_anterior": None, "estado_nuevo": None,
        "motivo": linea, "docente": clase.carga.docente.nombre if clase.carga.docente else None,
        "creado_en": None,
    } for indice, linea in enumerate(correcciones_legacy))
    tema_normalizado = (clase.tema_impartido or "").strip().lower()
    requiere_revision_clasificacion = clase.estado == "CERRADA" and any(
        frase in tema_normalizado for frase in ("no se impart", "no impartida", "no hubo clase")
    )
    return {
        "id": clase.id,
        "fecha": clase.fecha.isoformat(),
        "estado": clase.estado,
        "inicio": clase.inicio.isoformat() if clase.inicio else None,
        "fin": clase.fin.isoformat() if clase.fin else None,
        "observacion_general": observacion_limpia,
        "correcciones_asistencia": correcciones,
        "es_extemporanea": clase.es_extemporanea,
        "motivo_no_impartida": clase.motivo_no_impartida,
        "declarada_no_impartida_en": (
            clase.declarada_no_impartida_en.isoformat() if clase.declarada_no_impartida_en else None
        ),
        "es_reposicion": clase.es_reposicion,
        "clase_origen_id": clase.clase_origen_id,
        "fecha_original": clase.fecha_original.isoformat() if clase.fecha_original else None,
        "hora_inicio_reposicion": clase.hora_inicio_reposicion,
        "hora_fin_reposicion": clase.hora_fin_reposicion,
        "motivo_reposicion": clase.motivo_reposicion,
        "estado_reposicion": clase.estado_reposicion,
        "motivo_extemporaneo": clase.motivo_extemporaneo,
        "capturada_extemporanea_en": (
            clase.capturada_extemporanea_en.isoformat()
            if clase.capturada_extemporanea_en else None
        ),
        "requiere_revision_clasificacion": requiere_revision_clasificacion,
        "bitacora": {
            "tema_impartido": clase.tema_impartido,
            "avance_planeacion": clase.avance_planeacion,
            "actividades_realizadas": clase.actividades_realizadas,
            "tarea_asignada": clase.tarea_asignada,
            "incidencias": clase.incidencias,
            "incidencia_tipo": clase.incidencia_tipo,
            "incidencia_requiere_seguimiento": clase.incidencia_requiere_seguimiento,
            "incidencia_solicita_justificacion": clase.incidencia_solicita_justificacion,
            "tema_pendiente": clase.tema_pendiente,
        },
        "carga": {
            "id": carga.id,
            "periodo_id": carga.periodo_id,
            "periodo": carga.periodo.clave if carga.periodo else None,
            "actividad_nombre": carga.actividad_nombre,
            "grupo": (
                f"{carga.grupo_academico.cuatrimestre}° {carga.grupo_academico.grupo}"
                if carga.grupo_academico else None
            ),
            "carrera": carga.grupo_academico.carrera if carga.grupo_academico else None,
            "espacio_nombre": carga.laboratorio.nombre if carga.laboratorio else carga.espacio_nombre,
            "hora_inicio": carga.hora_inicio,
            "hora_fin": carga.hora_fin,
        },
        "resumen": {"total": len(asistencias), **{k.lower(): v for k, v in conteos.items()}},
        "alumnos": [
            {
                "asistencia_id": a.id,
                "alumno_id": a.alumno_id,
                "matricula": a.alumno.matricula,
                "nombre": (
                    f"{a.alumno.apellido_paterno} {a.alumno.apellido_materno} "
                    f"{a.alumno.nombres}"
                ).strip(),
                "estado": a.estado,
                "observacion": a.observacion,
            }
            for a in sorted(asistencias, key=lambda x: (x.alumno.apellido_paterno, x.alumno.nombres))
        ],
    }


def _fecha_programada_carga(carga: CargaDocente, fecha: datetime.date, *, usar_fin: bool = False):
    hora = datetime.time.fromisoformat(carga.hora_fin if usar_fin else carga.hora_inicio)
    return datetime.datetime.combine(fecha, hora, tzinfo=MX)


PLAZO_CAPTURA_EXTEMPORANEA = datetime.timedelta(days=7)


def _validar_ventana_extemporanea(db: Session, carga: CargaDocente, fecha: datetime.date):
    ahora = _ahora_mx()
    inicio_programado = _fecha_programada_carga(carga, fecha)
    fin_programado = _fecha_programada_carga(carga, fecha, usar_fin=True)
    if carga.dia_semana != fecha.weekday():
        raise HTTPException(422, "La fecha no corresponde al día programado de esta clase")
    estado_fecha = estado_fecha_academica(db, carga.periodo_id, fecha)
    if not estado_fecha["requiere_asistencia"]:
        raise HTTPException(409, f"No se requiere asistencia: {estado_fecha['motivo']}")
    if fin_programado > ahora:
        raise HTTPException(409, "La clase todavía está en curso")
    if ahora - fin_programado > PLAZO_CAPTURA_EXTEMPORANEA:
        raise HTTPException(409, "El plazo de 7 días para capturar esta asistencia ya venció")
    return inicio_programado


@router.get("/catalogos")
def catalogos_docente(
    periodo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _solo_docente(current_user)
    periodos = db.query(PeriodoEscolar).filter(PeriodoEscolar.activo == True).order_by(
        PeriodoEscolar.id.desc()
    ).all()
    actual = _periodo_actual(db)
    elegido = periodo_id or (actual.id if actual else None)
    if not elegido and periodos:
        hoy = _ahora_mx()
        bloque = "ENE-ABR" if hoy.month <= 4 else "MAY-AGO" if hoy.month <= 8 else "SEP-DIC"
        esperado = f"{bloque}{hoy.year}"
        elegido = next(
            (p.id for p in periodos if "".join(ch for ch in p.clave.upper() if ch.isalnum()) == esperado),
            periodos[0].id,
        )
    grupos_q = db.query(GrupoAcademico).filter(GrupoAcademico.activo == True)
    materias_q = db.query(CatalogoMateria).filter(CatalogoMateria.activo == True)
    periodo = db.get(PeriodoEscolar, elegido) if elegido else None
    if elegido:
        grupos_q = grupos_q.filter(GrupoAcademico.periodo_id == elegido)
    grupos = grupos_q.order_by(
        GrupoAcademico.carrera, GrupoAcademico.cuatrimestre, GrupoAcademico.grupo
    ).all()
    materias_catalogo, alias_materias = _catalogo_materias_unico(materias_q.all())
    asignaciones = {}
    if elegido:
        cargas_asignadas = db.query(CargaDocente).filter(
            CargaDocente.activo == True,
            CargaDocente.tipo_actividad == "CLASE",
            CargaDocente.periodo_id == elegido,
            CargaDocente.materia_id.isnot(None),
            CargaDocente.grupo_academico_id.isnot(None),
        ).all()
        for carga in cargas_asignadas:
            materia_canonica_id = alias_materias.get(carga.materia_id, carga.materia_id)
            clave = (materia_canonica_id, carga.grupo_academico_id)
            asignaciones.setdefault(clave, {
                "materia_id": materia_canonica_id,
                "grupo_academico_id": carga.grupo_academico_id,
                "docente_id": carga.docente_id,
                "docente": carga.docente.nombre if carga.docente else "Docente",
                "es_propia": carga.docente_id == current_user.id,
            })
    grupos_tutorados = db.query(GrupoTutorado).filter(
        GrupoTutorado.tutor_id == current_user.id,
        GrupoTutorado.activo == True,
        GrupoTutorado.estado.in_(["ACTIVO", "PREPARACION"]),
        GrupoTutorado.periodo == (periodo.clave if elegido and periodo else ""),
    ).order_by(GrupoTutorado.carrera, GrupoTutorado.cuatrimestre, GrupoTutorado.grupo).all()
    return {
        "periodo_sugerido_id": elegido,
        "periodos": [{
            "id": p.id,
            "clave": p.clave,
            "es_actual": bool(actual and p.id == actual.id and not db.query(CierreAcademicoPeriodo.id).filter_by(periodo_id=p.id, estado="CERRADO").first()),
            "es_actual_configurado": p.es_actual,
        } for p in periodos],
        "grupos": [{
            "id": g.id, "carrera": g.carrera, "cuatrimestre": g.cuatrimestre,
            "grupo": g.grupo, "label": f"{g.cuatrimestre}° {g.grupo} · {g.carrera}",
            "total_alumnos": sum(1 for i in g.inscripciones if i.estado == "ACTIVO"),
        } for g in grupos],
        "materias": [{
            "id": m.id, "nombre": m.nombre, "carrera": m.carrera,
            "cuatrimestre_oficial": m.cuatrimestre_oficial,
        } for m in materias_catalogo],
        "asignaciones_materias": list(asignaciones.values()),
        "grupos_tutorados": [{
            "id": g.id, "carrera": g.carrera, "cuatrimestre": g.cuatrimestre,
            "grupo": g.grupo, "label": f"{g.cuatrimestre}° {g.grupo} · {g.carrera}",
        } for g in grupos_tutorados],
        "laboratorios": [{"id": l.id, "nombre": l.nombre} for l in db.query(Laboratorio).filter(Laboratorio.activo == True).all()],
        "espacios": [{"id": e.id, "nombre": e.nombre} for e in db.query(EspacioInstitucional).filter(EspacioInstitucional.activo == True).all()],
    }


@router.get("/horario")
def mi_horario(
    periodo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _solo_docente(current_user)
    q = db.query(CargaDocente).filter(
        CargaDocente.docente_id == _docente_objetivo(current_user),
        CargaDocente.activo == True,
    )
    elegido = periodo_id
    if not elegido:
        actual = _periodo_actual(db)
        elegido = actual.id if actual else None
    if elegido:
        q = q.filter(CargaDocente.periodo_id == elegido)
    else:
        return []
    return [_serializar_carga(c, db) for c in q.order_by(
        CargaDocente.dia_semana, CargaDocente.hora_inicio
    ).all()]


def _filtro_funcion_docente():
    # El rol principal persiste aunque la persona cambie al modo Docente.
    return or_(
        Usuario.rol == RolUsuario.DOCENTE,
        Usuario.roles_adicionales.contains('"DOCENTE"'),
    )


def _permiso_consulta_horarios(usuario):
    if usuario.rol != RolUsuario.DOCENTE:
        raise HTTPException(403, "Esta consulta está disponible únicamente entre docentes")


def _periodo_consulta_horarios(db, periodo_id):
    if periodo_id is None:
        return _periodo_actual(db)
    periodo = db.query(PeriodoEscolar).filter(PeriodoEscolar.id == periodo_id).first()
    if not periodo:
        raise HTTPException(404, "Periodo no encontrado")
    return periodo


def _detalle_horario_publico(carga):
    privado = carga.tipo_actividad in {"RECESO", "DESCARGA"}
    grupo = carga.grupo_academico or carga.grupo_tutorado
    return {
        "id": carga.id, "dia_semana": carga.dia_semana,
        "tipo_actividad": carga.tipo_actividad,
        "actividad": carga.actividad_nombre,
        "docente": carga.docente.nombre,
        "grupo": None if privado or not grupo else f"{grupo.cuatrimestre}° {grupo.grupo}",
        "carrera": None if privado or not grupo else grupo.carrera,
        "salon": None if privado else (carga.laboratorio.nombre if carga.laboratorio else carga.espacio_nombre),
        "hora_inicio": carga.hora_inicio, "hora_fin": carga.hora_fin,
    }


def _resumen_horario_publico(cargas, periodo, ahora, calendario, es_actual):
    semana = [_detalle_horario_publico(c) for c in cargas]
    hoy = [c for c in semana if c["dia_semana"] == ahora.weekday()] if es_actual else []
    lectivo = calendario[0]["requiere_asistencia"] and calendario[0]["permite_iniciar_clase"]
    hora = ahora.strftime("%H:%M")
    actuales = [c for c in hoy if lectivo and c["hora_inicio"] <= hora < c["hora_fin"]]
    siguiente = None
    if es_actual:
        for dias, estado in enumerate(calendario):
            fecha = ahora.date() + datetime.timedelta(days=dias)
            if not estado["requiere_asistencia"] or not estado["permite_iniciar_clase"]:
                continue
            # No atribuir el horario de este cuatrimestre al siguiente.
            if (fecha.year, (fecha.month - 1) // 4) != (ahora.year, (ahora.month - 1) // 4):
                break
            opciones = [c for c in semana if c["dia_semana"] == fecha.weekday() and (dias > 0 or c["hora_inicio"] > hora)]
            if opciones:
                siguiente = {**min(opciones, key=lambda c: c["hora_inicio"]), "fecha": fecha.isoformat()}
                break
    return {
        "actividad_actual": actuales[0] if actuales else None,
        "actividades_actuales": actuales,
        "siguiente_actividad": siguiente,
        "estado": actuales[0]["tipo_actividad"] if actuales else "SIN_ACTIVIDAD",
        "jornada": hoy, "semana": semana,
    }


def _contexto_consulta_horarios(db, periodo):
    ahora = _ahora_mx()
    actual = _periodo_actual(db)
    es_actual = bool(periodo and actual and periodo.id == actual.id)
    calendario = [estado_fecha_academica(db, periodo.id if periodo else None, ahora.date() + datetime.timedelta(days=d)) for d in range(8)]
    return ahora, calendario, {
        "fecha": ahora.date().isoformat(), "hora_consulta": ahora.strftime("%H:%M"),
        "periodo": periodo.clave if periodo else None, "es_actual": es_actual,
        "calendario_hoy": calendario[0] if es_actual else None,
    }


@router.get("/ubicacion-docentes")
def buscar_ubicacion_docentes(
    q: str = "", periodo_id: Optional[int] = None,
    db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user),
):
    """Consulta docentes con función principal o adicional, sin ubicaciones privadas."""
    _permiso_consulta_horarios(current_user)
    termino = " ".join(q.strip().split())
    if len(termino) < 2:
        raise HTTPException(422, "Escribe al menos 2 caracteres del nombre del docente")
    docentes = db.query(Usuario).filter(
        _filtro_funcion_docente(), Usuario.activo == True,
        Usuario.nombre.ilike(f"%{termino}%"),
    ).order_by(Usuario.nombre).limit(20).all()
    periodo = _periodo_consulta_horarios(db, periodo_id)
    ahora, calendario, contexto = _contexto_consulta_horarios(db, periodo)
    cargas = db.query(CargaDocente).filter(
        CargaDocente.docente_id.in_([d.id for d in docentes]),
        CargaDocente.periodo_id == periodo.id,
        CargaDocente.activo == True, CargaDocente.estado == "ACTIVO",
    ).order_by(CargaDocente.dia_semana, CargaDocente.hora_inicio).all() if periodo and docentes else []
    return {**contexto, "resultados": [{
        "docente_id": d.id, "nombre": d.nombre,
        "numero_empleado": d.numero_empleado,
        "departamento": d.departamento.nombre if d.departamento else None,
        **_resumen_horario_publico([c for c in cargas if c.docente_id == d.id], periodo, ahora, calendario, contexto["es_actual"]),
    } for d in docentes]}


@router.get("/consulta-horarios/grupos")
def grupos_consulta_horarios(
    periodo_id: Optional[int] = None,
    db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user),
):
    _permiso_consulta_horarios(current_user)
    periodo = _periodo_consulta_horarios(db, periodo_id)
    if not periodo:
        return []
    grupos = db.query(GrupoAcademico).filter(
        GrupoAcademico.periodo_id == periodo.id, GrupoAcademico.activo == True,
    ).order_by(GrupoAcademico.carrera, GrupoAcademico.cuatrimestre, GrupoAcademico.grupo).all()
    return [{"id": g.id, "carrera": g.carrera, "grupo": f"{g.cuatrimestre}° {g.grupo}", "turno": g.turno} for g in grupos]


@router.get("/consulta-horarios/grupos/{grupo_id}")
def horario_publico_grupo(
    grupo_id: int, periodo_id: Optional[int] = None,
    db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user),
):
    _permiso_consulta_horarios(current_user)
    periodo = _periodo_consulta_horarios(db, periodo_id)
    grupo = db.query(GrupoAcademico).filter(
        GrupoAcademico.id == grupo_id, GrupoAcademico.activo == True,
        GrupoAcademico.periodo_id == (periodo.id if periodo else -1),
    ).first()
    if not grupo:
        raise HTTPException(404, "Grupo no encontrado en el periodo seleccionado")
    cargas = db.query(CargaDocente).join(Usuario, Usuario.id == CargaDocente.docente_id).filter(
        CargaDocente.grupo_academico_id == grupo.id, CargaDocente.periodo_id == periodo.id,
        CargaDocente.tipo_actividad == "CLASE", CargaDocente.activo == True,
        CargaDocente.estado == "ACTIVO", Usuario.activo == True, _filtro_funcion_docente(),
    ).order_by(CargaDocente.dia_semana, CargaDocente.hora_inicio).all()
    ahora, calendario, contexto = _contexto_consulta_horarios(db, periodo)
    return {**contexto, "resultados": [{
        "grupo_id": grupo.id, "nombre": f"{grupo.cuatrimestre}° {grupo.grupo}", "carrera": grupo.carrera,
        **_resumen_horario_publico(cargas, periodo, ahora, calendario, contexto["es_actual"]),
    }]}


@router.post("/horario")
def crear_carga(
    data: CargaInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _solo_docente(current_user)
    periodo = _validar_periodo_actual(db, data.periodo_id)
    _validar_identidad_academica(db, data)
    _validar_asignacion_materia(db, data, current_user.id)
    grupo_tutoria = _validar_asignacion_tutoria(db, data, current_user.id, periodo)
    if grupo_tutoria:
        data.actividad_nombre = f"Tutoría grupal · {grupo_tutoria.cuatrimestre}° {grupo_tutoria.grupo}"
    carga = CargaDocente(docente_id=current_user.id, estado="BORRADOR", **data.model_dump())
    db.add(carga)
    db.flush()
    avisos = _advertencias(db, carga, carga.id)
    _commit_asignacion(db)
    db.refresh(carga)
    return {"carga": _serializar_carga(carga, db), "advertencias": avisos}


@router.put("/horario/{carga_id}")
def actualizar_carga(
    carga_id: int,
    data: CargaInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id, CargaDocente.docente_id == current_user.id
    ).first()
    if not carga:
        raise HTTPException(404, "Actividad no encontrada")
    _validar_carga_actual(db, carga)
    periodo = _validar_periodo_actual(db, data.periodo_id)
    _validar_identidad_academica(db, data)
    _validar_asignacion_materia(db, data, current_user.id, carga.id)
    grupo_tutoria = _validar_asignacion_tutoria(db, data, current_user.id, periodo)
    if grupo_tutoria:
        data.actividad_nombre = f"Tutoría grupal · {grupo_tutoria.cuatrimestre}° {grupo_tutoria.grupo}"
    tiene_historial = db.query(ClaseDocente.id).filter(
        ClaseDocente.carga_docente_id == carga.id,
    ).first() is not None
    cambia_laboratorio = any(
        getattr(carga, campo) != getattr(data, campo)
        for campo in ("laboratorio_id", "dia_semana", "hora_inicio", "hora_fin", "periodo_id")
    )
    if cambia_laboratorio:
        _cancelar_reservas_carga(db, carga.id)
    if tiene_historial:
        # Una carga ya utilizada es parte del acta histórica de sus clases. Crear
        # una nueva versión evita que un cambio de día u hora convierta una clase
        # cerrada del horario anterior en la sesión pendiente del horario nuevo.
        carga.activo = False
        carga.estado = "RETIRADO"
        nueva_carga = CargaDocente(
            docente_id=current_user.id,
            estado="BORRADOR",
            **data.model_dump(),
        )
        db.add(nueva_carga)
        db.flush()
        avisos = _advertencias(db, nueva_carga, nueva_carga.id)
        avisos.insert(0, "Se conservó el horario anterior en el historial y se creó una nueva versión.")
        _commit_asignacion(db)
        db.refresh(nueva_carga)
        return {"carga": _serializar_carga(nueva_carga, db), "advertencias": avisos}
    for campo, valor in data.model_dump().items():
        setattr(carga, campo, valor)
    carga.estado = "BORRADOR"
    avisos = _advertencias(db, carga, carga.id)
    _commit_asignacion(db)
    db.refresh(carga)
    return {"carga": _serializar_carga(carga, db), "advertencias": avisos}


@router.post("/horario/{carga_id}/activar")
def activar_carga(
    carga_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id, CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
    ).first()
    if not carga:
        raise HTTPException(404, "Actividad no encontrada")
    _validar_periodo_actual(db, carga.periodo_id)
    _validar_carga_actual(db, carga)
    avisos = _advertencias(db, carga, carga.id)
    carga.estado = "ACTIVO"
    db.commit()
    return {"carga": _serializar_carga(carga, db), "advertencias": avisos}


@router.delete("/horario/{carga_id}")
def eliminar_carga(
    carga_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id, CargaDocente.docente_id == current_user.id
    ).first()
    if not carga:
        raise HTTPException(404, "Actividad no encontrada")
    _validar_carga_actual(db, carga)
    _cancelar_reservas_carga(db, carga.id)
    carga.activo = False
    db.commit()
    return {"mensaje": "Actividad retirada del horario"}


@router.post("/horario/verificar-laboratorio")
def verificar_laboratorio_carga(
    data: CargaInput,
    carga_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _solo_docente(current_user)
    if carga_id:
        propia = db.query(CargaDocente.id).filter(
            CargaDocente.id == carga_id, CargaDocente.docente_id == current_user.id,
        ).first()
        if not propia:
            raise HTTPException(404, "Actividad no encontrada")
    return _estado_laboratorio(db, data, current_user, carga_id)


@router.get("/horario/{carga_id}/estado-laboratorio")
def estado_laboratorio_carga(
    carga_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id, CargaDocente.docente_id == current_user.id,
    ).first()
    if not carga:
        raise HTTPException(404, "Actividad no encontrada")
    data = CargaInput.model_validate({
        campo: getattr(carga, campo) for campo in CargaInput.model_fields
    })
    return _estado_laboratorio(db, data, current_user, carga.id)


@router.post("/horario/{carga_id}/reservar-laboratorio")
def reservar_laboratorio_carga(
    carga_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id,
        CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
        CargaDocente.tipo_actividad.in_(["CLASE", "TUTORIA"]),
    ).first()
    grupo = carga.grupo_academico if carga else None
    grupo_tutorado = carga.grupo_tutorado if carga else None
    if not carga or not carga.laboratorio_id or not (grupo or grupo_tutorado):
        raise HTTPException(422, "La actividad debe tener laboratorio y grupo")
    _validar_carga_actual(db, carga)
    data = CargaInput.model_validate({
        campo: getattr(carga, campo) for campo in CargaInput.model_fields
    })
    disponibilidad = _estado_laboratorio(db, data, current_user, carga.id)
    if disponibilidad["estado"] == "RESERVADO":
        return disponibilidad
    if disponibilidad["estado"] != "DISPONIBLE":
        raise HTTPException(409, {
            "mensaje": "El laboratorio no está disponible en todo el horario.",
            "disponibilidad": disponibilidad,
        })
    for slot in disponibilidad["slots"]:
        db.add(Reservacion(
            horario_id=slot["id"],
            laboratorio_id=carga.laboratorio_id,
            docente_id=current_user.id,
            carga_docente_id=carga.id,
            materia=carga.actividad_nombre,
            carrera=(grupo.carrera if grupo else grupo_tutorado.carrera),
            cuatrimestre=carga.periodo.clave,
            periodo_id=carga.periodo_id,
            cuatrimestre_materia=str(grupo.cuatrimestre if grupo else grupo_tutorado.cuatrimestre),
            grupo=f"{grupo.cuatrimestre}° {grupo.grupo}" if grupo else f"{grupo_tutorado.cuatrimestre}° {grupo_tutorado.grupo}",
            estado="PROGRAMADA",
            creado_por=current_user.id,
            observaciones=f"Reservación de {'tutoría' if carga.tipo_actividad == 'TUTORIA' else 'clase'} vinculada desde Mi horario docente",
            tipo_actividad=carga.tipo_actividad,
        ))
    db.commit()
    disponibilidad["estado"] = "RESERVADO"
    return disponibilidad


@router.get("/capturas-extemporaneas/disponibles")
def capturas_extemporaneas_disponibles(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _solo_docente(current_user)
    actual = _periodo_actual(db)
    if not actual:
        return []
    ahora = _ahora_mx()
    cargas = db.query(CargaDocente).filter(
        CargaDocente.docente_id == current_user.id,
        CargaDocente.periodo_id == actual.id,
        CargaDocente.activo == True,
        CargaDocente.estado == "ACTIVO",
        CargaDocente.tipo_actividad == "CLASE",
        CargaDocente.grupo_academico_id.isnot(None),
    ).all()
    opciones = []
    for dias_atras in range(0, PLAZO_CAPTURA_EXTEMPORANEA.days + 1):
        fecha = ahora.date() - datetime.timedelta(days=dias_atras)
        for carga in cargas:
            if carga.dia_semana != fecha.weekday():
                continue
            estado_fecha = estado_fecha_academica(db, carga.periodo_id, fecha)
            if not estado_fecha["requiere_asistencia"] or not estado_fecha["genera_alertas"]:
                continue
            fin_programado = _fecha_programada_carga(carga, fecha, usar_fin=True)
            if fin_programado > ahora or ahora - fin_programado > PLAZO_CAPTURA_EXTEMPORANEA:
                continue
            existe = db.query(ClaseDocente.id).filter(
                ClaseDocente.carga_docente_id == carga.id,
                ClaseDocente.fecha == fecha,
            ).first()
            if existe:
                continue
            opciones.append({
                "carga_id": carga.id,
                "fecha": fecha.isoformat(),
                "materia": carga.actividad_nombre,
                "grupo": (
                    f"{carga.grupo_academico.cuatrimestre}° {carga.grupo_academico.grupo}"
                    if carga.grupo_academico else None
                ),
                "carrera": carga.grupo_academico.carrera if carga.grupo_academico else None,
                "hora_inicio": carga.hora_inicio,
                "hora_fin": carga.hora_fin,
                "vence_en": (fin_programado + PLAZO_CAPTURA_EXTEMPORANEA).isoformat(),
            })
    opciones.sort(key=lambda item: (item["fecha"], item["hora_inicio"]), reverse=True)
    return opciones


@router.post("/horario/{carga_id}/captura-extemporanea")
def crear_captura_extemporanea(
    carga_id: int,
    data: CapturaExtemporaneaInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id,
        CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
        CargaDocente.estado == "ACTIVO",
        CargaDocente.tipo_actividad == "CLASE",
    ).first()
    if not carga or not carga.grupo_academico_id:
        raise HTTPException(404, "Clase programada no encontrada")
    _validar_carga_actual(db, carga)
    _validar_ventana_extemporanea(db, carga, data.fecha)
    existente = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id == carga.id,
        ClaseDocente.fecha == data.fecha,
    ).first()
    if existente:
        raise HTTPException(409, "Esta clase ya tiene un registro de asistencia")
    clase = ClaseDocente(
        carga_docente_id=carga.id,
        fecha=data.fecha,
        estado="ABIERTA",
        es_extemporanea=True,
        motivo_extemporaneo=data.motivo.strip(),
        capturada_extemporanea_en=datetime.datetime.utcnow(),
    )
    db.add(clase)
    db.flush()
    inscripciones = db.query(InscripcionAlumno).filter(
        InscripcionAlumno.grupo_academico_id == carga.grupo_academico_id,
        InscripcionAlumno.estado == "ACTIVO",
    ).all()
    for inscripcion in inscripciones:
        db.add(AsistenciaDocente(
            clase_docente_id=clase.id,
            alumno_id=inscripcion.alumno_id,
            estado="PRESENTE",
        ))
    db.commit()
    db.refresh(clase)
    return _serializar_clase(clase)


@router.post("/horario/{carga_id}/no-impartida")
def declarar_clase_no_impartida(
    carga_id: int,
    data: ClaseNoImpartidaInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id,
        CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
        CargaDocente.estado == "ACTIVO",
        CargaDocente.tipo_actividad == "CLASE",
    ).first()
    if not carga or not carga.grupo_academico_id:
        raise HTTPException(404, "Clase programada no encontrada")
    _validar_carga_actual(db, carga)
    _validar_ventana_extemporanea(db, carga, data.fecha)
    if db.query(ClaseDocente.id).filter(
        ClaseDocente.carga_docente_id == carga.id,
        ClaseDocente.fecha == data.fecha,
    ).first():
        raise HTTPException(409, "Esta clase ya tiene un registro")

    original = ClaseDocente(
        carga_docente_id=carga.id,
        fecha=data.fecha,
        estado="NO_IMPARTIDA",
        fin=datetime.datetime.utcnow(),
        motivo_no_impartida=data.motivo.strip(),
        declarada_no_impartida_en=datetime.datetime.utcnow(),
        observacion_general=None,
        tema_pendiente=(data.tema or "").strip() or None,
        estado_reposicion="PENDIENTE" if data.requiere_reposicion else "NO_REQUERIDA",
    )
    db.add(original)
    db.flush()

    reposicion = None
    if data.programar_reposicion:
        if data.fecha_reposicion < _ahora_mx().date():
            raise HTTPException(422, "La reposicion no puede programarse en una fecha pasada")
        if db.query(ClaseDocente.id).filter(
            ClaseDocente.carga_docente_id == carga.id,
            ClaseDocente.fecha == data.fecha_reposicion,
        ).first():
            raise HTTPException(409, "La materia ya tiene una clase o reposicion en esa fecha")
        reposicion = ClaseDocente(
            carga_docente_id=carga.id,
            fecha=data.fecha_reposicion,
            estado="PROGRAMADA",
            es_reposicion=True,
            clase_origen_id=original.id,
            fecha_original=data.fecha,
            hora_inicio_reposicion=data.hora_inicio,
            hora_fin_reposicion=data.hora_fin,
            motivo_reposicion=data.motivo.strip(),
            tema_pendiente=(data.tema or "").strip() or None,
            estado_reposicion="PROGRAMADA",
        )
        db.add(reposicion)
        db.flush()
    db.commit()
    return {
        "clase_original": _serializar_clase(original),
        "reposicion": _serializar_clase(reposicion) if reposicion else None,
        "mensaje": (
            "Clase no impartida y reposicion programada"
            if reposicion else "Clase no impartida; la reposicion queda pendiente"
            if data.requiere_reposicion else "Clase registrada como no exigible; no requiere reposicion"
        ),
    }


@router.get("/hoy")
def clases_de_hoy(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _solo_docente(current_user)
    hoy = _ahora_mx()
    actual = _periodo_actual(db)
    if not actual:
        return []
    cargas = db.query(CargaDocente).filter(
        CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
        CargaDocente.estado == "ACTIVO",
        CargaDocente.dia_semana == hoy.weekday(),
        CargaDocente.periodo_id == actual.id,
    ).order_by(CargaDocente.hora_inicio).all()
    resultado = []
    for carga in cargas:
        clase = db.query(ClaseDocente).filter(
            ClaseDocente.carga_docente_id == carga.id,
            ClaseDocente.fecha == hoy.date(),
        ).first()
        item = _serializar_carga(carga, db)
        item["clase_id"] = clase.id if clase else None
        item["clase_estado"] = clase.estado if clase else None
        item["calendario"] = estado_fecha_academica(db, carga.periodo_id, hoy.date())
        resultado.append(item)
    reposiciones = db.query(ClaseDocente).join(CargaDocente).filter(
        CargaDocente.docente_id == current_user.id,
        ClaseDocente.es_reposicion == True,
        ClaseDocente.fecha == hoy.date(),
        ClaseDocente.estado_reposicion != "CANCELADA",
    ).all()
    for clase in reposiciones:
        if any(item.get("clase_id") == clase.id for item in resultado):
            continue
        item = _serializar_carga(clase.carga, db)
        item.update({
            "clase_id": clase.id, "clase_estado": clase.estado,
            "es_reposicion": True, "fecha_original": clase.fecha_original.isoformat(),
            "motivo_reposicion": clase.motivo_reposicion,
            "hora_inicio": clase.hora_inicio_reposicion,
            "hora_fin": clase.hora_fin_reposicion,
            "calendario": estado_fecha_academica(db, clase.carga.periodo_id, hoy.date()),
        })
        resultado.append(item)
    resultado.sort(key=lambda item: item["hora_inicio"])
    return resultado


@router.post("/horario/{carga_id}/reposiciones")
def programar_reposicion(carga_id: int, data: ReposicionInput, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id, CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True, CargaDocente.tipo_actividad == "CLASE",
    ).first()
    if not carga:
        raise HTTPException(404, "Materia no encontrada")
    _validar_carga_actual(db, carga)
    if data.fecha_original.weekday() != carga.dia_semana:
        raise HTTPException(422, "La fecha original no corresponde al horario oficial de la materia")
    if data.fecha < _ahora_mx().date():
        raise HTTPException(422, "La reposición no puede programarse en una fecha pasada")
    if db.query(ClaseDocente).filter(ClaseDocente.carga_docente_id == carga.id, ClaseDocente.fecha == data.fecha).first():
        raise HTTPException(409, "La materia ya tiene una clase o reposición en esa fecha")
    original = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id == carga.id,
        ClaseDocente.fecha == data.fecha_original,
        ClaseDocente.estado == "NO_IMPARTIDA",
    ).first()
    if not original:
        raise HTTPException(409, "Primero registra la clase original como no impartida")
    reposicion_activa = db.query(ClaseDocente.id).filter(
        ClaseDocente.clase_origen_id == original.id,
        ClaseDocente.es_reposicion == True,
        ClaseDocente.estado_reposicion != "CANCELADA",
    ).first()
    if reposicion_activa:
        raise HTTPException(409, "Esta clase no impartida ya tiene una reposición programada")
    clase = ClaseDocente(
        carga_docente_id=carga.id, fecha=data.fecha, estado="PROGRAMADA",
        es_reposicion=True, clase_origen_id=original.id,
        fecha_original=data.fecha_original,
        hora_inicio_reposicion=data.hora_inicio, hora_fin_reposicion=data.hora_fin,
        motivo_reposicion=data.motivo.strip(), tema_pendiente=(data.tema or "").strip() or None,
        estado_reposicion="PROGRAMADA",
    )
    db.add(clase); db.commit(); db.refresh(clase)
    return _serializar_clase(clase)


@router.get("/reposiciones/pendientes")
def reposiciones_pendientes(db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Sesiones no impartidas del periodo actual que todavía pueden reponerse."""
    _solo_docente(current_user)
    actual = _periodo_actual(db)
    if not actual:
        return []
    originales = db.query(ClaseDocente).join(CargaDocente).filter(
        CargaDocente.docente_id == current_user.id,
        CargaDocente.periodo_id == actual.id,
        CargaDocente.activo == True,
        ClaseDocente.estado == "NO_IMPARTIDA",
        ClaseDocente.es_reposicion == False,
        or_(ClaseDocente.estado_reposicion.is_(None), ClaseDocente.estado_reposicion != "NO_REQUERIDA"),
    ).order_by(ClaseDocente.fecha.desc()).all()
    ids_con_reposicion = {
        row[0] for row in db.query(ClaseDocente.clase_origen_id).filter(
            ClaseDocente.clase_origen_id.in_([clase.id for clase in originales]),
            ClaseDocente.es_reposicion == True,
            ClaseDocente.estado_reposicion != "CANCELADA",
        ).all()
    } if originales else set()
    return [{
        "clase_id": clase.id,
        "carga_id": clase.carga_docente_id,
        "fecha_original": clase.fecha.isoformat(),
        "materia": clase.carga.actividad_nombre,
        "grupo": (
            f"{clase.carga.grupo_academico.cuatrimestre}° {clase.carga.grupo_academico.grupo}"
            if clase.carga.grupo_academico else None
        ),
        "hora_inicio": clase.carga.hora_inicio,
        "hora_fin": clase.carga.hora_fin,
        "motivo": clase.motivo_no_impartida,
        "tema": clase.tema_pendiente,
    } for clase in originales if clase.id not in ids_con_reposicion]


@router.post("/reposiciones/pendientes/{clase_id}/no-requerida")
def marcar_reposicion_no_requerida(
    clase_id: int,
    data: CorreccionInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    original = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        ClaseDocente.estado == "NO_IMPARTIDA",
        ClaseDocente.es_reposicion == False,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not original:
        raise HTTPException(404, "Clase no impartida no encontrada")
    activa = db.query(ClaseDocente.id).filter(
        ClaseDocente.clase_origen_id == original.id,
        ClaseDocente.es_reposicion == True,
        ClaseDocente.estado_reposicion != "CANCELADA",
    ).first()
    if activa:
        raise HTTPException(409, "Cancela primero la reposicion programada")
    original.estado_reposicion = "NO_REQUERIDA"
    nota = data.motivo.strip()
    original.observacion_general = f"{original.observacion_general or 'Clase no impartida'}\nReposicion no requerida: {nota}"
    db.commit()
    return {"mensaje": "La clase queda cerrada sin reposicion; el motivo permanece en el historial"}


@router.post("/reposiciones/{clase_id}/cancelar")
def cancelar_reposicion(clase_id: int, data: CorreccionInput, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id, ClaseDocente.es_reposicion == True,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Reposición no encontrada")
    if clase.estado not in {"PROGRAMADA"}:
        raise HTTPException(409, "Solo una reposición pendiente puede cancelarse")
    clase.estado = "CANCELADA"; clase.estado_reposicion = "CANCELADA"
    clase.cancelada_en = datetime.datetime.utcnow()
    clase.observacion_general = f"Reposición cancelada: {data.motivo.strip()}"
    if clase.clase_origen:
        clase.clase_origen.estado_reposicion = "NO_REQUERIDA"
        clase.clase_origen.observacion_general = f"{clase.clase_origen.observacion_general or 'Clase no impartida'}\nReposicion no requerida: {data.motivo.strip()}"
    db.commit()
    return {"mensaje": "Reposición cancelada; permanece en el historial"}


@router.post("/reposiciones/{clase_id}/iniciar")
def iniciar_reposicion(clase_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    hoy = _ahora_mx().date()
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id, ClaseDocente.es_reposicion == True,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Reposición no encontrada")
    if clase.fecha != hoy or clase.estado_reposicion == "CANCELADA":
        raise HTTPException(409, "La reposición no está vigente hoy")
    if clase.estado == "PROGRAMADA":
        clase.estado = "ABIERTA"; clase.estado_reposicion = "EN_CURSO"; clase.inicio = datetime.datetime.utcnow()
        inscripciones = db.query(InscripcionAlumno).filter(
            InscripcionAlumno.grupo_academico_id == clase.carga.grupo_academico_id,
            InscripcionAlumno.estado == "ACTIVO",
        ).all()
        for inscripcion in inscripciones:
            db.add(AsistenciaDocente(clase_docente_id=clase.id, alumno_id=inscripcion.alumno_id, estado="PRESENTE"))
        db.commit(); db.refresh(clase)
    return _serializar_clase(clase)


@router.post("/horario/{carga_id}/iniciar")
def iniciar_clase(
    carga_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    hoy = _ahora_mx()
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id,
        CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
    ).first()
    if not carga:
        raise HTTPException(404, "Actividad no encontrada")
    _validar_carga_actual(db, carga)
    if carga.tipo_actividad != "CLASE" or not carga.grupo_academico_id:
        raise HTTPException(400, "Solo una clase con grupo puede generar asistencia")
    if carga.estado != "ACTIVO":
        raise HTTPException(409, "Activa primero este bloque del horario")
    if carga.dia_semana != hoy.weekday():
        raise HTTPException(409, "Esta clase no corresponde al día de hoy")
    estado_fecha = estado_fecha_academica(db, carga.periodo_id, hoy.date())
    if not estado_fecha["permite_iniciar_clase"] or not estado_fecha["requiere_asistencia"]:
        raise HTTPException(409, f"No se puede iniciar la clase: {estado_fecha['motivo']}")
    existente = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id == carga.id, ClaseDocente.fecha == hoy.date()
    ).first()
    if existente:
        return _serializar_clase(existente)
    tolerancia = datetime.timedelta(minutes=15)
    inicio_programado = datetime.datetime.combine(
        hoy.date(), datetime.time.fromisoformat(carga.hora_inicio), tzinfo=MX,
    )
    fin_programado = datetime.datetime.combine(
        hoy.date(), datetime.time.fromisoformat(carga.hora_fin), tzinfo=MX,
    )
    apertura = inicio_programado - tolerancia
    cierre_ventana = fin_programado + tolerancia
    if hoy < apertura:
        raise HTTPException(
            409,
            f"Esta clase inicia a las {carga.hora_inicio}. Podrás iniciarla desde las {apertura.strftime('%H:%M')}.",
        )
    if hoy > cierre_ventana:
        raise HTTPException(
            409,
            f"La ventana para iniciar esta clase terminó a las {cierre_ventana.strftime('%H:%M')}. Revisa las opciones de clase pendiente.",
        )
    clase = ClaseDocente(carga_docente_id=carga.id, fecha=hoy.date(), estado="ABIERTA")
    db.add(clase)
    db.flush()
    inscripciones = db.query(InscripcionAlumno).filter(
        InscripcionAlumno.grupo_academico_id == carga.grupo_academico_id,
        InscripcionAlumno.estado == "ACTIVO",
    ).all()
    for inscripcion in inscripciones:
        db.add(AsistenciaDocente(
            clase_docente_id=clase.id,
            alumno_id=inscripcion.alumno_id,
            estado="PRESENTE",
        ))
    db.commit()
    db.refresh(clase)
    return _serializar_clase(clase)


@router.get("/clases/{clase_id}")
def detalle_clase(
    clase_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Clase no encontrada")
    return _serializar_clase(clase)


@router.patch("/clases/{clase_id}/asistencia/{asistencia_id}")
def cambiar_asistencia(
    clase_id: int,
    asistencia_id: int,
    data: AsistenciaInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    estado = data.estado.upper()
    if estado not in ESTADOS_ASISTENCIA:
        raise HTTPException(422, "Estado de asistencia no válido")
    if estado == "JUSTIFICADA":
        raise HTTPException(
            422,
            "Las faltas se justifican desde Seguimiento de grupos con el documento validado por División de Carrera",
        )
    asistencia = db.query(AsistenciaDocente).join(ClaseDocente).join(CargaDocente).filter(
        AsistenciaDocente.id == asistencia_id,
        AsistenciaDocente.clase_docente_id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not asistencia:
        raise HTTPException(404, "Registro de asistencia no encontrado")
    _validar_carga_actual(db, asistencia.clase.carga)
    if asistencia.clase.estado not in {"ABIERTA", "CORRECCION"}:
        raise HTTPException(409, "La asistencia de esta clase ya está cerrada")
    estado_anterior = asistencia.estado
    observacion_anterior = asistencia.observacion
    es_correccion = asistencia.clase.estado == "CORRECCION"
    asistencia.estado = estado
    asistencia.observacion = data.observacion
    if (
        (es_correccion or estado == "JUSTIFICADA" or estado_anterior == "JUSTIFICADA")
        and (estado_anterior != estado or observacion_anterior != data.observacion)
    ):
        motivo_apertura = None
        if es_correccion:
            apertura = db.query(CorreccionAsistenciaDocente).filter(
                CorreccionAsistenciaDocente.clase_docente_id == clase_id,
                CorreccionAsistenciaDocente.tipo == "APERTURA",
            ).order_by(CorreccionAsistenciaDocente.creado_en.desc()).first()
            motivo_apertura = apertura.motivo if apertura else None
        db.add(CorreccionAsistenciaDocente(
            clase_docente_id=clase_id,
            asistencia_id=asistencia.id,
            alumno_id=asistencia.alumno_id,
            docente_id=current_user.id,
            tipo="CAMBIO",
            estado_anterior=estado_anterior,
            estado_nuevo=estado,
            motivo=data.observacion or motivo_apertura or "Ajuste de asistencia",
        ))
    db.commit()
    return {"id": asistencia.id, "estado": asistencia.estado, "observacion": asistencia.observacion}


@router.post("/clases/{clase_id}/cerrar")
def cerrar_clase(
    clase_id: int,
    data: CierreInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Clase no encontrada")
    _validar_carga_actual(db, clase.carga)
    if clase.estado not in {"ABIERTA", "CORRECCION"}:
        raise HTTPException(409, "La asistencia ya está cerrada")
    clase.estado = "CERRADA"
    clase.fin = datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None)
    if data.observacion_general:
        clase.observacion_general = data.observacion_general
    for campo in (
        "tema_impartido", "avance_planeacion", "actividades_realizadas",
        "tarea_asignada", "incidencias", "incidencia_tipo",
        "incidencia_requiere_seguimiento", "tema_pendiente",
        "incidencia_solicita_justificacion",
    ):
        valor = getattr(data, campo)
        if valor is not None:
            setattr(clase, campo, valor)
    canalizacion = None
    if clase.incidencia_requiere_seguimiento and clase.incidencias:
        grupo_tutorado = grupo_tutoria_para_academico(db, clase.carga.grupo_academico_id)
        tutor = db.query(Usuario).filter(
            Usuario.id == grupo_tutorado.tutor_id,
            Usuario.activo == True,
        ).first() if grupo_tutorado and grupo_tutorado.tutor_id else None
        reporte = db.query(ReporteTutor).filter(
            ReporteTutor.clase_docente_id == clase.id,
        ).first()
        es_nuevo = reporte is None
        tutor_anterior_id = reporte.tutor_destinatario_id if reporte else None
        if es_nuevo:
            reporte = ReporteTutor(
                alumno_id=None,
                reportado_por_id=current_user.id,
                clase_docente_id=clase.id,
                carga_docente_id=clase.carga_docente_id,
            )
            db.add(reporte)
        reporte.tutor_destinatario_id = tutor.id if tutor else None
        reporte.grupo_tutorado_id = grupo_tutorado.id if grupo_tutorado else None
        reporte.categoria = {
            "ACADEMICA": "ACADEMICO", "DISCIPLINA": "CONDUCTA",
        }.get(clase.incidencia_tipo, "OTRO")
        reporte.prioridad = "MEDIA"
        reporte.titulo = f"Nota de clase · {clase.carga.actividad_nombre}"
        grupo_nombre = (
            f"{clase.carga.grupo_academico.cuatrimestre}° {clase.carga.grupo_academico.grupo}"
            if clase.carga.grupo_academico else "grupo sin identificar"
        )
        reporte.detalle = (
            f"Grupo: {grupo_nombre}\nFecha: {clase.fecha.isoformat()}\n"
            f"Tipo: {(clase.incidencia_tipo or 'OTRA').replace('_', ' ').title()}\n\n"
            f"{clase.incidencias.strip()}"
        )
        reporte.confidencial = False
        if es_nuevo or reporte.estado == "SIN_TUTOR":
            reporte.estado = "ENVIADO" if tutor else "SIN_TUTOR"
        db.flush()

        if es_nuevo or (tutor and tutor.id != tutor_anterior_id):
            destinatarios = [tutor] if tutor else db.query(Usuario).filter(
                Usuario.rol.in_([RolUsuario.TUTORIA_ADMIN, RolUsuario.SUPER_ADMIN]),
                Usuario.activo == True,
            ).all()
            for destinatario in destinatarios:
                crear_notificacion(
                    db, destinatario.id, "tutoria_incidencia_clase",
                    "Nota de clase que requiere seguimiento",
                    f"{current_user.nombre} solicitó seguimiento para una nota de {clase.carga.actividad_nombre}, grupo {grupo_nombre}.",
                    "/docente/mis-tutorados?tab=reportes" if tutor else "/admin/tutoria?tab=reportes-tutor",
                    enviar_email=False,
                )
        canalizacion = {
            "reporte_tutor_id": reporte.id,
            "estado": reporte.estado,
            "tutor_id": tutor.id if tutor else None,
            "tutor_nombre": tutor.nombre if tutor else None,
            "mensaje": (
                f"Nota enviada a {tutor.nombre}."
                if tutor else
                "Nota registrada; el grupo no tiene tutor asignado y Tutoría fue notificada."
            ),
        }
    db.commit()
    respuesta = _serializar_clase(clase)
    respuesta["canalizacion_tutoria"] = canalizacion
    return respuesta


@router.post("/clases/{clase_id}/habilitar-correccion")
def habilitar_correccion(
    clase_id: int,
    data: CorreccionInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Clase no encontrada")
    _validar_carga_actual(db, clase.carga)
    if clase.estado != "CERRADA":
        raise HTTPException(409, "Solo se puede corregir una asistencia cerrada")
    db.add(CorreccionAsistenciaDocente(
        clase_docente_id=clase.id,
        docente_id=current_user.id,
        tipo="APERTURA",
        motivo=data.motivo.strip(),
    ))
    clase.estado = "CORRECCION"
    db.commit()
    return _serializar_clase(clase)


@router.post("/clases/{clase_id}/no-impartida")
def reclasificar_clase_no_impartida(
    clase_id: int,
    data: ReclasificarNoImpartidaInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Clase no encontrada")
    _validar_carga_actual(db, clase.carga)
    if clase.estado not in {"ABIERTA", "CORRECCION", "CERRADA"}:
        raise HTTPException(409, "Esta clase no se puede reclasificar")
    estado_anterior = clase.estado
    db.add(CorreccionAsistenciaDocente(
        clase_docente_id=clase.id,
        docente_id=current_user.id,
        tipo="CLASIFICACION",
        estado_anterior=estado_anterior,
        estado_nuevo="NO_IMPARTIDA",
        motivo=data.motivo.strip(),
    ))
    clase.estado = "NO_IMPARTIDA"
    clase.fin = datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None)
    clase.motivo_no_impartida = data.motivo.strip()
    clase.declarada_no_impartida_en = datetime.datetime.utcnow()
    clase.estado_reposicion = "PENDIENTE" if data.requiere_reposicion else "NO_REQUERIDA"
    db.commit()
    db.refresh(clase)
    return _serializar_clase(clase)


@router.patch("/clases/{clase_id}/incidencia")
def registrar_incidencia_clase(
    clase_id: int,
    data: IncidenciaClaseInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Clase no encontrada")
    _validar_carga_actual(db, clase.carga)
    if clase.estado not in {"ABIERTA", "CORRECCION"}:
        raise HTTPException(409, "La asistencia debe estar abierta para registrar una nota")
    solicitar_antes = bool(clase.incidencia_solicita_justificacion)
    seguridad_notificada = clase.incidencia_tipo == "SEGURIDAD"
    clase.incidencia_tipo = data.tipo
    clase.incidencias = data.descripcion
    clase.incidencia_requiere_seguimiento = data.requiere_seguimiento
    clase.incidencia_solicita_justificacion = data.solicita_justificacion
    grupo_nombre = (
        f"{clase.carga.grupo_academico.cuatrimestre}° {clase.carga.grupo_academico.grupo}"
        if clase.carga.grupo_academico else "grupo sin identificar"
    )
    if data.tipo == "SEGURIDAD" and not seguridad_notificada:
        destinatarios = db.query(Usuario).filter(
            Usuario.rol.in_([RolUsuario.TUTORIA_ADMIN, RolUsuario.SUPER_ADMIN]),
            Usuario.activo == True,
        ).all()
        for destinatario in destinatarios:
            crear_notificacion(
                db, destinatario.id, "seguridad_nota_clase", "Nota urgente de seguridad",
                f"{current_user.nombre} registró una situación de seguridad en {clase.carga.actividad_nombre}, grupo {grupo_nombre}: {data.descripcion[:220]}",
                "/admin/tutoria", enviar_email=False,
            )
    if data.solicita_justificacion and not solicitar_antes:
        responsables = db.query(Usuario).filter(
            Usuario.rol == RolUsuario.SUPER_ADMIN, Usuario.activo == True,
        ).all()
        if current_user.departamento and current_user.departamento.responsable_id:
            responsable = db.get(Usuario, current_user.departamento.responsable_id)
            if responsable and responsable.activo and responsable.id != current_user.id:
                responsables.append(responsable)
        for destinatario in {usuario.id: usuario for usuario in responsables}.values():
            crear_notificacion(
                db, destinatario.id, "justificacion_colectiva_solicitada",
                "Solicitud de justificación colectiva",
                f"{current_user.nombre} solicita revisar la justificación de las faltas de {clase.carga.actividad_nombre}, grupo {grupo_nombre}, del {clase.fecha.isoformat()}.",
                f"/docente/seguimiento?carga={clase.carga_docente_id}", enviar_email=False,
            )
    db.commit()
    return _serializar_clase(clase)


@router.get("/seguimiento/{carga_id}")
def seguimiento_grupo(
    carga_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id,
        CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
        CargaDocente.tipo_actividad == "CLASE",
    ).first()
    if not carga or not carga.grupo_academico_id:
        raise HTTPException(404, "Carga docente con grupo no encontrada")
    cargas_equivalentes = _cargas_equivalentes(db, carga)
    carga_ids = [item.id for item in cargas_equivalentes]
    clases_historial = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id.in_(carga_ids),
    ).order_by(ClaseDocente.fecha.desc()).all()
    clases_sin_cerrar = [
        clase for clase in clases_historial
        if clase.estado in {"ABIERTA", "CORRECCION"}
    ]
    # Los registros NO_IMPARTIDA y las reposiciones aún PROGRAMADA documentan
    # decisiones del calendario, pero no son sesiones de asistencia. Una clase
    # ABIERTA tampoco entra al porcentaje hasta que el docente termina la captura.
    clases = [
        clase for clase in clases_historial
        if clase.estado in {"CERRADA", "CORRECCION"}
        and not clase.motivo_no_impartida
    ]
    inscripciones = db.query(InscripcionAlumno).filter(
        InscripcionAlumno.grupo_academico_id == carga.grupo_academico_id,
        InscripcionAlumno.estado == "ACTIVO",
    ).all()
    registros = db.query(SeguimientoAlumnoDocente).filter(
        SeguimientoAlumnoDocente.carga_docente_id.in_(carga_ids),
        SeguimientoAlumnoDocente.docente_id == current_user.id,
    ).order_by(SeguimientoAlumnoDocente.creado_en.desc()).all()
    registros_por_alumno = {}
    for registro in registros:
        registros_por_alumno.setdefault(registro.alumno_id, []).append(registro)
    alumnos = {
        i.alumno_id: {
            "alumno_id": i.alumno_id,
            "matricula": i.alumno.matricula,
            "nombre": (
                f"{i.alumno.apellido_paterno} {i.alumno.apellido_materno} {i.alumno.nombres}"
            ).strip(),
            "presente": 0, "falta": 0, "retardo": 0, "justificada": 0,
        }
        for i in inscripciones
    }
    for clase in clases:
        for asistencia in clase.asistencias:
            fila = alumnos.get(asistencia.alumno_id)
            if fila:
                clave = asistencia.estado.lower()
                fila[clave] = fila.get(clave, 0) + 1
    total_clases = len(clases)
    filas = []
    for fila in alumnos.values():
        asistio = fila["presente"] + fila["retardo"] + fila["justificada"]
        fila["porcentaje_asistencia"] = round(
            asistio / total_clases * 100, 1
        ) if total_clases else None
        estados_recientes = []
        for clase in clases:
            asistencia = next((a for a in clase.asistencias if a.alumno_id == fila["alumno_id"]), None)
            if asistencia:
                estados_recientes.append(asistencia.estado)
        faltas_consecutivas = 0
        for estado in estados_recientes:
            if estado != "FALTA":
                break
            faltas_consecutivas += 1
        regs_alumno = registros_por_alumno.get(fila["alumno_id"], [])
        calificaciones = [r.calificacion for r in reversed(regs_alumno) if r.tipo == "CALIFICACION" and r.calificacion is not None]
        descenso = (
            len(calificaciones) >= 4
            and (sum(calificaciones[-2:]) / 2) <= (sum(calificaciones[-4:-2]) / 2) - 1
        )
        alertas = []
        # Una racha solo representa un patrón después de tres faltas y cuando
        # existe al menos una sesión previa con otra condición. En las primeras
        # sesiones, "faltó a todas" describe mejor el dato que una falsa tendencia.
        if faltas_consecutivas >= 3 and faltas_consecutivas < total_clases:
            alertas.append({"tipo": "FALTAS_CONSECUTIVAS", "nivel": "ALTO" if faltas_consecutivas >= 3 else "MEDIO", "mensaje": f"{faltas_consecutivas} faltas consecutivas", "accion": "Contactar al alumno y documentar el motivo."})
        muestra_suficiente = total_clases >= 5
        riesgo_proyectado = muestra_suficiente and ((asistio / (total_clases + 1)) * 100) < 80
        if muestra_suficiente and (fila["porcentaje_asistencia"] < 80 or riesgo_proyectado):
            alertas.append({"tipo": "RIESGO_INASISTENCIA", "nivel": "ALTO", "mensaje": f"Faltó a {fila['falta']} de {total_clases} clases", "accion": "Revisar justificantes y canalizar a tutoría si corresponde."})
        if calificaciones and (calificaciones[-1] < 7 or descenso):
            alertas.append({"tipo": "DESEMPENO", "nivel": "MEDIO", "mensaje": "Descenso o desempeño académico bajo", "accion": "Acordar una actividad de recuperación y fecha de revisión."})
        requiere_tutoria = any(r.tipo == "TUTORIA" and r.estado not in {"ATENDIDO", "CERRADO"} for r in regs_alumno)
        if requiere_tutoria:
            alertas.append({"tipo": "TUTORIA", "nivel": "MEDIO", "mensaje": "Seguimiento de tutoría pendiente", "accion": "Confirmar atención y registrar el acuerdo alcanzado."})
        fila["faltas_consecutivas"] = faltas_consecutivas
        fila["alertas"] = alertas
        fila["alerta"] = bool(alertas)
        fila["ultima_calificacion"] = calificaciones[-1] if calificaciones else None
        filas.append(fila)
    filas.sort(key=lambda x: x["nombre"])
    promedio = round(
        sum(f["porcentaje_asistencia"] for f in filas) / len(filas), 1
    ) if filas and total_clases else None
    return {
        "carga": _serializar_carga(carga, db),
        "bloques_semanales": len(cargas_equivalentes),
        "total_clases": total_clases,
        "muestra_suficiente": total_clases >= 5,
        "total_alumnos": len(filas),
        "promedio_asistencia": promedio,
        "alumnos_en_alerta": sum(1 for f in filas if f["alerta"]),
        "clases_sin_cerrar": [
            {"id": c.id, "fecha": c.fecha.isoformat(), "estado": c.estado}
            for c in clases_sin_cerrar
        ],
        "alumnos": filas,
        "clases": [_serializar_clase(c) for c in clases[:12]],
    }


def _carga_y_alumno_docente(db: Session, carga_id: int, alumno_id: int, docente_id: int):
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id,
        CargaDocente.docente_id == docente_id,
        CargaDocente.activo == True,
        CargaDocente.tipo_actividad == "CLASE",
    ).first()
    if not carga:
        raise HTTPException(404, "Alumno no encontrado en este grupo")
    inscripcion = db.query(InscripcionAlumno).filter(
        InscripcionAlumno.grupo_academico_id == carga.grupo_academico_id,
        InscripcionAlumno.alumno_id == alumno_id,
        InscripcionAlumno.estado == "ACTIVO",
    ).first()
    if not inscripcion:
        raise HTTPException(404, "Alumno no encontrado en este grupo")
    return carga, inscripcion.alumno


def _cargas_equivalentes(db: Session, carga: CargaDocente):
    """Agrupa los bloques semanales de una misma materia, grupo y periodo."""
    candidatas = db.query(CargaDocente).filter(
        CargaDocente.docente_id == carga.docente_id,
        CargaDocente.periodo_id == carga.periodo_id,
        CargaDocente.grupo_academico_id == carga.grupo_academico_id,
        CargaDocente.tipo_actividad == "CLASE",
        CargaDocente.activo == True,
    ).all()
    if carga.materia_id:
        return [item for item in candidatas if item.materia_id == carga.materia_id]
    nombre = (carga.actividad_nombre or "").strip().casefold()
    return [
        item for item in candidatas
        if not item.materia_id and (item.actividad_nombre or "").strip().casefold() == nombre
    ]


def _contexto_alumno_docente(db: Session, carga: CargaDocente, alumno: CatalogoAlumno, docente_id: int):
    cargas_grupo = db.query(CargaDocente).filter(
        CargaDocente.grupo_academico_id == carga.grupo_academico_id,
        CargaDocente.periodo_id == carga.periodo_id,
        CargaDocente.activo == True,
        CargaDocente.tipo_actividad == "CLASE",
    ).all()
    carga_ids = [item.id for item in cargas_grupo]
    clases = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id.in_(carga_ids),
    ).all() if carga_ids else []
    clase_ids = [clase.id for clase in clases]
    asistencias = db.query(AsistenciaDocente).filter(
        AsistenciaDocente.alumno_id == alumno.id,
        AsistenciaDocente.clase_docente_id.in_(clase_ids),
    ).all() if clase_ids else []
    asistio = sum(1 for item in asistencias if item.estado in {"PRESENTE", "RETARDO", "JUSTIFICADA"})
    porcentaje_global = (asistio / len(asistencias) * 100) if asistencias else 100
    calificacion_baja = db.query(SeguimientoAlumnoDocente.id).filter(
        SeguimientoAlumnoDocente.alumno_id == alumno.id,
        SeguimientoAlumnoDocente.carga_docente_id.in_(carga_ids),
        SeguimientoAlumnoDocente.tipo == "CALIFICACION",
        SeguimientoAlumnoDocente.calificacion < 7,
    ).first() if carga_ids else None
    riesgo_global = bool(
        (len(asistencias) >= 4 and porcentaje_global < 80)
        or calificacion_baja
    )
    canalizacion_activa = db.query(Canalizacion.id).filter(
        Canalizacion.alumno_id == alumno.id,
        Canalizacion.estado.in_(["PENDIENTE", "EN_SEGUIMIENTO"]),
    ).first() is not None
    reporte_activo = db.query(ReporteTutor.id).filter(
        ReporteTutor.alumno_id == alumno.id,
        ReporteTutor.estado.in_(["ENVIADO", "RECIBIDO", "EN_SEGUIMIENTO", "CANALIZADO", "SIN_TUTOR"]),
    ).first() is not None
    desde = datetime.datetime.utcnow() - datetime.timedelta(days=7)
    alerta_reciente = db.query(ReporteTutor).filter(
        ReporteTutor.alumno_id == alumno.id,
        ReporteTutor.reportado_por_id == docente_id,
        ReporteTutor.carga_docente_id == carga.id,
        ReporteTutor.creado_en >= desde,
        ReporteTutor.estado.notin_(["CERRADO", "ATENDIDO", "CANCELADO"]),
    ).order_by(ReporteTutor.creado_en.desc()).first()
    asignacion = db.query(AsignacionTutoria).filter(
        AsignacionTutoria.alumno_id == alumno.id,
        AsignacionTutoria.activo == True,
    ).order_by(AsignacionTutoria.asignado_en.desc()).first()
    grupo_tutorado = db.query(GrupoTutorado).filter(
        GrupoTutorado.id == asignacion.grupo_tutorado_id,
        GrupoTutorado.activo == True,
    ).first() if asignacion else None
    tutor = db.query(Usuario).filter(
        Usuario.id == grupo_tutorado.tutor_id,
        Usuario.activo == True,
    ).first() if grupo_tutorado else None
    return {
        "alumno_id": alumno.id,
        "riesgo_global": riesgo_global,
        "canalizacion_activa": canalizacion_activa,
        "seguimiento_activo": reporte_activo,
        "tutor_asignado": tutor.nombre if tutor else None,
        "alerta_reciente": ({
            "id": alerta_reciente.id,
            "estado": alerta_reciente.estado,
            "categoria": alerta_reciente.categoria,
            "creado_en": alerta_reciente.creado_en.isoformat(),
        } if alerta_reciente else None),
    }


@router.get("/seguimiento/{carga_id}/alumnos/{alumno_id}/contexto")
def contexto_alumno_docente(
    carga_id: int,
    alumno_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga, alumno = _carga_y_alumno_docente(db, carga_id, alumno_id, current_user.id)
    return _contexto_alumno_docente(db, carga, alumno, current_user.id)


@router.get("/clases/{clase_id}/contexto-alumnos")
def contexto_alumnos_clase(
    clase_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Clase no encontrada")
    return {
        str(asistencia.alumno_id): _contexto_alumno_docente(
            db, clase.carga, asistencia.alumno, current_user.id,
        )
        for asistencia in clase.asistencias
    }


@router.post("/seguimiento/{carga_id}/alumnos/{alumno_id}/alerta-temprana")
def crear_alerta_temprana(
    carga_id: int,
    alumno_id: int,
    data: AlertaTempranaInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga, alumno = _carga_y_alumno_docente(db, carga_id, alumno_id, current_user.id)
    _validar_carga_actual(db, carga)
    categorias = {
        "INASISTENCIA": ("ASISTENCIA", "Alerta por inasistencia"),
        "BAJO_DESEMPENO": ("ACADEMICO", "Alerta por bajo desempeño"),
        "CAMBIO_CONDUCTA": ("CONDUCTA", "Alerta por cambio de conducta"),
        "FALTA_PARTICIPACION": ("ACADEMICO", "Alerta por falta de participación"),
        "SITUACION_PERSONAL": ("PERSONAL", "Posible situación personal"),
        "OTRO": ("OTRO", "Señal de atención observada"),
    }
    categoria, titulo = categorias[data.senal]
    desde = datetime.datetime.utcnow() - datetime.timedelta(days=7)
    duplicado = db.query(ReporteTutor).filter(
        ReporteTutor.alumno_id == alumno.id,
        ReporteTutor.reportado_por_id == current_user.id,
        ReporteTutor.carga_docente_id == carga.id,
        ReporteTutor.categoria == categoria,
        ReporteTutor.creado_en >= desde,
        ReporteTutor.estado.notin_(["CERRADO", "ATENDIDO", "CANCELADO"]),
    ).first()
    if duplicado:
        raise HTTPException(409, "Ya enviaste una alerta similar durante los últimos 7 días. Puedes revisar su estado en la ficha del alumno.")
    grupo_tutorado = grupo_tutoria_para_academico(db, carga.grupo_academico_id)
    tutor = db.query(Usuario).filter(
        Usuario.id == grupo_tutorado.tutor_id,
        Usuario.activo == True,
    ).first() if grupo_tutorado else None
    prioridad = {"OBSERVACION": "BAJA", "ATENCION": "MEDIA", "URGENTE": "ALTA"}[data.nivel]
    detalle = data.comentario or f"Señal registrada desde {carga.actividad_nombre}."
    registro = SeguimientoAlumnoDocente(
        docente_id=current_user.id,
        carga_docente_id=carga.id,
        alumno_id=alumno.id,
        tipo="TUTORIA",
        titulo=titulo,
        detalle=detalle,
        estado="PENDIENTE",
        fecha_revision=(_ahora_mx() + datetime.timedelta(days=7)).date(),
    )
    db.add(registro)
    db.flush()
    reporte = ReporteTutor(
        alumno_id=alumno.id,
        reportado_por_id=current_user.id,
        tutor_destinatario_id=tutor.id if tutor else None,
        grupo_tutorado_id=grupo_tutorado.id if grupo_tutorado else None,
        carga_docente_id=carga.id,
        seguimiento_docente_id=registro.id,
        categoria=categoria,
        prioridad=prioridad,
        titulo=titulo,
        detalle=detalle,
        confidencial=data.senal in {"SITUACION_PERSONAL", "OTRO"},
        estado="ENVIADO" if tutor else "SIN_TUTOR",
    )
    db.add(reporte)
    db.flush()
    alumno_nombre = f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombres}".strip()
    destinatarios = [tutor] if tutor else db.query(Usuario).filter(
        Usuario.rol.in_([RolUsuario.TUTORIA_ADMIN, RolUsuario.SUPER_ADMIN]),
        Usuario.activo == True,
    ).all()
    for destinatario in destinatarios:
        crear_notificacion(
            db, destinatario.id, "tutoria_alerta_temprana",
            "Nueva alerta temprana",
            f"{current_user.nombre} registró una señal de atención para {alumno_nombre}.",
            "/docente/mis-tutorados?tab=reportes" if tutor else "/admin/tutoria?tab=reportes-tutor",
            enviar_email=False,
        )
    db.commit()
    return {
        "id": reporte.id,
        "estado": reporte.estado,
        "destinatario": tutor.nombre if tutor else "Responsable de Tutoría",
        "mensaje": f"Alerta enviada a {tutor.nombre}" if tutor else "Alerta enviada al Responsable de Tutoría",
    }


@router.get("/seguimiento/{carga_id}/alumnos/{alumno_id}/faltas")
def faltas_justificables(
    carga_id: int,
    alumno_id: int,
    fecha_inicio: datetime.date,
    fecha_fin: datetime.date,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga, alumno = _carga_y_alumno_docente(
        db, carga_id, alumno_id, current_user.id,
    )
    _validar_carga_actual(db, carga)
    cargas_equivalentes = _cargas_equivalentes(db, carga)
    carga_ids = [item.id for item in cargas_equivalentes]
    horarios = {item.id: f"{item.hora_inicio}–{item.hora_fin}" for item in cargas_equivalentes}
    if fecha_fin < fecha_inicio:
        raise HTTPException(422, "La fecha final debe ser igual o posterior a la inicial")
    faltas = (
        db.query(AsistenciaDocente, ClaseDocente)
        .join(ClaseDocente, ClaseDocente.id == AsistenciaDocente.clase_docente_id)
        .filter(
            ClaseDocente.carga_docente_id.in_(carga_ids),
            ClaseDocente.fecha.between(fecha_inicio, fecha_fin),
            AsistenciaDocente.alumno_id == alumno.id,
            AsistenciaDocente.estado == "FALTA",
        )
        .order_by(ClaseDocente.fecha.asc())
        .all()
    )
    return {
        "alumno": {
            "id": alumno.id,
            "matricula": alumno.matricula,
            "nombre": (
                f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombres}"
            ).strip(),
        },
        "materia": carga.actividad_nombre,
        "faltas": [
            {
                "asistencia_id": asistencia.id,
                "clase_id": clase.id,
                "fecha": clase.fecha.isoformat(),
                "horario": horarios.get(clase.carga_docente_id, ""),
                "estado": asistencia.estado,
            }
            for asistencia, clase in faltas
        ],
    }


@router.post("/seguimiento/{carga_id}/alumnos/{alumno_id}/justificar-faltas")
def justificar_faltas_multiples(
    carga_id: int,
    alumno_id: int,
    data: JustificacionMultipleInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga, alumno = _carga_y_alumno_docente(
        db, carga_id, alumno_id, current_user.id,
    )
    _validar_carga_actual(db, carga)
    carga_ids = [item.id for item in _cargas_equivalentes(db, carga)]
    registros = (
        db.query(AsistenciaDocente, ClaseDocente)
        .join(ClaseDocente, ClaseDocente.id == AsistenciaDocente.clase_docente_id)
        .filter(
            AsistenciaDocente.id.in_(data.asistencia_ids),
            AsistenciaDocente.alumno_id == alumno.id,
            ClaseDocente.carga_docente_id.in_(carga_ids),
            ClaseDocente.fecha.between(data.fecha_inicio, data.fecha_fin),
        )
        .all()
    )
    if len(registros) != len(data.asistencia_ids):
        raise HTTPException(
            422, "Una o más asistencias no pertenecen al alumno, materia o periodo indicado",
        )
    no_justificables = [
        asistencia.id for asistencia, _ in registros
        if asistencia.estado != "FALTA"
    ]
    if no_justificables:
        raise HTTPException(
            409, "Solo se pueden justificar registros que actualmente sean faltas",
        )

    justificacion = JustificacionAsistenciaDocente(
        docente_id=current_user.id,
        carga_docente_id=carga.id,
        alumno_id=alumno.id,
        fecha_inicio=data.fecha_inicio,
        fecha_fin=data.fecha_fin,
        motivo=data.motivo,
        folio=data.folio,
    )
    db.add(justificacion)
    db.flush()
    marca = _ahora_mx().strftime("%d/%m/%Y %H:%M")
    referencia = f" · Folio {data.folio}" if data.folio else ""
    nota = f"[Justificada {marca}{referencia}] {data.motivo}"
    for asistencia, _ in registros:
        db.add(DetalleJustificacionAsistencia(
            justificacion_id=justificacion.id,
            asistencia_id=asistencia.id,
            estado_anterior=asistencia.estado,
            estado_nuevo="JUSTIFICADA",
        ))
        db.add(CorreccionAsistenciaDocente(
            clase_docente_id=asistencia.clase_docente_id,
            asistencia_id=asistencia.id,
            alumno_id=alumno.id,
            docente_id=current_user.id,
            tipo="JUSTIFICACION",
            estado_anterior=asistencia.estado,
            estado_nuevo="JUSTIFICADA",
            motivo=f"{data.motivo}{referencia}",
        ))
        asistencia.estado = "JUSTIFICADA"
        asistencia.observacion = "\n".join(
            parte for parte in [asistencia.observacion, nota] if parte
        )
    db.commit()
    return {
        "id": justificacion.id,
        "alumno_id": alumno.id,
        "faltas_justificadas": len(registros),
        "fecha_inicio": data.fecha_inicio.isoformat(),
        "fecha_fin": data.fecha_fin.isoformat(),
        "motivo": data.motivo,
        "folio": data.folio,
    }


@router.get("/seguimiento/{carga_id}/alumnos/{alumno_id}")
def ficha_alumno_docente(
    carga_id: int,
    alumno_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga, alumno = _carga_y_alumno_docente(db, carga_id, alumno_id, current_user.id)
    carga_ids = [item.id for item in _cargas_equivalentes(db, carga)]
    clases = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id.in_(carga_ids),
    ).order_by(ClaseDocente.fecha.desc()).all()
    asistencias = []
    conteos = {estado.lower(): 0 for estado in ESTADOS_ASISTENCIA}
    for clase in clases:
        asistencia = next((a for a in clase.asistencias if a.alumno_id == alumno.id), None)
        if asistencia:
            conteos[asistencia.estado.lower()] += 1
            asistencias.append({
                "clase_id": clase.id,
                "fecha": clase.fecha.isoformat(),
                "hora_inicio": clase.carga.hora_inicio,
                "hora_fin": clase.carga.hora_fin,
                "estado": asistencia.estado,
                "observacion": asistencia.observacion,
            })
    registros = db.query(SeguimientoAlumnoDocente).filter(
        SeguimientoAlumnoDocente.carga_docente_id.in_(carga_ids),
        SeguimientoAlumnoDocente.alumno_id == alumno.id,
        SeguimientoAlumnoDocente.docente_id == current_user.id,
    ).order_by(SeguimientoAlumnoDocente.creado_en.desc()).all()
    total = len(asistencias)
    asistio = conteos["presente"] + conteos["retardo"] + conteos["justificada"]
    seguimiento = seguimiento_grupo(carga_id, db, current_user)
    alertas = next(
        (fila["alertas"] for fila in seguimiento["alumnos"] if fila["alumno_id"] == alumno.id),
        [],
    )
    return {
        "alumno": {
            "id": alumno.id, "matricula": alumno.matricula,
            "nombre": f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombres}".strip(),
            "carrera": alumno.carrera, "cuatrimestre": alumno.cuatrimestre, "grupo": alumno.grupo,
        },
        "carga": _serializar_carga(carga, db),
        "resumen": {**conteos, "total": total, "porcentaje_asistencia": round((asistio / total * 100) if total else 0, 1)},
        "alertas": alertas,
        "asistencias": asistencias,
        "registros": [{
            "id": r.id, "tipo": r.tipo, "titulo": r.titulo, "detalle": r.detalle,
            "calificacion": r.calificacion, "estado": r.estado,
            "fecha_limite": r.fecha_limite.isoformat() if r.fecha_limite else None,
            "fecha_revision": r.fecha_revision.isoformat() if r.fecha_revision else None,
            "resultado_atencion": r.resultado_atencion,
            "atendido_en": r.atendido_en.isoformat() if r.atendido_en else None,
            "creado_en": r.creado_en.isoformat(),
        } for r in registros],
    }


@router.post("/seguimiento/{carga_id}/alumnos/{alumno_id}/registros")
def registrar_seguimiento_alumno(
    carga_id: int,
    alumno_id: int,
    data: SeguimientoInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    carga, alumno = _carga_y_alumno_docente(db, carga_id, alumno_id, current_user.id)
    _validar_carga_actual(db, carga)
    registro = SeguimientoAlumnoDocente(
        docente_id=current_user.id, carga_docente_id=carga.id, alumno_id=alumno.id,
        tipo=data.tipo, titulo=data.titulo.strip(), detalle=data.detalle,
        calificacion=data.calificacion, estado=data.estado, fecha_limite=data.fecha_limite,
        fecha_revision=data.fecha_revision,
    )
    db.add(registro)
    db.flush()

    reporte = None
    tutor = None
    if data.tipo == "TUTORIA":
        grupo_tutorado = grupo_tutoria_para_academico(db, carga.grupo_academico_id)
        tutor = (
            db.query(Usuario).filter(Usuario.id == grupo_tutorado.tutor_id, Usuario.activo == True).first()
            if grupo_tutorado else None
        )
        reporte = ReporteTutor(
            alumno_id=alumno.id,
            reportado_por_id=current_user.id,
            tutor_destinatario_id=tutor.id if tutor else None,
            grupo_tutorado_id=grupo_tutorado.id if grupo_tutorado else None,
            carga_docente_id=carga.id,
            seguimiento_docente_id=registro.id,
            categoria=data.categoria_reporte,
            prioridad=data.prioridad_reporte,
            titulo=data.titulo.strip(),
            detalle=data.detalle,
            confidencial=data.confidencial,
            estado="ENVIADO" if tutor else "SIN_TUTOR",
        )
        db.add(reporte)
        db.flush()

        alumno_nombre = f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombres}".strip()
        if tutor:
            crear_notificacion(
                db, tutor.id, "tutoria_reporte",
                "Nuevo reporte de un docente",
                f"{current_user.nombre} reportó un caso de {alumno_nombre}: {data.titulo.strip()}.",
                "/docente/mis-tutorados?tab=reportes", enviar_email=False,
            )
        else:
            responsables = db.query(Usuario).filter(
                Usuario.rol.in_([RolUsuario.TUTORIA_ADMIN, RolUsuario.SUPER_ADMIN]),
                Usuario.activo == True,
            ).all()
            for responsable in responsables:
                crear_notificacion(
                    db, responsable.id, "tutoria_reporte_sin_tutor",
                    "Reporte sin tutor asignado",
                    f"{alumno_nombre} tiene un reporte pendiente, pero no cuenta con tutor activo.",
                    "/admin/tutoria?tab=reportes-tutor", enviar_email=False,
                )
    db.commit()
    db.refresh(registro)
    return {
        "id": registro.id,
        "reporte_tutor_id": reporte.id if reporte else None,
        "destinatario": tutor.nombre if tutor else ("Responsable de Tutoría" if reporte else None),
        "estado_envio": reporte.estado if reporte else None,
        "mensaje": (
            f"Reporte enviado a {tutor.nombre}" if reporte and tutor
            else "Reporte enviado al Responsable de Tutoría para asignación" if reporte
            else "Seguimiento registrado"
        ),
    }


@router.patch("/seguimiento/registros/{registro_id}")
def actualizar_estado_seguimiento(
    registro_id: int,
    data: EstadoSeguimientoInput,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    registro = db.query(SeguimientoAlumnoDocente).filter(
        SeguimientoAlumnoDocente.id == registro_id,
        SeguimientoAlumnoDocente.docente_id == current_user.id,
    ).first()
    if not registro:
        raise HTTPException(404, "Registro de seguimiento no encontrado")
    _validar_carga_actual(db, registro.carga)
    if registro.tipo == "TUTORIA":
        raise HTTPException(409, "El reporte debe ser atendido por el tutor destinatario")
    if registro.tipo != "ACUERDO":
        raise HTTPException(409, "Este registro no maneja estado")
    registro.estado = data.estado
    registro.resultado_atencion = data.resultado_atencion.strip() if data.resultado_atencion else None
    if data.estado == "REPROGRAMADO":
        registro.fecha_limite = data.fecha_limite
        registro.fecha_revision = data.fecha_revision
    estados_cerrados = {"ATENDIDO", "CUMPLIDO", "CUMPLIDO_PARCIAL", "NO_CUMPLIDO", "CERRADO"}
    registro.atendido_en = datetime.datetime.utcnow() if data.estado in estados_cerrados else None
    db.commit()
    return {"id": registro.id, "estado": registro.estado}


@router.get("/seguimiento/{carga_id}/exportar.xlsx")
def exportar_seguimiento_excel(
    carga_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Genera el concentrado de asistencia de una materia y grupo del docente."""
    carga = db.query(CargaDocente).filter(
        CargaDocente.id == carga_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not carga or not carga.grupo_academico_id:
        raise HTTPException(404, "Carga académica no encontrada")
    cargas = _cargas_equivalentes(db, carga)
    clases = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id.in_([item.id for item in cargas]),
        ClaseDocente.estado == "CERRADA",
    ).order_by(ClaseDocente.fecha.asc()).all()
    inscripciones = db.query(InscripcionAlumno).filter(
        InscripcionAlumno.grupo_academico_id == carga.grupo_academico_id,
        InscripcionAlumno.estado == "ACTIVO",
    ).all()
    alumnos = sorted(
        [item.alumno for item in inscripciones],
        key=lambda a: (a.apellido_paterno or "", a.apellido_materno or "", a.nombres or ""),
    )
    por_clase = {
        clase.id: {asistencia.alumno_id: asistencia.estado for asistencia in clase.asistencias}
        for clase in clases
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Concentrado"
    grupo = carga.grupo_academico
    ahora_mx = _ahora_mx()
    fecha_corte = ahora_mx.date()
    ws.append(["SIGA · Concentrado de asistencia"])
    ws.append(["Materia", carga.actividad_nombre, "Grupo", f"{grupo.cuatrimestre}° {grupo.grupo}", "Docente", current_user.nombre])
    ws.append(["Carrera", grupo.carrera, "", "", "Periodo", carga.periodo.clave if carga.periodo else ""])
    leyenda = "P = Presente · F = Falta · R = Retardo · J = Justificada"
    if clases:
        ws.append([leyenda])
    else:
        ws.append([f"Sin clases registradas en el periodo al {fecha_corte.strftime('%d/%m/%Y')} · {leyenda}"])
    encabezados = ["Matrícula", "Alumno"] + [clase.fecha.strftime("%d/%m/%Y") for clase in clases] + ["P", "F", "R", "J", "% asistencia"]
    ws.append(encabezados)
    abreviatura = {"PRESENTE": "P", "FALTA": "F", "RETARDO": "R", "JUSTIFICADA": "J"}
    for alumno in alumnos:
        estados = [por_clase[clase.id].get(alumno.id, "") for clase in clases]
        conteos = {estado: estados.count(estado) for estado in ESTADOS_ASISTENCIA}
        asistio = conteos["PRESENTE"] + conteos["RETARDO"] + conteos["JUSTIFICADA"]
        porcentaje = asistio / len(estados) if estados else None
        nombre = f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombres}".strip()
        ws.append([alumno.matricula, nombre] + [abreviatura.get(e, "") for e in estados] + [
            conteos["PRESENTE"], conteos["FALTA"], conteos["RETARDO"], conteos["JUSTIFICADA"], porcentaje,
        ])
    primera_fila_alumnos = 6
    ultima_fila_alumnos = ws.max_row
    total_presentes = sum(1 for clase in clases for asistencia in clase.asistencias if asistencia.estado == "PRESENTE")
    total_faltas = sum(1 for clase in clases for asistencia in clase.asistencias if asistencia.estado == "FALTA")
    total_retardos = sum(1 for clase in clases for asistencia in clase.asistencias if asistencia.estado == "RETARDO")
    total_justificadas = sum(1 for clase in clases for asistencia in clase.asistencias if asistencia.estado == "JUSTIFICADA")
    promedio_grupo = (
        sum((a["PRESENTE"] + a["RETARDO"] + a["JUSTIFICADA"]) / len(clases) for a in [
            {estado: [por_clase[c.id].get(alumno.id, "") for c in clases].count(estado) for estado in ESTADOS_ASISTENCIA}
            for alumno in alumnos
        ]) / len(alumnos)
        if clases and alumnos else None
    )
    total_col = len(encabezados)
    ws.append(["Resumen", f"{len(alumnos)} alumnos"] + [""] * len(clases) + [
        total_presentes, total_faltas, total_retardos, total_justificadas, promedio_grupo,
    ])
    fila_resumen = ws.max_row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(encabezados)))
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="047857")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max(2, len(encabezados)))
    ws["A4"].font = Font(italic=True, color="64748B")
    ws["A4"].alignment = Alignment(wrap_text=True)
    for cell in ws[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "C6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(encabezados))}{max(5, ultima_fila_alumnos)}"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 44
    for index in range(3, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 12
    ws.column_dimensions[get_column_letter(total_col)].width = 15
    for fila in range(primera_fila_alumnos, ultima_fila_alumnos + 1):
        ws.cell(fila, total_col).number_format = "0.0%"
    ws.cell(fila_resumen, total_col).number_format = "0.0%"
    for cell in ws[fila_resumen]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")
    if len(clases) >= 3 and ultima_fila_alumnos >= primera_fila_alumnos:
        rango_porcentaje = f"{get_column_letter(total_col)}{primera_fila_alumnos}:{get_column_letter(total_col)}{ultima_fila_alumnos}"
        ws.conditional_formatting.add(
            rango_porcentaje,
            CellIsRule(operator="lessThan", formula=["0.8"], fill=PatternFill("solid", fgColor="FECACA")),
        )

    info = wb.create_sheet("Información")
    info.append(["Información de generación", "Valor"])
    info.append(["Docente", current_user.nombre])
    info.append(["Exportado por", current_user.nombre])
    info.append(["Generado", ahora_mx.replace(tzinfo=None)])
    info.append(["Fecha de corte", fecha_corte])
    info.append(["Periodo", carga.periodo.clave if carga.periodo else ""])
    info.append(["Formato", "Concentrado de asistencia SIGA"])
    info.append(["Versión", "1"])
    info.append(["Leyenda", leyenda])
    info["B4"].number_format = "dd/mm/yyyy hh:mm"
    info["B5"].number_format = "dd/mm/yyyy"
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 58
    for cell in info[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="047857")
    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    nombre = _nombre_archivo(f"Asistencia_{carga.actividad_nombre}_{grupo.cuatrimestre}{grupo.grupo}_{fecha_corte.isoformat()}")
    return StreamingResponse(
        salida,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}.xlsx"'},
    )


@router.get("/clases/{clase_id}/exportar.pdf")
def exportar_clase_pdf(
    clase_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Genera una lista de asistencia imprimible para una sesión docente."""
    clase = db.query(ClaseDocente).join(CargaDocente).filter(
        ClaseDocente.id == clase_id,
        CargaDocente.docente_id == current_user.id,
    ).first()
    if not clase:
        raise HTTPException(404, "Clase no encontrada")
    salida = io.BytesIO()
    doc = SimpleDocTemplate(salida, pagesize=landscape(letter), leftMargin=1.2 * cm, rightMargin=1.2 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    estilos = getSampleStyleSheet()
    carga = clase.carga
    serializada = _serializar_clase(clase)
    materia = escape(carga.actividad_nombre or "—")
    grupo_texto = escape(serializada["carga"]["grupo"] or "—")
    docente_texto = escape(carga.docente.nombre if carga.docente else "—")
    espacio_texto = escape(serializada["carga"]["espacio_nombre"] or "—")
    contenido = [
        Paragraph("SIGA · Lista de asistencia", estilos["Title"]),
        Paragraph(f"<b>Materia:</b> {materia} &nbsp;&nbsp; <b>Grupo:</b> {grupo_texto} &nbsp;&nbsp; <b>Fecha:</b> {clase.fecha.strftime('%d/%m/%Y')}", estilos["Normal"]),
        Paragraph(f"<b>Docente:</b> {docente_texto} &nbsp;&nbsp; <b>Horario:</b> {carga.hora_inicio}–{carga.hora_fin} &nbsp;&nbsp; <b>Espacio:</b> {espacio_texto}", estilos["Normal"]),
        Spacer(1, 0.35 * cm),
    ]
    filas = [["N.º", "Matrícula", "Alumno", "Estado", "Observación", "Firma"]]
    for indice, asistencia in enumerate(sorted(clase.asistencias, key=lambda a: (a.alumno.apellido_paterno, a.alumno.nombres)), 1):
        alumno = asistencia.alumno
        filas.append([indice, alumno.matricula, f"{alumno.apellido_paterno} {alumno.apellido_materno} {alumno.nombres}".strip(), asistencia.estado, asistencia.observacion or "", ""])
    tabla = Table(filas, repeatRows=1, colWidths=[1.1 * cm, 2.8 * cm, 7.3 * cm, 2.7 * cm, 7.2 * cm, 4.2 * cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#047857")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#94A3B8")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
    ]))
    contenido.extend([tabla, Spacer(1, 0.5 * cm), Paragraph("Firma del docente: ______________________________________________", estilos["Normal"])])
    doc.build(contenido)
    salida.seek(0)
    nombre = _nombre_archivo(f"Lista_{carga.actividad_nombre}_{clase.fecha.isoformat()}")
    return StreamingResponse(salida, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{nombre}.pdf"'})


@router.get("/historial")
def historial_clases(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    clases = db.query(ClaseDocente).join(CargaDocente).filter(
        CargaDocente.docente_id == current_user.id
    ).order_by(ClaseDocente.fecha.desc(), ClaseDocente.inicio.desc()).limit(500).all()
    return [_serializar_clase(clase) for clase in clases]


@router.get("/dashboard", summary="Resumen operativo del panel docente")
def dashboard_docente(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _solo_docente(current_user)
    ahora = _ahora_mx()
    hoy = ahora.date()
    actual = _periodo_actual(db)
    if not actual:
        return {
            "periodo": None,
            "resumen": {
                "clases_hoy": 0, "clases_cerradas": 0, "asistencias_pendientes": 0,
                "grupos_activos": 0, "alumnos_atencion": 0, "acuerdos_pendientes": 0,
                "acuerdos_vencidos": 0,
            },
            "jornada": [], "grupos": [], "alumnos_prioritarios": [],
        }
    cargas = db.query(CargaDocente).filter(
        CargaDocente.docente_id == current_user.id,
        CargaDocente.activo == True,
        CargaDocente.estado == "ACTIVO",
        CargaDocente.periodo_id == actual.id,
    ).order_by(CargaDocente.hora_inicio).all()
    cargas_clase = [carga for carga in cargas if carga.tipo_actividad == "CLASE"]
    clases = db.query(ClaseDocente).filter(
        ClaseDocente.carga_docente_id.in_([carga.id for carga in cargas_clase]),
    ).all() if cargas_clase else []
    clase_por_carga_fecha = {
        (clase.carga_docente_id, clase.fecha): clase for clase in clases
    }

    jornada = []
    for carga in cargas:
        if carga.dia_semana != hoy.weekday():
            continue
        estado_fecha = estado_fecha_academica(db, carga.periodo_id, hoy)
        clase = clase_por_carga_fecha.get((carga.id, hoy))
        hora_fin = datetime.datetime.combine(
            hoy, datetime.time.fromisoformat(carga.hora_fin),
            tzinfo=MX,
        )
        if clase:
            estado = {
                "ABIERTA": "EN_CURSO",
                "CORRECCION": "CORRECCION",
                "CERRADA": "CERRADA",
            }.get(clase.estado, clase.estado)
        elif not estado_fecha["requiere_asistencia"]:
            estado = "NO_LECTIVA"
        elif carga.tipo_actividad != "CLASE":
            hora_inicio = datetime.datetime.combine(
                hoy, datetime.time.fromisoformat(carga.hora_inicio), tzinfo=MX,
            )
            estado = "EN_CURSO" if hora_inicio <= ahora <= hora_fin else "FINALIZADA" if ahora > hora_fin else "PROGRAMADA"
        elif ahora > hora_fin:
            estado = "SIN_REGISTRO"
        else:
            estado = "PROGRAMADA"
        jornada.append({
            "carga_id": carga.id,
            "clase_id": clase.id if clase else None,
            "tipo_actividad": carga.tipo_actividad,
            "grupo_tutorado_id": carga.grupo_tutorado_id,
            "materia": carga.actividad_nombre,
            "grupo": (
                f"{carga.grupo_academico.cuatrimestre}° {carga.grupo_academico.grupo}"
                if carga.grupo_academico else
                f"{carga.grupo_tutorado.cuatrimestre}° {carga.grupo_tutorado.grupo}"
                if carga.grupo_tutorado else carga.tipo_actividad.title()
            ),
            "carrera": carga.grupo_academico.carrera if carga.grupo_academico else carga.grupo_tutorado.carrera if carga.grupo_tutorado else None,
            "espacio": (
                carga.laboratorio.nombre if carga.laboratorio
                else carga.espacio_nombre or "Sin espacio"
            ),
            "hora_inicio": carga.hora_inicio,
            "hora_fin": carga.hora_fin,
            "estado": estado,
            "calendario": estado_fecha,
            "resumen": _serializar_clase(clase)["resumen"] if clase else None,
        })

    # Una materia puede tener varios bloques semanales. El dashboard debe
    # analizarla una sola vez, porque seguimiento_grupo ya consolida sus bloques.
    cargas_seguimiento = []
    claves_seguimiento = set()
    for carga in cargas:
        if carga.tipo_actividad != "CLASE" or not carga.grupo_academico_id:
            continue
        identidad_materia = (
            f"materia:{carga.materia_id}" if carga.materia_id
            else f"nombre:{(carga.actividad_nombre or '').strip().casefold()}"
        )
        clave = (carga.grupo_academico_id, identidad_materia)
        if clave not in claves_seguimiento:
            claves_seguimiento.add(clave)
            cargas_seguimiento.append(carga)

    grupos = []
    alumnos_prioritarios_por_id = {}
    acuerdos_pendientes = 0
    acuerdos_vencidos = 0
    for carga in cargas_seguimiento:
        seguimiento = seguimiento_grupo(carga.id, db, current_user)
        cargas_equivalentes = _cargas_equivalentes(db, carga)
        cargas_equivalentes_ids = [item.id for item in cargas_equivalentes]
        coincidencia_periodo = re.search(r"(ENE-ABR|MAY-AGO|SEP-DIC)\s+(\d{4})", actual.clave or "", re.I)
        if coincidencia_periodo:
            mes_inicio = {"ENE-ABR": 1, "MAY-AGO": 5, "SEP-DIC": 9}[coincidencia_periodo.group(1).upper()]
            fecha_inicio_periodo = datetime.date(int(coincidencia_periodo.group(2)), mes_inicio, 1)
        else:
            fecha_inicio_periodo = min((clase.fecha for clase in clases), default=hoy)
        clases_esperadas = 0
        fecha_revision = fecha_inicio_periodo
        while fecha_revision <= hoy:
            for bloque in cargas_equivalentes:
                if bloque.dia_semana != fecha_revision.weekday():
                    continue
                estado_fecha = estado_fecha_academica(db, bloque.periodo_id, fecha_revision)
                if not estado_fecha["requiere_asistencia"]:
                    continue
                fin_bloque = datetime.datetime.combine(
                    fecha_revision, datetime.time.fromisoformat(bloque.hora_fin), tzinfo=MX,
                )
                if fin_bloque <= ahora:
                    clases_esperadas += 1
            fecha_revision += datetime.timedelta(days=1)
        acuerdos_activos = db.query(SeguimientoAlumnoDocente).filter(
            SeguimientoAlumnoDocente.carga_docente_id.in_(cargas_equivalentes_ids),
            SeguimientoAlumnoDocente.docente_id == current_user.id,
            SeguimientoAlumnoDocente.tipo == "ACUERDO",
            SeguimientoAlumnoDocente.estado.in_(["PENDIENTE", "REPROGRAMADO"]),
        ).all()
        pendientes_carga = len(acuerdos_activos)
        vencidos_carga = sum(1 for acuerdo in acuerdos_activos if acuerdo.fecha_revision and acuerdo.fecha_revision < hoy)
        acuerdos_pendientes += pendientes_carga
        acuerdos_vencidos += vencidos_carga
        grupos.append({
            "carga_id": carga.id,
            "materia": carga.actividad_nombre,
            "grupo": seguimiento["carga"]["grupo"],
            "carrera": seguimiento["carga"]["carrera"],
            "total_alumnos": seguimiento["total_alumnos"],
            "total_clases": seguimiento["total_clases"],
            "clases_esperadas": clases_esperadas,
            "asistencia_promedio": seguimiento["promedio_asistencia"],
            "alumnos_alerta": seguimiento["alumnos_en_alerta"],
            "acuerdos_pendientes": pendientes_carga,
            "acuerdos_vencidos": vencidos_carga,
            "ultima_clase": max(
                (clase.fecha for clase in clases if clase.carga_docente_id in cargas_equivalentes_ids),
                default=None,
            ),
        })
        for alumno in seguimiento["alumnos"]:
            if not alumno["alertas"]:
                continue
            candidato = {
                "alumno_id": alumno["alumno_id"],
                "carga_id": carga.id,
                "nombre": alumno["nombre"],
                "matricula": alumno["matricula"],
                "materia": carga.actividad_nombre,
                "grupo": seguimiento["carga"]["grupo"],
                "asistencia": alumno["porcentaje_asistencia"],
                "faltas": alumno["falta"],
                "faltas_consecutivas": alumno["faltas_consecutivas"],
                "motivos": [alerta["mensaje"] for alerta in alumno["alertas"]],
                "prioridad": (
                    "ALTA" if any(alerta["nivel"] == "ALTO" for alerta in alumno["alertas"])
                    else "MEDIA"
                ),
            }
            existente = alumnos_prioritarios_por_id.get(alumno["alumno_id"])
            if not existente:
                candidato["materias"] = [carga.actividad_nombre]
                alumnos_prioritarios_por_id[alumno["alumno_id"]] = candidato
            else:
                if carga.actividad_nombre not in existente["materias"]:
                    existente["materias"].append(carga.actividad_nombre)
                existente["motivos"] = list(dict.fromkeys(existente["motivos"] + candidato["motivos"]))
                existente["faltas"] += candidato["faltas"]
                existente["faltas_consecutivas"] = max(existente["faltas_consecutivas"], candidato["faltas_consecutivas"])
                existente["asistencia"] = min(existente["asistencia"], candidato["asistencia"])
                if candidato["prioridad"] == "ALTA":
                    existente["prioridad"] = "ALTA"
    alumnos_prioritarios = list(alumnos_prioritarios_por_id.values())
    alumnos_prioritarios.sort(key=lambda item: (
        0 if item["prioridad"] == "ALTA" else 1,
        item["asistencia"],
        item["nombre"],
    ))
    # La tarjeta del panel debe poder abrir la captura concreta. Incluye clases
    # que quedaron abiertas durante la ventana de recaptura, aunque no sean de hoy.
    pendientes_asistencia_detalle = []
    for clase in sorted(clases, key=lambda item: (item.fecha, item.id)):
        if clase.estado not in {"ABIERTA", "CORRECCION"}:
            continue
        if clase.fecha < hoy - PLAZO_CAPTURA_EXTEMPORANEA:
            continue
        carga = next((item for item in cargas_clase if item.id == clase.carga_docente_id), None)
        if not carga:
            continue
        pendientes_asistencia_detalle.append({
            "clase_id": clase.id,
            "carga_id": carga.id,
            "fecha": clase.fecha.isoformat(),
            "materia": carga.actividad_nombre,
            "grupo": (
                f"{carga.grupo_academico.cuatrimestre}° {carga.grupo_academico.grupo}"
                if carga.grupo_academico else "Sin grupo"
            ),
            "estado": "CORRECCION" if clase.estado == "CORRECCION" else "ABIERTA",
            "accion": "Continuar corrección" if clase.estado == "CORRECCION" else "Continuar asistencia",
        })
    clases_pendientes_ids = {item["clase_id"] for item in pendientes_asistencia_detalle}
    for item in jornada:
        if item["tipo_actividad"] != "CLASE" or item["estado"] not in {"EN_CURSO", "CORRECCION", "SIN_REGISTRO"}:
            continue
        if item["clase_id"] in clases_pendientes_ids:
            continue
        pendientes_asistencia_detalle.append({
            "clase_id": item["clase_id"],
            "carga_id": item["carga_id"],
            "fecha": hoy.isoformat(),
            "materia": item["materia"],
            "grupo": item["grupo"],
            "estado": item["estado"],
            "accion": "Registrar asistencia" if item["estado"] == "SIN_REGISTRO" else "Continuar asistencia",
        })
    pendientes_asistencia_detalle.sort(key=lambda item: (item["fecha"], item["clase_id"] or 0))
    pendientes_asistencia = len(pendientes_asistencia_detalle)
    actividades_suspendidas_hoy = sum(1 for item in jornada if item["estado"] == "NO_LECTIVA")
    clases_exigibles_hoy = sum(1 for item in jornada if item["tipo_actividad"] == "CLASE" and item["estado"] != "NO_LECTIVA")

    # Próxima clase realmente lectiva. El horario semanal propone candidatos,
    # pero el calendario oficial decide si esa fecha exige asistencia.
    proxima_clase = None
    for desplazamiento in range(0, 184):
        fecha_candidata = hoy + datetime.timedelta(days=desplazamiento)
        candidatas = []
        for carga in cargas:
            if carga.tipo_actividad not in {"CLASE", "TUTORIA"}:
                continue
            if carga.dia_semana != fecha_candidata.weekday():
                continue
            estado_candidata = estado_fecha_academica(db, carga.periodo_id, fecha_candidata)
            if not estado_candidata["requiere_asistencia"] or not estado_candidata["permite_iniciar_clase"]:
                continue
            inicio_candidata = datetime.datetime.combine(
                fecha_candidata, datetime.time.fromisoformat(carga.hora_inicio), tzinfo=MX,
            )
            if inicio_candidata <= ahora:
                continue
            candidatas.append((inicio_candidata, carga))
        if candidatas:
            inicio_candidata, carga_candidata = min(candidatas, key=lambda item: item[0])
            proxima_clase = {
                "carga_id": carga_candidata.id,
                "tipo_actividad": carga_candidata.tipo_actividad,
                "grupo_tutorado_id": carga_candidata.grupo_tutorado_id,
                "fecha": inicio_candidata.date().isoformat(),
                "inicio": inicio_candidata.isoformat(),
                "materia": carga_candidata.actividad_nombre,
                "grupo": (
                    f"{carga_candidata.grupo_academico.cuatrimestre}° {carga_candidata.grupo_academico.grupo}"
                    if carga_candidata.grupo_academico else
                    f"{carga_candidata.grupo_tutorado.cuatrimestre}° {carga_candidata.grupo_tutorado.grupo}"
                    if carga_candidata.grupo_tutorado else "Sin grupo"
                ),
                "laboratorio_nombre": (
                    carga_candidata.laboratorio.nombre if carga_candidata.laboratorio
                    else carga_candidata.espacio_nombre or "Sin espacio"
                ),
                "hora_inicio": carga_candidata.hora_inicio,
                "hora_fin": carga_candidata.hora_fin,
            }
            break

    inicio_semana = hoy - datetime.timedelta(days=hoy.weekday())
    clases_semana_lectivas = 0
    for desplazamiento in range(7):
        fecha_semana = inicio_semana + datetime.timedelta(days=desplazamiento)
        for carga in cargas:
            if carga.tipo_actividad != "CLASE":
                continue
            if carga.dia_semana != fecha_semana.weekday():
                continue
            estado_semana = estado_fecha_academica(db, carga.periodo_id, fecha_semana)
            if estado_semana["requiere_asistencia"]:
                clases_semana_lectivas += 1

    calendario_hoy = next(
        (item["calendario"] for item in jornada if item["estado"] == "NO_LECTIVA"), None,
    )
    return {
        "fecha": hoy.isoformat(),
        "periodo": {"id": actual.id, "clave": actual.clave},
        "resumen": {
            "clases_hoy": clases_exigibles_hoy,
            "clases_cerradas": sum(1 for item in jornada if item["estado"] == "CERRADA"),
            "asistencias_pendientes": pendientes_asistencia,
            "actividades_suspendidas_hoy": actividades_suspendidas_hoy,
            "clases_semana_lectivas": clases_semana_lectivas,
            "grupos_activos": len({carga.grupo_academico_id for carga in cargas_seguimiento}),
            "alumnos_atencion": len(alumnos_prioritarios),
            "acuerdos_pendientes": acuerdos_pendientes,
            "acuerdos_vencidos": acuerdos_vencidos,
        },
        "jornada": jornada,
        "asistencias_pendientes": pendientes_asistencia_detalle,
        "calendario_hoy": calendario_hoy,
        "proxima_clase": proxima_clase,
        "grupos": grupos,
        "alumnos_prioritarios": alumnos_prioritarios[:8],
    }
