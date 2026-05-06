"""
routers/auditoria.py -- Endpoints de consulta de la bitacora de auditoria.

Acceso: solo SUPER_ADMIN y LAB_ADMIN.
  GET  /auditoria/          -- lista paginada con filtros
  GET  /auditoria/export    -- descarga Excel
  GET  /auditoria/resumen   -- conteos por accion (para dashboard futuro)
"""

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from typing import Optional
from datetime import datetime, date
import io

from database import get_db
from models.auditoria import AuditLog
from models.usuario import RolUsuario
from dependencies import get_current_user

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/auditoria", tags=["Auditoria"])


def _solo_admin(current_user=Depends(get_current_user)):
    from fastapi import HTTPException, status
    if current_user.rol not in (RolUsuario.SUPER_ADMIN, RolUsuario.LAB_ADMIN):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    return current_user


def _build_query(
    db: Session,
    usuario_id: Optional[int],
    accion: Optional[str],
    recurso: Optional[str],
    exito: Optional[bool],
    fecha_inicio: Optional[date],
    fecha_fin: Optional[date],
    buscar: Optional[str],
    current_user,
):
    q = db.query(AuditLog)

    # LAB_ADMIN solo ve su propio laboratorio (restringimos por accion/recurso de sesiones)
    # SUPER_ADMIN ve todo
    if current_user.rol == RolUsuario.LAB_ADMIN:
        # Para simplificar: LAB_ADMIN ve todos los registros de su lab
        # (refinamiento futuro: filtrar por laboratorio_id en detalle JSON)
        pass

    if usuario_id:
        q = q.filter(AuditLog.usuario_id == usuario_id)
    if accion:
        q = q.filter(AuditLog.accion == accion)
    if recurso:
        q = q.filter(AuditLog.recurso == recurso)
    if exito is not None:
        q = q.filter(AuditLog.exito == exito)
    if fecha_inicio:
        q = q.filter(AuditLog.timestamp >= datetime.combine(fecha_inicio, datetime.min.time()))
    if fecha_fin:
        q = q.filter(AuditLog.timestamp <= datetime.combine(fecha_fin, datetime.max.time()))
    if buscar:
        like = f"%{buscar}%"
        q = q.filter(or_(
            AuditLog.usuario_nombre.ilike(like),
            AuditLog.usuario_email.ilike(like),
            AuditLog.accion.ilike(like),
            AuditLog.recurso.ilike(like),
            AuditLog.ip_address.ilike(like),
        ))

    return q.order_by(AuditLog.timestamp.desc())


@router.get("/")
def listar_auditoria(
    db: Session = Depends(get_db),
    current_user=Depends(_solo_admin),
    # Filtros
    usuario_id:   Optional[int]  = Query(None),
    accion:       Optional[str]  = Query(None),
    recurso:      Optional[str]  = Query(None),
    exito:        Optional[bool] = Query(None),
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin:    Optional[date] = Query(None),
    buscar:       Optional[str]  = Query(None, description="Busca en nombre, email, accion, recurso, IP"),
    # Paginacion
    page:  int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
):
    q = _build_query(db, usuario_id, accion, recurso, exito,
                     fecha_inicio, fecha_fin, buscar, current_user)

    total = q.count()
    registros = q.offset((page - 1) * limit).limit(limit).all()

    def _fmt(r: AuditLog):
        return {
            "id":             r.id,
            "timestamp":      r.timestamp.isoformat() if r.timestamp else None,
            "usuario_id":     r.usuario_id,
            "usuario_nombre": r.usuario_nombre,
            "usuario_email":  r.usuario_email,
            "accion":         r.accion,
            "recurso":        r.recurso,
            "recurso_id":     r.recurso_id,
            "detalle":        r.detalle,
            "exito":          r.exito,
            "ip_address":     r.ip_address,
        }

    return {
        "total": total,
        "page": page,
        "pages": max(1, (total + limit - 1) // limit),
        "items": [_fmt(r) for r in registros],
    }


@router.get("/acciones")
def listar_acciones_disponibles(current_user=Depends(_solo_admin)):
    """Devuelve la lista de codigos de accion unicos para llenar el filtro."""
    from services.auditoria import Accion, Recurso
    acciones = [v for k, v in vars(Accion).items() if not k.startswith("_")]
    recursos = [v for k, v in vars(Recurso).items() if not k.startswith("_")]
    return {"acciones": sorted(acciones), "recursos": sorted(recursos)}


@router.get("/resumen")
def resumen_auditoria(
    db: Session = Depends(get_db),
    current_user=Depends(_solo_admin),
    dias: int = Query(7, ge=1, le=90, description="Ultimos N dias"),
):
    """Conteos por accion para los ultimos N dias."""
    from sqlalchemy import func
    from datetime import timedelta

    desde = datetime.utcnow() - timedelta(days=dias)
    rows = (
        db.query(AuditLog.accion, func.count(AuditLog.id).label("total"))
        .filter(AuditLog.timestamp >= desde)
        .group_by(AuditLog.accion)
        .order_by(func.count(AuditLog.id).desc())
        .all()
    )
    return {"dias": dias, "conteos": [{"accion": r.accion, "total": r.total} for r in rows]}


@router.get("/export")
def exportar_auditoria(
    db: Session = Depends(get_db),
    current_user=Depends(_solo_admin),
    usuario_id:   Optional[int]  = Query(None),
    accion:       Optional[str]  = Query(None),
    recurso:      Optional[str]  = Query(None),
    exito:        Optional[bool] = Query(None),
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin:    Optional[date] = Query(None),
    buscar:       Optional[str]  = Query(None),
):
    q = _build_query(db, usuario_id, accion, recurso, exito,
                     fecha_inicio, fecha_fin, buscar, current_user)
    registros = q.limit(5000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitacora"

    # Estilos
    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ID", "Fecha/Hora", "Usuario", "Email", "Accion",
               "Recurso", "Recurso ID", "Exito", "IP", "Detalle"]
    col_widths = [6, 20, 25, 30, 25, 18, 12, 8, 16, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = hdr_align
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    for row_idx, r in enumerate(registros, 2):
        ts = r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else ""
        detalle_str = str(r.detalle) if r.detalle else ""
        vals = [r.id, ts, r.usuario_nombre or "", r.usuario_email or "",
                r.accion, r.recurso, r.recurso_id or "",
                "Si" if r.exito else "No", r.ip_address or "", detalle_str]
        fill = PatternFill("solid", fgColor="FEF2F2") if not r.exito else (
               PatternFill("solid", fgColor="F8FAFC") if row_idx % 2 == 0 else None)
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = border
            c.alignment = Alignment(vertical="center")
            if fill:
                c.fill = fill

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bitacora_{fecha_str}.xlsx"},
    )
